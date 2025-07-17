import sys
from pathlib import Path
import tkinter as tk
import customtkinter as ctk
from tkinter import filedialog, messagebox, scrolledtext
import subprocess
import os
import logging
import threading
import multiprocessing
import pandas as pd
from typing import Optional
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import json
import time
import re
import importlib
import importlib.util
import warnings
import xlwings as xw
import shutil
import psutil

# Add project root to Python path for imports - must be done before local imports
project_root = Path(__file__).parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

# Import main modules (after path setup) - imports must be here, not at top
# pylint: disable=wrong-import-position
# flake8: noqa: E402
# type: ignore
from config.app_config import ProcessingConfig, AppConstants  # noqa: E402
from modules.file_processor import FileProcessor  # noqa: E402
from modules.template_processor import TemplateProcessor  # noqa: E402
from modules.data_processor import DataProcessor  # noqa: E402
from modules.process_manager import ProcessManager  # noqa: E402
from modules.ui_builder import UIBuilder  # noqa: E402
from modules.log_manager import LogManager, ThemeController  # noqa: E402
from utils.utils import write_shared_log  # noqa: E402
from modules.audit_helper import (  # noqa: E402
    log_file_access,
    log_process_action,
    log_system_error,
    log_file_error,
)


# Excel COM check
XLWINGS_AVAILABLE = importlib.util.find_spec("xlwings") is not None
EXCEL_COM_AVAILABLE = importlib.util.find_spec("win32com.client") is not None

# Logging setup
logging.basicConfig(
    filename="repricing_log.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


class ConfigManager:
    def __init__(self):
        self.config = {}
        if AppConstants.CONFIG_FILE.exists():
            self.load()
        else:
            self.save_default()

    def save_default(self):
        self.config = {"last_folder": str(Path.cwd())}
        self.save()

    def load(self):
        with open(AppConstants.CONFIG_FILE, "r") as f:
            self.config = json.load(f)

    def save(self):
        with open(AppConstants.CONFIG_FILE, "w") as f:
            json.dump(self.config, f, indent=4)


class App:
    def __init__(self, root):
        self.root = root
        self._initialize_variables()
        self._initialize_processors()
        self.ui_builder.build_complete_ui()
        self.theme_controller.apply_initial_theme()
        self.log_manager.initialize_logging()
        self.log_manager.log_application_start()

        # Set up proper window close handler for audit logging
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def on_closing(self):
        """Handle application closing with proper audit logging."""
        try:
            self.log_manager.log_application_shutdown()
        except Exception as e:
            logging.error(f"Failed to log shutdown: {e}")
        finally:
            self.root.destroy()

    def _initialize_variables(self):
        """Initialize all instance variables."""
        self.file1_path = None
        self.file2_path = None
        self.template_file_path = None
        self.cancel_event = threading.Event()
        self.start_time = None
        self.progress_var = tk.DoubleVar(value=0.0)
        self.progress_label_var = tk.StringVar(value="0%")
        self.config_manager = ConfigManager()
        self.selected_disruption_type = tk.StringVar(value="Tier")
        self.file1_label: Optional[ctk.CTkLabel] = None
        self.file2_label: Optional[ctk.CTkLabel] = None
        self.template_label: Optional[ctk.CTkLabel] = None
        self.toggle_theme_button: Optional[ctk.CTkButton] = None
        self.progress_bar: Optional[ctk.CTkProgressBar] = None
        self.progress_label: Optional[ctk.CTkLabel] = None
        self.processed_claim_data = None  # Store processed data for CSV generation

    def _initialize_processors(self):
        self.file_processor = FileProcessor(self)
        self.template_processor = TemplateProcessor(self)
        self.data_processor = DataProcessor(self)
        self.process_manager = ProcessManager(self)
        self.ui_builder = UIBuilder(self)
        self.log_manager = LogManager(self)
        self.theme_controller = ThemeController(self)
        # No explicit type annotation, so fallback and imported classes are interchangeable

    # The following methods are moved to their respective manager classes for better cohesion:
    # - apply_theme_colors -> ThemeController
    # - check_template -> FileProcessor
    # - sharx_lbl, epls_lbl -> ProcessManager
    # - show_shared_log_viewer -> LogManager

    # Example: Remove apply_theme_colors from App, and use self.theme_controller.apply_theme_colors instead.

    def import_file1(self):
        """Import the first file with template validation using guard clauses."""
        file_path = None
        try:
            file_path = self._get_file_path("Select File Uploaded to Tool")
            if not file_path:
                return  # User cancelled

            self._set_file1_path(file_path)
            self._validate_gross_cost_template(file_path)

        except Exception as e:
            error_msg = f"Failed to import File1: {str(e)}"
            log_file_error("File1Import", file_path or "Unknown", error_msg)
            messagebox.showerror(
                "File Import Error", f"Could not import File1:\n{error_msg}"
            )
            logging.error(f"File1 import failed: {e}")

    def _get_file_path(self, title):
        """Get file path from file dialog."""
        return filedialog.askopenfilename(
            title=title,
            filetypes=ProcessingConfig.FILE_TYPES,
        )

    def _set_file1_path(self, file_path):
        """Set the file1 path and update UI."""
        self.file1_path = file_path
        if self.file1_label:
            self.file1_label.configure(text=os.path.basename(file_path))
        self.file_processor.check_template(file_path)
        log_file_access("File1Import", file_path, "IMPORTED")
        write_shared_log("File1 imported", file_path)

    def _validate_gross_cost_template(self, file_path):
        """Validate GrossCost column and suggest template type using data processor."""
        template_suggestion = self.data_processor.validate_gross_cost_template(
            file_path
        )

        # Always show the template recommendation to the user
        if template_suggestion:
            messagebox.showinfo("Template Selection Guide", template_suggestion)
        else:
            # Fallback message if validation fails
            messagebox.showinfo(
                "Template Selection Guide",
                "File imported successfully!\n\n"
                "Template Selection:\n"
                "• Use BLANK template if your data has no cost information\n"
                "• Use STANDARD template if your data contains cost values\n\n"
                "Check your GrossCost column to determine which template to use.",
            )

    def import_file2(self):
        """Import the second file."""
        file_path = None
        try:
            file_path = self._get_file_path("Select File From Tool")
            if not file_path:
                return  # User cancelled

            self.file2_path = file_path
            if self.file2_label:
                self.file2_label.configure(text=os.path.basename(file_path))
            log_file_access("File2Import", file_path, "IMPORTED")
            write_shared_log("File2 imported", file_path)

        except Exception as e:
            error_msg = f"Failed to import File2: {str(e)}"
            log_file_error("File2Import", file_path or "Unknown", error_msg)
            messagebox.showerror(
                "File Import Error", f"Could not import File2:\n{error_msg}"
            )
            logging.error(f"File2 import failed: {e}")

    def import_template_file(self):
        """Import the template file."""
        file_path = None
        try:
            file_path = filedialog.askopenfilename(
                title="Select Template File",
                filetypes=ProcessingConfig.TEMPLATE_FILE_TYPES,
            )
            if not file_path:
                return  # User cancelled

            self.template_file_path = file_path
            if self.template_label:
                self.template_label.configure(text=os.path.basename(file_path))
            log_file_access("TemplateImport", file_path, "IMPORTED")
            write_shared_log("Template file imported", file_path)

        except Exception as e:
            error_msg = f"Failed to import template file: {str(e)}"
            log_file_error("TemplateImport", file_path or "Unknown", error_msg)
            messagebox.showerror(
                "Template Import Error", f"Could not import template file:\n{error_msg}"
            )
            logging.error(f"Template import failed: {e}")

    # Logging and notification methods
    # Removed duplicate write_audit_log method to resolve method name conflict.

    # Cancel during repricing
    def cancel_process(self):
        """Cancel the process using process manager."""
        self.process_manager.cancel_process()

    # Live log viewer (old version, renamed to avoid conflict)
    def show_log_viewer_old(self):
        win = tk.Toplevel(self.root)
        win.title("Log Viewer")
        txt = scrolledtext.ScrolledText(win, width=100, height=30)
        txt.pack(fill="both", expand=True)

        def refresh():
            with open("repricing_log.log", "r") as f:
                txt.delete("1.0", tk.END)
                txt.insert(tk.END, f.read())
            win.after(3000, refresh)

        refresh()

    def update_progress(self, value=None, message=None):
        """Update the progress bar and label with reduced complexity."""

        def do_update():
            if value is None:
                self._set_indeterminate_progress(message)
            else:
                self._set_determinate_progress(value, message)
            self.root.update_idletasks()

        self._schedule_ui_update(do_update)

    def _set_indeterminate_progress(self, message):
        """Set progress bar to indeterminate mode."""
        if self.progress_bar:
            self.progress_bar.configure(mode="indeterminate")
            self.progress_bar.start()
        display_message = message or "Processing... (unknown duration)"
        self.progress_label_var.set(display_message)

    def _set_determinate_progress(self, value, message):
        """Set progress bar to determinate mode with specific value."""
        if self.progress_bar:
            if self.progress_bar.cget("mode") != "determinate":
                self.progress_bar.stop()
                self.progress_bar.configure(mode="determinate")

            self.progress_bar.set(value)

        self.progress_var.set(value)

        if message:
            self.progress_label_var.set(message)
        else:
            self._set_calculated_progress_message(value)

    def _set_calculated_progress_message(self, value):
        """Calculate and set progress message with time estimates."""
        percent = int(value * 100)
        elapsed = time.time() - self.start_time if self.start_time else 0
        est = int((elapsed / value) * (1 - value)) if value > 0 else 0
        self.progress_label_var.set(f"Progress: {percent}% | Est. {est}s left")

    def _schedule_ui_update(self, update_func):
        """Schedule UI update on main thread or execute immediately."""
        if threading.current_thread() is threading.main_thread():
            update_func()
        else:
            self.root.after(0, update_func)

    def show_log_viewer(self):
        """Show log viewer using log manager."""
        self.log_manager.show_log_viewer()

    def show_shared_log_viewer(self):
        """Show shared log viewer using log manager."""
        self.log_manager.show_shared_log_viewer()

    def sharx_lbl(self):
        """Generate SHARx LBL (method not implemented in ProcessManager)."""
        messagebox.showerror(
            "Not Implemented", "SHARx LBL functionality is not available."
        )

    def epls_lbl(self):
        """Generate EPLS LBL (method not implemented in ProcessManager)."""
        messagebox.showerror(
            "Not Implemented", "EPLS LBL functionality is not available."
        )

    # Disruption and process methods
    # Removed select_disruption_type method since disruption_type_combobox does not exist.

    def start_disruption(self, disruption_type=None):
        """Start disruption processing using process manager."""
        self.process_manager.start_disruption(disruption_type)

    def start_process_threaded(self):
        """Start the repricing process using the process manager."""
        self.process_manager.start_process_threaded()

    def finish_notification(self):
        """Show completion notification using process manager."""
        self.process_manager.finish_notification()

    # Repricing workflow methods
    def paste_into_template(self, processed_file):
        """Paste processed data into Excel template using background threading."""

        def run_in_background():
            try:
                self._execute_template_paste(processed_file)
            except Exception as e:
                logger.exception("Error during paste with xlwings")
                self.root.after(
                    0,
                    lambda e=e: messagebox.showerror(
                        "Error", f"Template update failed:\n{e}"
                    ),
                )
                self.root.after(0, lambda: self.update_progress(0))

        threading.Thread(target=run_in_background, daemon=True).start()

    def _check_excel_availability(self):
        """Check if Excel is available and functioning properly."""
        try:
            # Try to create a temporary Excel application
            app = xw.App(visible=False, add_book=False)
            app.quit()
            return True, "Excel COM interface is available"
        except Exception as e:
            return False, f"Excel COM interface unavailable: {e}"

    def _execute_template_paste(self, processed_file):
        """Execute the template paste operation with proper error handling."""
        import time

        start_time = time.time()

        # Check Excel availability first
        excel_available, excel_message = self._check_excel_availability()
        logger.info(f"Excel check: {excel_message}")

        # Initialize progress
        self.root.after(
            0,
            lambda: self.update_progress(0.72, "Preparing to paste into template..."),
        )

        # Validate template path
        if not self.template_file_path:
            raise ValueError("Template file path is not set.")

        # Prepare data and paths
        self.root.after(
            0,
            lambda: self.update_progress(0.75, "Preparing template data..."),
        )
        paste_data = self._prepare_template_data(processed_file)
        paths = self._prepare_template_paths()

        # Create backup and setup output file
        self.root.after(
            0,
            lambda: self.update_progress(0.80, "Creating template backup..."),
        )
        self._create_template_backup(paths)

        # Execute Excel operations
        self.root.after(
            0,
            lambda: self.update_progress(0.85, "Opening template in Excel..."),
        )
        self._execute_excel_paste(paste_data, paths)

        # Finalize and notify
        elapsed = time.time() - start_time
        msg = f"Template updated successfully in {elapsed:.2f} seconds."
        logger.info(msg)
        self.root.after(0, lambda: self.update_progress(0.95, msg))

        # Generate Claim Detail CSV with Logic from completed template
        self.root.after(
            0, lambda: self._generate_claim_detail_csv_with_template_logic()
        )

        self.root.after(
            0,
            lambda: self.update_progress(1.0, "Process complete with Claim Detail CSV"),
        )
        self.root.after(0, lambda: self.show_toast("Process complete with updated CSV"))
        self.root.after(
            0,
            lambda: messagebox.showinfo(
                "Template Update Complete",
                "Pasting into the template is complete. Claim Detail CSV has been generated with Logic from the workflow. You may now review the updated files.",
            ),
        )

    def _prepare_template_data(self, processed_file):
        """Prepare data for template pasting with enhanced data cleaning."""
        df = pd.read_excel(processed_file)
        df = self.format_dataframe(df)

        # Enhanced data cleaning to prevent Excel errors
        df = self._clean_data_for_excel(df)

        # Only return the data values (not headers) since template already has headers
        # Ensure we only paste the first 39 columns (A:AM) to match template structure
        data_values = df.iloc[
            :, :39
        ].values  # Get data without headers, limit to 39 columns

        return {
            "data": data_values,
            "nrows": df.shape[0],
            "ncols": min(df.shape[1], 39),
        }

    def _clean_data_for_excel(self, df):
        """Clean data to prevent Excel errors during paste operations."""
        df_clean = df.copy()

        try:
            # Replace problematic values that can cause Excel errors
            for col in df_clean.columns:
                if df_clean[col].dtype == "object":
                    # Replace None/NaN with empty string
                    df_clean[col] = df_clean[col].fillna("")
                    # Convert to string and handle any remaining issues
                    df_clean[col] = df_clean[col].astype(str)
                    # Replace 'None' string with empty string
                    df_clean[col] = df_clean[col].replace("None", "")
                    # Truncate very long strings that might cause Excel issues
                    df_clean[col] = df_clean[col].apply(
                        lambda x: x[:32767] if len(str(x)) > 32767 else x
                    )
                elif pd.api.types.is_numeric_dtype(df_clean[col]):
                    # Handle infinite values
                    df_clean[col] = df_clean[col].replace([np.inf, -np.inf], np.nan)
                    # Fill NaN with 0 for numeric columns
                    df_clean[col] = df_clean[col].fillna(0)

            logger.info("Data cleaning for Excel completed successfully")

        except Exception as e:
            logger.warning(f"Error during data cleaning: {e}. Using original data.")
            return df

        return df_clean

    def _prepare_template_paths(self):
        """Prepare file paths for template operations using file processor."""
        # Pass None for opportunity_name to use default _Rx Repricing_wf.xlsx naming
        return self.file_processor.prepare_file_paths(self.template_file_path, None)

    def _create_template_backup(self, paths):
        """Create backup of template and prepare output file using template processor."""
        self.template_processor.create_template_backup(paths)

    def _execute_excel_paste(self, paste_data, paths):
        """Execute the Excel paste operation with improved error handling."""
        try:
            self._try_xlwings_paste(paste_data, paths)
        except Exception as xlwings_error:
            logger.warning(f"xlwings failed: {xlwings_error}")
            try:
                self._try_openpyxl_paste(paste_data, paths)
            except Exception as openpyxl_error:
                logger.error("Both xlwings and openpyxl failed")
                raise Exception(
                    f"Template update failed with both methods. xlwings: {xlwings_error}, openpyxl: {openpyxl_error}"
                )

    def _try_xlwings_paste(self, paste_data, paths):
        """Try pasting using xlwings with better error handling."""

        # Kill any existing Excel processes first
        try:
            subprocess.run(
                ["taskkill", "/F", "/IM", "excel.exe"],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                timeout=10,
            )

            time.sleep(2)  # Wait for processes to fully terminate
        except Exception as e:
            logger.warning(f"Could not kill Excel processes: {e}")

        # Start Excel session with timeout and retry logic
        self.root.after(
            0,
            lambda: self.update_progress(0.87, "Starting Excel session..."),
        )

        app = None
        wb = None
        try:
            app = xw.App(visible=False, add_book=False)
            app.api.DisplayAlerts = False
            wb = app.books.open(str(paths["output"]))
            ws = wb.sheets["Claims Table"]

            # Batch read formulas and prepare data
            self.root.after(
                0,
                lambda: self.update_progress(0.90, "Reading template formulas..."),
            )
            formulas = ws.range(
                (2, 1), (paste_data["nrows"] + 1, paste_data["ncols"])
            ).formula
            data_to_write = self._prepare_excel_data(paste_data, formulas)

            # Paste values with progress updates
            self.root.after(
                0,
                lambda: self.update_progress(
                    0.95, f"Pasting {paste_data['nrows']} rows of data..."
                ),
            )
            self._paste_data_with_progress(
                ws, data_to_write, paste_data["nrows"], paste_data["ncols"]
            )

            # Save and close
            self.root.after(
                0,
                lambda: self.update_progress(0.98, "Saving template file..."),
            )
            wb.save()

        finally:
            # Ensure Excel is properly closed
            if wb:
                try:
                    wb.close()
                except Exception:
                    pass
            if app:
                try:
                    app.quit()
                except Exception:
                    pass

    def _try_openpyxl_paste(self, paste_data, paths):
        """Fallback method using openpyxl when xlwings fails."""
        from openpyxl import load_workbook
        import pandas as pd

        self.root.after(
            0,
            lambda: self.update_progress(0.87, "Using fallback method (openpyxl)..."),
        )

        # Load workbook with openpyxl
        wb = load_workbook(str(paths["output"]))
        ws = wb["Claims Table"]

        # Convert paste_data to DataFrame for easier handling
        df = pd.DataFrame(paste_data["data"])

        self.root.after(
            0,
            lambda: self.update_progress(
                0.95, f"Writing {len(df)} rows with openpyxl..."
            ),
        )

        # Write data starting from row 2 (assuming row 1 has headers)
        for row_idx, row_data in enumerate(df.values, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        self.root.after(
            0,
            lambda: self.update_progress(0.98, "Saving with openpyxl..."),
        )

        # Save workbook
        wb.save(str(paths["output"]))
        wb.close()

    def _prepare_excel_data(self, paste_data, formulas):
        """Prepare data for Excel, preserving formulas."""
        data_to_write = []

        for i in range(paste_data["nrows"]):
            row = []
            for j in range(paste_data["ncols"]):
                if formulas[i][j] == "":
                    row.append(paste_data["data"][i][j])
                else:
                    row.append(None)
            data_to_write.append(row)

        return data_to_write

    def _paste_data_with_progress(self, ws, data_to_write, nrows, ncols):
        """Paste data to Excel with progress updates."""
        # Paste values - this happens as a batch operation
        ws.range((2, 1), (nrows + 1, ncols)).value = data_to_write

    def filter_template_columns(self, df):
        try:
            # Ensure 'ClientName' and 'Logic' columns exist and are in the correct order
            if "ClientName" in df.columns and "Logic" in df.columns:
                client_name_idx = df.columns.get_loc("ClientName")
                logic_idx = df.columns.get_loc("Logic")
                if client_name_idx <= logic_idx:
                    # Select columns from 'ClientName' to 'Logic' (inclusive)
                    selected_columns = df.columns[client_name_idx : logic_idx + 1]
                    logger.info(
                        f"Pasting only these columns: {selected_columns.tolist()}"
                    )
                    return df[selected_columns]
                else:
                    logger.warning(
                        "'Logic' column appears before 'ClientName'; returning full DataFrame."
                    )
                    return df
            else:
                raise ValueError(
                    "Required columns 'ClientName' or 'Logic' are missing."
                )
        except Exception as e:
            logger.warning(f"Error filtering columns: {e}. Using full DataFrame.")
            return df

    def format_dataframe(self, df):
        """Format DataFrame using template processor."""
        return self.template_processor.format_dataframe(df)

    def show_toast(self, message, duration=3000):
        """Show toast notification using template processor."""
        return self.template_processor.show_toast(message, duration)

    def perform_preflight_checks(self) -> tuple[bool, str]:
        """
        Perform pre-flight checks before starting any processing.
        Returns (is_ready, warning_message)
        """
        warnings = []
        errors = []

        # Check disk space
        try:
            stat = shutil.disk_usage(".")
            free_gb = stat.free / (1024 * 1024 * 1024)
            if free_gb < 1.0:
                errors.append(
                    f"Insufficient disk space: {free_gb:.1f}GB available (minimum 1GB required)"
                )
            elif free_gb < 2.0:
                warnings.append(
                    f"Low disk space: {free_gb:.1f}GB available (recommended 2GB+)"
                )
        except Exception:
            warnings.append("Could not check disk space")

        # Check for running Excel processes
        try:
            excel_processes = [
                p
                for p in psutil.process_iter(["name"])
                if "excel" in p.info["name"].lower()
            ]
            if excel_processes:
                warnings.append(
                    f"Found {len(excel_processes)} Excel process(es) running - this may cause file conflicts"
                )
        except ImportError:
            # psutil not available, skip this check
            pass
        except Exception:
            pass

        # Check for existing output files that might be locked
        output_patterns = ["*_Rx Repricing_wf.xlsx", "LBL for Disruption.xlsx"]
        locked_files = []
        for pattern in output_patterns:
            for file_path in Path(".").glob(pattern):
                try:
                    # Try to open the file in append mode to check if it's locked
                    with open(file_path, "r+b"):
                        pass
                except (PermissionError, IOError):
                    locked_files.append(str(file_path))

        if locked_files:
            warnings.append(
                f"Output files may be locked by Excel: {', '.join(locked_files)}"
            )

        # Check memory usage
        try:
            memory = psutil.virtual_memory()
            available_gb = memory.available / (1024 * 1024 * 1024)
            if available_gb < 2.0:
                warnings.append(
                    f"Low available memory: {available_gb:.1f}GB (recommended 2GB+)"
                )
        except Exception:
            pass

        # Check OneDrive sync status (if applicable)
        onedrive_path = os.path.expandvars("%OneDrive%")
        if (
            onedrive_path
            and onedrive_path != "%OneDrive%"
            and os.path.exists(onedrive_path)
        ):
            # Check if there are any sync conflicts or pending uploads
            sync_files = list(Path(onedrive_path).rglob("*- Copy.*"))
            if sync_files:
                warnings.append(f"Found {len(sync_files)} OneDrive sync conflict files")

        # Compile results
        if errors:
            return False, "Critical Issues Found:\n" + "\n".join(
                f"• {error}" for error in errors
            )

        if warnings:
            warning_msg = "Warnings (process can continue but may have issues):\n"
            warning_msg += "\n".join(f"• {warning}" for warning in warnings)
            warning_msg += "\n\nContinue anyway?"
            return True, warning_msg

        return True, ""

    def start_process(self):
        # Perform pre-flight checks
        is_ready, message = self.perform_preflight_checks()

        if not is_ready:
            messagebox.showerror("System Check Failed", message)
            return

        if message:  # Warnings present
            if not messagebox.askyesno("System Warnings", message):
                return

        threading.Thread(target=self._start_process_internal).start()
        log_process_action("RepricingProcess", "STARTED")
        write_shared_log("Repricing process started", "")

    def validate_merge_inputs(self):
        """Validate merge inputs using data processor."""
        is_valid, message = self.data_processor.validate_merge_inputs(
            self.file1_path, self.file2_path
        )
        if not is_valid:
            messagebox.showerror("Error", message)
        return is_valid

    def _start_process_internal(self):
        self.start_time = time.time()
        self.update_progress(0.05)

        # Extra safeguard: Remove any accidental LBL/disruption output during repricing
        os.environ["NO_LBL_OUTPUT"] = "1"

        if not self.file1_path or not self.file2_path:
            self.update_progress(0)
            messagebox.showerror("Error", "Please select both files before proceeding.")
            return

        if not self.validate_merge_inputs():
            self.update_progress(0)
            return

        try:
            self.update_progress(0.10)
            subprocess.run(
                ["taskkill", "/F", "/IM", "excel.exe"],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )
            self.update_progress(0.20)
            subprocess.run(
                ["python", "merge.py", self.file1_path, self.file2_path], check=True
            )
            self.update_progress(0.40)
            MERGED_FILE = "merged_file.xlsx"
            self.process_merged_file(MERGED_FILE)
            self.update_progress(0.70)
            # Progress will reach 100% when template pasting is complete
            # Ensure LBL scripts are NOT called here or in process_merged_file
        except subprocess.CalledProcessError as e:
            self.update_progress(0)
            error_message = f"Merge process failed: {str(e)}"
            log_system_error("MergeProcess", error_message)
            logger.exception("Failed to run merge.py")
            messagebox.showerror("Error", f"Failed to run merge.py: {e}")
        except Exception as e:
            self.update_progress(0)
            error_message = f"Unexpected error during merge: {str(e)}"
            log_system_error("MergeProcess", error_message)
            logger.exception("Unexpected error in merge process")
            messagebox.showerror("Error", f"Unexpected error during merge: {e}")

    def process_merged_file(self, file_path):
        """Process merged file with reduced complexity using helper methods."""
        try:
            # Initialize processing
            self._initialize_processing()

            # Load and validate data
            df = self._load_and_validate_data(file_path)

            # Process data using multiprocessing
            processed_df = self._process_data_multiprocessing(df)

            # Save outputs
            output_file = self._save_processed_outputs(processed_df)

            # Finalize processing
            self._finalize_processing(output_file)

        except Exception as e:
            error_message = f"Error processing merged file: {str(e)}"
            log_system_error("DataProcessing", error_message)
            logger.error(f"Error processing merged file: {e}")
            messagebox.showerror("Error", f"Processing failed: {e}")

    def _initialize_processing(self):
        """Initialize the processing environment."""
        self.update_progress(0.45)
        open("repricing_log.log", "w").close()
        logging.basicConfig(
            filename="repricing_log.log",
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
        logging.info("Starting merged file processing")

    def _load_and_validate_data(self, file_path):
        """Load and validate the merged file data with enhanced error handling."""
        try:
            df = pd.read_excel(file_path)
            logging.info(f"Loaded {len(df)} records from {file_path}")
            self.update_progress(0.50)

            # Validate that we have data
            if df.empty:
                raise ValueError(f"The file {file_path} is empty or contains no data")

            # Validate required columns using configuration with fallback
            try:
                ProcessingConfig.validate_required_columns(df)
            except Exception as config_error:
                logging.warning(f"Configuration validation failed: {config_error}")
                # Fallback validation for essential columns
                essential_columns = ["DATEFILLED", "SOURCERECORDID"]
                missing_essential = [
                    col for col in essential_columns if col not in df.columns
                ]
                if missing_essential:
                    raise ValueError(f"Missing essential columns: {missing_essential}")

            # Safe data preparation
            df = self._safe_prepare_dataframe(df)

            return df

        except Exception as e:
            error_msg = f"Error loading data from {file_path}: {str(e)}"
            logging.error(error_msg)
            raise Exception(f"Failed to load merged file: {error_msg}")

    def _safe_prepare_dataframe(self, df):
        """Safely prepare DataFrame for processing."""
        try:
            # Create a copy to avoid modifying the original
            df_processed = df.copy()

            # Remove existing RowID column if present
            if "RowID" in df_processed.columns:
                df_processed = df_processed.drop(columns=["RowID"])
                logging.info("Removed existing RowID column")

            # Add Logic column if missing
            if "Logic" not in df_processed.columns:
                df_processed["Logic"] = ""
                logging.info("Added missing Logic column")

            # Safe sorting
            sort_columns = ["DATEFILLED", "SOURCERECORDID"]
            available_sort_cols = [
                col for col in sort_columns if col in df_processed.columns
            ]

            if available_sort_cols:
                # Handle null values before sorting
                for col in available_sort_cols:
                    if df_processed[col].isnull().any():
                        if col == "DATEFILLED":
                            df_processed[col] = df_processed[col].fillna(
                                pd.Timestamp("1900-01-01")
                            )
                        else:
                            df_processed[col] = df_processed[col].fillna("UNKNOWN")
                        logging.warning(f"Filled null values in {col}")

                df_processed = df_processed.sort_values(
                    by=available_sort_cols, ascending=True
                )
                logging.info(f"Sorted by: {available_sort_cols}")

            # Safe RowID creation with multiple fallback methods
            try:
                df_processed["RowID"] = np.arange(len(df_processed))
                logging.info("Created RowID using np.arange")
            except Exception as e1:
                try:
                    df_processed["RowID"] = df_processed.index.values
                    logging.warning(f"np.arange failed ({e1}), used index instead")
                except Exception as e2:
                    df_processed["RowID"] = list(range(len(df_processed)))
                    logging.warning(
                        f"Both methods failed ({e1}, {e2}), used list comprehension"
                    )

            return df_processed

        except Exception as e:
            logging.error(f"Error in DataFrame preparation: {e}")
            # Minimal fallback
            if "Logic" not in df.columns:
                df["Logic"] = ""
            if "RowID" not in df.columns:
                df["RowID"] = df.index
            return df

    def _process_data_multiprocessing(self, df):
        """Process data using multiprocessing for improved performance."""
        self.update_progress(0.55)

        # Import and use multiprocessing helpers
        from modules import mp_helpers

        num_workers = ProcessingConfig.get_multiprocessing_workers()
        df_blocks = np.array_split(df, num_workers)
        out_queue = multiprocessing.Queue()
        processes = []

        # Start worker processes
        for block in df_blocks:
            p = multiprocessing.Process(
                target=mp_helpers.worker, args=(block, out_queue)
            )
            p.start()
            processes.append(p)

        # Collect results
        results = [out_queue.get() for _ in processes]
        for p in processes:
            p.join()

        processed_df = pd.concat(results)

        # Critical Fix: Recreate RowID after multiprocessing concat to prevent conflicts
        try:
            processed_df = processed_df.reset_index(drop=True)
            processed_df["RowID"] = np.arange(len(processed_df))
            logging.info(
                f"Recreated RowID after multiprocessing concat: {len(processed_df)} rows"
            )
        except Exception as e:
            logging.warning(f"Could not recreate RowID after concat: {e}")
            processed_df["RowID"] = processed_df.index.values

        self.update_progress(0.60)

        return processed_df

    def _save_processed_outputs(self, df):
        """Save processed data to various output formats with enhanced RowID error handling."""
        try:
            # Validate input DataFrame
            if df is None or df.empty:
                raise ValueError("Input DataFrame is None or empty")

            logging.info(f"Starting output processing with {len(df)} rows")

            # Ensure RowID integrity before processing
            if "RowID" not in df.columns:
                logging.warning("RowID missing in input DataFrame, recreating...")
                df = df.reset_index(drop=True)
                df["RowID"] = np.arange(len(df))

            # Sort and filter data with error handling
            try:
                df_sorted = pd.concat([df[df["Logic"] == ""], df[df["Logic"] == "OR"]])
                logging.info(f"Filtered and sorted data: {len(df_sorted)} rows")
            except Exception as e:
                logging.error(f"Error in data filtering: {e}")
                # Fallback: use all data
                df_sorted = df.copy()

            # Ensure RowID is still valid after concat
            if "RowID" in df_sorted.columns:
                # Check for any RowID issues and fix them
                rowid_issues = df_sorted["RowID"].isnull().sum()
                if rowid_issues > 0:
                    logging.warning(
                        f"Found {rowid_issues} null RowID values, fixing..."
                    )
                    df_sorted = df_sorted.reset_index(drop=True)
                    df_sorted["RowID"] = np.arange(len(df_sorted))

        except Exception as e:
            error_msg = (
                f"Critical error in _save_processed_outputs initialization: {str(e)}"
            )
            logging.error(error_msg)
            raise Exception(error_msg)

        # Ensure Logic column is positioned as column AM (39th column) and rename to "O's & R's Check"
        df_sorted = self._fix_column_positioning(df_sorted)

        # Prepare output directory and files
        output_dir = Path.cwd()

        # Get opportunity name and create custom filename
        opportunity_name = self._extract_opportunity_name()
        output_file = output_dir / f"{opportunity_name}_merged_file_with_OR.xlsx"

        # Create row mapping for highlighting with robust error handling
        try:
            # Ensure RowID exists and is accessible
            if "RowID" not in df_sorted.columns:
                logging.warning("RowID column missing in df_sorted, recreating...")
                df_sorted = df_sorted.reset_index(drop=True)
                df_sorted["RowID"] = np.arange(len(df_sorted))

            # Safely create row mapping
            row_mapping = {}
            for i, (_, row) in enumerate(df_sorted.iterrows()):
                try:
                    row_id = row["RowID"]
                    if pd.notna(row_id):  # Check for valid RowID
                        row_mapping[row_id] = i + 2
                except (KeyError, IndexError, TypeError) as e:
                    logging.warning(
                        f"Skipped row {i} in mapping due to RowID issue: {e}"
                    )
                    continue

            excel_rows_to_highlight = [
                row_mapping[rid] for rid in [] if rid in row_mapping
            ]  # Placeholder

            logging.info(f"Created row mapping with {len(row_mapping)} entries")

        except Exception as e:
            logging.error(f"Failed to create row mapping: {e}")
            row_mapping = {}
            excel_rows_to_highlight = []

        # Clean up data with safe RowID removal
        try:
            if "RowID" in df_sorted.columns:
                df_sorted = df_sorted.drop(columns=["RowID"])
                logging.info("Successfully removed RowID column before saving")
            else:
                logging.info("RowID column not present, skipping removal")
        except Exception as e:
            logging.warning(f"Could not remove RowID column: {e}")
            # Continue without removing RowID if there's an issue

        # Save to multiple formats with updated names
        self._save_to_parquet(df_sorted, output_dir, opportunity_name)
        self._save_to_excel(df_sorted, output_file)

        # Save unmatched reversals info
        self._save_unmatched_reversals(excel_rows_to_highlight, output_dir)

        # Store the processed data for later CSV generation after template completion
        self.processed_claim_data = df_sorted.copy()

        self.update_progress(0.65)
        return output_file

    def _save_to_parquet(self, df, output_dir, opportunity_name=None):
        """Save data to Parquet format for large DataFrames."""
        try:
            if opportunity_name:
                parquet_path = (
                    output_dir / f"{opportunity_name}_merged_file_with_OR.parquet"
                )
            else:
                parquet_path = output_dir / "merged_file_with_OR.parquet"
            df.drop_duplicates().to_parquet(parquet_path, index=False)
            logger.info(f"Saved intermediate Parquet file: {parquet_path}")
        except Exception as e:
            logger.warning(f"Could not save Parquet: {e}")

    def _save_to_excel(self, df, output_file):
        """Save data to Excel format."""
        df.drop_duplicates().to_excel(output_file, index=False)

    def _save_to_csv(self, df, output_dir):
        """Save data to CSV format with opportunity name."""
        opportunity_name = self._extract_opportunity_name()
        csv_path = output_dir / f"{opportunity_name} Claim Detail.csv"
        df.drop_duplicates().to_csv(csv_path, index=False)

    def _save_unmatched_reversals(self, excel_rows_to_highlight, output_dir):
        """Save unmatched reversals information."""
        unmatched_path = output_dir / "unmatched_reversals.txt"
        with open(unmatched_path, "w") as f:
            f.write(",".join(map(str, excel_rows_to_highlight)))

    def _extract_opportunity_name(self):
        """Extract opportunity name from file1_path."""
        opportunity_name = ProcessingConfig.DEFAULT_OPPORTUNITY_NAME
        try:
            if self.file1_path:
                if self.file1_path.lower().endswith(".xlsx"):
                    df_file1 = pd.read_excel(self.file1_path)
                else:
                    df_file1 = pd.read_csv(self.file1_path)

                if df_file1.shape[1] >= 2:
                    # Get the value from the first row, second column
                    raw_name = str(df_file1.iloc[0, 1])
                    # Clean for filename
                    opportunity_name = re.sub(r'[\\/*?:"<>|]', "_", raw_name)
        except Exception as e:
            logger.warning(f"Could not extract opportunity name from file1: {e}")

        return opportunity_name

    def _generate_claim_detail_csv_with_template_logic(self):
        """Generate Claim Detail CSV using Logic column from completed _Rx Repricing_wf template."""
        try:
            # Check if we have processed claim data stored
            if (
                not hasattr(self, "processed_claim_data")
                or self.processed_claim_data is None
            ):
                logger.warning("No processed claim data available for CSV generation")
                return

            # Read the completed template to get the Logic column
            # Use the updated template file that was created during the paste operation
            updated_template_path = Path.cwd() / "_Rx Repricing_wf.xlsx"

            if not updated_template_path.exists():
                logger.error(
                    f"Updated template file not found: {updated_template_path}"
                )
                return

            # Read the Logic column from the completed template
            template_df = pd.read_excel(
                updated_template_path, sheet_name="Claims Table"
            )
            if "Logic" not in template_df.columns:
                logger.error("Logic column not found in template")
                return

            # Prepare the claim detail data with updated Logic
            claim_detail_df = self.processed_claim_data.copy()

            # Merge Logic column from template based on common key(s)
            # Using SOURCERECORDID as the primary key for matching
            if (
                "SOURCERECORDID" in template_df.columns
                and "SOURCERECORDID" in claim_detail_df.columns
            ):
                # Create mapping of SOURCERECORDID to Logic from template
                logic_mapping = template_df.set_index("SOURCERECORDID")[
                    "Logic"
                ].to_dict()

                # Update Logic column in claim detail data
                claim_detail_df["Logic"] = (
                    claim_detail_df["SOURCERECORDID"].map(logic_mapping).fillna("")
                )

                # Save the updated CSV
                output_dir = Path.cwd()
                opportunity_name = self._extract_opportunity_name()
                csv_path = output_dir / f"{opportunity_name} Claim Detail.csv"

                claim_detail_df.drop_duplicates().to_csv(csv_path, index=False)
                logger.info(
                    f"Generated Claim Detail CSV with template Logic: {csv_path}"
                )

            else:
                logger.error(
                    "SOURCERECORDID not found in template or claim data for matching"
                )
                # Fallback: save without updated Logic
                self._save_to_csv(claim_detail_df, Path.cwd())

        except Exception as e:
            logger.error(
                f"Failed to generate Claim Detail CSV with template Logic: {e}"
            )
            # Fallback: save the original processed data as CSV
            try:
                self._save_to_csv(self.processed_claim_data, Path.cwd())
            except Exception as fallback_error:
                logger.error(f"Fallback CSV generation also failed: {fallback_error}")

    def _finalize_processing(self, output_file):
        """Finalize processing with highlighting and notifications."""
        self.highlight_unmatched_reversals(output_file)
        self.update_progress(0.70)

        messagebox.showinfo(
            "Success", f"Processing complete. File saved as {output_file}"
        )
        # Template pasting will handle progress from 70% to 100%
        self.paste_into_template(output_file)

    def highlight_unmatched_reversals(self, excel_file):
        """Highlight unmatched reversals in Excel file with reduced nesting complexity."""
        try:
            wb = load_workbook(excel_file)
            ws = wb.active

            # Early return if worksheet is invalid
            if ws is None:
                logger.warning("Worksheet is None, cannot highlight reversals")
                return

            # Early return if reversals file doesn't exist
            if not os.path.exists("unmatched_reversals.txt"):
                logger.info("No unmatched_reversals.txt file found")
                wb.save(excel_file)
                return

            # Apply highlighting to unmatched reversals
            self._apply_reversal_highlighting(ws, wb, excel_file)

        except Exception as e:
            logger.error(f"Failed to highlight unmatched reversals: {e}")

    def _apply_reversal_highlighting(self, ws, wb, excel_file):
        """Apply highlighting to rows specified in unmatched_reversals.txt."""
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        with open("unmatched_reversals.txt", "r") as f:
            rows = f.read().strip().split(",")

        for row_str in rows:
            if self._is_valid_row_number(row_str, ws):
                self._highlight_row(ws, int(row_str), fill)

        wb.save(excel_file)
        logger.info(f"Highlighted unmatched reversals in {excel_file}")

    def _is_valid_row_number(self, row_str, ws):
        """Check if row string represents a valid row number."""
        return row_str.isdigit() and 1 <= int(row_str) <= ws.max_row

    def _highlight_row(self, ws, row_num, fill):
        """Highlight all cells in a specific row."""
        row = ws[row_num]
        for cell in row:
            cell.fill = fill

    def _fix_column_positioning(self, df):
        """
        Ensure the 'Logic' column is the 39th column (AM) and rename it to "O's & R's Check".
        If there are fewer than 39 columns, pad with empty columns.
        """
        df = df.copy()
        # Rename 'Logic' column if present
        if "Logic" in df.columns:
            df = df.rename(columns={"Logic": "O's & R's Check"})

        # Ensure we have exactly 39 columns for template compatibility
        current_cols = len(df.columns)
        if current_cols < 39:
            # Add empty columns to reach 39 total
            for i in range(current_cols, 39):
                df[f"Column_{i + 1}"] = ""
        elif current_cols > 39:
            # Keep only first 39 columns
            df = df.iloc[:, :39]

        return df


warnings.filterwarnings("ignore", category=FutureWarning, message=".*swapaxes.*")
warnings.filterwarnings("ignore", category=FutureWarning, message=".*swapaxes.*")


def process_logic_block(df_block):
    """
    Vectorized numpy logic to mark 'OR' in 'Logic' for reversals with matching claims.
    Refactored to reduce nesting complexity and improve readability.
    """
    arr = df_block.to_numpy()
    col_idx = {col: i for i, col in enumerate(df_block.columns)}

    # Extract and prepare data
    logic_data = _extract_logic_data(arr, col_idx)

    # Early return if no reversals to process
    if not np.any(logic_data["is_reversal"]):
        return pd.DataFrame(arr, columns=df_block.columns)

    # Process reversals with reduced nesting
    _process_reversals(arr, col_idx, logic_data)

    return pd.DataFrame(arr, columns=df_block.columns)


def _extract_logic_data(arr, col_idx):
    """Extract and prepare data for logic processing."""
    qty = arr[:, col_idx["QUANTITY"]].astype(float)
    return {
        "qty": qty,
        "is_reversal": qty < 0,
        "is_claim": qty > 0,
        "ndc": arr[:, col_idx["NDC"]].astype(str),
        "member": arr[:, col_idx["MemberID"]].astype(str),
        "datefilled": pd.to_datetime(arr[:, col_idx["DATEFILLED"]], errors="coerce"),
        "abs_qty": np.abs(qty),
    }


def _process_reversals(arr, col_idx, logic_data):
    """Process reversals with matching logic, using guard clauses to reduce nesting."""
    rev_idx = np.where(logic_data["is_reversal"])[0]
    claim_idx = (
        np.where(logic_data["is_claim"])[0]
        if np.any(logic_data["is_claim"])
        else np.array([], dtype=int)
    )

    # Create context object to reduce function argument count
    match_context = MatchContext(arr, col_idx, logic_data, claim_idx)

    for i in rev_idx:
        found_match = _try_find_match(match_context, i)

        # Mark unmatched reversals as 'OR'
        if not found_match:
            arr[i, col_idx["Logic"]] = "OR"


class MatchContext:
    """Context object to encapsulate matching parameters and reduce argument count."""

    def __init__(self, arr, col_idx, logic_data, claim_idx):
        self.arr = arr
        self.col_idx = col_idx
        self.logic_data = logic_data
        self.claim_idx = claim_idx


def _try_find_match(context, reversal_idx):
    """Attempt to find a matching claim for a reversal. Returns True if match found."""
    # Guard clause: no claims to match against
    if context.claim_idx.size == 0:
        return False

    # Find potential matches
    matches = _find_matching_claims(context.logic_data, context.claim_idx, reversal_idx)

    # Guard clause: no matches found
    if not np.any(matches):
        return False

    # Mark both reversal and matching claim as 'OR'
    context.arr[reversal_idx, context.col_idx["Logic"]] = "OR"
    context.arr[context.claim_idx[matches][0], context.col_idx["Logic"]] = "OR"
    return True


def _find_matching_claims(logic_data, claim_idx, reversal_idx):
    """Find claims that match the reversal based on NDC, member, quantity, and date."""
    matches = (
        (logic_data["ndc"][claim_idx] == logic_data["ndc"][reversal_idx])
        & (logic_data["member"][claim_idx] == logic_data["member"][reversal_idx])
        & (logic_data["abs_qty"][claim_idx] == logic_data["abs_qty"][reversal_idx])
    )

    # Add date constraint (within 30 days)
    date_diffs = np.abs(
        (
            logic_data["datefilled"][claim_idx] - logic_data["datefilled"][reversal_idx]
        ).days
    )
    matches &= date_diffs <= 30

    return matches


if __name__ == "__main__":
    multiprocessing.freeze_support()
    ctk.set_appearance_mode("light")  # Start in light mode
    root = ctk.CTk()  # or tk.Tk() if not using customtkinter
    app = App(root)
    root.mainloop()
