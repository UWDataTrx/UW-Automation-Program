import pandas as pd
import sys
import re
import logging
from pathlib import Path
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.utils import (
    load_file_paths,
    standardize_pharmacy_ids,
    standardize_network_ids,
    merge_with_network,
    drop_duplicates_df,
    clean_logic_and_tier,
    filter_recent_date,
    filter_logic_and_maintenance,
    filter_products_and_alternative,
    write_shared_log,
)

# Set up logger
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

try:
    import importlib.util

    if importlib.util.find_spec("xlsxwriter") is None:
        print(
            "The 'xlsxwriter' module is not installed. Please install it using 'pip install xlsxwriter'."
        )
        sys.exit(1)
except Exception:
    print("Error checking for 'xlsxwriter' module.")
    sys.exit(1)

# Load NABP/NPI list
included_nabp_npi = {
    "4528874": "1477571404",
    "2365422": "1659313435",
    "3974157": "1972560688",
    "320793": "1164437406",
    "4591055": "1851463087",
    "2348046": "1942303110",
    "4023610": "1407879588",
    "4025385": "1588706212",
    "4025311": "1588705446",
    "4026806": "1285860312",
    "4931350": "1750330775",
    "4024585": "1396768461",
    "4028026": "1497022438",
    "2643749": "1326490376",
}


# ---------------------------------------------------------------------------
# Tier summarization helper
# ---------------------------------------------------------------------------
def summarize_by_tier(df, col, from_val, to_val):
    filtered = df[(df[col] == from_val) & (df["FormularyTier"] == to_val)]
    pt = pd.pivot_table(
        filtered,
        values=["Rxs", "MemberID"],
        index=["Product Name"],
        aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
    )
    pt = pt.rename(columns={"Rxs": "Total Rxs", "MemberID": "Unique Members"})
    rxs = filtered["Rxs"].sum()
    members = filtered["MemberID"].nunique()
    return pt, rxs, members


def load_tier_disruption_data(file_paths):
    """Load all required data files for tier disruption processing."""
    # Load claims with fallback
    if not file_paths.get("reprice"):
        write_shared_log(
            "tier_disruption.py",
            "No reprice/template file provided.",
            status="ERROR",
        )
        print("No reprice/template file provided. Skipping claims loading.")
        return None

    try:
        claims = pd.read_excel(file_paths["reprice"], sheet_name="Claims Table")
    except Exception as e:
        logger.error(f"Error loading claims: {e}")
        claims = pd.read_excel(file_paths["reprice"], sheet_name=0)

    print(f"claims shape: {claims.shape}")
    claims.info()

    # Load reference tables
    medi = pd.read_excel(
        file_paths["medi_span"], usecols=["NDC", "Maint Drug?", "Product Name"]
    )
    print(f"medi shape: {medi.shape}")

    u = pd.read_excel(
        file_paths["u_disrupt"], sheet_name="Universal NDC", usecols=["NDC", "Tier"]
    )
    print(f"u shape: {u.shape}")

    e = pd.read_excel(
        file_paths["e_disrupt"],
        sheet_name="Alternatives NDC",
        usecols=["NDC", "Tier", "Alternative"],
    )
    print(f"e shape: {e.shape}")

    network = pd.read_excel(
        file_paths["n_disrupt"],
        usecols=["pharmacy_npi", "pharmacy_nabp", "pharmacy_is_excluded"],
    )
    print(f"network shape: {network.shape}")

    return claims, medi, u, e, network


def process_tier_data_pipeline(claims, reference_data, network):
    """Process the data pipeline for tier disruption."""
    medi, u, e = reference_data
    
    # Merge reference data
    df = claims.merge(medi, on="NDC", how="left")
    print(f"After merge with medi: {df.shape}")

    df = df.merge(
        u.rename(columns={"Tier": "Universal Tier"}), on="NDC", how="left"
    )
    print(f"After merge with u: {df.shape}")

    df = df.merge(
        e.rename(columns={"Tier": "Exclusive Tier"}), on="NDC", how="left"
    )
    print(f"After merge with e: {df.shape}")

    # Standardize IDs and perform network merge
    df = standardize_pharmacy_ids(df)
    print(f"After standardize_pharmacy_ids: {df.shape}")
    network = standardize_network_ids(network)
    print(f"After standardize_network_ids: {network.shape}")
    df = merge_with_network(df, network)
    print(f"After merge_with_network: {df.shape}")

    # Date parsing, deduplication, type cleaning, and filters
    df["DATEFILLED"] = pd.to_datetime(df["DATEFILLED"], errors="coerce")
    print(f"After DATEFILLED to_datetime: {df.shape}")

    df = drop_duplicates_df(df)
    print(f"After drop_duplicates_df: {df.shape}")

    df = clean_logic_and_tier(df)
    print(f"After clean_logic_and_tier: {df.shape}")

    df = filter_recent_date(df)
    print(f"After filter_recent_date: {df.shape}")

    df = filter_logic_and_maintenance(df)
    print(f"After filter_logic_and_maintenance: {df.shape}")

    df = filter_products_and_alternative(df)
    print(f"After filter_products_and_alternative: {df.shape}")

    return df


def handle_tier_pharmacy_exclusions(df, file_paths):
    """Handle pharmacy exclusions for tier disruption."""
    # Ensure 'pharmacy_is_excluded' column contains actual boolean values with type inference
    if "pharmacy_is_excluded" in df.columns:
        df["pharmacy_is_excluded"] = (
            df["pharmacy_is_excluded"]
            .astype(str)
            .str.lower()
            .map({"true": True, "false": False})
            .fillna(False)
            .infer_objects(copy=False)
        )
        logger.info(
            f"pharmacy_is_excluded value counts: {df['pharmacy_is_excluded'].value_counts().to_dict()}"
        )

        # Identify rows where pharmacy_is_excluded is "NA"
        na_pharmacies = df[df["pharmacy_is_excluded"].isna()]
        logger.info(f"NA pharmacies count: {na_pharmacies.shape[0]}")

        if not na_pharmacies.empty:
            output_file_path = file_paths["pharmacy_validation"]
            na_pharmacies_output = na_pharmacies[["PHARMACYNPI", "NABP"]].fillna("N/A")
            
            # Update existing template or create new file
            try:
                from openpyxl import load_workbook
                import os
                
                # Check if template exists
                if os.path.exists(output_file_path):
                    # Load existing workbook
                    wb = load_workbook(output_file_path)
                    logger.info(f"Loading existing template: {output_file_path}")
                else:
                    # Create new workbook
                    from openpyxl import Workbook
                    wb = Workbook()
                    logger.info(f"Creating new validation file: {output_file_path}")
                
                # Check if Validations sheet exists, if not create it
                if "Validations" in wb.sheetnames:
                    ws = wb["Validations"]
                    # Clear existing data
                    ws.delete_rows(1, ws.max_row)
                else:
                    ws = wb.create_sheet("Validations")
                    # Remove default sheet if it exists and is empty
                    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                        wb.remove(wb["Sheet"])
                
                # Write headers
                ws.append(["PHARMACYNPI", "NABP"])
                
                # Write data
                for _, row in na_pharmacies_output.iterrows():
                    ws.append([row["PHARMACYNPI"], row["NABP"]])
                
                # Save the workbook
                wb.save(output_file_path)
                logger.info(f"NA pharmacies written to '{output_file_path}' Validations sheet.")
                
            except Exception as e:
                logger.error(f"Error updating pharmacy validation file: {e}")
                # Fallback to original method
                writer = pd.ExcelWriter(output_file_path, engine="openpyxl")
                na_pharmacies_output.to_excel(
                    writer, sheet_name="Validations", index=False, engine="openpyxl"
                )
                writer.close()
                logger.info(f"NA pharmacies written to '{output_file_path}' sheet (fallback).")

    return df


def create_tier_definitions():
    """Create the tier definitions for analysis."""
    return [
        ("Universal_Positive 2-1", "Universal Tier", 1, 2),
        ("Universal_Positive 3-1", "Universal Tier", 1, 3),
        ("Universal_Positive 3-2", "Universal Tier", 2, 3),
        ("Universal_Negative 1-2", "Universal Tier", 2, 1),
        ("Universal_Negative 1-3", "Universal Tier", 3, 1),
        ("Universal_Negative 2-3", "Universal Tier", 3, 2),
        ("Exclusive_Positive 2-1", "Exclusive Tier", 1, 2),
        ("Exclusive_Positive 3-1", "Exclusive Tier", 1, 3),
        ("Exclusive_Positive 3-2", "Exclusive Tier", 2, 3),
        ("Exclusive_Negative 1-2", "Exclusive Tier", 2, 1),
        ("Exclusive_Negative 1-3", "Exclusive Tier", 3, 1),
        ("Exclusive_Negative 2-3", "Exclusive Tier", 3, 2),
    ]


def process_tier_pivots(df, tiers):
    """Process tier pivot tables and collect statistics."""
    tab_members = {}
    tab_rxs = {}
    tier_pivots = []

    for name, col, from_val, to_val in tiers:
        pt, rxs, members = summarize_by_tier(df, col, from_val, to_val)
        tier_pivots.append((name, pt, members, rxs))
        tab_members[name] = members
        tab_rxs[name] = rxs

    return tier_pivots, tab_members, tab_rxs


def process_exclusions(df):
    """Process exclusions data and create pivot table."""
    exclusions = df[df["Exclusive Tier"] == "Nonformulary"]
    ex_pt = exclusions.pivot_table(
        values=["Rxs", "MemberID"],
        index=["Product Name", "Alternative"],
        aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
    )
    ex_pt = ex_pt.rename(columns={"Rxs": "Total Rxs", "MemberID": "Unique Members"})
    exc_rxs = exclusions["Rxs"].sum()
    exc_members = exclusions["MemberID"].nunique()
    
    return ex_pt, exc_rxs, exc_members


def create_summary_dataframe(tab_members, tab_rxs, total_claims, total_members):
    """Create the summary DataFrame with calculated statistics."""
    uni_pos_keys = [
        "Universal_Positive 2-1",
        "Universal_Positive 3-1",
        "Universal_Positive 3-2",
    ]
    uni_neg_keys = [
        "Universal_Negative 1-2",
        "Universal_Negative 1-3",
        "Universal_Negative 2-3",
    ]
    ex_pos_keys = [
        "Exclusive_Positive 2-1",
        "Exclusive_Positive 3-1",
        "Exclusive_Positive 3-2",
    ]
    ex_neg_keys = [
        "Exclusive_Negative 1-2",
        "Exclusive_Negative 1-3",
        "Exclusive_Negative 2-3",
    ]

    uni_pos_utilizers = sum(tab_members[k] for k in uni_pos_keys)
    uni_pos_claims = sum(tab_rxs[k] for k in uni_pos_keys)
    uni_pos_pct = uni_pos_claims / total_claims if total_claims else 0

    uni_neg_utilizers = sum(tab_members[k] for k in uni_neg_keys)
    uni_neg_claims = sum(tab_rxs[k] for k in uni_neg_keys)
    uni_neg_pct = uni_neg_claims / total_claims if total_claims else 0

    ex_pos_utilizers = sum(tab_members[k] for k in ex_pos_keys)
    ex_pos_claims = sum(tab_rxs[k] for k in ex_pos_keys)
    ex_pos_pct = ex_pos_claims / total_claims if total_claims else 0

    ex_neg_utilizers = sum(tab_members[k] for k in ex_neg_keys)
    ex_neg_claims = sum(tab_rxs[k] for k in ex_neg_keys)
    ex_neg_pct = ex_neg_claims / total_claims if total_claims else 0

    exc_utilizers = tab_members["Exclusions"]
    exc_claims = tab_rxs["Exclusions"]
    exc_pct = exc_claims / total_claims if total_claims else 0

    return pd.DataFrame(
        {
            "Formulary": [
                "Universal Positive",
                "Universal Negative",
                "Exclusive Positive",
                "Exclusive Negative",
                "Exclusions",
            ],
            "Utilizers": [
                uni_pos_utilizers,
                uni_neg_utilizers,
                ex_pos_utilizers,
                ex_neg_utilizers,
                exc_utilizers,
            ],
            "Rxs": [
                uni_pos_claims,
                uni_neg_claims,
                ex_pos_claims,
                ex_neg_claims,
                exc_claims,
            ],
            "% of claims": [
                uni_pos_pct,
                uni_neg_pct,
                ex_pos_pct,
                ex_neg_pct,
                exc_pct,
            ],
            "": ["", "", "", "", ""],
            "Totals": [
                f"Members: {total_members}",
                f"Claims: {total_claims}",
                "",
                "",
                "",
            ],
        }
    )


def create_network_analysis(df):
    """Create network analysis for excluded pharmacies."""
    network_df = df[df["pharmacy_is_excluded"]]
    logger.debug(f"Initial network_df shape: {network_df.shape}")
    logger.debug(f"Initial network_df contents: {network_df.head(10).to_dict()}")

    filter_phrases = [
        "CVS",
        "Walgreens",
        "Kroger",
        "Walmart",
        "Rite Aid",
        "Optum",
        "Express Scripts",
        "DMR",
        "Williams Bro",
        "Publix",
    ]
    filter_phrases_escaped = [
        re.escape(phrase.lower()) for phrase in filter_phrases
    ]
    regex_pattern = "|".join([f"\\b{p}\\b" for p in filter_phrases_escaped])
    network_df = network_df[
        ~network_df["Pharmacy Name"]
        .str.lower()
        .str.contains(regex_pattern, case=False, regex=True, na=False)
    ]
    logger.debug(f"Filtered network_df shape: {network_df.shape}")
    logger.debug(f"Filtered network_df contents: {network_df.head(10).to_dict()}")

    if {"PHARMACYNPI", "NABP", "Pharmacy Name"}.issubset(network_df.columns):
        network_df["PHARMACYNPI"] = network_df["PHARMACYNPI"].fillna("N/A")
        network_df["NABP"] = network_df["NABP"].fillna("NABP")
        # Create Unique Identifier based on whichever identifier is available (PHARMACYNPI or NABP)
        network_df["Unique Identifier"] = network_df.apply(
            lambda row: row["PHARMACYNPI"]
            if row["PHARMACYNPI"] != "N/A"
            else row["NABP"],
            axis=1,
        )
        network_pivot = pd.pivot_table(
            network_df,
            values=["Rxs", "MemberID"],
            index=["Unique Identifier"],
            aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
        )
        network_pivot = network_pivot.rename(
            columns={"Rxs": "Total Rxs", "MemberID": "Unique Members"}
        )
        network_pivot.reset_index(
            inplace=True
        )  # Ensure index columns are included in the output
        logger.debug(f"Network pivot shape: {network_pivot.shape}")
        logger.debug(f"Network pivot contents: {network_pivot.head(10).to_dict()}")
        
        return network_df, network_pivot
    
    return network_df, None


def write_excel_sheets(writer, df, summary_df, tier_pivots, ex_pt, exc_members, network_df, network_pivot):
    """Write all sheets to the Excel file."""
    # Write Summary sheet
    summary_df.to_excel(writer, sheet_name="Summary", index=False)

    # Write tier pivots
    for name, pt, members, _ in tier_pivots:
        pt.to_excel(writer, sheet_name=name)
        writer.sheets[name].write("F1", f"Total Members: {members}")

    # Write Exclusions sheet
    ex_pt.to_excel(writer, sheet_name="Exclusions")
    writer.sheets["Exclusions"].write("F1", f"Total Members: {exc_members}")

    # Write Data sheet
    data_sheet_df = df.copy()
    data_sheet_df.to_excel(writer, sheet_name="Data", index=False)

    # Write Network sheet
    if network_pivot is not None:
        network_pivot.to_excel(writer, sheet_name="Network", index=False)

    # Write filtered network data
    selected_columns = [
        "PHARMACYNPI",
        "NABP",
        "MemberID",
        "Pharmacy Name",
        "pharmacy_is_excluded",
        "Unique Identifier",
    ]
    network_df[selected_columns].to_excel(writer, sheet_name="Network", index=False)
    logger.info(
        f"Network sheet updated with {network_df.shape[0]} excluded pharmacy records (minus major chains) and selected columns"
    )


def reorder_excel_sheets(writer):
    """Reorder sheets so Summary follows Data."""
    sheets = writer.sheets  # This is a dict: {sheet_name: worksheet_object}
    names = list(sheets.keys())
    if "Data" in names and "Summary" in names:
        data_idx = names.index("Data")
        summary_idx = names.index("Summary")
        if summary_idx != data_idx + 1:
            # Move "Summary" sheet immediately after "Data"
            items = list(sheets.items())
            summary_item = items.pop(summary_idx)
            items.insert(data_idx + 1, summary_item)
            writer.sheets.clear()
            writer.sheets.update(dict(items))


def show_completion_message(output_path):
    """Show completion message and popup."""
    write_shared_log("tier_disruption.py", "Processing complete.")
    print(f"Processing complete. Output file: {output_path}")
    try:
        import tkinter as tk
        from tkinter import messagebox

        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo("Processing Complete", "Processing complete")
        root.destroy()
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Main processing pipeline
# ---------------------------------------------------------------------------
def process_data():
    write_shared_log("tier_disruption.py", "Processing started.")
    # Output filename from CLI arg or default
    output_filename = "LBL for Disruption.xlsx"
    if len(sys.argv) > 1:
        output_filename = sys.argv[1]
    output_path = Path(output_filename).resolve()
    try:
        # Get the config file path relative to the project root
        config_path = Path(__file__).parent.parent / "config" / "file_paths.json"
        file_paths = load_file_paths(str(config_path))

        result = load_tier_disruption_data(file_paths)
        if result is None:
            return  # Early exit if claims loading failed
        claims, medi, u, e, network = result

        reference_data = (medi, u, e)
        df = process_tier_data_pipeline(claims, reference_data, network)

        df = handle_tier_pharmacy_exclusions(df, file_paths)

        # Totals for summary
        total_claims = df["Rxs"].sum()
        total_members = df["MemberID"].nunique()

        # Excel writer setup
        writer = pd.ExcelWriter(output_path, engine="xlsxwriter")

        # Summary calculations (must be written immediately after Data)
        tiers = create_tier_definitions()

        tier_pivots, tab_members, tab_rxs = process_tier_pivots(df, tiers)

        # Exclusions sheet (Nonformulary)
        ex_pt, exc_rxs, exc_members = process_exclusions(df)
        tab_members["Exclusions"] = exc_members
        tab_rxs["Exclusions"] = exc_rxs

        # Summary calculations
        summary_df = create_summary_dataframe(tab_members, tab_rxs, total_claims, total_members)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # Write tier pivots and Exclusions after Summary
        for name, pt, members, _ in tier_pivots:
            pt.to_excel(writer, sheet_name=name)
            writer.sheets[name].write("F1", f"Total Members: {members}")

        ex_pt.to_excel(writer, sheet_name="Exclusions")
        writer.sheets["Exclusions"].write("F1", f"Total Members: {exc_members}")

        # Write the 'Data Sheet' with excluded and non-excluded pharmacies
        data_sheet_df = df.copy()
        data_sheet_df.to_excel(writer, sheet_name="Data", index=False)

        # Network summary for excluded pharmacies (pharmacy_is_excluded=True)
        network_df, network_pivot = create_network_analysis(df)
        logger.info(f"Total pharmacies in dataset: {df.shape[0]}")
        logger.info(
            f"Excluded pharmacies (pharmacy_is_excluded=True): {df['pharmacy_is_excluded'].sum()}"
        )
        logger.info(
            f"Non-excluded pharmacies (pharmacy_is_excluded=False): {(~df['pharmacy_is_excluded']).sum()}"
        )
        logger.info(
            f"Network sheet will show {network_df.shape[0]} excluded pharmacy records (minus major chains)"
        )

        # Write Network sheet
        if network_pivot is not None:
            network_pivot.to_excel(writer, sheet_name="Network", index=False)

        # Write filtered network data
        selected_columns = [
            "PHARMACYNPI",
            "NABP",
            "MemberID",
            "Pharmacy Name",
            "pharmacy_is_excluded",
            "Unique Identifier",
        ]
        network_df[selected_columns].to_excel(writer, sheet_name="Network", index=False)
        logger.info(
            f"Network sheet updated with {network_df.shape[0]} excluded pharmacy records (minus major chains) and selected columns"
        )

        # Reorder sheets so Summary follows Data
        reorder_excel_sheets(writer)

        writer.close()
        show_completion_message(output_path)
    except Exception as e:
        write_shared_log(
            "tier_disruption.py", f"Processing failed: {e}", status="ERROR"
        )
        raise


if __name__ == "__main__":
    process_data()
    print("Data processing complete")
