import sys
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
import logging
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.utils import write_shared_log

# Configure logging
logging.basicConfig(
    filename="merge_log.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

MERGED_FILENAME = "merged_file.xlsx"
REQUIRED_COLUMNS = [
    "DATEFILLED",
    "SOURCERECORDID",
    "QUANTITY",
    "DAYSUPPLY",
    "NDC",
    "MemberID",
    "Drug Name",
    "Pharmacy Name",
    "Total AWP (Historical)",
]


def merge_files(file1_path, file2_path):
    file1 = Path(file1_path)
    file2 = Path(file2_path)
    try:
        logger.info(f"Starting merge: {file1} + {file2}")
        write_shared_log("merge.py", f"Starting merge: {file1} + {file2}")

        if not file1.exists():
            logger.error(f"File not found: {file1}")
            write_shared_log("merge.py", f"File not found: {file1}", status="ERROR")
            return False
        if not file2.exists():
            logger.error(f"File not found: {file2}")
            write_shared_log("merge.py", f"File not found: {file2}", status="ERROR")
            return False

        # Load data (support Excel or CSV for both files)
        try:
            if file1.suffix == ".csv":
                df1 = pd.read_csv(file1, parse_dates=["DATEFILLED"], dayfirst=False)
            else:
                df1 = pd.read_excel(file1, parse_dates=["DATEFILLED"])
        except Exception as e:
            logger.error(f"Failed to load file1: {e}")
            write_shared_log("merge.py", f"Failed to load file1: {e}", status="ERROR")
            return False
        try:
            if file2.suffix == ".csv":
                df2 = pd.read_csv(file2)
            else:
                df2 = pd.read_excel(file2)
        except Exception as e:
            logger.error(f"Failed to load file2: {e}")
            write_shared_log("merge.py", f"Failed to load file2: {e}", status="ERROR")
            return False

        # Log data source details
        logger.info(f"df1 shape: {df1.shape}, df2 shape: {df2.shape}")
        write_shared_log("merge.py", f"df1 shape: {df1.shape}, df2 shape: {df2.shape}")
        logger.info(f"df1 columns: {list(df1.columns)}")
        logger.info(f"df2 columns: {list(df2.columns)}")

        # Clean up and standardize column names
        df2.columns = [col.strip() for col in df2.columns]
        if "Source Record ID" in df2.columns:
            df2.rename(columns={"Source Record ID": "SOURCERECORDID"}, inplace=True)

        # Merge
        try:
            df_merged = pd.merge(df1, df2, on="SOURCERECORDID", how="outer")
        except Exception as e:
            logger.error(f"Failed to merge: {e}")
            write_shared_log("merge.py", f"Failed to merge: {e}", status="ERROR")
            return False
        if "Total AWP (Historical)" in df_merged.columns:
            df_merged["Total AWP (Historical)"] = pd.to_numeric(
                df_merged["Total AWP (Historical)"], errors="coerce"
            ).round(2)
        else:
            df_merged["Total AWP (Historical)"] = 0.00
        df_merged["MemberID"] = df_merged["MemberID"].fillna(0)

        # Log merged file row count and checksum
        row_count = len(df_merged)
        try:
            import hashlib

            sample = df_merged.head(1000).to_csv(index=False).encode("utf-8")
            checksum = hashlib.md5(sample).hexdigest()
        except Exception as e:
            checksum = f"ERROR: {e}"
        logger.info(f"Merged row count: {row_count}, sample checksum: {checksum}")
        write_shared_log(
            "merge.py", f"Merged row count: {row_count}, sample checksum: {checksum}"
        )

        # Log missing required columns
        for col in REQUIRED_COLUMNS:
            if col not in df_merged.columns:
                logger.warning(f"Missing expected column: {col}")
                write_shared_log(
                    "merge.py", f"Missing expected column: {col}", status="WARNING"
                )

        # Drop the DATEFILLED_DIFF column if it exists before saving the merged file
        if "DATEFILLED_DIFF" in df_merged.columns:
            df_merged.drop(columns=["DATEFILLED_DIFF"], inplace=True)

        merged_path = Path.cwd() / MERGED_FILENAME
        try:
            df_merged.to_excel(merged_path, index=False)
        except Exception as e:
            logger.error(f"Failed to write merged Excel: {e}")
            write_shared_log(
                "merge.py", f"Failed to write merged Excel: {e}", status="ERROR"
            )
            return False
        logger.info(f"Merged file saved to: {merged_path}")
        write_shared_log("merge.py", f"Merged file saved to: {merged_path}")

        # Apply Excel formatting
        try:
            wb = load_workbook(merged_path)
            ws = wb.active
            date_style = NamedStyle(name="date_style", number_format="MM/DD/YYYY")

            if ws is not None and ws.max_row >= 1:
                header = [
                    cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))
                ]
                if "DATEFILLED" in header:
                    date_col_index = header.index("DATEFILLED") + 1
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=date_col_index).style = date_style
                    wb.save(merged_path)
                    logger.info("Applied date formatting successfully.")
                    write_shared_log(
                        "merge.py", "Applied date formatting successfully."
                    )
                else:
                    logger.warning("DATEFILLED column not found for formatting.")
                    write_shared_log(
                        "merge.py",
                        "DATEFILLED column not found for formatting.",
                        status="WARNING",
                    )
            else:
                logger.warning(
                    "Worksheet is empty or not loaded, cannot apply formatting."
                )
                write_shared_log(
                    "merge.py",
                    "Worksheet is empty or not loaded, cannot apply formatting.",
                    status="WARNING",
                )

        except Exception as ex:
            logger.warning(f"Failed to apply formatting: {ex}")
            write_shared_log(
                "merge.py", f"Failed to apply formatting: {ex}", status="WARNING"
            )

        return True
    except Exception as e:
        logger.exception(f"Merge failed: {e}")
        write_shared_log("merge.py", f"Merge failed: {e}", status="ERROR")
        return False


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python merge.py <file1_path> <file2_path>")
        sys.exit(1)

    success = merge_files(sys.argv[1], sys.argv[2])
    if not success:
        sys.exit(2)  # Exit with code 2 to indicate merge failure
