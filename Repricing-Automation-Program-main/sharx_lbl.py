import logging
import tkinter as tk
from pathlib import Path
from tkinter import messagebox

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from utils import load_file_paths

# Setup logging
logging.basicConfig(
    filename="sharx_lbl.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


def show_message(awp, ing, total, rxs):
    messagebox.showinfo(
        "Process Complete",
        f"SHARx LBL has been created\n\n"
        f"Total AWP: ${awp:.2f}\n\n"
        f"Total Ing Cost: ${ing:.2f}\n\n"
        f"Total Gross Cost: ${total:.2f}\n\n"
        f"Total Claim Count: {rxs}",
    )


def main():
    root = tk.Tk()
    root.withdraw()
    try:
        paths = load_file_paths()
        template_path = Path(paths["sharx"])

        df = pd.read_excel(paths["reprice"], sheet_name="Claims Table")

        # Debug print and log
        print("Columns in claims DataFrame:")
        print(df.columns)
        logger.info(f"Columns in claims DataFrame: {df.columns.tolist()}")

        df["Logic"] = pd.to_numeric(df["Logic"], errors="coerce")
        df = df[df["Logic"].between(1, 10)]

        awp = df["Total AWP (Historical)"].sum()
        ing = df["Rx Sense Ing Cost"].sum()
        total = df["RxSense Total Cost"].sum()
        rxs = df["Rxs"].sum()

        columns_to_keep = [
            "MONY",
            "Rxs",
            "Rx Sense Ing Cost",
            "RxSense Dispense Fee",
            "RxSense Total Cost",
            "Total AWP (Historical)",
            "GrossCost",
            "Pharmacy Name",
            "INPUTFILECHANNEL",
            "DATEFILLED",
            "MemberID",
            "DAYSUPPLY",
            "QUANTITY",
            "NDC",
            "DST Drug Name",
        ]
        df = df[columns_to_keep]

        wb = load_workbook(template_path)
        ws = wb["Line By Line"]

        for r_idx, row in enumerate(
            dataframe_to_rows(df, index=False, header=False), start=2
        ):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        output_path = Path("_Rx Claims for SHARx.xlsx")
        wb.save(output_path)

        logger.info("SHARx LBL file created successfully.")
        show_message(awp, ing, total, rxs)

    except Exception as e:
        logger.exception("An error occurred during SHARx LBL processing")
        messagebox.showerror("Error", f"An error occurred: {e}")
    finally:
        root.quit()


if __name__ == "__main__":
    main()
