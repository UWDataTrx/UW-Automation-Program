import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Optional, Tuple, Any


def detect_overwritten_cells(
    original_path: str,
    updated_path: str,
    sheet_names: Optional[List[str]] = None
) -> List[Tuple[str, int, int, Any, Any]]:
    """
    Compare two Excel workbooks and return a list of all cells
    where the 'updated' workbook has overwritten a non-empty value
    that existed in the 'original' workbook.

    Parameters:
        original_path   - path to the original/template workbook
        updated_path    - path to the modified workbook
        sheet_names     - list of sheet names to compare; if None, compare all sheets present in both

    Returns:
        A list of tuples: (sheet_name, row_idx, col_idx, original_value, updated_value)
        - row_idx and col_idx are 1-based indices (matching Excel’s numbering).
        - original_value is the value from the original workbook.
        - updated_value  is the value from the updated workbook.
    """
    overwritten_cells: List[Tuple[str, int, int, Any, Any]] = []

    # Load both workbooks (data_only=True to read computed values)
    wb_original = load_workbook(original_path, data_only=True)
    wb_updated = load_workbook(updated_path, data_only=True)

    # If no specific sheets were passed in, compare all sheets common to both workbooks
    if sheet_names is None:
        sheet_names = [s for s in wb_original.sheetnames if s in wb_updated.sheetnames]

    for sheet in sheet_names:
        # Read each sheet into a DataFrame with header=None so that
        # DataFrame indices match Excel’s 1-based rows and columns.
        df_orig = pd.read_excel(original_path, sheet_name=sheet, header=None, dtype=object)
        df_upd = pd.read_excel(updated_path, sheet_name=sheet, header=None, dtype=object)

        max_rows = max(df_orig.shape[0], df_upd.shape[0])
        max_cols = max(df_orig.shape[1], df_upd.shape[1])

        for r in range(max_rows):
            for c in range(max_cols):
                # Use DataFrame.iat with bounds checking
                val_orig = df_orig.iat[r, c] if (r < df_orig.shape[0] and c < df_orig.shape[1]) else None
                val_upd = df_upd.iat[r, c] if (r < df_upd.shape[0] and c < df_upd.shape[1]) else None

                # If there was a non-empty original and the updated differs, record it
                if pd.notna(val_orig) and val_orig != "" and val_orig != val_upd:
                    overwritten_cells.append((sheet, r + 1, c + 1, val_orig, val_upd))

    return overwritten_cells


def highlight_overwritten_cells(
    updated_path: str,
    overwritten_cells: List[Tuple[str, int, int, Any, Any]],
    output_path: Optional[str] = None
) -> None:
    """
    Highlight the list of overwritten cells in the 'updated' workbook by filling them yellow.
    Saves the modified workbook to output_path (or overwrites updated_path if output_path is None).

    Parameters:
        updated_path      - path to the workbook where cells have been overwritten
        overwritten_cells - list of tuples: (sheet_name, row_idx, col_idx, orig_val, new_val)
        output_path       - if provided, save to this path; otherwise overwrite updated_path
    """
    wb = load_workbook(updated_path)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for sheet_name, row_idx, col_idx, _, _ in overwritten_cells:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = yellow_fill

    save_target = output_path if output_path else updated_path
    wb.save(save_target)