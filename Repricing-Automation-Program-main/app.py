import tkinter as tk
import customtkinter as ctk
from tkinter import filedialog, messagebox, scrolledtext
import subprocess
import os
import shutil
import logging
import threading
import multiprocessing
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import win32com.client as win32
from pathlib import Path
import json
import csv
import time
from datetime import datetime
from plyer import notification
import re
import sys
# Excel COM check
try:
    import win32com.client as win32
    EXCEL_COM_AVAILABLE = True
except ImportError:
    EXCEL_COM_AVAILABLE = False

# Configuration and audit log files
CONFIG_FILE = Path('config.json')
AUDIT_LOG = Path('audit_log.csv')

# UI styling variables
font_select = ('Cambria', 20, 'bold')

# Color palettes
LIGHT_COLORS = {
    "dark_blue": "#D9EAF7",
    "grey_blue": "#A3B9CC",
    "mint": "#8FD9A8",
    "button_red": "#D52B2B"
}
DARK_COLORS = {
    "dark_blue": "#223354",
    "grey_blue": "#31476A",
    "mint": "#26A69A",
    "button_red": "#931D1D"
}

# Template handling constants
BACKUP_SUFFIX = '_backup.xlsx'
UPDATED_TEMPLATE_NAME = '_Rx Repricing_wf.xlsx'

# Logging setup
logging.basicConfig(filename='repricing_log.log', level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)

class ConfigManager:
    def __init__(self):
        self.config = {}
        if CONFIG_FILE.exists():
            self.load()
        else:
            self.save_default()

    def save_default(self):
        self.config = {'last_folder': str(Path.cwd())}
        self.save()

    def load(self):
        with open(CONFIG_FILE, 'r') as f:
            self.config = json.load(f)

    def save(self):
        with open(CONFIG_FILE, 'w') as f:
            json.dump(self.config, f, indent=4)

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Reprice Automation")
        self.root.configure(fg_color=LIGHT_COLORS["dark_blue"])
        self.root.resizable(True, True)
        # Ensure full UI visibility on launch
        self.root.geometry('900x900')
        # Or maximize window:
        # self.root.state('zoomed')

        # Paths
        self.file1_path = None
        self.file2_path = None
        self.template_file_path = None

        # Process control
        self.cancel_event = threading.Event()
        self.start_time = None

        # Config
        self.config_manager = ConfigManager()

        self.selected_disruption_type = tk.StringVar(value="Tier")

        self._build_ui()
        self.apply_theme_colors(LIGHT_COLORS)

        default_template = Path("_Rx Repricing_wf.xlsx")
        if default_template.exists():
            self.template_file_path = str(default_template)
        else:
            self.template_file_path = None

    def _build_ui(self):
        # Title
        self.title_label = ctk.CTkLabel(self.root, text='Repricing Automation', font=('Cambria', 26, 'bold'))
        self.title_label.grid(row=0, column=0, sticky='w', pady=20, padx=20)

        # File & action buttons frame
        self.button_frame = ctk.CTkFrame(self.root, fg_color=LIGHT_COLORS["grey_blue"])
        self.button_frame.grid(row=1, column=0, columnspan=3, sticky='ew', pady=10, padx=10)

        # Headers
        file_name_title = ctk.CTkLabel(self.button_frame, text='File Name', font=font_select)
        file_name_title.grid(row=0, column=2, pady=10, padx=10)
        
        # Notes
        self.notes_frame = ctk.CTkFrame(self.root, fg_color=LIGHT_COLORS["grey_blue"])
        self.notes_frame.grid(row=2, column=0, columnspan=4, sticky='ew', pady=10, padx=10)
        notes = ctk.CTkLabel(self.notes_frame, text=(
            'Note:\n\n'
            'Ensure FormularyTier is set before running disruption.\n'
            'Close any open Excel instances.\n'
            'Keep template name as _Rx Repricing_wf until done.'
        ), font=font_select, justify='left')
        notes.pack(padx=20, pady=10)


        # Import File 1
        self.file1_button = ctk.CTkButton(self.button_frame, text="Import File Uploaded to Tool",
            command=self.import_file1, font=font_select, height=40, fg_color=LIGHT_COLORS["mint"], text_color="#000000")
        self.file1_button.grid(row=1, column=0, pady=10, padx=10, sticky='ew')
        self.file1_label = ctk.CTkLabel(self.button_frame, text="", width=350)
        self.file1_label.grid(row=1, column=2, pady=20, padx=10)

        # Import File 2
        self.file2_button = ctk.CTkButton(self.button_frame, text="Import File From Tool",
            command=self.import_file2, font=font_select, height=40, fg_color=LIGHT_COLORS["mint"], text_color="#000000")
        self.file2_button.grid(row=2, column=0, pady=10, padx=10, sticky='ew')
        self.file2_label = ctk.CTkLabel(self.button_frame, text="")
        self.file2_label.grid(row=2, column=2, pady=20, padx=10)

        # Select Template
        self.template_button = ctk.CTkButton(self.button_frame, text="Select Template File",
            command=self.import_template_file, font=font_select, height=40, fg_color=LIGHT_COLORS["mint"], text_color="#000000")
        self.template_button.grid(row=3, column=0, pady=10, padx=10, sticky='ew')
        self.template_label = ctk.CTkLabel(self.button_frame, text="")
        self.template_label.grid(row=3, column=2, pady=20, padx=10)
        if self.template_file_path:
            self.template_label.configure(text=os.path.basename(self.template_file_path), font=font_select)

         # Cancel, Logs, Theme buttons
        self.cancel_button = ctk.CTkButton(
            self.button_frame,
            text="Cancel",
            command=self.cancel_process,
            font=font_select,
            height=40,
            fg_color=LIGHT_COLORS["button_red"],      # darker red
            text_color="#000000"     # black text
        )
        self.cancel_button.grid(row=4, column=0, pady=10, padx=10, sticky='ew')

        # View Logs button
        self.logs_button = ctk.CTkButton(
            self.button_frame,
            text="View Logs",
            command=self.show_log_viewer,
            font=font_select,
            height=40,
            fg_color=LIGHT_COLORS["mint"],
            text_color="#000000"
        )
        self.logs_button.grid(row=4, column=1, pady=10, padx=10, sticky='ew')

        # Toggle Dark Mode button
        self.toggle_theme_button = ctk.CTkButton(
            self.button_frame,
            text="Switch to Dark Mode",
            command=self.toggle_dark_mode,
            font=font_select,
            height=40,
            fg_color=LIGHT_COLORS["mint"],
            text_color="#000000"
        )
        
        self.toggle_theme_button.grid(row=4, column=2, pady=10, padx=10, sticky='ew')

        # SHARx and EPLS LBL buttons
        self.sharx_lbl_button = ctk.CTkButton(self.button_frame, text="Generate SHARx LBL", command=self.sharx_lbl,
            font=font_select, height=40, fg_color=LIGHT_COLORS["mint"], text_color="#000000")
        self.sharx_lbl_button.grid(row=5, column=0, pady=10, padx=10, sticky='ew')
        self.epls_lbl_button = ctk.CTkButton(self.button_frame, text="Generate EPLS LBL", command=self.epls_lbl,
            font=font_select, height=40, fg_color=LIGHT_COLORS["mint"], text_color="#000000")
        self.epls_lbl_button.grid(row=5, column=1, pady=10, padx=10, sticky='ew')
        self.start_process_button = ctk.CTkButton(self.button_frame, text="Start Repricing",
            command=self.start_process, font=font_select, height=40, width=200, fg_color=LIGHT_COLORS["mint"], text_color="#000000")
        self.start_process_button.grid(row=5, column=2, pady=10, padx=10, sticky='ew')

        self.start_disruption_button = ctk.CTkButton(self.root, text='Start Disruption',
            command=self.start_disruption, font=font_select, height=40, width=200, fg_color=LIGHT_COLORS["mint"], text_color="#000000")
        self.start_disruption_button.grid(row=5, column=2, pady=10, padx=10, sticky='e')
        
        # Disruption type selector
        self.dis_frame = ctk.CTkFrame(self.root, fg_color=LIGHT_COLORS["grey_blue"])
        self.dis_frame.grid(row=3, column=0, columnspan=4, sticky='ew', pady=10, padx=10)
        self.disruption_type_label = ctk.CTkLabel(self.dis_frame, text="Disruption Type:", font=font_select)
        self.disruption_type_label.grid(row=0, column=0, pady=10, padx=10)
        self.disruption_type_combobox = ctk.CTkComboBox(self.dis_frame,
            values=["Tier","B/G","OpenMDF (Tier)","OpenMDF (B/G)"],
            variable=self.selected_disruption_type, width=200,
            dropdown_font=font_select, fg_color=LIGHT_COLORS["grey_blue"],
            dropdown_fg_color=LIGHT_COLORS["grey_blue"], dropdown_hover_color=LIGHT_COLORS["mint"],
            font=font_select, button_color=LIGHT_COLORS["mint"],
            button_hover_color=LIGHT_COLORS["mint"], border_color=LIGHT_COLORS["mint"])
        self.disruption_type_combobox.grid(row=0, column=1, padx=10, pady=10)
        self.disruption_type_combobox.bind("<<ComboboxSelected>>", self.select_disruption_type)


        # Progress
        self.prog_frame = ctk.CTkFrame(self.root, fg_color=LIGHT_COLORS["grey_blue"])
        self.prog_frame.grid(row=6, column=0, columnspan=3, sticky='ew', pady=10, padx=10)
        self.progress_bar = ctk.CTkProgressBar(self.prog_frame, orientation="horizontal", mode="determinate")
        self.progress_bar.set(0)
        self.progress_bar.pack(padx=10, pady=(10,5), fill="x")
        self.progress_label = ctk.CTkLabel(self.prog_frame, text="Progress: 0%")
        self.progress_label.pack(padx=10, pady=(0,10), anchor='e')

        # Exit button
        self.exit_button = ctk.CTkButton(
            self.button_frame,
            text="Exit",
            command=self.root.quit,
            font=font_select,
            height=40,
            fg_color=LIGHT_COLORS["button_red"],      # darker red
            text_color="#000000"     # black text
        )
        self.exit_button.grid(row=6, column=2, pady=10, padx=10, sticky='ew')

    # File import methods
    def import_file1(self):
        file_path = filedialog.askopenfilename(title="Select File Uploaded to Tool", filetypes=[("All files", "*.*"), ("CSV files", "*.csv"), ("Excel files", "*.xlsx")])
        if file_path:
            self.file1_path = file_path
            self.file1_label.configure(text=os.path.basename(file_path), font=font_select)
            self.check_template(file_path)
        if not file_path:
            return  # User cancelled, do nothing
        df = pd.read_csv(file_path)
        if 'GrossCost' in df.columns:
            if df['GrossCost'].isna().all() or (df['GrossCost'] == 0).all():
                messagebox.showinfo("Template Selection", "The GrossCost column is blank or all zero. Please use the Blind template.")
            else:
                messagebox.showinfo("Template Selection", "The GrossCost column contains data. Please use the Standard template.")

    def import_file2(self):
        file_path = filedialog.askopenfilename(title="Select File From Tool", filetypes=[("All files", "*.*"), ("CSV files", "*.csv"), ("Excel files", "*.xlsx")])
        if file_path:
            self.file2_path = file_path
            self.file2_label.configure(text=os.path.basename(file_path), font=font_select)

    def import_template_file(self):
        file_path = filedialog.askopenfilename(title="Select Template File", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            self.template_file_path = file_path
            self.template_label.configure(text=os.path.basename(file_path), font=font_select) 
        if self.template_file_path:
            self.template_label.configure(text=os.path.basename(self.template_file_path), font=font_select)

    # Logging and notification methods
    def write_audit_log(self, file1, file2, status):
        entry = [datetime.now().isoformat(), str(file1), str(file2), status]
        write_header = not AUDIT_LOG.exists()
        with open(AUDIT_LOG, 'a', newline='') as f:
            writer = csv.writer(f)
            if write_header:
                writer.writerow(['Timestamp','File1','File2','Status'])
            writer.writerow(entry)

    # Cancel during repricing
    def cancel_process(self):
        # Cancel not implemented for repricing loop, placeholder
        logger.info("Cancel pressed")
        messagebox.showinfo("Cancelled","Repricing cancellation requested.")

    # Live log viewer
    def show_log_viewer(self):
        win = tk.Toplevel(self.root)
        win.title("Log Viewer")
        txt = scrolledtext.ScrolledText(win, width=100, height=30)
        txt.pack(fill='both',expand=True)
        def refresh():
            with open('repricing_log.log','r') as f:
                txt.delete('1.0', tk.END)
                txt.insert(tk.END, f.read())
            win.after(3000, refresh)
        refresh()

    def toggle_dark_mode(self):
        current = ctk.get_appearance_mode().lower()

        if current == 'light':
            # Switch *into* Dark mode
            ctk.set_appearance_mode('dark')
            self.apply_theme_colors(DARK_COLORS)
            # Now the button’s job is to switch *back* to Light
            self.toggle_theme_button.configure(text="Switch to Light Mode")

        else:
            # Switch *into* Light mode
            ctk.set_appearance_mode('light')
            self.apply_theme_colors(LIGHT_COLORS)
            # Now the button’s job is to switch *into* Dark
            self.toggle_theme_button.configure(text="Switch to Dark Mode")

    def update_progress(self, value):
        self.progress_bar.set(value)
        percent = int(value * 100)
        elapsed = time.time() - self.start_time if self.start_time else 0
        estimated = int((elapsed / value) * (1 - value)) if value > 0 else 0
        self.progress_label.configure(text=f"Progress: {percent}% | Est. {estimated}s left")
        self.root.update_idletasks()

    def write_audit_log(self, file1, file2, status):
        entry = [datetime.now().isoformat(), str(file1), str(file2), status]
        write_header = not AUDIT_LOG.exists()
        with open(AUDIT_LOG, 'a', newline='') as csvfile:
            writer = csv.writer(csvfile)
            if write_header:
                writer.writerow(['Timestamp', 'File1', 'File2', 'Status'])
            writer.writerow(entry)

    def show_log_viewer(self):
        log_win = tk.Toplevel(self.root)
        log_win.title("Live Log Viewer")
        text_area = scrolledtext.ScrolledText(log_win, width=100, height=30)
        text_area.pack(fill='both', expand=True)
        def update_logs():
            with open('repricing_log.log', 'r') as f:
                text_area.delete(1.0, tk.END)
                text_area.insert(tk.END, f.read())
            log_win.after(3000, update_logs)
        update_logs()

    def finish_notification(self):
        notification.notify(title='Repricing Automation', message='Batch processing completed.', timeout=5)
        messagebox.showinfo("Completed", "Batch processing finished!")

    # Disruption and process methods
    def select_disruption_type(self, event):
        self.selected_disruption_type.set(self.disruption_type_combobox.get())

    def start_disruption(self):
        selected_type = self.selected_disruption_type.get().strip()
        program_files = {
            'Tier': 'tier_disruption.py',
            'B/G': 'bg_disruption.py',
            'OpenMDF (Tier)': 'openmdf_tier.py',
            'OpenMDF (B/G)': 'openmdf_bg.py',
            'Full Claims File': 'full_claims.py'
        }
        if selected_type in program_files:
            program_name = program_files[selected_type]
            try:
                # Only pass template_file_path if it exists, otherwise just run the script
                args = ['python', program_name]
                if self.template_file_path:
                    args.append(str(self.template_file_path))
                subprocess.run(args, check=True)
                messagebox.showinfo("Success", f"{selected_type} disruption completed successfully.")
            except subprocess.CalledProcessError as e:
                logger.exception(f"Failed to run {program_name}")
                messagebox.showerror("Error", f"{selected_type} disruption failed: {e}")
                
    # Repricing workflow
    def paste_into_template(self, processed_file):
        try:
            self.update_progress(0.92)
            output_path = "_Rx Repricing_wf.xlsx"
            output_dir = os.path.dirname(os.path.abspath(output_path))
            backup_name = os.path.basename(self.template_file_path).replace('.xlsx', '_backup.xlsx')
            backup_path = os.path.join(output_dir, backup_name)
            shutil.copy(self.template_file_path, backup_path)
            df = pd.read_excel(processed_file).fillna("")
            for col in df.select_dtypes(include=['datetime', 'datetimetz']):
                df[col] = df[col].astype(str)
            data = df.values.tolist()

            self.update_progress(0.94)

            excel = win32.Dispatch("Excel.Application")
            excel.Visible = False
            wb_template = excel.Workbooks.Open(self.template_file_path)
            ws = wb_template.Worksheets("Claims Table")

            start_row, start_col = 2, 1
            end_row = start_row + len(data) - 1
            end_col = start_col + len(data[0]) - 1

            col_letter = chr(64 + len(data[0]))
            logger.info(f"Pasting data into range: A2:{col_letter}{end_row}")

            excel_range = ws.Range(ws.Cells(start_row, start_col), ws.Cells(end_row, end_col))
            existing = excel_range.VALUE
            new_data = []
            for row_old, row_new in zip(existing, data):
                new_row = []
                for old_val, new_val in zip(row_old, row_new):
                    if isinstance(old_val, str) and old_val.strip().startswith("="):
                        new_row.append(old_val) # Keep formulas as is
                    else:
                        new_row.append(new_val)
                new_data.append(new_row)
            excel_range.Value = new_data

            self.update_progress(0.97)

            wb_template.SaveAs(os.path.abspath(output_path))
            wb_template.Close(SaveChanges=True)
            excel.Quit()

            os.startfile(output_path)
            self.show_toast("Data was pasted into template successfully.")
            self.update_progress(1.0)
            
        except Exception as e:
            logger.error(f"Template paste error: {e}")
            messagebox.showerror("Error", f"Template update failed: {e}")
            self.update_progress(0)

    def filter_template_columns(self, df):
        columns = list(df.columns)
        try:
            client_name_idx = columns.index('Client Name')
            logic_idx = columns.index('Logic')
            selected_columns = columns[client_name_idx:logic_idx+1]
            logger.info(f"Pasting only these columns: {selected_columns}")
            return df[selected_columns]
        except ValueError:
            logger.warning("Could not locate 'Client Name' to 'Logic' columns. Using full DataFrame.")
            return df

    def format_dataframe(self, df):
        datetime_columns = df.select_dtypes(include=['datetime64']).columns
        for col in datetime_columns:
            df[col] = df[col].dt.strftime('%Y-%m-%d %H:%M:%S')
        return df.fillna("")

    def show_toast(self, message, duration=3000):
        toast = tk.Toplevel(self.root)
        toast.overrideredirect(True)
        toast.configure(bg='black')
        
        # Position bottom-right
        self.root.update_idletasks()
        screen_width = toast.winfo_screenwidth()
        screen_height = toast.winfo_screenheight()
        x = screen_width - 320
        y = screen_height - 100
        toast.geometry(f"300x50+{x}+{y}")

        label = tk.Label(toast, text=message, bg='black', fg='white', font=('Arial', 11))
        label.pack(fill='both', expand=True)

        toast.after(duration, toast.destroy)

    def start_process(self):
        threading.Thread(target=self._start_process_internal).start()
        
    def  _start_process_internal(self):
        self.start_time = time.time()
        self.update_progress(0.05)
        if not self.file1_path or not self.file2_path:
            messagebox.showerror("Error", "Please select both files before proceeding.")
            self.update_progress(0)
            return

        result = messagebox.askyesno("Confirmation", "Are you sure you want to start the repricing process?")
        if not result:
            messagebox.showinfo("Cancelled", "Repricing process has been cancelled.")
            self.update_progress(0)
            return

        try:
            self.update_progress(0.10)
            subprocess.run(["taskkill", "/F", "/IM", "excel.exe"], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            self.update_progress(0.20)
            subprocess.run(['python', 'merge.py', self.file1_path, self.file2_path], check=True)
            self.update_progress(0.50)
            MERGED_FILE = 'merged_file.xlsx'
            self.process_merged_file(MERGED_FILE)
            self.update_progress(0.90)
            # After all processing is done
            self.update_progress(1.0)
        except subprocess.CalledProcessError as e:
            logger.exception("Failed to run merge.py")
            messagebox.showerror("Error", f"Failed to run merge.py: {e}")
            self.update_progress(0)


    def process_merged_file(self, file_path):
        import numpy as np

        try:
            # 1. Start processing
            self.update_progress(0.55)
            open('repricing_log.log', 'w').close()
            logging.basicConfig(filename='repricing_log.log', level=logging.INFO, 
                                format='%(asctime)s - %(levelname)s - %(message)s',
                                datefmt='%Y-%m-%d %H:%M:%S')

            logging.info("Starting merged file processing")
            df = pd.read_excel(file_path)
            logging.info(f"Loaded {len(df)} records from {file_path}")

            # 2. After loading file
            self.update_progress(0.60)

            # Ensure required columns exist
            for col in ['DATEFILLED', 'SOURCERECORDID', 'QUANTITY', 'DAYSUPPLY', 'NDC', 'MemberID', 'Drug Name', 'Pharmacy Name', 'Total AWP (Historical)']:
                if col not in df.columns:
                    raise ValueError(f"Missing required column: {col}")

            df = df.sort_values(by=['DATEFILLED', 'SOURCERECORDID'], ascending=True)
            df['Logic'] = ""
            df['RowID'] = np.arange(len(df))  # Use numpy for speed

            # 3. After initial DataFrame prep
            self.update_progress(0.65)

            # Build a multi-index for fast matching
            df['key'] = df['NDC'].astype(str) + '_' + df['MemberID'].astype(str) + '_' + df['QUANTITY'].abs().astype(str)
            positive_claims = df[(df['QUANTITY'] > 0) & (df['Logic'] == "")]
            positive_lookup = positive_claims.set_index('key')

            unmatched_reversals = []
            matched_count = 0

            def match_conditions(row, candidate):
                return (
                    row['QUANTITY'] == -candidate['QUANTITY'] and
                    row['DAYSUPPLY'] == -candidate['DAYSUPPLY'] and
                    row['NDC'] == candidate['NDC'] and
                    row['MemberID'] == candidate['MemberID'] and
                    row['Drug Name'] == candidate['Drug Name'] and
                    row['Pharmacy Name'] == candidate['Pharmacy Name'] and
                    candidate['DATEFILLED'] <= row['DATEFILLED'] and
                    abs(row['Total AWP (Historical)'] + candidate['Total AWP (Historical)']) < 0.01
                )

            # 4. Matching reversals (can be slow)
            total_reversals = len(df[(df['QUANTITY'] < 0) & (df['Logic'] == "")])
            processed_reversals = 0
            for idx, row in df[(df['QUANTITY'] < 0) & (df['Logic'] == "")].iterrows():
                candidates = df[(df['QUANTITY'] > 0) & (df['Logic'] == "") &
                                (df['DATEFILLED'] <= row['DATEFILLED']) &
                                (df['NDC'] == row['NDC']) &
                                (df['MemberID'] == row['MemberID']) &
                                (df['QUANTITY'] == -row['QUANTITY'])]
                matched = candidates[candidates.apply(lambda c: match_conditions(row, c), axis=1)]

                if not matched.empty:
                    closest = matched.loc[(row['DATEFILLED'] - matched['DATEFILLED']).abs().idxmin()]
                    df.at[idx, 'Logic'] = 'OR'
                    df.at[closest.name, 'Logic'] = 'OR'
                    matched_count += 1
                else:
                    df.at[idx, 'Logic'] = 'OR'
                    unmatched_reversals.append(df.at[idx, 'RowID'])

                # Update progress during matching (optional, for large files)
                processed_reversals += 1
                if total_reversals > 0 and processed_reversals % max(1, total_reversals // 10) == 0:
                    # Progress between 0.65 and 0.75
                    self.update_progress(0.65 + 0.1 * processed_reversals / total_reversals)

            # 5. After matching
            self.update_progress(0.75)

            df_sorted = pd.concat([df[df['Logic'] == ""], df[df['Logic'] == "OR"]])
            output_file = 'merged_file_with_OR.xlsx'
            row_mapping = {row['RowID']: i+2 for i, (_, row) in enumerate(df_sorted.iterrows())}
            excel_rows_to_highlight = [row_mapping[rid] for rid in unmatched_reversals if rid in row_mapping]

            df_sorted.drop(columns=['RowID', 'key'], inplace=True)
            df_sorted.drop_duplicates().to_excel(output_file, index=False)

            # --- Get opportunity name from second column of file1_path ---
            opportunity_name = "claims detail PCU"
            try:
                # Try reading as Excel first, fallback to CSV
                if self.file1_path.lower().endswith('.xlsx'):
                    df_file1 = pd.read_excel(self.file1_path)
                else:
                    df_file1 = pd.read_csv(self.file1_path)
                if df_file1.shape[1] >= 2:
                    # Get the value from the first row, second column
                    raw_name = str(df_file1.iloc[0, 1])
                    # Clean for filename
                    opportunity_name = re.sub(r'[\\/*?:"<>|]', "_", raw_name)
            except Exception as e:
                logger.warning(f"Could not extract opportunity name from file1: {e}")
            # -------------------------------------------------------------

            # Save as CSV 
            df_sorted.drop_duplicates().to_csv(f'{opportunity_name} Claim Detail.csv', index=False)

            with open('unmatched_reversals.txt', 'w') as f:
                f.write(','.join(map(str, excel_rows_to_highlight)))

            # 6. After saving files
            self.update_progress(0.80)

            self.highlight_unmatched_reversals(output_file)
            self.update_progress(0.85)

            messagebox.showinfo("Success", f"Processing complete. File saved as {output_file}")
            self.paste_into_template(output_file)
            self.update_progress(0.90)

        except Exception as e:
            logger.error(f"Error processing merged file: {e}")
            messagebox.showerror("Error", f"Processing failed: {e}")    

    def highlight_unmatched_reversals(self, file_path):
        logging.info("Highlighting unmatched reversals")
        
        # Load the workbook and the specific worksheet
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Define a fill color for highlighting (yellow)
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Read the unmatched reversals from file
        try:
            with open('unmatched_reversals.txt', 'r') as f:
                content = f.read().strip()
                unmatched_rows = list(map(int, content.split(','))) if content else []
        except FileNotFoundError:
            unmatched_rows = []

        # Find the "Logic" column by its header name
        logic_column = None
        for col in sheet.iter_cols(1, sheet.max_column):
            if col[0].value == "Logic":
                logic_column = col[0].column
                break

        if logic_column is None:
            raise ValueError("Logic column not found in the file.")

        # Apply highlighting only to unmatched reversal rows
        for row_num in range(2, sheet.max_row + 1):
            if row_num in unmatched_rows:
                for cell in sheet[row_num]:
                    cell.fill = highlight_fill

        # Save the workbook with the highlighted rows
        workbook.save(file_path)
        logging.info(f"Applied highlighting to {len(unmatched_rows)} rows in {file_path}")
        
        # Clean up the temporary file
        try:
            os.remove('unmatched_reversals.txt')
        except FileNotFoundError:
            pass
            
        logging.info("Completed highlighting process")
        

    def open_file(self, file_path):
        try:
            os.startfile(file_path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open file: {str(e)}")

    def sharx_lbl(self):
        try:
            subprocess.run(['python', 'sharx_lbl.py'], check=True)
        except subprocess.CalledProcessError as e:
            logger.exception("SHARx LBL generation failed")
            messagebox.showerror("Error", f"SHARx LBL generation failed: {e}")

    def epls_lbl(self):
        try:
            subprocess.run(['python', 'epls_lbl.py'], check=True)
        except subprocess.CalledProcessError as e:
            logger.exception("EPLS LBL generation failed")
            messagebox.showerror("Error", f"EPLS LBL generation failed: {e}")

    def check_template(self, file_path):
        # Add your logic here
        print(f"Checking template for: {file_path}")

    def apply_theme_colors(self, colors):
        # Main frames
        self.root.configure(fg_color=colors["dark_blue"])
        self.button_frame.configure(fg_color=colors["grey_blue"])
        self.notes_frame.configure(fg_color=colors["grey_blue"])
        self.dis_frame.configure(fg_color=colors["grey_blue"])
        self.prog_frame.configure(fg_color=colors["grey_blue"])

        # Buttons in button_frame
        button_widgets = [
            "file1_button", "file2_button", "template_button",
            "cancel_button", "logs_button", "toggle_theme_button",
            "sharx_lbl_button", "epls_lbl_button", "start_process_button"
        ]
        for btn_name in button_widgets:
            btn = getattr(self, btn_name, None)
            if btn:
                btn.configure(fg_color=colors["mint"], text_color="#000000")
        # Exit button (darker red)
        if hasattr(self, "exit_button"):
            self.exit_button.configure(fg_color=colors["button_red"], text_color="#000000")

        # Start Disruption button (outside button_frame)
        if hasattr(self, "start_disruption_button"):
            self.start_disruption_button.configure(fg_color=colors["mint"], text_color="#000000")

        # Disruption type combobox
        if hasattr(self, "disruption_type_combobox"):
            self.disruption_type_combobox.configure(
                fg_color=colors["grey_blue"],
                dropdown_fg_color=colors["grey_blue"],
                dropdown_hover_color=colors["mint"],
                button_color=colors["mint"],
                button_hover_color=colors["mint"],
                border_color=colors["mint"],
                text_color="#000000" if colors is LIGHT_COLORS else "#FFFFFF"
            )

        # Progress frame and label
        if hasattr(self, "prog_frame"):
            self.prog_frame.configure(fg_color=colors["grey_blue"])
        if hasattr(self, "progress_label"):
            self.progress_label.configure(bg_color=colors["grey_blue"], text_color="#000000")

if __name__ == "__main__":
    ctk.set_appearance_mode("light")  # Start in light mode
    root = ctk.CTk()  # or tk.Tk() if not using customtkinter
    app = App(root)
    root.mainloop()