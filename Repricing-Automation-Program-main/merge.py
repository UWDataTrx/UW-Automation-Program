
import sys
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
import logging

# Configure logging
logging.basicConfig(
    filename='merge_log.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

MERGED_FILENAME = "merged_file.xlsx"
REQUIRED_COLUMNS = [
    'DATEFILLED', 'SOURCERECORDID', 'QUANTITY', 'DAYSUPPLY', 'NDC',
    'MemberID', 'Drug Name', 'Pharmacy Name', 'Total AWP (Historical)'
]

def merge_files(file1_path, file2_path):
    try:
        file1 = Path(file1_path)
        file2 = Path(file2_path)

        logger.info(f"Starting merge: {file1} + {file2}")

        if not file1.exists():
            logger.error(f"File not found: {file1}")
            return
        if not file2.exists():
            logger.error(f"File not found: {file2}")
            return

        # Load data
        df1 = pd.read_csv(file1, parse_dates=['DATEFILLED'], dayfirst=False)
        df2 = pd.read_csv(file2) if file2.suffix == ".csv" else pd.read_excel(file2)

        # Clean up and standardize column names
        df2.columns = [col.strip() for col in df2.columns]
        if 'Source Record ID' in df2.columns:
            df2.rename(columns={"Source Record ID": "SOURCERECORDID"}, inplace=True)

        # Merge
        df_merged = pd.merge(df1, df2, on="SOURCERECORDID", how="outer")
        if 'Total AWP (Historical)' in df_merged.columns:
            df_merged['Total AWP (Historical)'] = pd.to_numeric(df_merged['Total AWP (Historical)'], errors='coerce').round(2)
        else:
            df_merged['Total AWP (Historical)'] = 0.00
        df_merged['MemberID'] = df_merged['MemberID'].fillna(0)

        # Log missing required columns
        for col in REQUIRED_COLUMNS:
            if col not in df_merged.columns:
                logger.warning(f"Missing expected column: {col}")

        merged_path = Path.cwd() / MERGED_FILENAME
        df_merged.to_excel(merged_path, index=False)
        logger.info(f"Merged file saved to: {merged_path}")

        # Apply Excel formatting
        try:
            wb = load_workbook(merged_path)
            ws = wb.active
            date_style = NamedStyle(name="date_style", number_format='MM/DD/YYYY')

            if ws is not None and ws.max_row >= 1:
                header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                if 'DATEFILLED' in header:
                    date_col_index = header.index('DATEFILLED') + 1
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=date_col_index).style = date_style
                    wb.save(merged_path)
                    logger.info("Applied date formatting successfully.")
                else:
                    logger.warning("DATEFILLED column not found for formatting.")
            else:
                logger.warning("Worksheet is empty or not loaded, cannot apply formatting.")

        except Exception as ex:
            logger.warning(f"Failed to apply formatting: {ex}")

    except Exception as e:
        logger.exception(f"Merge failed: {e}")
        sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python merge.py <file1_path> <file2_path>")
        sys.exit(1)

    merge_files(sys.argv[1], sys.argv[2])
