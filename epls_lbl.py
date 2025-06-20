import pandas as pd
import tkinter as tk
from tkinter import messagebox
import logging
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from utils import (
    load_file_paths,
    write_shared_log
)

# Setup logging
logging.basicConfig(
    filename='epls_lbl.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

def show_message(awp, ing, total, rxs):
    messagebox.showinfo(
        'Process Complete',
        f'EPLS LBL has been created\n\n'
        f'Total AWP: ${awp:.2f}\n\n'
        f'Total Ing Cost: ${ing:.2f}\n\n'
        f'Total Gross Cost: ${total:.2f}\n\n'
        f'Total Claim Count: {rxs}'
    )

def main():
    root = tk.Tk()
    root.withdraw()
    write_shared_log("epls_lbl.py", "Processing started.")
    try:
        paths = load_file_paths()
        template_path = Path(paths['epls'])

        df = pd.read_excel(paths['reprice'], sheet_name='Claims Table')

        # Debug print and log
        print("Columns in claims DataFrame:")
        print(df.columns)
        logger.info(f"Columns in claims DataFrame: {df.columns.tolist()}")

        df['Logic'] = pd.to_numeric(df['Logic'], errors='coerce')
        df = df[df['Logic'].between(1, 10)]

        awp = df['Total AWP (Historical)'].sum()
        ing = df['Rx Sense Ing Cost'].sum()
        total = df['RxSense Total Cost'].sum()
        rxs = df['Rxs'].sum()

        columns_to_keep = [
            'MONY', 'Rxs', 'Rx Sense Ing Cost', 'RxSense Dispense Fee', 'RxSense Total Cost',
            'Total AWP (Historical)', 'Pharmacy Name', 'INPUTFILECHANNEL', 'DATEFILLED',
            'MemberID', 'DAYSUPPLY', 'QUANTITY', 'NDC', 'DST Drug Name', 'GrossCost',
            'Universal Rebates', 'Exclusive Rebates', 'Specialty Vlookup'
        ]
        df = df[columns_to_keep]

        df['Specialty Vlookup'] = df['Specialty Vlookup'].apply(lambda x: 'N' if x == 'No' else 'Y')

        wb = load_workbook(template_path)
        ws = wb['Line By Line']

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        output_path = Path('_Rx Claims for EPLS.xlsx')
        wb.save(output_path)

        logger.info("EPLS LBL file created successfully.")
        write_shared_log("epls_lbl.py", "EPLS LBL file created successfully.")
        show_message(awp, ing, total, rxs)

    except Exception as e:
        logger.exception("An error occurred during EPLS LBL processing")
        write_shared_log("epls_lbl.py", f"An error occurred: {e}", status="ERROR")
        messagebox.showerror("Error", f"An error occurred: {e}")
    finally:
        root.quit()

if __name__ == '__main__':
    main()
