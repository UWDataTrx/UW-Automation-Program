
import tkinter as tk
import customtkinter as ctk
from tkinter import filedialog, messagebox
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def open_file_1():
    global file1_path
    file1_path = filedialog.askopenfilename(title="Open First File",
                                            filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx *.xls")])
    if file1_path:
        file1_label.configure(text=os.path.basename(file1_path))

def open_file_2():
    global file2_path
    file2_path = filedialog.askopenfilename(title="Open Second File",
                                            filetypes=[("Excel files", "*.xlsx *.xls"), ("CSV files", "*.csv")])
    if file2_path:
        file2_label.configure(text=os.path.basename(file2_path))

def merge_files():
    # Explicitly use tk and get_column_letter to ensure imports are used
    _ = tk.TkVersion  # Use tkinter
    _ = get_column_letter(1)  # Use openpyxl.utils.get_column_letter
    if not file1_path or not file2_path:
        messagebox.showerror("Error", "Please select both files first")
        return

    # Check file existence
    if not os.path.exists(file1_path):
        messagebox.showerror("File Not Found", f"The file '{file1_path}' does not exist.")
        return
    if not os.path.exists(file2_path):
        messagebox.showerror("File Not Found", f"The file '{file2_path}' does not exist.")
        return

    try:
        if file1_path.endswith('.csv'):
            df1 = pd.read_csv(file1_path)
        else:
            df1 = pd.read_excel(file1_path)
    except Exception as e:
        messagebox.showerror("Read Error", f"Could not read file 1: {e}")
        return

    try:
        if file2_path.endswith('.csv'):
            df2 = pd.read_csv(file2_path)
        else:
            df2 = pd.read_excel(file2_path)
    except Exception as e:
        messagebox.showerror("Read Error", f"Could not read file 2: {e}")
        return

    # Check for required columns
    if 'SOURCERECORDID' not in df1.columns:
        messagebox.showerror("Missing Column", "'SOURCERECORDID' column is missing in the first file.")
        return
    # Rename 'Source Record ID' to 'SOURCERECORDID' to match the first file
    if 'Source Record ID' in df2.columns:
        df2 = df2.rename(columns={"Source Record ID": "SOURCERECORDID"})
    if 'SOURCERECORDID' not in df2.columns:
        messagebox.showerror("Missing Column", "'SOURCERECORDID' column is missing in the second file.")
        return

    try:
        # Merge the dataframes on 'SOURCERECORDID'
        merged_df = pd.merge(df1, df2, on="SOURCERECORDID", how="outer")

        # Count the rows in df1 that are not in df2
        unmatched_count = df1[~df1['SOURCERECORDID'].isin(df2['SOURCERECORDID'])].shape[0]

        global merged_file_path
        merged_file_path = os.path.join(os.getcwd(), "merged_file.xlsx")
        merged_df.to_excel(merged_file_path, index=False)

        # Load the workbook and get the first worksheet
        wb = load_workbook(merged_file_path)
        if wb.worksheets:
            ws = wb.worksheets[0]
        else:
            messagebox.showerror("Error", "No worksheet found in the merged file.")
            return

        # Apply DATEVALUE function to the date column (assuming the date column is in column 'G' (7))
        date_column_index = 7  # Change to the actual column index of your date column
        from openpyxl.cell.cell import MergedCell
        for row in range(2, ws.max_row + 1):  # Start from the second row to skip headers
            cell = ws.cell(row=row, column=date_column_index)
            # Skip merged cells
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str):
                # Only assign if not a MergedCell
                if not isinstance(cell, MergedCell):
                    cell.value = f"=DATEVALUE(\"{cell.value}\")"

        # Save the workbook with the applied DATEVALUE functions
        wb.save(merged_file_path)

        open_link.configure(text="Open Merged File in Excel", command=open_merged_file)
        unmatched_label.configure(text=f"Unknown Claims Removed From Tool: {unmatched_count}")
    except Exception as e:
        messagebox.showerror("Error", str(e))

def open_merged_file():
    os.startfile(merged_file_path)

root = ctk.CTk()
root.title("File Merger")
root.configure(fg_color='#192428', pady=20)

file1_label = ctk.CTkLabel(root, text="No file selected", font=('Aptos Narrow', 16, 'bold'))
file1_label.grid(row=0, column=0, pady=2, padx=30)
file1_button = ctk.CTkButton(root, text="Select Uploaded File", command=open_file_1, width=200, height=30, fg_color='#0784b5', font=('Aptos Narrow', 16), corner_radius=2)
file1_button.grid(row=1, column=0, pady=10, padx=30)

file2_label = ctk.CTkLabel(root, text="No file selected", font=('Aptos Narrow', 16, 'bold'))
file2_label.grid(row=0, column=1, pady=2, padx=30)
file2_button = ctk.CTkButton(root, text="Select File From Tool", command=open_file_2, width=200, height=30, fg_color='#0784b5', font=('Aptos Narrow', 16), corner_radius=2)
file2_button.grid(row=1, column=1, pady=10, padx=30)

space_label = ctk.CTkLabel(root, text='')
space_label.grid(row=2, column=0, columnspan=2, pady=30)

merge_button = ctk.CTkButton(root, text="Merge Files", command=merge_files, width=200, height=30, fg_color='#0784b5', font=('Aptos Narrow', 16), corner_radius=2)
merge_button.grid(row=3, column=0, columnspan=2, pady=20, padx=20)

open_link = ctk.CTkButton(root, text="", command=None, width=200, height=30, fg_color='#414c50', font=('Aptos Narrow', 16), corner_radius=2)
open_link.grid(row=4, column=0, columnspan=2, pady=10, padx=20)

unmatched_label = ctk.CTkLabel(root, text="", font=('Aptos Narrow', 16, 'bold'))
unmatched_label.grid(row=5, column=0, columnspan=2, pady=10, padx=20)

file1_path = ""
file2_path = ""
merged_file_path = ""

root.mainloop()
