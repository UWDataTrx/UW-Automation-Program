import unittest
import pandas as pd
from pathlib import Path
from utils.excel_utils import write_df_to_template
import openpyxl


class TestTemplatePasting(unittest.TestCase):
    def setUp(self):
        # Create a dummy template Excel file
        self.template_path = Path("test_template.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "Sheet1"
            ws.cell(row=1, column=1, value="Header1")
            ws.cell(row=1, column=2, value="Header2")
        wb.save(self.template_path)
        # Create a dummy DataFrame
        self.df = pd.DataFrame({"Header1": [1, 2], "Header2": [3, 4]})
        self.output_path = Path("test_output.xlsx")

    def tearDown(self):
        # Clean up test files
        if self.template_path.exists():
            self.template_path.unlink()
        if self.output_path.exists():
            self.output_path.unlink()

    def test_write_df_to_template(self):
        # Paste DataFrame into template
        write_df_to_template(
            template_path=self.template_path,
            output_path=self.output_path,
            sheet_name="Sheet1",
            df=self.df,
            start_cell="A2",
            header=False,
            index=False,
            visible=False,
            open_file=False,
        )
        # Verify output
        wb = openpyxl.load_workbook(self.output_path)
        ws = wb["Sheet1"]
        self.assertEqual(ws["A2"].value, 1)
        self.assertEqual(ws["B2"].value, 3)
        self.assertEqual(ws["A3"].value, 2)
        self.assertEqual(ws["B3"].value, 4)


if __name__ == "__main__":
    unittest.main()
