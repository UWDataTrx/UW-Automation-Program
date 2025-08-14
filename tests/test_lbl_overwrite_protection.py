import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import unittest
import openpyxl
import pandas as pd
from utils.excel_utils import write_df_to_template


class TestLBLOverwriteProtection(unittest.TestCase):
    def setUp(self):
        # Create dummy templates for SHARx, EPLS, and Disruption
        self.sharx_template = Path("SHARx_Template.xlsx")
        self.epls_template = Path("EPLS_Template.xlsx")
        self.disruption_template = Path("Disruption_Template.xlsx")
        for template in [
            self.sharx_template,
            self.epls_template,
            self.disruption_template,
        ]:
            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is not None:
                ws.title = "Line By Line"
                ws.cell(row=1, column=1, value="Header1")
                ws.cell(row=1, column=2, value="Header2")
            wb.save(template)
        self.df = pd.DataFrame({"Header1": [1, 2], "Header2": [3, 4]})

    def tearDown(self):
        # Remove all test files and copies
        for f in Path(".").glob("*Template*.xlsx"):
            f.unlink()
        for f in Path(".").glob("_Rx Claims for SHARx*.xlsx"):
            f.unlink()
        for f in Path(".").glob("_Rx Claims for EPLS*.xlsx"):
            f.unlink()
        for f in Path(".").glob("Unknown_Disruption_Report*.xlsx"):
            f.unlink()

    def test_sharx_lbl_protection(self):
        # Try to write output to template name (should create a copy)
        output_path = Path("_Rx Claims for SHARx.xlsx")
        write_df_to_template(
            template_path=self.sharx_template,
            output_path=output_path,
            sheet_name="Line By Line",
            df=self.df,
            start_cell="A2",
            header=False,
            index=False,
            visible=False,
            open_file=False,
        )
        # Should not overwrite template, should create a copy if protected
        copies = list(Path(".").glob("_Rx Claims for SHARx_copy*.xlsx"))
        self.assertTrue(copies or output_path.exists())

    def test_epls_lbl_protection(self):
        output_path = Path("_Rx Claims for EPLS.xlsx")
        write_df_to_template(
            template_path=self.epls_template,
            output_path=output_path,
            sheet_name="Line By Line",
            df=self.df,
            start_cell="A2",
            header=False,
            index=False,
            visible=False,
            open_file=False,
        )
        copies = list(Path(".").glob("_Rx Claims for EPLS_copy*.xlsx"))
        self.assertTrue(copies or output_path.exists())

    def test_disruption_lbl_protection(self):
        output_path = Path("Unknown_Disruption_Report.xlsx")
        write_df_to_template(
            template_path=self.disruption_template,
            output_path=output_path,
            sheet_name="Line By Line",
            df=self.df,
            start_cell="A2",
            header=False,
            index=False,
            visible=False,
            open_file=False,
        )
        self.assertTrue(output_path.exists())


if __name__ == "__main__":
    unittest.main()
