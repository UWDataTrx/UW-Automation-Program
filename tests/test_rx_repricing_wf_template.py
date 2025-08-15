import unittest
from pathlib import Path

import openpyxl
import pandas as pd

from utils.excel_utils import write_df_to_template


class TestRxRepricingWFTemplate(unittest.TestCase):
    def setUp(self):
        # Create a dummy _Rx Repricing_wf.xlsx template
        self.template_path = Path("_Rx Repricing_wf.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "Sheet1"
            ws.cell(row=1, column=1, value="Header1")
            ws.cell(row=1, column=2, value="Header2")
        wb.save(self.template_path)
        # Create a dummy DataFrame
        self.df = pd.DataFrame({"Header1": [10, 20], "Header2": [30, 40]})
        self.output_path = Path("_Rx Repricing_wf.xlsx")

    def tearDown(self):
        # Clean up test files and any copies
        for f in Path(".").glob("_Rx Repricing_wf*.xlsx"):
            f.unlink()

    def test_write_df_to_template_actual_file(self):
        # Paste DataFrame into template
        write_df_to_template(
            template_path=self.template_path,
            output_path=self.output_path,
            sheet_name="Sheet1",
            df=self.df,
            start_cell="A2",
            header=False,
            index=False,
            visible=False,
            open_file=False,
        )
        # Check if a copy was made
        copies = list(Path(".").glob("_Rx Repricing_wf_copy*.xlsx"))
        if copies:
            # If a copy was made, check the copy
            output_file = copies[0]
        else:
            output_file = self.output_path
        # Verify output
        wb = openpyxl.load_workbook(output_file)
        ws = wb["Sheet1"]
        # Check that the DataFrame is not empty
        self.assertFalse(self.df.empty, "DataFrame is empty!")
        # Check that data is present in expected cells
        self.assertEqual(ws["A2"].value, 10)
        self.assertEqual(ws["B2"].value, 30)
        self.assertEqual(ws["A3"].value, 20)
        self.assertEqual(ws["B3"].value, 40)
        # Diagnostic: If cells are blank, print sheet values
        if ws["A2"].value is None or ws["B2"].value is None:
            print("Diagnostic dump:")
            for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=2):
                print([cell.value for cell in row])


if __name__ == "__main__":
    unittest.main()

if __name__ == "__main__":
    unittest.main()
