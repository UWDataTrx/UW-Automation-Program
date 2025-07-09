 import tkinter as tk
import customtkinter as ctk
from tkinter import filedialog, messagebox, scrolledtext
import subprocess
import os
from utils.utils import write_shared_log
import logging
import threading
import multiprocessing
import pandas as pd
from typing import Optional
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pathlib import Path
import json
import time
import re
import importlib
import importlib.util
import warnings

# Import custom modules
# Theme colors now handled by UIBuilder
from config.app_config import ProcessingConfig, AppConstants
from modules.file_processor import FileProcessor
from modules.template_processor import TemplateProcessor
from modules.data_processor import DataProcessor
from modules.process_manager import ProcessManager
from modules.ui_builder import UIBuilder
from modules.log_manager import LogManager, ThemeController

# Excel COM check
XLWINGS_AVAILABLE = importlib.util.find_spec("xlwings") is not None
EXCEL_COM_AVAILABLE = importlib.util.find_spec("win32com.client") is not None

# Logging setup
logging.basicConfig(
    filename="repricing_log.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


class ConfigManager:
    def __init__(self):
        self.config = {}
        if AppConstants.CONFIG_FILE.exists():
            self.load()
        else:
            self.save_default()

    def save_default(self):
        self.config = {"last_folder": str(Path.cwd())}
        self.save()

    def load(self):
        with open(AppConstants.CONFIG_FILE, "r") as f:
            self.config = json.load(f)

    def save(self):
        with open(AppConstants.CONFIG_FILE, "w") as f:
            json.dump(self.config, f, indent=4)


class App:
    def __init__(self, root):
        self.root = root
        self._initialize_variables()
        self._initialize_processors()
        self.ui_builder.build_complete_ui()
        self.theme_controller.apply_initial_theme()
        self.log_manager.initialize_logging()
        self.log_manager.log_application_start()

    def _initialize_variables(self):
        """Initialize all instance variables."""
        self.file1_path = None
        self.file2_path = None
        self.template_file_path = None
        self.cancel_event = threading.Event()
        self.start_time = None
        self.progress_var = tk.DoubleVar(value=0.0)
        self.progress_label_var = tk.StringVar(value="0%")
        self.config_manager = ConfigManager()
        self.selected_disruption_type = tk.StringVar(value="Tier")
        self.file1_label: Optional[ctk.CTkLabel] = None
        self.file2_label: Optional[ctk.CTkLabel] = None
        self.template_label: Optional[ctk.CTkLabel] = None
        self.toggle_theme_button: Optional[ctk.CTkButton] = None
        self.progress_bar: Optional[ctk.CTkProgressBar] = None
        self.progress_label: Optional[ctk.CTkLabel] = None

    def _initialize_processors(self):
        """Initialize all processor and manager instances."""
        self.file_processor = FileProcessor(self)
        self.template_processor = TemplateProcessor(self)
        self.data_processor = DataProcessor(self)
        self.process_manager = ProcessManager(self)
        self.ui_builder = UIBuilder(self)
        self.log_manager = LogManager(self)
        self.theme_controller = ThemeController(self)

    # The following methods are moved to their respective manager classes for better cohesion:
    # - apply_theme_colors -> ThemeController
    # - check_template -> FileProcessor
    # - sharx_lbl, epls_lbl -> ProcessManager
    # - show_shared_log_viewer -> LogManager

    # Example: Remove apply_theme_colors from App, and use self.theme_controller.apply_theme_colors instead.



    def import_file1(self):
        """Import the first file with template validation using guard clauses."""
        file_path = self._get_file_path("Select File Uploaded to Tool")
        if not file_path:
            return  # User cancelled
        
        self._set_file1_path(file_path)
        self._validate_gross_cost_template(file_path)

    def _get_file_path(self, title):
        """Get file path from file dialog."""
        return filedialog.askopenfilename(
            title=title,
            filetypes=ProcessingConfig.FILE_TYPES,
        )

    def _set_file1_path(self, file_path):
        """Set the file1 path and update UI."""
        self.file1_path = file_path
        if self.file1_label:
            self.file1_label.configure(text=os.path.basename(file_path))
        self.file_processor.check_template(file_path)
        write_shared_log("File1 imported", file_path)

    def _validate_gross_cost_template(self, file_path):
        """Validate GrossCost column and suggest template type using data processor."""
        template_suggestion = self.data_processor.validate_gross_cost_template(file_path)
        if template_suggestion:
            messagebox.showinfo("Template Selection", template_suggestion)

    def import_file2(self):
        """Import the second file."""
        file_path = self._get_file_path("Select File From Tool")
        if not file_path:
            return  # User cancelled
        
        self.file2_path = file_path
        if self.file2_label:
            self.file2_label.configure(text=os.path.basename(file_path))
        write_shared_log("File2 imported", file_path)

    def import_template_file(self):
        """Import the template file."""
        file_path = filedialog.askopenfilename(
            title="Select Template File", filetypes=ProcessingConfig.TEMPLATE_FILE_TYPES
        )
        if not file_path:
            return  # User cancelled
        
        self.template_file_path = file_path
        if self.template_label:
            self.template_label.configure(text=os.path.basename(file_path))
        write_shared_log("Template file imported", file_path)

    # Logging and notification methods
    # Removed duplicate write_audit_log method to resolve method name conflict.

    # Cancel during repricing
    def cancel_process(self):
        """Cancel the process using process manager."""
        self.process_manager.cancel_process()

    # Live log viewer (old version, renamed to avoid conflict)
    def show_log_viewer_old(self):
        win = tk.Toplevel(self.root)
        win.title("Log Viewer")
        txt = scrolledtext.ScrolledText(win, width=100, height=30)
        txt.pack(fill="both", expand=True)

        def refresh():
            with open("repricing_log.log", "r") as f:
                txt.delete("1.0", tk.END)
                txt.insert(tk.END, f.read())
            win.after(3000, refresh)

        refresh()

    def update_progress(self, value=None, message=None):
        """Update the progress bar and label with reduced complexity."""
        def do_update():
            if value is None:
                self._set_indeterminate_progress(message)
            else:
                self._set_determinate_progress(value, message)
            self.root.update_idletasks()

        self._schedule_ui_update(do_update)

    def _set_indeterminate_progress(self, message):
        """Set progress bar to indeterminate mode."""
        if self.progress_bar:
            self.progress_bar.configure(mode="indeterminate")
            self.progress_bar.start()
        display_message = message or "Processing... (unknown duration)"
        self.progress_label_var.set(display_message)

    def _set_determinate_progress(self, value, message):
        """Set progress bar to determinate mode with specific value."""
        if self.progress_bar:
            if self.progress_bar.cget("mode") != "determinate":
                self.progress_bar.stop()
                self.progress_bar.configure(mode="determinate")
            
            self.progress_bar.set(value)
        
        self.progress_var.set(value)
        
        if message:
            self.progress_label_var.set(message)
        else:
            self._set_calculated_progress_message(value)

    def _set_calculated_progress_message(self, value):
        """Calculate and set progress message with time estimates."""
        percent = int(value * 100)
        elapsed = time.time() - self.start_time if self.start_time else 0
        est = int((elapsed / value) * (1 - value)) if value > 0 else 0
        self.progress_label_var.set(f"Progress: {percent}% | Est. {est}s left")

    def _schedule_ui_update(self, update_func):
        """Schedule UI update on main thread or execute immediately."""
        if threading.current_thread() is threading.main_thread():
            update_func()
        else:
            self.root.after(0, update_func)

    def write_audit_log(self, file1, file2, status):
        """Write audit log entry using file processor."""
        return self.file_processor.write_audit_log(file1, file2, status)

    def show_log_viewer(self):
        """Show log viewer using log manager."""
        self.log_manager.show_log_viewer()

    # Disruption and process methods
    # Removed select_disruption_type method since disruption_type_combobox does not exist.

    def start_disruption(self, disruption_type=None):
        """Start disruption processing using process manager."""
        self.process_manager.start_disruption(disruption_type)

    def start_process_threaded(self):
        """Start the repricing process using the process manager."""
        self.process_manager.start_process_threaded()

    def finish_notification(self):
        """Show completion notification using process manager."""
        self.process_manager.finish_notification()

    # Repricing workflow methods
    def paste_into_template(self, processed_file):
        """Paste processed data into Excel template using background threading."""
        def run_in_background():
            try:
                self._execute_template_paste(processed_file)
            except Exception as e:
                logger.exception("Error during paste with xlwings")
                self.root.after(
                    0,
                    lambda e=e: messagebox.showerror(
                        "Error", f"Template update failed:\n{e}"
                    ),
                )
                self.root.after(0, lambda: self.update_progress(0))

        threading.Thread(target=run_in_background, daemon=True).start()

    def _execute_template_paste(self, processed_file):
        """Execute the template paste operation with proper error handling."""
        import time

        start_time = time.time()
        
        # Initialize progress
        self.root.after(
            0,
            lambda: self.update_progress(None, "Preparing to paste into template..."),
        )

        # Validate template path
        if not self.template_file_path:
            raise ValueError("Template file path is not set.")

        # Prepare data and paths
        paste_data = self._prepare_template_data(processed_file)
        paths = self._prepare_template_paths()
        
        # Create backup and setup output file
        self._create_template_backup(paths)
        
        # Execute Excel operations
        self._execute_excel_paste(paste_data, paths)
        
        # Finalize and notify
        elapsed = time.time() - start_time
        msg = f"Template updated successfully in {elapsed:.2f} seconds."
        logger.info(msg)
        self.root.after(0, lambda: self.update_progress(1.0, msg))
        self.root.after(0, lambda: self.show_toast(msg))
        self.root.after(
            0,
            lambda: messagebox.showinfo(
                "Template Update Complete",
                "Pasting into the template is complete. You may now review the updated file.",
            ),
        )

    def _prepare_template_data(self, processed_file):
        """Prepare data for template pasting."""
        df = pd.read_excel(processed_file)
        df = self.format_dataframe(df)
        return {
            "data": df.values,
            "nrows": df.shape[0],
            "ncols": df.shape[1]
        }

    def _prepare_template_paths(self):
        """Prepare file paths for template operations using file processor."""
        return self.file_processor.prepare_file_paths(self.template_file_path)

    def _create_template_backup(self, paths):
        """Create backup of template and prepare output file using template processor."""
        self.template_processor.create_template_backup(paths)

    def _execute_excel_paste(self, paste_data, paths):
        """Execute the Excel paste operation."""
        import xlwings as xw
        
        # Start Excel session
        app = xw.App(visible=False)
        wb = app.books.open(str(paths["output"]))
        ws = wb.sheets["Claims Table"]

        try:
            # Batch read formulas and prepare data
            formulas = ws.range((2, 1), (paste_data["nrows"] + 1, paste_data["ncols"])).formula
            data_to_write = self._prepare_excel_data(paste_data, formulas)
            
            # Paste values with progress updates
            self._paste_data_with_progress(ws, data_to_write, paste_data["nrows"], paste_data["ncols"])
            
            # Save and close
            wb.save()
            wb.close()
            app.quit()
            
        except Exception as e:
            # Ensure Excel is closed even on error
            try:
                wb.close()
                app.quit()
            except Exception:
                pass
            raise e

    def _prepare_excel_data(self, paste_data, formulas):
        """Prepare data for Excel, preserving formulas."""
        data_to_write = []
        
        for i in range(paste_data["nrows"]):
            row = []
            for j in range(paste_data["ncols"]):
                if formulas[i][j] == "":
                    row.append(paste_data["data"][i][j])
                else:
                    row.append(None)
            data_to_write.append(row)
        
        return data_to_write

    def _paste_data_with_progress(self, ws, data_to_write, nrows, ncols):
        """Paste data to Excel with progress updates."""
        # Paste values
        ws.range((2, 1), (nrows + 1, ncols)).value = data_to_write
        
        # Update progress periodically
        for i in range(0, nrows, 250):
            percent = 0.94 + 0.04 * (i / max(1, nrows))
            msg = f"Pasting row {i + 1} of {nrows}..."
            self.root.after(
                0, lambda v=percent, m=msg: self.update_progress(v, m)
            )

    def filter_template_columns(self, df):
        try:
            # Ensure 'ClientName' and 'Logic' columns exist and are in the correct order
            if "ClientName" in df.columns and "Logic" in df.columns:
                client_name_idx = df.columns.get_loc("ClientName")
                logic_idx = df.columns.get_loc("Logic")
                if client_name_idx <= logic_idx:
                    # Select columns from 'ClientName' to 'Logic' (inclusive)
                    selected_columns = df.columns[client_name_idx : logic_idx + 1]
                    logger.info(
                        f"Pasting only these columns: {selected_columns.tolist()}"
                    )
                    return df[selected_columns]
                else:
                    logger.warning(
                        "'Logic' column appears before 'ClientName'; returning full DataFrame."
                    )
                    return df
            else:
                raise ValueError(
                    "Required columns 'ClientName' or 'Logic' are missing."
                )
        except Exception as e:
            logger.warning(f"Error filtering columns: {e}. Using full DataFrame.")
            return df

    def format_dataframe(self, df):
        """Format DataFrame using template processor."""
        return self.template_processor.format_dataframe(df)

    def show_toast(self, message, duration=3000):
        """Show toast notification using template processor."""
        return self.template_processor.show_toast(message, duration)

    def start_process(self):
        threading.Thread(target=self._start_process_internal).start()
        write_shared_log("Repricing process started", "")

    def validate_merge_inputs(self):
        """Validate merge inputs using data processor."""
        is_valid, message = self.data_processor.validate_merge_inputs(
            self.file1_path, self.file2_path
        )
        if not is_valid:
            messagebox.showerror("Error", message)
        return is_valid

    def _start_process_internal(self):
        self.start_time = time.time()
        self.update_progress(0.05)

        # Extra safeguard: Remove any accidental LBL/disruption output during repricing
        os.environ["NO_LBL_OUTPUT"] = "1"

        if not self.file1_path or not self.file2_path:
            self.update_progress(0)
            messagebox.showerror("Error", "Please select both files before proceeding.")
            return

        if not self.validate_merge_inputs():
            self.update_progress(0)
            return

        try:
            self.update_progress(0.10)
            subprocess.run(
                ["taskkill", "/F", "/IM", "excel.exe"],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )
            self.update_progress(0.20)
            subprocess.run(
                ["python", "merge.py", self.file1_path, self.file2_path], check=True
            )
            self.update_progress(0.50)
            MERGED_FILE = "merged_file.xlsx"
            self.process_merged_file(MERGED_FILE)
            self.update_progress(0.90)
            # After all processing is done
            self.update_progress(1.0)
            # Ensure LBL scripts are NOT called here or in process_merged_file
        except subprocess.CalledProcessError as e:
            self.update_progress(0)
            logger.exception("Failed to run merge.py")
            messagebox.showerror("Error", f"Failed to run merge.py: {e}")

    def process_merged_file(self, file_path):
        """Process merged file with reduced complexity using helper methods."""
        try:
            # Initialize processing
            self._initialize_processing()
            
            # Load and validate data
            df = self._load_and_validate_data(file_path)
            
            # Process data using multiprocessing
            processed_df = self._process_data_multiprocessing(df)
            
            # Save outputs
            output_file = self._save_processed_outputs(processed_df)
            
            # Finalize processing
            self._finalize_processing(output_file)
            
        except Exception as e:
            logger.error(f"Error processing merged file: {e}")
            messagebox.showerror("Error", f"Processing failed: {e}")

    def _initialize_processing(self):
        """Initialize the processing environment."""
        self.update_progress(0.55)
        open("repricing_log.log", "w").close()
        logging.basicConfig(
            filename="repricing_log.log",
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
        logging.info("Starting merged file processing")

    def _load_and_validate_data(self, file_path):
        """Load and validate the merged file data."""
        df = pd.read_excel(file_path)
        logging.info(f"Loaded {len(df)} records from {file_path}")
        self.update_progress(0.60)

        # Validate required columns using configuration
        ProcessingConfig.validate_required_columns(df)

        # Prepare data for processing
        df = df.sort_values(by=["DATEFILLED", "SOURCERECORDID"], ascending=True)
        df["Logic"] = ""
        df["RowID"] = np.arange(len(df))
        
        return df

    def _process_data_multiprocessing(self, df):
        """Process data using multiprocessing for improved performance."""
        self.update_progress(0.65)
        
        # Import and use multiprocessing helpers
        from modules import mp_helpers

        num_workers = ProcessingConfig.get_multiprocessing_workers()
        df_blocks = np.array_split(df, num_workers)
        out_queue = multiprocessing.Queue()
        processes = []
        
        # Start worker processes
        for block in df_blocks:
            p = multiprocessing.Process(
                target=mp_helpers.worker, args=(block, out_queue)
            )
            p.start()
            processes.append(p)
        
        # Collect results
        results = [out_queue.get() for _ in processes]
        for p in processes:
            p.join()
        
        processed_df = pd.concat(results)
        self.update_progress(0.75)
        
        return processed_df

    def _save_processed_outputs(self, df):
        """Save processed data to various output formats."""
        # Sort and filter data
        df_sorted = pd.concat([df[df["Logic"] == ""], df[df["Logic"] == "OR"]])
        
        # Prepare output directory and files
        output_dir = Path.cwd()
        output_file = output_dir / "merged_file_with_OR.xlsx"
        
        # Create row mapping for highlighting
        row_mapping = {
            row["RowID"]: i + 2 for i, (_, row) in enumerate(df_sorted.iterrows())
        }
        excel_rows_to_highlight = [
            row_mapping[rid] for rid in [] if rid in row_mapping
        ]  # Placeholder
        
        # Clean up data
        df_sorted.drop(columns=["RowID"], inplace=True, errors="ignore")
        
        # Save to multiple formats
        self._save_to_parquet(df_sorted, output_dir)
        self._save_to_excel(df_sorted, output_file)
        self._save_to_csv(df_sorted, output_dir)
        
        # Save unmatched reversals info
        self._save_unmatched_reversals(excel_rows_to_highlight, output_dir)
        
        self.update_progress(0.80)
        return output_file

    def _save_to_parquet(self, df, output_dir):
        """Save data to Parquet format for large DataFrames."""
        try:
            parquet_path = output_dir / "merged_file_with_OR.parquet"
            df.drop_duplicates().to_parquet(parquet_path, index=False)
            logger.info(f"Saved intermediate Parquet file: {parquet_path}")
        except Exception as e:
            logger.warning(f"Could not save Parquet: {e}")

    def _save_to_excel(self, df, output_file):
        """Save data to Excel format."""
        df.drop_duplicates().to_excel(output_file, index=False)

    def _save_to_csv(self, df, output_dir):
        """Save data to CSV format with opportunity name."""
        opportunity_name = self._extract_opportunity_name()
        csv_path = output_dir / f"{opportunity_name} Claim Detail.csv"
        df.drop_duplicates().to_csv(csv_path, index=False)

    def _save_unmatched_reversals(self, excel_rows_to_highlight, output_dir):
        """Save unmatched reversals information."""
        unmatched_path = output_dir / "unmatched_reversals.txt"
        with open(unmatched_path, "w") as f:
            f.write(",".join(map(str, excel_rows_to_highlight)))

    def _extract_opportunity_name(self):
        """Extract opportunity name from file1_path."""
        opportunity_name = ProcessingConfig.DEFAULT_OPPORTUNITY_NAME
        try:
            if self.file1_path:
                if self.file1_path.lower().endswith(".xlsx"):
                    df_file1 = pd.read_excel(self.file1_path)
                else:
                    df_file1 = pd.read_csv(self.file1_path)
                
                if df_file1.shape[1] >= 2:
                    # Get the value from the first row, second column
                    raw_name = str(df_file1.iloc[0, 1])
                    # Clean for filename
                    opportunity_name = re.sub(r'[\\/*?:"<>|]', "_", raw_name)
        except Exception as e:
            logger.warning(f"Could not extract opportunity name from file1: {e}")
        
        return opportunity_name

    def _finalize_processing(self, output_file):
        """Finalize processing with highlighting and notifications."""
        self.highlight_unmatched_reversals(output_file)
        self.update_progress(0.85)

        messagebox.showinfo(
            "Success", f"Processing complete. File saved as {output_file}"
        )
        self.paste_into_template(output_file)
        self.update_progress(0.90)

    def highlight_unmatched_reversals(self, excel_file):
        """Highlight unmatched reversals in Excel file with reduced nesting complexity."""
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            
            # Early return if worksheet is invalid
            if ws is None:
                logger.warning("Worksheet is None, cannot highlight reversals")
                return
            
            # Early return if reversals file doesn't exist
            if not os.path.exists("unmatched_reversals.txt"):
                logger.info("No unmatched_reversals.txt file found")
                wb.save(excel_file)
                return
            
            # Apply highlighting to unmatched reversals
            self._apply_reversal_highlighting(ws, wb, excel_file)
            
        except Exception as e:
            logger.error(f"Failed to highlight unmatched reversals: {e}")

    def _apply_reversal_highlighting(self, ws, wb, excel_file):
        """Apply highlighting to rows specified in unmatched_reversals.txt."""
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        with open("unmatched_reversals.txt", "r") as f:
            rows = f.read().strip().split(",")
        
        for row_str in rows:
            if self._is_valid_row_number(row_str, ws):
                self._highlight_row(ws, int(row_str), fill)
        
        wb.save(excel_file)
        logger.info(f"Highlighted unmatched reversals in {excel_file}")

    def _is_valid_row_number(self, row_str, ws):
        """Check if row string represents a valid row number."""
        return row_str.isdigit() and 1 <= int(row_str) <= ws.max_row

    def _highlight_row(self, ws, row_num, fill):
        """Highlight all cells in a specific row."""
        row = ws[row_num]
        for cell in row:
            cell.fill = fill


warnings.filterwarnings("ignore", category=FutureWarning, message=".*swapaxes.*")
warnings.filterwarnings("ignore", category=FutureWarning, message=".*swapaxes.*")


def process_logic_block(df_block):
    """
    Vectorized numpy logic to mark 'OR' in 'Logic' for reversals with matching claims.
    Refactored to reduce nesting complexity and improve readability.
    """
    arr = df_block.to_numpy()
    col_idx = {col: i for i, col in enumerate(df_block.columns)}
    
    # Extract and prepare data
    logic_data = _extract_logic_data(arr, col_idx)
    
    # Early return if no reversals to process
    if not np.any(logic_data["is_reversal"]):
        return pd.DataFrame(arr, columns=df_block.columns)
    
    # Process reversals with reduced nesting
    _process_reversals(arr, col_idx, logic_data)
    
    return pd.DataFrame(arr, columns=df_block.columns)


def _extract_logic_data(arr, col_idx):
    """Extract and prepare data for logic processing."""
    qty = arr[:, col_idx["QUANTITY"]].astype(float)
    return {
        "qty": qty,
        "is_reversal": qty < 0,
        "is_claim": qty > 0,
        "ndc": arr[:, col_idx["NDC"]].astype(str),
        "member": arr[:, col_idx["MemberID"]].astype(str),
        "datefilled": pd.to_datetime(arr[:, col_idx["DATEFILLED"]], errors="coerce"),
        "abs_qty": np.abs(qty)
    }


def _process_reversals(arr, col_idx, logic_data):
    """Process reversals with matching logic, using guard clauses to reduce nesting."""
    rev_idx = np.where(logic_data["is_reversal"])[0]
    claim_idx = (
        np.where(logic_data["is_claim"])[0] 
        if np.any(logic_data["is_claim"]) 
        else np.array([], dtype=int)
    )
    
    # Create context object to reduce function argument count
    match_context = MatchContext(arr, col_idx, logic_data, claim_idx)
    
    for i in rev_idx:
        found_match = _try_find_match(match_context, i)
        
        # Mark unmatched reversals as 'OR'
        if not found_match:
            arr[i, col_idx["Logic"]] = "OR"


class MatchContext:
    """Context object to encapsulate matching parameters and reduce argument count."""
    
    def __init__(self, arr, col_idx, logic_data, claim_idx):
        self.arr = arr
        self.col_idx = col_idx
        self.logic_data = logic_data
        self.claim_idx = claim_idx


def _try_find_match(context, reversal_idx):
    """Attempt to find a matching claim for a reversal. Returns True if match found."""
    # Guard clause: no claims to match against
    if context.claim_idx.size == 0:
        return False
    
    # Find potential matches
    matches = _find_matching_claims(context.logic_data, context.claim_idx, reversal_idx)
    
    # Guard clause: no matches found
    if not np.any(matches):
        return False
    
    # Mark both reversal and matching claim as 'OR'
    context.arr[reversal_idx, context.col_idx["Logic"]] = "OR"
    context.arr[context.claim_idx[matches][0], context.col_idx["Logic"]] = "OR"
    return True


def _find_matching_claims(logic_data, claim_idx, reversal_idx):
    """Find claims that match the reversal based on NDC, member, quantity, and date."""
    matches = (
        (logic_data["ndc"][claim_idx] == logic_data["ndc"][reversal_idx])
        & (logic_data["member"][claim_idx] == logic_data["member"][reversal_idx])
        & (logic_data["abs_qty"][claim_idx] == logic_data["abs_qty"][reversal_idx])
    )
    
    # Add date constraint (within 30 days)
    date_diffs = np.abs(
        (logic_data["datefilled"][claim_idx] - logic_data["datefilled"][reversal_idx]).days
    )
    matches &= date_diffs <= 30
    
    return matches


if __name__ == "__main__":
    multiprocessing.freeze_support()
    ctk.set_appearance_mode("light")  # Start in light mode
    root = ctk.CTk()  # or tk.Tk() if not using customtkinter
    app = App(root)
    root.mainloop()
import os

import pytest
from openpyxl import Workbook


@pytest.fixture(scope="session", autouse=True)
def create_dummy_excel():
    filename = "./_Rx Repricing_wf.xlsx"
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Claims Table"
        ws.append(
            ["pharmacy_npi", "pharmacy_nabp", "pharmacy_id"]
        )  # add columns as needed
        ws.append([1234567890, "NABP123", "123"])  # add dummy data as needed
        wb.save(filename)
    yield
    # Optionally, remove the file after tests
    # os.remove(filename)
import json
import os
import tkinter as tk

import pandas as pd
import pytest

from app import App, ConfigManager


@pytest.fixture
def tmp_work_dir(tmp_path, monkeypatch):
    """
    Create a temporary working directory and chdir into it, so that ConfigManager
    will create its config.json there.
    """
    monkeypatch.chdir(tmp_path)
    return tmp_path


def test_save_default_creates_config(tmp_work_dir):
    # When no config.json exists, ConfigManager.save_default() should create one
    # under the current directory (tmp_work_dir).
    cm = ConfigManager()
    config_path = tmp_work_dir / "config.json"

    assert config_path.exists(), "Config file was not created."
    with open(config_path, "r") as f:
        data = json.load(f)
    # By default, ConfigManager sets 'last_folder' to the cwd
    assert "last_folder" in data
    assert data["last_folder"] == str(tmp_work_dir)


def test_load_existing_config(tmp_work_dir):
    # Write a custom config.json, then ensure ConfigManager.load() honors it.
    config_path = tmp_work_dir / "config.json"
    custom = {"last_folder": "C:/some/path/xyz"}
    with open(config_path, "w") as f:
        json.dump(custom, f)

    cm = ConfigManager()
    assert (
        cm.config == custom
    ), "ConfigManager did not load the existing config.json correctly."


def test_filter_template_columns_extracts_correct_range():
    # Build a sample DataFrame where columns go: ['A','B','Client Name','X','Y','Logic','Z','W']
    df = pd.DataFrame(
        {
            "A": [1],
            "B": [2],
            "Client Name": ["foo"],
            "X": [3],
            "Y": [4],
            "Logic": [5],
            "Z": [6],
            "W": [7],
        }
    )

    # We only expect columns from 'Client Name' up through 'Logic' (inclusive).
    root = tk.Tk()
    root.withdraw()
    app = App(root)
    filtered = app.filter_template_columns(df)
    root.destroy()

    assert list(filtered.columns) == [
        "Client Name",
        "X",
        "Y",
        "Logic",
    ], f"Expected columns from 'Client Name' to 'Logic', got {list(filtered.columns)}"


def test_filter_template_columns_fallback_to_full_df_if_missing_logic():
    # If 'Client Name' or 'Logic' aren't found, it should return the full DataFrame unmodified
    df = pd.DataFrame({"Foo": [1, 2], "Bar": [3, 4]})

    root = tk.Tk()
    root.withdraw()
    app = App(root)
    result = app.filter_template_columns(df)
    root.destroy()

    # Since 'Client Name' or 'Logic' are not present, filter_template_columns should catch ValueError
    # and return the original DataFrame
    pd.testing.assert_frame_equal(result, df)


def test_format_dataframe_converts_datetimes_and_handles_na():
    # Build a DataFrame with one datetime column and one column containing a None
    orig = pd.DataFrame(
        {
            "dt1": [
                pd.to_datetime("2020-12-31 13:45:00"),
                pd.to_datetime("2021-01-01 00:00:00"),
            ],
            "value": [10, None],
        }
    )

    root = tk.Tk()
    root.withdraw()
    app = App(root)
    formatted = app.format_dataframe(orig)
    root.destroy()

    # 'dt1' should now be strings in format '%Y-%m-%d %H:%M:%S'
    assert formatted["dt1"].dtype == object
    assert formatted["dt1"].iloc[0] == "2020-12-31 13:45:00"
    assert formatted["dt1"].iloc[1] == "2021-01-01 00:00:00"

    # The None in 'value' should become an empty string
    assert formatted["value"].iloc[1] == ""


# (Optional) If you want to guard GUI‐dependent tests to skip automatically when no display is available:
def is_display_available():
    try:
        root = tk.Tk()
        root.destroy()
        return True
    except tk.TclError:
        return False


@pytest.mark.skipif(not is_display_available(), reason="Tkinter display not available")
def test_app_instantiation_and_basic_attributes():
    # A minimal smoke–test to ensure that App(root) does not crash immediately,
    # and that certain attributes exist.
    root = tk.Tk()
    root.withdraw()
    app = App(root)

    # Basic sanity checks:
    assert hasattr(app, "file1_path")
    assert hasattr(app, "file2_path")
    assert hasattr(app, "template_file_path")
    assert isinstance(app.progress_bar, type(app.progress_bar))

    root.destroy()
from bg_disruption import process_data


def test_process_bg_runs():
    process_data()
from epls_lbl import main


def test_epls_main_runs():
    main()
from pathlib import Path

import pytest

from merge import merge_files


def test_merge_files(tmp_path):
    file1 = tmp_path / "f1.csv"
    file2 = tmp_path / "f2.csv"
    file1.write_text("DATEFILLED,SOURCERECORDID\n2020-01-01,1")
    file2.write_text("SOURCERECORDID,Total AWP (Historical)\n1,50.0")
    merge_files(str(file1), str(file2))
    assert Path("merged_file.xlsx").exists()
from openmdf_bg import process_data


def test_process_openmdf_bg_runs():
    process_data()
"""
UI Builder module for constructing the main application interface.
Extracted from app.py to reduce file size and improve organization.
"""

import customtkinter as ctk
from tkinter import messagebox
import random
import pyjokes
import emoji
import getpass
from pathlib import Path

from ui.ui_components import UIFactory, LIGHT_COLORS, DARK_COLORS
from config.app_config import DisruptionConfig, AppConstants


class UIBuilder:
    """Handles the construction of the main application UI."""
    
    def __init__(self, app_instance):
        self.app = app_instance
        
    def build_complete_ui(self):
        """Build the complete user interface."""
        self._setup_window()
        self._build_ui_components()
        self._setup_default_template()
        self._show_welcome_message()
        
    def _setup_window(self):
        """Configure the main window properties."""
        self.app.root.title("Reprice Automation")
        self.app.root.configure(fg_color=LIGHT_COLORS["dark_blue"])
        self.app.root.resizable(True, True)
        self.app.root.geometry("900x900")
        
    def _build_ui_components(self):
        """Build all UI components."""
        self._create_title()
        self._create_button_frame()
        self._create_notes_frame()
        self._create_disruption_frame()
        self._create_progress_frame()
        
    def _create_title(self):
        """Create the title label."""
        self.app.title_label = ctk.CTkLabel(
            self.app.root, text="Repricing Automation", font=("Cambria", 26, "bold")
        )
        self.app.title_label.grid(row=0, column=0, sticky="w", pady=20, padx=20)

    def _create_button_frame(self):
        """Create the main button frame with all action buttons."""
        self.app.button_frame = UIFactory.create_standard_frame(self.app.root)
        self.app.button_frame.grid(
            row=2, column=0, columnspan=3, sticky="ew", pady=10, padx=10
        )

        # Headers
        file_name_title = UIFactory.create_standard_label(
            self.app.button_frame, "File Name"
        )
        file_name_title.grid(row=0, column=2, pady=10, padx=10)

        # Create sub-components
        self._create_file_import_buttons()
        self._create_action_buttons()
        self._create_process_buttons()

    def _create_file_import_buttons(self):
        """Create file import buttons and labels."""
        # Import File 1
        self.app.file1_button = UIFactory.create_standard_button(
            self.app.button_frame, "Import File Uploaded to Tool", self.app.import_file1
        )
        self.app.file1_button.grid(row=1, column=0, pady=10, padx=10, sticky="ew")
        self.app.file1_label = UIFactory.create_standard_label(
            self.app.button_frame, "", width=350
        )
        self.app.file1_label.grid(row=1, column=2, pady=20, padx=10)

        # Import File 2
        self.app.file2_button = UIFactory.create_standard_button(
            self.app.button_frame, "Import File From Tool", self.app.import_file2
        )
        self.app.file2_button.grid(row=2, column=0, pady=10, padx=10, sticky="ew")
        self.app.file2_label = UIFactory.create_standard_label(self.app.button_frame, "")
        self.app.file2_label.grid(row=2, column=2, pady=20, padx=10)

        # Select Template
        self.app.template_button = UIFactory.create_standard_button(
            self.app.button_frame, "Select Template File", self.app.import_template_file
        )
        self.app.template_button.grid(row=3, column=0, pady=10, padx=10, sticky="ew")
        self.app.template_label = UIFactory.create_standard_label(self.app.button_frame, "")
        self.app.template_label.grid(row=3, column=2, pady=20, padx=10)

    def _create_action_buttons(self):
        """Create action buttons (cancel, logs, theme)."""
        # Cancel button
        self.app.cancel_button = UIFactory.create_red_button(
            self.app.button_frame, "Cancel", self.app.cancel_process
        )
        self.app.cancel_button.grid(row=4, column=0, pady=10, padx=10, sticky="ew")

        # View Logs button
        self.app.logs_button = UIFactory.create_standard_button(
            self.app.button_frame, "View Logs", self.app.show_log_viewer
        )
        self.app.logs_button.grid(row=4, column=1, pady=10, padx=10, sticky="ew")

        # Toggle Dark Mode button
        self.app.toggle_theme_button = UIFactory.create_standard_button(
            self.app.button_frame, "Switch to Dark Mode", self.toggle_dark_mode
        )
        self.app.toggle_theme_button.grid(row=4, column=2, pady=10, padx=10, sticky="ew")

        # Shared Log Button
        self.app.shared_log_button = UIFactory.create_standard_button(
            self.app.button_frame, "Shared Audit Log", self.app.show_shared_log_viewer
        )
        self.app.shared_log_button.grid(row=6, column=1, pady=10, padx=10, sticky="ew")

        # Exit button
        self.app.exit_button = UIFactory.create_red_button(
            self.app.button_frame, "Exit", self.app.root.quit
        )
        self.app.exit_button.grid(row=6, column=2, pady=10, padx=10, sticky="ew")

    def _create_process_buttons(self):
        """Create processing and LBL generation buttons."""
        # SHARx LBL button
        self.app.sharx_lbl_button = UIFactory.create_standard_button(
            self.app.button_frame, "Generate SHARx LBL", self.app.sharx_lbl
        )
        self.app.sharx_lbl_button.grid(row=5, column=0, pady=10, padx=10, sticky="ew")
        
        # EPLS LBL button
        self.app.epls_lbl_button = UIFactory.create_standard_button(
            self.app.button_frame, "Generate EPLS LBL", self.app.epls_lbl
        )
        self.app.epls_lbl_button.grid(row=5, column=1, pady=10, padx=10, sticky="ew")
        
        # Start Process button
        self.app.start_process_button = ctk.CTkButton(
            self.app.button_frame,
            text="Start Repricing",
            command=self.app.start_process_threaded,
            font=("Cambria", 20, "bold"),
            height=40,
            width=200,
            fg_color=LIGHT_COLORS["mint"],
            text_color="#000000",
        )
        self.app.start_process_button.grid(row=5, column=2, pady=10, padx=10, sticky="ew")

    def _create_notes_frame(self):
        """Create the notes frame with important information."""
        self.app.notes_frame = UIFactory.create_standard_frame(self.app.root)
        self.app.notes_frame.grid(
            row=3, column=0, columnspan=4, sticky="ew", pady=10, padx=10
        )
        notes = UIFactory.create_standard_label(
            self.app.notes_frame, AppConstants.NOTES_TEXT
        )
        notes.configure(justify="left")
        notes.pack(padx=20, pady=10)

    def _create_disruption_frame(self):
        """Create the disruption type selector frame."""
        self.app.dis_frame = UIFactory.create_standard_frame(self.app.root)
        self.app.dis_frame.grid(
            row=4, column=0, columnspan=4, sticky="ew", pady=10, padx=10
        )

        # Create disruption buttons using configuration
        disruption_labels = DisruptionConfig.get_disruption_labels()
        for idx, label in enumerate(disruption_labels):
            btn = UIFactory.create_standard_button(
                self.app.dis_frame, label, 
                lambda label_text=label: self.app.start_disruption(label_text)
            )
            btn.grid(row=0, column=idx, padx=10, pady=10, sticky="ew")

    def _create_progress_frame(self):
        """Create the progress bar frame."""
        self.app.prog_frame = UIFactory.create_standard_frame(self.app.root)
        self.app.prog_frame.grid(
            row=5, column=0, columnspan=4, sticky="ew", pady=10, padx=10
        )
        self.app.progress_bar = ctk.CTkProgressBar(
            self.app.prog_frame, orientation="horizontal", mode="determinate"
        )
        self.app.progress_bar.set(self.app.progress_var.get())
        self.app.progress_bar.pack(padx=10, pady=(10, 5), fill="x")
        self.app.progress_label = ctk.CTkLabel(
            self.app.prog_frame, textvariable=self.app.progress_label_var
        )
        self.app.progress_label.pack(padx=10, pady=(0, 10), anchor="w")
        
    def _setup_default_template(self):
        """Set up default template if it exists."""
        default_template = Path("_Rx Repricing_wf.xlsx")
        if default_template.exists():
            self.app.template_file_path = str(default_template)
            if hasattr(self.app, 'template_label'):
                self.app.template_label.configure(
                    text=default_template.name
                )
        else:
            self.app.template_file_path = None
            
    def _show_welcome_message(self):
        """Show a personalized welcome message with a random joke and emoji."""
        user = getpass.getuser()
        welcome_messages = AppConstants.WELCOME_MESSAGES
        
        msg = welcome_messages.get(
            user, f"Welcome, {user}! Ready to use the Repricing Automation Toolkit?"
        )

        # Add a random joke and emoji
        try:
            joke = pyjokes.get_joke()
        except Exception:
            joke = "Have a great day!"
            
        chosen_emoji = emoji.emojize(random.choice(AppConstants.EMOJIS), language="alias")
        full_msg = f"{msg}\n\n{joke} {chosen_emoji}"
        
        # Show after UI is built
        self.app.root.after(500, lambda: messagebox.showinfo("Welcome", full_msg))
        
    def toggle_dark_mode(self):
        """Toggle between light and dark modes."""
        current = ctk.get_appearance_mode().lower()

        if current == "light":
            self._switch_to_dark_mode()
        else:
            self._switch_to_light_mode()
            
    def _switch_to_dark_mode(self):
        """Switch to dark mode."""
        ctk.set_appearance_mode("dark")
        self.app.apply_theme_colors(DARK_COLORS)
        if self.app.toggle_theme_button:
            self.app.toggle_theme_button.configure(text="Switch to Light Mode")
            
    def _switch_to_light_mode(self):
        """Switch to light mode."""
        ctk.set_appearance_mode("light")
        self.app.apply_theme_colors(LIGHT_COLORS)
        if self.app.toggle_theme_button:
            self.app.toggle_theme_button.configure(text="Switch to Dark Mode")
from openmdf_tier import process_data


def test_process_openmdf_tier_runs():
    process_data()
from tier_disruption import process_data


def test_process_tier_runs():
    process_data()
import json
import os

import pandas as pd

# Standardize IDs


def standardize_pharmacy_ids(df):
    if "PHARMACYNPI" in df.columns:
        df["PHARMACYNPI"] = df["PHARMACYNPI"].astype(str).str.zfill(10)
    if "NABP" in df.columns:
        df["NABP"] = df["NABP"].astype(str).str.zfill(7)
    return df


def standardize_network_ids(network):
    if "pharmacy_npi" in network.columns:
        network["pharmacy_npi"] = network["pharmacy_npi"].astype(str).str.zfill(10)
    if "pharmacy_nabp" in network.columns:
        network["pharmacy_nabp"] = network["pharmacy_nabp"].astype(str).str.zfill(7)
    return network


def merge_with_network(df, network):
    return df.merge(
        network,
        left_on=["PHARMACYNPI", "NABP"],
        right_on=["pharmacy_npi", "pharmacy_nabp"],
        how="left",
    )


def load_file_paths(json_file):
    with open(json_file, "r") as f:
        file_paths = json.load(f)
    return file_paths


def summarize_by_tier(df, col, from_val, to_val):
    filtered = df[(df[col] == from_val) & (df["FormularyTier"] == to_val)]
    pt = pd.pivot_table(
        filtered,
        values=["Rxs", "MemberID"],
        index=["Product Name"],
        aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
    )
    rxs = filtered["Rxs"].sum()
    members = filtered["MemberID"].nunique()
    return pt, rxs, members


def process_data():
    file_paths = load_file_paths("file_paths.json")

    # Only load claims if the path is present and not empty
    claims = None
    if "reprice" in file_paths and file_paths["reprice"]:
        claims = pd.read_excel(file_paths["reprice"], sheet_name="Claims Table")
    else:
        print("No reprice/template file provided. Skipping claims loading.")
        # You can decide to return, raise, or continue with alternate logic here
        return

    medi = pd.read_excel(file_paths["medi_span"])[
        ["NDC", "Maint Drug?", "Product Name"]
    ]
    u = pd.read_excel(file_paths["u_disrupt"], sheet_name="Universal NDC")[
        ["NDC", "Tier"]
    ]
    e = pd.read_excel(file_paths["e_disrupt"], sheet_name="Alternatives NDC")[
        ["NDC", "Tier", "Alternative"]
    ]
    network = pd.read_excel(file_paths["n_disrupt"])[
        ["pharmacy_npi", "pharmacy_nabp", "pharmacy_is_excluded"]
    ]

    df = claims.merge(medi, on="NDC", how="left")
    df = df.merge(u.rename(columns={"Tier": "Universal Tier"}), on="NDC", how="left")
    df = df.merge(e.rename(columns={"Tier": "Exclusive Tier"}), on="NDC", how="left")

    df = standardize_pharmacy_ids(df)
    network = standardize_network_ids(network)
    df = merge_with_network(df, network)

    df["DATEFILLED"] = pd.to_datetime(df["DATEFILLED"], errors="coerce")
    df = df.drop_duplicates()
    df["Logic"] = pd.to_numeric(df["Logic"], errors="coerce")
    df["FormularyTier"] = pd.to_numeric(df["FormularyTier"], errors="coerce")

    latest_date = df["DATEFILLED"].max()
    starting_point = latest_date - pd.DateOffset(months=6) + pd.DateOffset(days=1)
    df = df[(df["DATEFILLED"] >= starting_point) & (df["DATEFILLED"] <= latest_date)]
    df = df[(df["Logic"] >= 5) & (df["Logic"] <= 10) & (df["Maint Drug?"] == "Y")]
    df = df[~df["Product Name"].str.contains(r"\balbuterol\b", case=False)]
    df = df[~df["Product Name"].str.contains(r"\bventolin\b", case=False)]
    df = df[~df["Product Name"].str.contains(r"\bepinephrine\b", case=False)]
    df = df[
        ~df["Alternative"]
        .astype(str)
        .str.contains("Covered|Use different NDC", case=False, regex=True)
    ]

    total_claims = df["Rxs"].sum()
    total_members = df["MemberID"].nunique()

    writer = pd.ExcelWriter("LBL for Disruption.xlsx", engine="xlsxwriter")
    df.to_excel(writer, sheet_name="Data", index=False)

    tiers = [
        ("Universal_Positive 2-1", "Universal Tier", 2, 1),
        ("Universal_Positive 3-1", "Universal Tier", 3, 1),
        ("Universal_Positive 3-2", "Universal Tier", 3, 2),
        ("Universal_Negative 1-2", "Universal Tier", 1, 2),
        ("Universal_Negative 1-3", "Universal Tier", 1, 3),
        ("Universal_Negative 2-3", "Universal Tier", 2, 3),
        ("Exclusive_Positive 2-1", "Exclusive Tier", 2, 1),
        ("Exclusive_Positive 3-1", "Exclusive Tier", 3, 1),
        ("Exclusive_Positive 3-2", "Exclusive Tier", 3, 2),
        ("Exclusive_Negative 1-2", "Exclusive Tier", 1, 2),
        ("Exclusive_Negative 1-3", "Exclusive Tier", 1, 3),
        ("Exclusive_Negative 2-3", "Exclusive Tier", 2, 3),
    ]

    summary_rows = []
    tab_members = {}

    for name, col, from_val, to_val in tiers:
        pt, rxs, members = summarize_by_tier(df, col, from_val, to_val)
        pt.to_excel(writer, sheet_name=name)
        summary_rows.append((name.replace("_", " "), members, rxs))
        tab_members[name] = members

    exclusions = df[df["pharmacy_is_excluded"] == True]
    ex_pt = exclusions.pivot_table(
        values=["Rxs", "MemberID"],
        index=["Product Name"],
        aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
    )
    exc_rxs = exclusions["Rxs"].sum()
    exc_members = exclusions["MemberID"].nunique()
    ex_pt.to_excel(writer, sheet_name="Exclusions")

    summary_rows.append(("Exclusions", exc_members, exc_rxs))
    tab_members["Exclusions"] = exc_members

    # Aggregate utilizer and claims counts for summary
    uni_pos_utilizers = (
        tab_members.get("Universal_Positive 2-1", 0)
        + tab_members.get("Universal_Positive 3-1", 0)
        + tab_members.get("Universal_Positive 3-2", 0)
    )
    uni_neg_utilizers = (
        tab_members.get("Universal_Negative 1-2", 0)
        + tab_members.get("Universal_Negative 1-3", 0)
        + tab_members.get("Universal_Negative 2-3", 0)
    )
    ex_pos_utilizers = (
        tab_members.get("Exclusive_Positive 2-1", 0)
        + tab_members.get("Exclusive_Positive 3-1", 0)
        + tab_members.get("Exclusive_Positive 3-2", 0)
    )
    ex_neg_utilizers = (
        tab_members.get("Exclusive_Negative 1-2", 0)
        + tab_members.get("Exclusive_Negative 1-3", 0)
        + tab_members.get("Exclusive_Negative 2-3", 0)
    )
    exc_utilizers = tab_members.get("Exclusions", 0)

    # Similarly, aggregate claims for each group
    uni_pos_claims = 0
    uni_neg_claims = 0
    ex_pos_claims = 0
    ex_neg_claims = 0
    exc_claims = exc_rxs

    for name, _, _, _ in tiers:
        if name.startswith("Universal_Positive"):
            uni_pos_claims += df[
                (df["Universal Tier"] == int(name.split()[1][0]))
                & (df["FormularyTier"] == int(name.split()[1][2]))
            ]["Rxs"].sum()
        elif name.startswith("Universal_Negative"):
            uni_neg_claims += df[
                (df["Universal Tier"] == int(name.split()[1][0]))
                & (df["FormularyTier"] == int(name.split()[1][2]))
            ]["Rxs"].sum()
        elif name.startswith("Exclusive_Positive"):
            ex_pos_claims += df[
                (df["Exclusive Tier"] == int(name.split()[1][0]))
                & (df["FormularyTier"] == int(name.split()[1][2]))
            ]["Rxs"].sum()
        elif name.startswith("Exclusive_Negative"):
            ex_neg_claims += df[
                (df["Exclusive Tier"] == int(name.split()[1][0]))
                & (df["FormularyTier"] == int(name.split()[1][2]))
            ]["Rxs"].sum()

    # Calculate percentages
    uni_pos_pct = uni_pos_claims / total_claims if total_claims else 0
    uni_neg_pct = uni_neg_claims / total_claims if total_claims else 0
    ex_pos_pct = ex_pos_claims / total_claims if total_claims else 0
    ex_neg_pct = ex_neg_claims / total_claims if total_claims else 0
    exc_pct = exc_claims / total_claims if total_claims else 0

    summary_df = pd.DataFrame(
        {
            "Formulary": [
                "Universal Positive",
                "Universal Negative",
                "Exclusive Positive",
                "Exclusive Negative",
                "Exclusions",
            ],
            "Utilizers": [
                uni_pos_utilizers,
                uni_neg_utilizers,
                ex_pos_utilizers,
                ex_neg_utilizers,
                exc_utilizers,
            ],
            "Rxs": [
                uni_pos_claims,
                uni_neg_claims,
                ex_pos_claims,
                ex_neg_claims,
                exc_claims,
            ],
            "% of claims": [uni_pos_pct, uni_neg_pct, ex_pos_pct, ex_neg_pct, exc_pct],
            "": ["", "", "", "", ""],
            "Totals": [
                f"Members: {total_members}",
                f"Claims: {total_claims}",
                "",
                "",
                "",
            ],
        }
    )

    summary_df.to_excel(writer, sheet_name="Summary", index=False)

    for sheet_name, value in tab_members.items():
        worksheet = writer.sheets[sheet_name]
        worksheet.write("F1", f"Total Members: {value}")

    # Network summary for non-excluded pharmacies
    network_df = df[df["pharmacy_is_excluded"].isna()]
    filter_phrases = [
        "CVS",
        "Walgreens",
        "Kroger",
        "Walmart",
        "Rite Aid",
        "Optum",
        "Express Scripts",
        "DMR",
        "Williams Bro",
        "Publix",
    ]
    regex_pattern = "|".join([f"\b{phrase}\b" for phrase in filter_phrases])
    network_df = network_df[
        ~network_df["Pharmacy Name"].str.contains(regex_pattern, case=False, regex=True)
    ]

    if (
        "PHARMACYNPI" in network_df.columns
        and "NABP" in network_df.columns
        and "Pharmacy Name" in network_df.columns
    ):
        network_pivot = pd.pivot_table(
            network_df,
            values=["Rxs", "MemberID"],
            index=["PHARMACYNPI", "NABP", "Pharmacy Name"],
            aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
        )
        network_pivot.to_excel(writer, sheet_name="Network")
    else:
        print(
            "PHARMACYNPI, NABP, or Pharmacy Name column missing in the data dataframe."
        )

    # Reorder sheets so 'Summary' is right after 'Data'
    workbook = writer.book
    sheets = workbook.worksheets()
    sheet_names = [ws.get_name() for ws in sheets]

    # Move 'Summary' after 'Data'
    if "Data" in sheet_names and "Summary" in sheet_names:
        data_idx = sheet_names.index("Data")
        summary_idx = sheet_names.index("Summary")
        if summary_idx != data_idx + 1:
            summary_ws = sheets[summary_idx]
            sheets.pop(summary_idx)
            sheets.insert(data_idx + 1, summary_ws)

    # Save once at the end
    writer._save()


if __name__ == "__main__":
    process_data()
    print("Data processing complete")
"""
UI Components module for the Repricing Automation application.
This module contains UI-related classes and utilities to improve code organization.
"""

import customtkinter as ctk


# UI styling variables
FONT_SELECT = ("Cambria", 20, "bold")

# Color palettes
LIGHT_COLORS = {
    "dark_blue": "#D9EAF7",
    "grey_blue": "#A3B9CC",
    "mint": "#8FD9A8",
    "button_red": "#D52B2B",
}

DARK_COLORS = {
    "dark_blue": "#223354",
    "grey_blue": "#31476A",
    "mint": "#26A69A",
    "button_red": "#931D1D",
}


class UIFactory:
    """Factory class to create UI components and reduce code duplication."""
    
    @staticmethod
    def _create_button_base(parent, text, command, fg_color):
        """Base method for creating buttons with common styling."""
        return ctk.CTkButton(
            parent,
            text=text,
            command=command,
            font=FONT_SELECT,
            height=40,
            fg_color=fg_color,
            text_color="#000000"
        )
    
    @staticmethod
    def create_standard_button(parent, text, command):
        """Create a standardized button with common styling."""
        return UIFactory._create_button_base(parent, text, command, LIGHT_COLORS["mint"])
    
    @staticmethod
    def create_red_button(parent, text, command):
        """Create a red button (for cancel/exit actions)."""
        return UIFactory._create_button_base(parent, text, command, LIGHT_COLORS["button_red"])
    
    @staticmethod
    def create_standard_frame(parent):
        """Create a standardized frame with common styling."""
        return ctk.CTkFrame(parent, fg_color=LIGHT_COLORS["grey_blue"])
    
    @staticmethod
    def create_standard_label(parent, text, width=None):
        """Create a standardized label."""
        if width:
            return ctk.CTkLabel(parent, text=text, font=FONT_SELECT, width=width)
        return ctk.CTkLabel(parent, text=text, font=FONT_SELECT)


class ThemeManager:
    """Manages theme colors and application of themes to UI components."""
    
    @staticmethod
    def apply_theme_colors(app_instance, colors):
        """Apply theme colors to all UI components."""
        ThemeManager._apply_root_colors(app_instance, colors)
        ThemeManager._apply_frame_colors(app_instance, colors)
        ThemeManager._apply_button_colors(app_instance, colors)
        ThemeManager._apply_special_component_colors(app_instance, colors)
    
    @staticmethod
    def _apply_root_colors(app_instance, colors):
        """Apply colors to the root window."""
        app_instance.root.configure(fg_color=colors["dark_blue"])
    
    @staticmethod
    def _apply_frame_colors(app_instance, colors):
        """Apply colors to frames."""
        frames = ["button_frame", "notes_frame", "dis_frame", "prog_frame"]
        for frame_name in frames:
            frame = getattr(app_instance, frame_name, None)
            if frame:
                frame.configure(fg_color=colors["grey_blue"])
    
    @staticmethod
    def _apply_button_colors(app_instance, colors):
        """Apply colors to standard buttons."""
        button_widgets = [
            "file1_button", "file2_button", "template_button", "cancel_button",
            "logs_button", "toggle_theme_button", "sharx_lbl_button", 
            "epls_lbl_button", "start_process_button"
        ]
        
        for btn_name in button_widgets:
            btn = getattr(app_instance, btn_name, None)
            if btn:
                btn.configure(fg_color=colors["mint"], text_color="#000000")
    
    @staticmethod
    def _apply_special_component_colors(app_instance, colors):
        """Apply colors to special components."""
        # Apply colors to special buttons
        if hasattr(app_instance, "exit_button"):
            app_instance.exit_button.configure(
                fg_color=colors["button_red"], text_color="#000000"
            )
        
        # Apply colors to progress components
        if hasattr(app_instance, "progress_label"):
            app_instance.progress_label.configure(
                bg_color=colors["grey_blue"], text_color="#000000"
            )


class ProgressManager:
    """Manages progress bar updates and calculations."""
    
    @staticmethod
    def calculate_time_estimates(value, start_time):
        """Calculate progress percentage and time estimates."""
        import time
        
        percent = int(value * 100)
        elapsed = time.time() - start_time if start_time else 0
        est = int((elapsed / value) * (1 - value)) if value > 0 else 0
        return percent, est
    
    @staticmethod
    def format_progress_message(percent, estimated_seconds):
        """Format progress message with percentage and time estimate."""
        return f"Progress: {percent}% | Est. {estimated_seconds}s left"
import os
import shutil
import logging
import xlwings as xw
import importlib.util
from typing import Any, Tuple
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed

logger = logging.getLogger(__name__)

# COM fallback via pywin32
EXCEL_COM_AVAILABLE = importlib.util.find_spec("win32com.client") is not None


def open_workbook(path: str, visible: bool = False) -> Tuple[Any, Any, bool]:
    """
    Open workbook via xlwings or COM fallback.
    Returns (wb, app_obj, use_com).
    """
    import time

    max_retries = 3
    delay = 2
    last_exc = None
    for attempt in range(max_retries):
        try:
            app = xw.App(visible=visible, add_book=False)  # Ensure no new book is added
            try:
                wb = app.books.open(path)
            except TypeError:
                # Try with password if provided in path (e.g., path='file.xlsx::password')
                if "::" in path:
                    file_path, password = path.split("::", 1)
                    wb = app.books.open(file_path, password=password)
                else:
                    raise
            return wb, app, False
        except Exception as e:
            last_exc = e
            logger.warning(
                f"Failed to open workbook (attempt {attempt + 1}/{max_retries}): {e}"
            )
            time.sleep(delay)
    if EXCEL_COM_AVAILABLE:
        import win32com.client as win32

        excel: Any = win32.Dispatch("Excel.Application")
        excel.Visible = visible  # Ensure Excel remains hidden
        excel.DisplayAlerts = False  # Suppress alerts
        try:
            if "::" in path:
                file_path, password = path.split("::", 1)
                wb: Any = excel.Workbooks.Open(
                    os.path.abspath(file_path), False, False, None, password
                )
            else:
                wb: Any = excel.Workbooks.Open(os.path.abspath(path))
        except Exception as e:
            logger.error(f"COM fallback failed to open workbook: {e}")
            raise
        return wb, excel, True
    logger.error(f"Failed to open workbook after {max_retries} attempts: {last_exc}")
    if last_exc is not None:
        raise last_exc
    # Should never reach here, but raise as a safeguard
    raise RuntimeError("Failed to open workbook and no exception was captured.")


def write_df_to_sheet_async(
    path: str,
    sheet_name: str,
    df: pd.DataFrame,
    start_cell: str = "A2",
    header: bool = False,
    index: bool = False,
    clear: bool = True,
    visible: bool = False,
    clear_by_label: bool = False,
    max_workers: int = 4,
) -> None:
    """
    Async version of write_df_to_sheet for large DataFrames (xlwings only).
    Splits DataFrame into row blocks and writes in parallel threads.
    """
    logger.info(
        f"[ASYNC] Writing to {path} in sheet '{sheet_name}' from cell {start_cell} with {max_workers} workers"
    )
    wb, app, use_com = open_workbook(path, visible)
    if use_com:
        # COM automation is not thread-safe; fallback to sync
        logger.warning("COM fallback does not support async writes. Using sync write.")
        return write_df_to_sheet(
            path,
            sheet_name,
            df,
            start_cell,
            header,
            index,
            clear,
            visible,
            clear_by_label,
        )
    try:
        ws = wb.sheets[sheet_name]
        cell = ws.range(start_cell)
    except Exception as e:
        close_workbook(wb, app, save=False, use_com=use_com)
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook.") from e

    start_row = cell.row
    start_col = cell.column
    n_rows, n_cols = df.shape
    total_rows = n_rows + (1 if header else 0)
    end_row = start_row + total_rows - 1
    end_col = start_col + n_cols - 1

    # Optionally clear before writing
    target = ws.range((start_row, start_col), (end_row, end_col))
    if clear:
        if clear_by_label:
            for idx, col in enumerate(df.columns, start_col):
                col_range = ws.range((start_row, idx), (end_row, idx))
                col_range.clear_contents()
        else:
            target.clear_contents()

    # Write header if needed
    if header:
        for j, h in enumerate(df.columns, start_col):
            ws.Cells(start_row, j).Value = h
        data_start = start_row + 1
    else:
        data_start = start_row

    # Split DataFrame into blocks for parallel writing
    block_size = max(100, n_rows // max_workers)
    blocks = [(i, min(i + block_size, n_rows)) for i in range(0, n_rows, block_size)]

    def write_block(start, stop):
        for i, row in enumerate(df.values[start:stop], data_start + start):
            for j, val in enumerate(row, start_col):
                ws.Cells(i, j).Value = val

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(write_block, start, stop) for start, stop in blocks]
        for f in as_completed(futures):
            f.result()

    close_workbook(wb, app, save=True, use_com=use_com)


def close_workbook(
    wb: Any, app_obj: Any, save: bool = True, use_com: bool = False
) -> None:
    """
    Close the workbook and quit the application.
    """
    if not use_com:
        if save:
            wb.save()
        wb.close()
        app_obj.quit()
    else:
        if save:
            wb.Save()
        wb.Close(SaveChanges=save)
        app_obj.Quit()


def write_df_to_sheet(
    path: str,
    sheet_name: str,
    df: pd.DataFrame,
    start_cell: str = "A2",
    header: bool = False,
    index: bool = False,
    clear: bool = True,
    visible: bool = False,
    clear_by_label: bool = False,
) -> None:
    """
    Write DataFrame to an Excel sheet without removing any formatting.
    Only clears the cells where values will be written.
    """
    logger.info(f"Writing to {path} in sheet '{sheet_name}' from cell {start_cell}")

    wb, app, use_com = open_workbook(path, visible)

    try:
        if not use_com:
            ws = wb.sheets[sheet_name]
            cell = ws.range(start_cell)

            def clear_func(rng):
                rng.clear_contents()
        else:
            ws: Any = wb.Worksheets(sheet_name)
            cell: Any = ws.Range(start_cell)

            def clear_func(rng):
                rng.ClearContents()
    except Exception as e:
        close_workbook(wb, app, save=False, use_com=use_com)
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook.") from e

    # Determine start row/col and target range
    start_row = cell.row if not use_com else cell.Row
    start_col = cell.column if not use_com else cell.Column
    n_rows, n_cols = df.shape
    total_rows = n_rows + (1 if header else 0)
    end_row = start_row + total_rows - 1
    end_col = start_col + n_cols - 1

    if not use_com:
        target = ws.range((start_row, start_col), (end_row, end_col))
        if clear:
            if clear_by_label:
                # Clear by column label (header row)
                for idx, col in enumerate(df.columns, start_col):
                    col_range = ws.range((start_row, idx), (end_row, idx))
                    col_range.clear_contents()
            else:
                clear_func(target)
        target.options(index=index, header=header).value = df
    else:
        target = ws.Range(ws.Cells(start_row, start_col), ws.Cells(end_row, end_col))
        if clear:
            if clear_by_label:
                # Clear by column label (header row)
                for idx, col in enumerate(df.columns, start_col):
                    col_rng = ws.Range(ws.Cells(start_row, idx), ws.Cells(end_row, idx))
                    col_rng.ClearContents()
            else:
                clear_func(target)
        data_start = start_row
        if header:
            for j, h in enumerate(df.columns, start_col):
                ws.Cells(start_row, j).Value = h
            data_start += 1
        for i, row in enumerate(df.values.tolist(), data_start):
            for j, val in enumerate(row, start_col):
                ws.Cells(i, j).Value = val

    close_workbook(wb, app, save=True, use_com=use_com)


def write_df_to_template(
    template_path: str,
    output_path: str,
    sheet_name: str,
    df: pd.DataFrame,
    start_cell: str = "A2",
    header: bool = False,
    index: bool = False,
    visible: bool = False,
    open_file: bool = False,
) -> None:
    """
    Copy an Excel template and write a DataFrame into it without altering
    any existing formatting, charts, tables, or objects.

    If open_file is True, launch the filled workbook in Excel after writing.
    """
    shutil.copy(template_path, output_path)
    write_df_to_sheet(
        path=output_path,
        sheet_name=sheet_name,
        df=df,
        start_cell=start_cell,
        header=header,
        index=index,
        clear=True,
        visible=visible,
    )
    if open_file:
        os.startfile(output_path)
"""
Logic processing utilities extracted from app.py
Following CodeScene ACE principles for better code organization
"""

import logging
import warnings
from dataclasses import dataclass
from typing import Dict

import numpy as np
import pandas as pd

logger = logging.getLogger(__name__)

# Filter out specific warnings
warnings.filterwarnings("ignore", category=FutureWarning, message=".*swapaxes.*")


@dataclass
class LogicData:
    """Data class to encapsulate logic processing data."""
    qty: np.ndarray
    is_reversal: np.ndarray
    is_claim: np.ndarray
    ndc: np.ndarray
    member: np.ndarray
    datefilled: pd.DatetimeIndex
    abs_qty: np.ndarray


@dataclass
class MatchContext:
    """Context object to encapsulate matching parameters."""
    arr: np.ndarray
    col_idx: Dict[str, int]
    logic_data: LogicData
    claim_idx: np.ndarray


class LogicProcessor:
    """Handles logic processing for reversal matching."""
    
    @staticmethod
    def process_logic_block(df_block: pd.DataFrame) -> pd.DataFrame:
        """
        Vectorized numpy logic to mark 'OR' in 'Logic' for reversals with matching claims.
        Refactored to reduce nesting complexity and improve readability.
        """
        arr = df_block.to_numpy()
        col_idx = {col: i for i, col in enumerate(df_block.columns)}
        
        # Extract and prepare data
        logic_data = LogicProcessor._extract_logic_data(arr, col_idx)
        
        # Early return if no reversals to process
        if not np.any(logic_data.is_reversal):
            return pd.DataFrame(arr, columns=df_block.columns)
        
        # Process reversals with reduced nesting
        LogicProcessor._process_reversals(arr, col_idx, logic_data)
        
        return pd.DataFrame(arr, columns=df_block.columns)

    @staticmethod
    def _extract_logic_data(arr: np.ndarray, col_idx: Dict[str, int]) -> LogicData:
        """Extract and prepare data for logic processing."""
        qty = arr[:, col_idx["QUANTITY"]].astype(float)
        
        return LogicData(
            qty=qty,
            is_reversal=qty < 0,
            is_claim=qty > 0,
            ndc=arr[:, col_idx["NDC"]].astype(str),
            member=arr[:, col_idx["MemberID"]].astype(str),
            datefilled=pd.to_datetime(arr[:, col_idx["DATEFILLED"]], errors="coerce"),
            abs_qty=np.abs(qty)
        )

    @staticmethod
    def _process_reversals(arr: np.ndarray, col_idx: Dict[str, int], logic_data: LogicData):
        """Process reversals with matching logic, using guard clauses to reduce nesting."""
        rev_idx = np.where(logic_data.is_reversal)[0]
        claim_idx = (
            np.where(logic_data.is_claim)[0] 
            if np.any(logic_data.is_claim) 
            else np.array([], dtype=int)
        )
        
        # Create context object to reduce function argument count
        match_context = MatchContext(arr, col_idx, logic_data, claim_idx)
        
        for i in rev_idx:
            found_match = LogicProcessor._try_find_match(match_context, i)
            
            # Mark unmatched reversals as 'OR'
            if not found_match:
                arr[i, col_idx["Logic"]] = "OR"

    @staticmethod
    def _try_find_match(context: MatchContext, reversal_idx: int) -> bool:
        """Attempt to find a matching claim for a reversal. Returns True if match found."""
        # Guard clause: no claims to match against
        if context.claim_idx.size == 0:
            return False
        
        # Find potential matches
        matches = LogicProcessor._find_matching_claims(
            context.logic_data, context.claim_idx, reversal_idx
        )
        
        # Guard clause: no matches found
        if not np.any(matches):
            return False
        
        # Mark both reversal and matching claim as 'OR'
        context.arr[reversal_idx, context.col_idx["Logic"]] = "OR"
        context.arr[context.claim_idx[matches][0], context.col_idx["Logic"]] = "OR"
        return True

    @staticmethod
    def _find_matching_claims(logic_data: LogicData, claim_idx: np.ndarray, 
                            reversal_idx: int) -> np.ndarray:
        """Find claims that match the reversal based on NDC, member, quantity, and date."""
        matches = (
            (logic_data.ndc[claim_idx] == logic_data.ndc[reversal_idx])
            & (logic_data.member[claim_idx] == logic_data.member[reversal_idx])
            & (logic_data.abs_qty[claim_idx] == logic_data.abs_qty[reversal_idx])
        )
        
        # Add date constraint (within 30 days)
        try:
            date_diffs = np.abs(
                (logic_data.datefilled[claim_idx] - logic_data.datefilled[reversal_idx]).days
            )
            matches &= date_diffs <= 30
        except Exception as e:
            logger.warning(f"Date filtering failed: {e}")
            # Continue without date constraint
        
        return matches


# Backwards compatibility functions
def process_logic_block(df_block: pd.DataFrame) -> pd.DataFrame:
    """Backwards compatibility wrapper."""
    return LogicProcessor.process_logic_block(df_block)
from sharx_lbl import main


def test_sharx_main_runs():
    main()
import pandas as pd
import sys
import re
import logging
from pathlib import Path
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.utils import (
    load_file_paths,
    standardize_pharmacy_ids,
    standardize_network_ids,
    merge_with_network,
    drop_duplicates_df,
    clean_logic_and_tier,
    filter_recent_date,
    filter_logic_and_maintenance,
    filter_products_and_alternative,
    write_shared_log,
)

# Set up logger
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

try:
    import importlib.util

    if importlib.util.find_spec("xlsxwriter") is None:
        print(
            "The 'xlsxwriter' module is not installed. Please install it using 'pip install xlsxwriter'."
        )
        sys.exit(1)
except Exception:
    print("Error checking for 'xlsxwriter' module.")
    sys.exit(1)

# Load NABP/NPI list
included_nabp_npi = {
    "4528874": "1477571404",
    "2365422": "1659313435",
    "3974157": "1972560688",
    "320793": "1164437406",
    "4591055": "1851463087",
    "2348046": "1942303110",
    "4023610": "1407879588",
    "4025385": "1588706212",
    "4025311": "1588705446",
    "4026806": "1285860312",
    "4931350": "1750330775",
    "4024585": "1396768461",
    "4028026": "1497022438",
    "2643749": "1326490376",
}


# ---------------------------------------------------------------------------
# Tier summarization helper
# ---------------------------------------------------------------------------
def summarize_by_tier(df, col, from_val, to_val):
    filtered = df[(df[col] == from_val) & (df["FormularyTier"] == to_val)]
    pt = pd.pivot_table(
        filtered,
        values=["Rxs", "MemberID"],
        index=["Product Name"],
        aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
    )
    pt = pt.rename(columns={"Rxs": "Total Rxs", "MemberID": "Unique Members"})
    rxs = filtered["Rxs"].sum()
    members = filtered["MemberID"].nunique()
    return pt, rxs, members


def load_tier_disruption_data(file_paths):
    """Load all required data files for tier disruption processing."""
    # Load claims with fallback
    if not file_paths.get("reprice"):
        write_shared_log(
            "tier_disruption.py",
            "No reprice/template file provided.",
            status="ERROR",
        )
        print("No reprice/template file provided. Skipping claims loading.")
        return None

    try:
        claims = pd.read_excel(file_paths["reprice"], sheet_name="Claims Table")
    except Exception as e:
        logger.error(f"Error loading claims: {e}")
        claims = pd.read_excel(file_paths["reprice"], sheet_name=0)

    print(f"claims shape: {claims.shape}")
    claims.info()

    # Load reference tables
    medi = pd.read_excel(
        file_paths["medi_span"], usecols=["NDC", "Maint Drug?", "Product Name"]
    )
    print(f"medi shape: {medi.shape}")

    u = pd.read_excel(
        file_paths["u_disrupt"], sheet_name="Universal NDC", usecols=["NDC", "Tier"]
    )
    print(f"u shape: {u.shape}")

    e = pd.read_excel(
        file_paths["e_disrupt"],
        sheet_name="Alternatives NDC",
        usecols=["NDC", "Tier", "Alternative"],
    )
    print(f"e shape: {e.shape}")

    network = pd.read_excel(
        file_paths["n_disrupt"],
        usecols=["pharmacy_npi", "pharmacy_nabp", "pharmacy_is_excluded"],
    )
    print(f"network shape: {network.shape}")

    return claims, medi, u, e, network


def process_tier_data_pipeline(claims, reference_data, network):
    """Process the data pipeline for tier disruption."""
    medi, u, e = reference_data
    
    # Merge reference data
    df = claims.merge(medi, on="NDC", how="left")
    print(f"After merge with medi: {df.shape}")

    df = df.merge(
        u.rename(columns={"Tier": "Universal Tier"}), on="NDC", how="left"
    )
    print(f"After merge with u: {df.shape}")

    df = df.merge(
        e.rename(columns={"Tier": "Exclusive Tier"}), on="NDC", how="left"
    )
    print(f"After merge with e: {df.shape}")

    # Standardize IDs and perform network merge
    df = standardize_pharmacy_ids(df)
    print(f"After standardize_pharmacy_ids: {df.shape}")
    network = standardize_network_ids(network)
    print(f"After standardize_network_ids: {network.shape}")
    df = merge_with_network(df, network)
    print(f"After merge_with_network: {df.shape}")

    # Date parsing, deduplication, type cleaning, and filters
    df["DATEFILLED"] = pd.to_datetime(df["DATEFILLED"], errors="coerce")
    print(f"After DATEFILLED to_datetime: {df.shape}")

    df = drop_duplicates_df(df)
    print(f"After drop_duplicates_df: {df.shape}")

    df = clean_logic_and_tier(df)
    print(f"After clean_logic_and_tier: {df.shape}")

    df = filter_recent_date(df)
    print(f"After filter_recent_date: {df.shape}")

    df = filter_logic_and_maintenance(df)
    print(f"After filter_logic_and_maintenance: {df.shape}")

    df = filter_products_and_alternative(df)
    print(f"After filter_products_and_alternative: {df.shape}")

    return df


def handle_tier_pharmacy_exclusions(df, file_paths):
    """Handle pharmacy exclusions for tier disruption."""
    # Ensure 'pharmacy_is_excluded' column contains actual boolean values with type inference
    if "pharmacy_is_excluded" in df.columns:
        df["pharmacy_is_excluded"] = (
            df["pharmacy_is_excluded"]
            .astype(str)
            .str.lower()
            .map({"true": True, "false": False})
            .fillna(False)
            .infer_objects(copy=False)
        )
        logger.info(
            f"pharmacy_is_excluded value counts: {df['pharmacy_is_excluded'].value_counts().to_dict()}"
        )

        # Identify rows where pharmacy_is_excluded is "NA"
        na_pharmacies = df[df["pharmacy_is_excluded"].isna()]
        logger.info(f"NA pharmacies count: {na_pharmacies.shape[0]}")

        if not na_pharmacies.empty:
            output_file_path = file_paths["pharmacy_validation"]
            na_pharmacies_output = na_pharmacies[["PHARMACYNPI", "NABP"]].fillna("N/A")
            
            # Update existing template or create new file
            try:
                from openpyxl import load_workbook
                import os
                
                # Check if template exists
                if os.path.exists(output_file_path):
                    # Load existing workbook
                    wb = load_workbook(output_file_path)
                    logger.info(f"Loading existing template: {output_file_path}")
                else:
                    # Create new workbook
                    from openpyxl import Workbook
                    wb = Workbook()
                    logger.info(f"Creating new validation file: {output_file_path}")
                
                # Check if Validations sheet exists, if not create it
                if "Validations" in wb.sheetnames:
                    ws = wb["Validations"]
                    # Clear existing data
                    ws.delete_rows(1, ws.max_row)
                else:
                    ws = wb.create_sheet("Validations")
                    # Remove default sheet if it exists and is empty
                    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                        wb.remove(wb["Sheet"])
                
                # Write headers
                ws.append(["PHARMACYNPI", "NABP"])
                
                # Write data
                for _, row in na_pharmacies_output.iterrows():
                    ws.append([row["PHARMACYNPI"], row["NABP"]])
                
                # Save the workbook
                wb.save(output_file_path)
                logger.info(f"NA pharmacies written to '{output_file_path}' Validations sheet.")
                
            except Exception as e:
                logger.error(f"Error updating pharmacy validation file: {e}")
                # Fallback to original method
                writer = pd.ExcelWriter(output_file_path, engine="openpyxl")
                na_pharmacies_output.to_excel(
                    writer, sheet_name="Validations", index=False, engine="openpyxl"
                )
                writer.close()
                logger.info(f"NA pharmacies written to '{output_file_path}' sheet (fallback).")

    return df


def create_tier_definitions():
    """Create the tier definitions for analysis."""
    return [
        ("Universal_Positive 2-1", "Universal Tier", 1, 2),
        ("Universal_Positive 3-1", "Universal Tier", 1, 3),
        ("Universal_Positive 3-2", "Universal Tier", 2, 3),
        ("Universal_Negative 1-2", "Universal Tier", 2, 1),
        ("Universal_Negative 1-3", "Universal Tier", 3, 1),
        ("Universal_Negative 2-3", "Universal Tier", 3, 2),
        ("Exclusive_Positive 2-1", "Exclusive Tier", 1, 2),
        ("Exclusive_Positive 3-1", "Exclusive Tier", 1, 3),
        ("Exclusive_Positive 3-2", "Exclusive Tier", 2, 3),
        ("Exclusive_Negative 1-2", "Exclusive Tier", 2, 1),
        ("Exclusive_Negative 1-3", "Exclusive Tier", 3, 1),
        ("Exclusive_Negative 2-3", "Exclusive Tier", 3, 2),
    ]


def process_tier_pivots(df, tiers):
    """Process tier pivot tables and collect statistics."""
    tab_members = {}
    tab_rxs = {}
    tier_pivots = []

    for name, col, from_val, to_val in tiers:
        pt, rxs, members = summarize_by_tier(df, col, from_val, to_val)
        tier_pivots.append((name, pt, members, rxs))
        tab_members[name] = members
        tab_rxs[name] = rxs

    return tier_pivots, tab_members, tab_rxs


def process_exclusions(df):
    """Process exclusions data and create pivot table."""
    exclusions = df[df["Exclusive Tier"] == "Nonformulary"]
    ex_pt = exclusions.pivot_table(
        values=["Rxs", "MemberID"],
        index=["Product Name", "Alternative"],
        aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
    )
    ex_pt = ex_pt.rename(columns={"Rxs": "Total Rxs", "MemberID": "Unique Members"})
    exc_rxs = exclusions["Rxs"].sum()
    exc_members = exclusions["MemberID"].nunique()
    
    return ex_pt, exc_rxs, exc_members


def create_summary_dataframe(tab_members, tab_rxs, total_claims, total_members):
    """Create the summary DataFrame with calculated statistics."""
    uni_pos_keys = [
        "Universal_Positive 2-1",
        "Universal_Positive 3-1",
        "Universal_Positive 3-2",
    ]
    uni_neg_keys = [
        "Universal_Negative 1-2",
        "Universal_Negative 1-3",
        "Universal_Negative 2-3",
    ]
    ex_pos_keys = [
        "Exclusive_Positive 2-1",
        "Exclusive_Positive 3-1",
        "Exclusive_Positive 3-2",
    ]
    ex_neg_keys = [
        "Exclusive_Negative 1-2",
        "Exclusive_Negative 1-3",
        "Exclusive_Negative 2-3",
    ]

    uni_pos_utilizers = sum(tab_members[k] for k in uni_pos_keys)
    uni_pos_claims = sum(tab_rxs[k] for k in uni_pos_keys)
    uni_pos_pct = uni_pos_claims / total_claims if total_claims else 0

    uni_neg_utilizers = sum(tab_members[k] for k in uni_neg_keys)
    uni_neg_claims = sum(tab_rxs[k] for k in uni_neg_keys)
    uni_neg_pct = uni_neg_claims / total_claims if total_claims else 0

    ex_pos_utilizers = sum(tab_members[k] for k in ex_pos_keys)
    ex_pos_claims = sum(tab_rxs[k] for k in ex_pos_keys)
    ex_pos_pct = ex_pos_claims / total_claims if total_claims else 0

    ex_neg_utilizers = sum(tab_members[k] for k in ex_neg_keys)
    ex_neg_claims = sum(tab_rxs[k] for k in ex_neg_keys)
    ex_neg_pct = ex_neg_claims / total_claims if total_claims else 0

    exc_utilizers = tab_members["Exclusions"]
    exc_claims = tab_rxs["Exclusions"]
    exc_pct = exc_claims / total_claims if total_claims else 0

    return pd.DataFrame(
        {
            "Formulary": [
                "Universal Positive",
                "Universal Negative",
                "Exclusive Positive",
                "Exclusive Negative",
                "Exclusions",
            ],
            "Utilizers": [
                uni_pos_utilizers,
                uni_neg_utilizers,
                ex_pos_utilizers,
                ex_neg_utilizers,
                exc_utilizers,
            ],
            "Rxs": [
                uni_pos_claims,
                uni_neg_claims,
                ex_pos_claims,
                ex_neg_claims,
                exc_claims,
            ],
            "% of claims": [
                uni_pos_pct,
                uni_neg_pct,
                ex_pos_pct,
                ex_neg_pct,
                exc_pct,
            ],
            "": ["", "", "", "", ""],
            "Totals": [
                f"Members: {total_members}",
                f"Claims: {total_claims}",
                "",
                "",
                "",
            ],
        }
    )


def create_network_analysis(df):
    """Create network analysis for excluded pharmacies."""
    network_df = df[df["pharmacy_is_excluded"]]
    logger.debug(f"Initial network_df shape: {network_df.shape}")
    logger.debug(f"Initial network_df contents: {network_df.head(10).to_dict()}")

    filter_phrases = [
        "CVS",
        "Walgreens",
        "Kroger",
        "Walmart",
        "Rite Aid",
        "Optum",
        "Express Scripts",
        "DMR",
        "Williams Bro",
        "Publix",
    ]
    filter_phrases_escaped = [
        re.escape(phrase.lower()) for phrase in filter_phrases
    ]
    regex_pattern = "|".join([f"\\b{p}\\b" for p in filter_phrases_escaped])
    network_df = network_df[
        ~network_df["Pharmacy Name"]
        .str.lower()
        .str.contains(regex_pattern, case=False, regex=True, na=False)
    ]
    logger.debug(f"Filtered network_df shape: {network_df.shape}")
    logger.debug(f"Filtered network_df contents: {network_df.head(10).to_dict()}")

    if {"PHARMACYNPI", "NABP", "Pharmacy Name"}.issubset(network_df.columns):
        network_df["PHARMACYNPI"] = network_df["PHARMACYNPI"].fillna("N/A")
        network_df["NABP"] = network_df["NABP"].fillna("NABP")
        # Create Unique Identifier based on whichever identifier is available (PHARMACYNPI or NABP)
        network_df["Unique Identifier"] = network_df.apply(
            lambda row: row["PHARMACYNPI"]
            if row["PHARMACYNPI"] != "N/A"
            else row["NABP"],
            axis=1,
        )
        network_pivot = pd.pivot_table(
            network_df,
            values=["Rxs", "MemberID"],
            index=["Unique Identifier"],
            aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
        )
        network_pivot = network_pivot.rename(
            columns={"Rxs": "Total Rxs", "MemberID": "Unique Members"}
        )
        network_pivot.reset_index(
            inplace=True
        )  # Ensure index columns are included in the output
        logger.debug(f"Network pivot shape: {network_pivot.shape}")
        logger.debug(f"Network pivot contents: {network_pivot.head(10).to_dict()}")
        
        return network_df, network_pivot
    
    return network_df, None


def write_excel_sheets(writer, df, summary_df, tier_pivots, ex_pt, exc_members, network_df, network_pivot):
    """Write all sheets to the Excel file."""
    # Write Summary sheet
    summary_df.to_excel(writer, sheet_name="Summary", index=False)

    # Write tier pivots
    for name, pt, members, _ in tier_pivots:
        pt.to_excel(writer, sheet_name=name)
        writer.sheets[name].write("F1", f"Total Members: {members}")

    # Write Exclusions sheet
    ex_pt.to_excel(writer, sheet_name="Exclusions")
    writer.sheets["Exclusions"].write("F1", f"Total Members: {exc_members}")

    # Write Data sheet
    data_sheet_df = df.copy()
    data_sheet_df.to_excel(writer, sheet_name="Data", index=False)

    # Write Network sheet
    if network_pivot is not None:
        network_pivot.to_excel(writer, sheet_name="Network", index=False)

    # Write filtered network data
    selected_columns = [
        "PHARMACYNPI",
        "NABP",
        "MemberID",
        "Pharmacy Name",
        "pharmacy_is_excluded",
        "Unique Identifier",
    ]
    network_df[selected_columns].to_excel(writer, sheet_name="Network", index=False)
    logger.info(
        f"Network sheet updated with {network_df.shape[0]} excluded pharmacy records (minus major chains) and selected columns"
    )


def reorder_excel_sheets(writer):
    """Reorder sheets so Summary follows Data."""
    sheets = writer.sheets  # This is a dict: {sheet_name: worksheet_object}
    names = list(sheets.keys())
    if "Data" in names and "Summary" in names:
        data_idx = names.index("Data")
        summary_idx = names.index("Summary")
        if summary_idx != data_idx + 1:
            # Move "Summary" sheet immediately after "Data"
            items = list(sheets.items())
            summary_item = items.pop(summary_idx)
            items.insert(data_idx + 1, summary_item)
            writer.sheets.clear()
            writer.sheets.update(dict(items))


def show_completion_message(output_path):
    """Show completion message and popup."""
    write_shared_log("tier_disruption.py", "Processing complete.")
    print(f"Processing complete. Output file: {output_path}")
    try:
        import tkinter as tk
        from tkinter import messagebox

        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo("Processing Complete", "Processing complete")
        root.destroy()
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Main processing pipeline
# ---------------------------------------------------------------------------
def process_data():
    write_shared_log("tier_disruption.py", "Processing started.")
    # Output filename from CLI arg or default
    output_filename = "LBL for Disruption.xlsx"
    if len(sys.argv) > 1:
        output_filename = sys.argv[1]
    output_path = Path(output_filename).resolve()
    try:
        # Get the config file path relative to the project root
        config_path = Path(__file__).parent.parent / "config" / "file_paths.json"
        file_paths = load_file_paths(str(config_path))

        result = load_tier_disruption_data(file_paths)
        if result is None:
            return  # Early exit if claims loading failed
        claims, medi, u, e, network = result

        reference_data = (medi, u, e)
        df = process_tier_data_pipeline(claims, reference_data, network)

        df = handle_tier_pharmacy_exclusions(df, file_paths)

        # Totals for summary
        total_claims = df["Rxs"].sum()
        total_members = df["MemberID"].nunique()

        # Excel writer setup
        writer = pd.ExcelWriter(output_path, engine="xlsxwriter")

        # Summary calculations (must be written immediately after Data)
        tiers = create_tier_definitions()

        tier_pivots, tab_members, tab_rxs = process_tier_pivots(df, tiers)

        # Exclusions sheet (Nonformulary)
        ex_pt, exc_rxs, exc_members = process_exclusions(df)
        tab_members["Exclusions"] = exc_members
        tab_rxs["Exclusions"] = exc_rxs

        # Summary calculations
        summary_df = create_summary_dataframe(tab_members, tab_rxs, total_claims, total_members)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # Write tier pivots and Exclusions after Summary
        for name, pt, members, _ in tier_pivots:
            pt.to_excel(writer, sheet_name=name)
            writer.sheets[name].write("F1", f"Total Members: {members}")

        ex_pt.to_excel(writer, sheet_name="Exclusions")
        writer.sheets["Exclusions"].write("F1", f"Total Members: {exc_members}")

        # Write the 'Data Sheet' with excluded and non-excluded pharmacies
        data_sheet_df = df.copy()
        data_sheet_df.to_excel(writer, sheet_name="Data", index=False)

        # Network summary for excluded pharmacies (pharmacy_is_excluded=True)
        network_df, network_pivot = create_network_analysis(df)
        logger.info(f"Total pharmacies in dataset: {df.shape[0]}")
        logger.info(
            f"Excluded pharmacies (pharmacy_is_excluded=True): {df['pharmacy_is_excluded'].sum()}"
        )
        logger.info(
            f"Non-excluded pharmacies (pharmacy_is_excluded=False): {(~df['pharmacy_is_excluded']).sum()}"
        )
        logger.info(
            f"Network sheet will show {network_df.shape[0]} excluded pharmacy records (minus major chains)"
        )

        # Write Network sheet
        if network_pivot is not None:
            network_pivot.to_excel(writer, sheet_name="Network", index=False)

        # Write filtered network data
        selected_columns = [
            "PHARMACYNPI",
            "NABP",
            "MemberID",
            "Pharmacy Name",
            "pharmacy_is_excluded",
            "Unique Identifier",
        ]
        network_df[selected_columns].to_excel(writer, sheet_name="Network", index=False)
        logger.info(
            f"Network sheet updated with {network_df.shape[0]} excluded pharmacy records (minus major chains) and selected columns"
        )

        # Reorder sheets so Summary follows Data
        reorder_excel_sheets(writer)

        writer.close()
        show_completion_message(output_path)
    except Exception as e:
        write_shared_log(
            "tier_disruption.py", f"Processing failed: {e}", status="ERROR"
        )
        raise


if __name__ == "__main__":
    process_data()
    print("Data processing complete")
"""
Template processing module for handling Excel template operations.
Extracted from app.py to improve cohesion and reduce file size.

This module provides:
- Template backup creation
- Excel data formatting
- Column filtering for templates
- Data preparation for Excel export
"""

import pandas as pd
import shutil
import os
import sys
from pathlib import Path
import logging

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.utils import write_shared_log


class TemplateProcessor:
    """
    Handles Excel template operations with a focus on simplicity and reliability.
    
    This class manages template backup, data formatting, and Excel export operations
    while maintaining separation of concerns from the main application logic.
    """
    
    def __init__(self, app_instance):
        """Initialize with reference to the main application instance."""
        self.app = app_instance
        
    def create_template_backup(self, paths):
        """Create backup of template and prepare output file."""
        try:
            # Backup original template
            shutil.copy(paths["template"], paths["backup"])
            logging.info(f"Template backed up to {paths['backup']}")
            
            # Remove old output if it exists
            if paths["output"].exists():
                try:
                    os.remove(paths["output"])
                except PermissionError:
                    raise RuntimeError(
                        f"Cannot overwrite {paths['output']} — please close it in Excel."
                    )
            
            # Copy template to output location
            shutil.copy(paths["template"], paths["output"])
            write_shared_log("TemplateProcessor", f"Template backup created: {paths['backup']}")
            
        except Exception as e:
            error_msg = f"Failed to create template backup: {str(e)}"
            logging.error(error_msg)
            write_shared_log("TemplateProcessor", error_msg, "ERROR")
            raise
    
    def format_dataframe(self, df):
        """Format DataFrame for Excel export."""
        # Format datetime columns
        datetime_columns = df.select_dtypes(include=["datetime64"]).columns
        for col in datetime_columns:
            df[col] = df[col].dt.strftime("%Y-%m-%d %H:%M:%S")
        
        # Fill NaN values
        return df.fillna("")
    
    def filter_template_columns(self, df):
        """Filter columns for template pasting."""
        try:
            # Ensure 'ClientName' and 'Logic' columns exist and are in the correct order
            if "ClientName" in df.columns and "Logic" in df.columns:
                client_name_idx = df.columns.get_loc("ClientName")
                logic_idx = df.columns.get_loc("Logic")
                
                if client_name_idx <= logic_idx:
                    # Select columns from 'ClientName' to 'Logic' (inclusive)
                    selected_columns = df.columns[client_name_idx : logic_idx + 1]
                    logging.info(f"Pasting only these columns: {selected_columns.tolist()}")
                    return df[selected_columns]
                else:
                    logging.warning(
                        "'Logic' column appears before 'ClientName'; returning full DataFrame."
                    )
                    return df
            else:
                raise ValueError("Required columns 'ClientName' or 'Logic' are missing.")
                
        except Exception as e:
            logging.warning(f"Error filtering columns: {e}. Using full DataFrame.")
            return df
    
    def prepare_template_data(self, processed_file):
        """Prepare data for template pasting."""
        try:
            df = pd.read_excel(processed_file)
            df = self.format_dataframe(df)
            
            return {
                "data": df.values,
                "nrows": df.shape[0],
                "ncols": df.shape[1]
            }
        except Exception as e:
            error_msg = f"Failed to prepare template data: {str(e)}"
            logging.error(error_msg)
            write_shared_log("TemplateProcessor", error_msg, "ERROR")
            raise
    
    def prepare_excel_data(self, paste_data, formulas):
        """Prepare data for Excel, preserving formulas."""
        data_to_write = []
        
        for i in range(paste_data["nrows"]):
            row = []
            for j in range(paste_data["ncols"]):
                if formulas[i][j] == "":
                    row.append(paste_data["data"][i][j])
                else:
                    row.append(None)
            data_to_write.append(row)
        
        return data_to_write
    
    def validate_template_requirements(self, template_path):
        """Validate that template meets requirements."""
        if not template_path:
            raise ValueError("Template file path is not set.")
            
        template = Path(template_path)
        if not template.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")
            
        if template.suffix != '.xlsx':
            raise ValueError("Template must be an Excel file (.xlsx)")
            
        return True
    
    def show_toast(self, message, duration=3000):
        """Show a toast notification."""
        try:
            import tkinter as tk
            from tkinter import messagebox
            
            toast = tk.Toplevel(self.app.root)
            toast.overrideredirect(True)
            toast.configure(bg="black")

            # Position bottom-right
            self.app.root.update_idletasks()
            screen_width = toast.winfo_screenwidth()
            screen_height = toast.winfo_screenheight()
            x = screen_width - 320
            y = screen_height - 100
            toast.geometry(f"300x50+{x}+{y}")

            label = tk.Label(
                toast, text=message, bg="black", fg="white", font=("Arial", 11)
            )
            label.pack(fill="both", expand=True)

            toast.after(duration, toast.destroy)
            
        except Exception as e:
            logging.warning(f"Toast notification failed: {e}")
            # Fallback to messagebox
            from tkinter import messagebox
            messagebox.showinfo("Notification", message)
import logging
import tkinter as tk
from pathlib import Path
from tkinter import messagebox
import os
import sys

import pandas as pd
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.excel_utils import write_df_to_template
from utils.utils import load_file_paths, write_shared_log

CLAIMS_SHEET = "Claims Table"
OUTPUT_SHEET = "Line By Line"

# Setup logging
logging.basicConfig(
    filename="sharx_lbl.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


def show_message(awp: float, ing: float, total: float, rxs: int) -> None:
    messagebox.showinfo(
        "Process Complete",
        f"SHARx LBL has been created\n\n"
        f"Total AWP: ${awp:.2f}\n\n"
        f"Total Ing Cost: ${ing:.2f}\n\n"
        f"Total Gross Cost: ${total:.2f}\n\n"
        f"Total Claim Count: {rxs}",
    )


def main():
    root = tk.Tk()
    root.withdraw()
    write_shared_log("sharx_lbl.py", "Processing started.")

    try:
        # Get the config file path relative to the project root
        config_path = Path(__file__).parent.parent / "config" / "file_paths.json"
        paths = load_file_paths(str(config_path))
        # Fallback to file dialogs if required keys are missing
        if "reprice" not in paths:
            from tkinter import filedialog

            paths["reprice"] = filedialog.askopenfilename(title="Select Claims File")
        if "sharx" not in paths:
            from tkinter import filedialog

            paths["sharx"] = filedialog.askopenfilename(
                title="Select SHARx Template File"
            )

        template_path = Path(paths["sharx"])
        try:
            df = pd.read_excel(paths["reprice"], sheet_name=CLAIMS_SHEET)
        except FileNotFoundError:
            logger.error(f"Claims file not found: {paths['reprice']}")
            raise FileNotFoundError(f"Claims file not found: {paths['reprice']}")
        except ValueError as e:
            logger.error(f"Sheet loading failed: {e}")
            raise ValueError(f"Sheet loading failed: {e}")

        # Debug print and log
        print("Columns in claims DataFrame:")
        print(df.columns)
        logger.info(f"Columns in claims DataFrame: {df.columns.tolist()}")

        df["Logic"] = pd.to_numeric(df["Logic"], errors="coerce")
        df = df[df["Logic"].between(1, 10)]

        awp = df["Total AWP (Historical)"].sum()
        ing = df["Rx Sense Ing Cost"].sum()
        total = df["RxSense Total Cost"].sum()
        rxs = df["Rxs"].sum()

        columns_to_keep = [
            "MONY",
            "Rxs",
            "Rx Sense Ing Cost",
            "RxSense Dispense Fee",
            "RxSense Total Cost",
            "Total AWP (Historical)",
            "GrossCost",
            "Pharmacy Name",
            "INPUTFILECHANNEL",
            "DATEFILLED",
            "MemberID",
            "DAYSUPPLY",
            "QUANTITY",
            "NDC",
            "DST Drug Name",
        ]
        missing_cols = [col for col in columns_to_keep if col not in df.columns]
        if missing_cols:
            raise ValueError(f"Missing columns in input data: {missing_cols}")
        df = df[columns_to_keep]

        output_path = Path("_Rx Claims for SHARx.xlsx")

        write_df_to_template(
            str(template_path),
            str(output_path),
            sheet_name=OUTPUT_SHEET,
            df=df,
            start_cell="A2",
            header=False,
            index=False,
            open_file=False,
            visible=False,
        )

        logger.info(f"SHARx output saved to: {output_path}")
        logger.info("SHARx LBL file created successfully.")
        write_shared_log("sharx_lbl.py", "SHARx LBL file created successfully.")
        show_message(awp, ing, total, rxs)
        messagebox.showinfo("Processing Complete", "Processing complete")

    except Exception as e:
        logger.exception("An error occurred during SHARx LBL processing")
        write_shared_log("sharx_lbl.py", f"An error occurred: {e}", status="ERROR")
        messagebox.showerror("Error", f"An error occurred: {e}")
    finally:
        root.quit()


if __name__ == "__main__":
    main()
 import tkinter as tk
import customtkinter as ctk
from tkinter import filedialog, messagebox, scrolledtext
import subprocess
import os
from utils.utils import write_shared_log
import logging
import threading
import multiprocessing
import pandas as pd
from typing import Optional
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pathlib import Path
import json
import time
import re
import importlib
import importlib.util
import warnings

# Import custom modules
# Theme colors now handled by UIBuilder
from config.app_config import ProcessingConfig, AppConstants
from modules.file_processor import FileProcessor
from modules.template_processor import TemplateProcessor
from modules.data_processor import DataProcessor
from modules.process_manager import ProcessManager
from modules.ui_builder import UIBuilder
from modules.log_manager import LogManager, ThemeController

# Excel COM check
XLWINGS_AVAILABLE = importlib.util.find_spec("xlwings") is not None
EXCEL_COM_AVAILABLE = importlib.util.find_spec("win32com.client") is not None

# Logging setup
logging.basicConfig(
    filename="repricing_log.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


class ConfigManager:
    def __init__(self):
        self.config = {}
        if AppConstants.CONFIG_FILE.exists():
            self.load()
        else:
            self.save_default()

    def save_default(self):
        self.config = {"last_folder": str(Path.cwd())}
        self.save()

    def load(self):
        with open(AppConstants.CONFIG_FILE, "r") as f:
            self.config = json.load(f)

    def save(self):
        with open(AppConstants.CONFIG_FILE, "w") as f:
            json.dump(self.config, f, indent=4)


class App:
    def __init__(self, root):
        self.root = root
        self._initialize_variables()
        self._initialize_processors()
        self.ui_builder.build_complete_ui()
        self.theme_controller.apply_initial_theme()
        self.log_manager.initialize_logging()
        self.log_manager.log_application_start()

    def _initialize_variables(self):
        """Initialize all instance variables."""
        self.file1_path = None
        self.file2_path = None
        self.template_file_path = None
        self.cancel_event = threading.Event()
        self.start_time = None
        self.progress_var = tk.DoubleVar(value=0.0)
        self.progress_label_var = tk.StringVar(value="0%")
        self.config_manager = ConfigManager()
        self.selected_disruption_type = tk.StringVar(value="Tier")
        self.file1_label: Optional[ctk.CTkLabel] = None
        self.file2_label: Optional[ctk.CTkLabel] = None
        self.template_label: Optional[ctk.CTkLabel] = None
        self.toggle_theme_button: Optional[ctk.CTkButton] = None
        self.progress_bar: Optional[ctk.CTkProgressBar] = None
        self.progress_label: Optional[ctk.CTkLabel] = None

    def _initialize_processors(self):
        """Initialize all processor and manager instances."""
        self.file_processor = FileProcessor(self)
        self.template_processor = TemplateProcessor(self)
        self.data_processor = DataProcessor(self)
        self.process_manager = ProcessManager(self)
        self.ui_builder = UIBuilder(self)
        self.log_manager = LogManager(self)
        self.theme_controller = ThemeController(self)

    # The following methods are moved to their respective manager classes for better cohesion:
    # - apply_theme_colors -> ThemeController
    # - check_template -> FileProcessor
    # - sharx_lbl, epls_lbl -> ProcessManager
    # - show_shared_log_viewer -> LogManager

    # Example: Remove apply_theme_colors from App, and use self.theme_controller.apply_theme_colors instead.



    def import_file1(self):
        """Import the first file with template validation using guard clauses."""
        file_path = self._get_file_path("Select File Uploaded to Tool")
        if not file_path:
            return  # User cancelled
        
        self._set_file1_path(file_path)
        self._validate_gross_cost_template(file_path)

    def _get_file_path(self, title):
        """Get file path from file dialog."""
        return filedialog.askopenfilename(
            title=title,
            filetypes=ProcessingConfig.FILE_TYPES,
        )

    def _set_file1_path(self, file_path):
        """Set the file1 path and update UI."""
        self.file1_path = file_path
        if self.file1_label:
            self.file1_label.configure(text=os.path.basename(file_path))
        self.file_processor.check_template(file_path)
        write_shared_log("File1 imported", file_path)

    def _validate_gross_cost_template(self, file_path):
        """Validate GrossCost column and suggest template type using data processor."""
        template_suggestion = self.data_processor.validate_gross_cost_template(file_path)
        if template_suggestion:
            messagebox.showinfo("Template Selection", template_suggestion)

    def import_file2(self):
        """Import the second file."""
        file_path = self._get_file_path("Select File From Tool")
        if not file_path:
            return  # User cancelled
        
        self.file2_path = file_path
        if self.file2_label:
            self.file2_label.configure(text=os.path.basename(file_path))
        write_shared_log("File2 imported", file_path)

    def import_template_file(self):
        """Import the template file."""
        file_path = filedialog.askopenfilename(
            title="Select Template File", filetypes=ProcessingConfig.TEMPLATE_FILE_TYPES
        )
        if not file_path:
            return  # User cancelled
        
        self.template_file_path = file_path
        if self.template_label:
            self.template_label.configure(text=os.path.basename(file_path))
        write_shared_log("Template file imported", file_path)

    # Logging and notification methods
    # Removed duplicate write_audit_log method to resolve method name conflict.

    # Cancel during repricing
    def cancel_process(self):
        """Cancel the process using process manager."""
        self.process_manager.cancel_process()

    # Live log viewer (old version, renamed to avoid conflict)
    def show_log_viewer_old(self):
        win = tk.Toplevel(self.root)
        win.title("Log Viewer")
        txt = scrolledtext.ScrolledText(win, width=100, height=30)
        txt.pack(fill="both", expand=True)

        def refresh():
            with open("repricing_log.log", "r") as f:
                txt.delete("1.0", tk.END)
                txt.insert(tk.END, f.read())
            win.after(3000, refresh)

        refresh()

    def update_progress(self, value=None, message=None):
        """Update the progress bar and label with reduced complexity."""
        def do_update():
            if value is None:
                self._set_indeterminate_progress(message)
            else:
                self._set_determinate_progress(value, message)
            self.root.update_idletasks()

        self._schedule_ui_update(do_update)

    def _set_indeterminate_progress(self, message):
        """Set progress bar to indeterminate mode."""
        if self.progress_bar:
            self.progress_bar.configure(mode="indeterminate")
            self.progress_bar.start()
        display_message = message or "Processing... (unknown duration)"
        self.progress_label_var.set(display_message)

    def _set_determinate_progress(self, value, message):
        """Set progress bar to determinate mode with specific value."""
        if self.progress_bar:
            if self.progress_bar.cget("mode") != "determinate":
                self.progress_bar.stop()
                self.progress_bar.configure(mode="determinate")
            
            self.progress_bar.set(value)
        
        self.progress_var.set(value)
        
        if message:
            self.progress_label_var.set(message)
        else:
            self._set_calculated_progress_message(value)

    def _set_calculated_progress_message(self, value):
        """Calculate and set progress message with time estimates."""
        percent = int(value * 100)
        elapsed = time.time() - self.start_time if self.start_time else 0
        est = int((elapsed / value) * (1 - value)) if value > 0 else 0
        self.progress_label_var.set(f"Progress: {percent}% | Est. {est}s left")

    def _schedule_ui_update(self, update_func):
        """Schedule UI update on main thread or execute immediately."""
        if threading.current_thread() is threading.main_thread():
            update_func()
        else:
            self.root.after(0, update_func)

    def write_audit_log(self, file1, file2, status):
        """Write audit log entry using file processor."""
        return self.file_processor.write_audit_log(file1, file2, status)

    def show_log_viewer(self):
        """Show log viewer using log manager."""
        self.log_manager.show_log_viewer()

    # Disruption and process methods
    # Removed select_disruption_type method since disruption_type_combobox does not exist.

    def start_disruption(self, disruption_type=None):
        """Start disruption processing using process manager."""
        self.process_manager.start_disruption(disruption_type)

    def start_process_threaded(self):
        """Start the repricing process using the process manager."""
        self.process_manager.start_process_threaded()

    def finish_notification(self):
        """Show completion notification using process manager."""
        self.process_manager.finish_notification()

    # Repricing workflow methods
    def paste_into_template(self, processed_file):
        """Paste processed data into Excel template using background threading."""
        def run_in_background():
            try:
                self._execute_template_paste(processed_file)
            except Exception as e:
                logger.exception("Error during paste with xlwings")
                self.root.after(
                    0,
                    lambda e=e: messagebox.showerror(
                        "Error", f"Template update failed:\n{e}"
                    ),
                )
                self.root.after(0, lambda: self.update_progress(0))

        threading.Thread(target=run_in_background, daemon=True).start()

    def _execute_template_paste(self, processed_file):
        """Execute the template paste operation with proper error handling."""
        import time

        start_time = time.time()
        
        # Initialize progress
        self.root.after(
            0,
            lambda: self.update_progress(None, "Preparing to paste into template..."),
        )

        # Validate template path
        if not self.template_file_path:
            raise ValueError("Template file path is not set.")

        # Prepare data and paths
        paste_data = self._prepare_template_data(processed_file)
        paths = self._prepare_template_paths()
        
        # Create backup and setup output file
        self._create_template_backup(paths)
        
        # Execute Excel operations
        self._execute_excel_paste(paste_data, paths)
        
        # Finalize and notify
        elapsed = time.time() - start_time
        msg = f"Template updated successfully in {elapsed:.2f} seconds."
        logger.info(msg)
        self.root.after(0, lambda: self.update_progress(1.0, msg))
        self.root.after(0, lambda: self.show_toast(msg))
        self.root.after(
            0,
            lambda: messagebox.showinfo(
                "Template Update Complete",
                "Pasting into the template is complete. You may now review the updated file.",
            ),
        )

    def _prepare_template_data(self, processed_file):
        """Prepare data for template pasting."""
        df = pd.read_excel(processed_file)
        df = self.format_dataframe(df)
        return {
            "data": df.values,
            "nrows": df.shape[0],
            "ncols": df.shape[1]
        }

    def _prepare_template_paths(self):
        """Prepare file paths for template operations using file processor."""
        return self.file_processor.prepare_file_paths(self.template_file_path)

    def _create_template_backup(self, paths):
        """Create backup of template and prepare output file using template processor."""
        self.template_processor.create_template_backup(paths)

    def _execute_excel_paste(self, paste_data, paths):
        """Execute the Excel paste operation."""
        import xlwings as xw
        
        # Start Excel session
        app = xw.App(visible=False)
        wb = app.books.open(str(paths["output"]))
        ws = wb.sheets["Claims Table"]

        try:
            # Batch read formulas and prepare data
            formulas = ws.range((2, 1), (paste_data["nrows"] + 1, paste_data["ncols"])).formula
            data_to_write = self._prepare_excel_data(paste_data, formulas)
            
            # Paste values with progress updates
            self._paste_data_with_progress(ws, data_to_write, paste_data["nrows"], paste_data["ncols"])
            
            # Save and close
            wb.save()
            wb.close()
            app.quit()
            
        except Exception as e:
            # Ensure Excel is closed even on error
            try:
                wb.close()
                app.quit()
            except Exception:
                pass
            raise e

    def _prepare_excel_data(self, paste_data, formulas):
        """Prepare data for Excel, preserving formulas."""
        data_to_write = []
        
        for i in range(paste_data["nrows"]):
            row = []
            for j in range(paste_data["ncols"]):
                if formulas[i][j] == "":
                    row.append(paste_data["data"][i][j])
                else:
                    row.append(None)
            data_to_write.append(row)
        
        return data_to_write

    def _paste_data_with_progress(self, ws, data_to_write, nrows, ncols):
        """Paste data to Excel with progress updates."""
        # Paste values
        ws.range((2, 1), (nrows + 1, ncols)).value = data_to_write
        
        # Update progress periodically
        for i in range(0, nrows, 250):
            percent = 0.94 + 0.04 * (i / max(1, nrows))
            msg = f"Pasting row {i + 1} of {nrows}..."
            self.root.after(
                0, lambda v=percent, m=msg: self.update_progress(v, m)
            )

    def filter_template_columns(self, df):
        try:
            # Ensure 'ClientName' and 'Logic' columns exist and are in the correct order
            if "ClientName" in df.columns and "Logic" in df.columns:
                client_name_idx = df.columns.get_loc("ClientName")
                logic_idx = df.columns.get_loc("Logic")
                if client_name_idx <= logic_idx:
                    # Select columns from 'ClientName' to 'Logic' (inclusive)
                    selected_columns = df.columns[client_name_idx : logic_idx + 1]
                    logger.info(
                        f"Pasting only these columns: {selected_columns.tolist()}"
                    )
                    return df[selected_columns]
                else:
                    logger.warning(
                        "'Logic' column appears before 'ClientName'; returning full DataFrame."
                    )
                    return df
            else:
                raise ValueError(
                    "Required columns 'ClientName' or 'Logic' are missing."
                )
        except Exception as e:
            logger.warning(f"Error filtering columns: {e}. Using full DataFrame.")
            return df

    def format_dataframe(self, df):
        """Format DataFrame using template processor."""
        return self.template_processor.format_dataframe(df)

    def show_toast(self, message, duration=3000):
        """Show toast notification using template processor."""
        return self.template_processor.show_toast(message, duration)

    def start_process(self):
        threading.Thread(target=self._start_process_internal).start()
        write_shared_log("Repricing process started", "")

    def validate_merge_inputs(self):
        """Validate merge inputs using data processor."""
        is_valid, message = self.data_processor.validate_merge_inputs(
            self.file1_path, self.file2_path
        )
        if not is_valid:
            messagebox.showerror("Error", message)
        return is_valid

    def _start_process_internal(self):
        self.start_time = time.time()
        self.update_progress(0.05)

        # Extra safeguard: Remove any accidental LBL/disruption output during repricing
        os.environ["NO_LBL_OUTPUT"] = "1"

        if not self.file1_path or not self.file2_path:
            self.update_progress(0)
            messagebox.showerror("Error", "Please select both files before proceeding.")
            return

        if not self.validate_merge_inputs():
            self.update_progress(0)
            return

        try:
            self.update_progress(0.10)
            subprocess.run(
                ["taskkill", "/F", "/IM", "excel.exe"],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )
            self.update_progress(0.20)
            subprocess.run(
                ["python", "merge.py", self.file1_path, self.file2_path], check=True
            )
            self.update_progress(0.50)
            MERGED_FILE = "merged_file.xlsx"
            self.process_merged_file(MERGED_FILE)
            self.update_progress(0.90)
            # After all processing is done
            self.update_progress(1.0)
            # Ensure LBL scripts are NOT called here or in process_merged_file
        except subprocess.CalledProcessError as e:
            self.update_progress(0)
            logger.exception("Failed to run merge.py")
            messagebox.showerror("Error", f"Failed to run merge.py: {e}")

    def process_merged_file(self, file_path):
        """Process merged file with reduced complexity using helper methods."""
        try:
            # Initialize processing
            self._initialize_processing()
            
            # Load and validate data
            df = self._load_and_validate_data(file_path)
            
            # Process data using multiprocessing
            processed_df = self._process_data_multiprocessing(df)
            
            # Save outputs
            output_file = self._save_processed_outputs(processed_df)
            
            # Finalize processing
            self._finalize_processing(output_file)
            
        except Exception as e:
            logger.error(f"Error processing merged file: {e}")
            messagebox.showerror("Error", f"Processing failed: {e}")

    def _initialize_processing(self):
        """Initialize the processing environment."""
        self.update_progress(0.55)
        open("repricing_log.log", "w").close()
        logging.basicConfig(
            filename="repricing_log.log",
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
        logging.info("Starting merged file processing")

    def _load_and_validate_data(self, file_path):
        """Load and validate the merged file data."""
        df = pd.read_excel(file_path)
        logging.info(f"Loaded {len(df)} records from {file_path}")
        self.update_progress(0.60)

        # Validate required columns using configuration
        ProcessingConfig.validate_required_columns(df)

        # Prepare data for processing
        df = df.sort_values(by=["DATEFILLED", "SOURCERECORDID"], ascending=True)
        df["Logic"] = ""
        df["RowID"] = np.arange(len(df))
        
        return df

    def _process_data_multiprocessing(self, df):
        """Process data using multiprocessing for improved performance."""
        self.update_progress(0.65)
        
        # Import and use multiprocessing helpers
        from modules import mp_helpers

        num_workers = ProcessingConfig.get_multiprocessing_workers()
        df_blocks = np.array_split(df, num_workers)
        out_queue = multiprocessing.Queue()
        processes = []
        
        # Start worker processes
        for block in df_blocks:
            p = multiprocessing.Process(
                target=mp_helpers.worker, args=(block, out_queue)
            )
            p.start()
            processes.append(p)
        
        # Collect results
        results = [out_queue.get() for _ in processes]
        for p in processes:
            p.join()
        
        processed_df = pd.concat(results)
        self.update_progress(0.75)
        
        return processed_df

    def _save_processed_outputs(self, df):
        """Save processed data to various output formats."""
        # Sort and filter data
        df_sorted = pd.concat([df[df["Logic"] == ""], df[df["Logic"] == "OR"]])
        
        # Prepare output directory and files
        output_dir = Path.cwd()
        output_file = output_dir / "merged_file_with_OR.xlsx"
        
        # Create row mapping for highlighting
        row_mapping = {
            row["RowID"]: i + 2 for i, (_, row) in enumerate(df_sorted.iterrows())
        }
        excel_rows_to_highlight = [
            row_mapping[rid] for rid in [] if rid in row_mapping
        ]  # Placeholder
        
        # Clean up data
        df_sorted.drop(columns=["RowID"], inplace=True, errors="ignore")
        
        # Save to multiple formats
        self._save_to_parquet(df_sorted, output_dir)
        self._save_to_excel(df_sorted, output_file)
        self._save_to_csv(df_sorted, output_dir)
        
        # Save unmatched reversals info
        self._save_unmatched_reversals(excel_rows_to_highlight, output_dir)
        
        self.update_progress(0.80)
        return output_file

    def _save_to_parquet(self, df, output_dir):
        """Save data to Parquet format for large DataFrames."""
        try:
            parquet_path = output_dir / "merged_file_with_OR.parquet"
            df.drop_duplicates().to_parquet(parquet_path, index=False)
            logger.info(f"Saved intermediate Parquet file: {parquet_path}")
        except Exception as e:
            logger.warning(f"Could not save Parquet: {e}")

    def _save_to_excel(self, df, output_file):
        """Save data to Excel format."""
        df.drop_duplicates().to_excel(output_file, index=False)

    def _save_to_csv(self, df, output_dir):
        """Save data to CSV format with opportunity name."""
        opportunity_name = self._extract_opportunity_name()
        csv_path = output_dir / f"{opportunity_name} Claim Detail.csv"
        df.drop_duplicates().to_csv(csv_path, index=False)

    def _save_unmatched_reversals(self, excel_rows_to_highlight, output_dir):
        """Save unmatched reversals information."""
        unmatched_path = output_dir / "unmatched_reversals.txt"
        with open(unmatched_path, "w") as f:
            f.write(",".join(map(str, excel_rows_to_highlight)))

    def _extract_opportunity_name(self):
        """Extract opportunity name from file1_path."""
        opportunity_name = ProcessingConfig.DEFAULT_OPPORTUNITY_NAME
        try:
            if self.file1_path:
                if self.file1_path.lower().endswith(".xlsx"):
                    df_file1 = pd.read_excel(self.file1_path)
                else:
                    df_file1 = pd.read_csv(self.file1_path)
                
                if df_file1.shape[1] >= 2:
                    # Get the value from the first row, second column
                    raw_name = str(df_file1.iloc[0, 1])
                    # Clean for filename
                    opportunity_name = re.sub(r'[\\/*?:"<>|]', "_", raw_name)
        except Exception as e:
            logger.warning(f"Could not extract opportunity name from file1: {e}")
        
        return opportunity_name

    def _finalize_processing(self, output_file):
        """Finalize processing with highlighting and notifications."""
        self.highlight_unmatched_reversals(output_file)
        self.update_progress(0.85)

        messagebox.showinfo(
            "Success", f"Processing complete. File saved as {output_file}"
        )
        self.paste_into_template(output_file)
        self.update_progress(0.90)

    def highlight_unmatched_reversals(self, excel_file):
        """Highlight unmatched reversals in Excel file with reduced nesting complexity."""
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            
            # Early return if worksheet is invalid
            if ws is None:
                logger.warning("Worksheet is None, cannot highlight reversals")
                return
            
            # Early return if reversals file doesn't exist
            if not os.path.exists("unmatched_reversals.txt"):
                logger.info("No unmatched_reversals.txt file found")
                wb.save(excel_file)
                return
            
            # Apply highlighting to unmatched reversals
            self._apply_reversal_highlighting(ws, wb, excel_file)
            
        except Exception as e:
            logger.error(f"Failed to highlight unmatched reversals: {e}")

    def _apply_reversal_highlighting(self, ws, wb, excel_file):
        """Apply highlighting to rows specified in unmatched_reversals.txt."""
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        with open("unmatched_reversals.txt", "r") as f:
            rows = f.read().strip().split(",")
        
        for row_str in rows:
            if self._is_valid_row_number(row_str, ws):
                self._highlight_row(ws, int(row_str), fill)
        
        wb.save(excel_file)
        logger.info(f"Highlighted unmatched reversals in {excel_file}")

    def _is_valid_row_number(self, row_str, ws):
        """Check if row string represents a valid row number."""
        return row_str.isdigit() and 1 <= int(row_str) <= ws.max_row

    def _highlight_row(self, ws, row_num, fill):
        """Highlight all cells in a specific row."""
        row = ws[row_num]
        for cell in row:
            cell.fill = fill


warnings.filterwarnings("ignore", category=FutureWarning, message=".*swapaxes.*")
warnings.filterwarnings("ignore", category=FutureWarning, message=".*swapaxes.*")


def process_logic_block(df_block):
    """
    Vectorized numpy logic to mark 'OR' in 'Logic' for reversals with matching claims.
    Refactored to reduce nesting complexity and improve readability.
    """
    arr = df_block.to_numpy()
    col_idx = {col: i for i, col in enumerate(df_block.columns)}
    
    # Extract and prepare data
    logic_data = _extract_logic_data(arr, col_idx)
    
    # Early return if no reversals to process
    if not np.any(logic_data["is_reversal"]):
        return pd.DataFrame(arr, columns=df_block.columns)
    
    # Process reversals with reduced nesting
    _process_reversals(arr, col_idx, logic_data)
    
    return pd.DataFrame(arr, columns=df_block.columns)


def _extract_logic_data(arr, col_idx):
    """Extract and prepare data for logic processing."""
    qty = arr[:, col_idx["QUANTITY"]].astype(float)
    return {
        "qty": qty,
        "is_reversal": qty < 0,
        "is_claim": qty > 0,
        "ndc": arr[:, col_idx["NDC"]].astype(str),
        "member": arr[:, col_idx["MemberID"]].astype(str),
        "datefilled": pd.to_datetime(arr[:, col_idx["DATEFILLED"]], errors="coerce"),
        "abs_qty": np.abs(qty)
    }


def _process_reversals(arr, col_idx, logic_data):
    """Process reversals with matching logic, using guard clauses to reduce nesting."""
    rev_idx = np.where(logic_data["is_reversal"])[0]
    claim_idx = (
        np.where(logic_data["is_claim"])[0] 
        if np.any(logic_data["is_claim"]) 
        else np.array([], dtype=int)
    )
    
    # Create context object to reduce function argument count
    match_context = MatchContext(arr, col_idx, logic_data, claim_idx)
    
    for i in rev_idx:
        found_match = _try_find_match(match_context, i)
        
        # Mark unmatched reversals as 'OR'
        if not found_match:
            arr[i, col_idx["Logic"]] = "OR"


class MatchContext:
    """Context object to encapsulate matching parameters and reduce argument count."""
    
    def __init__(self, arr, col_idx, logic_data, claim_idx):
        self.arr = arr
        self.col_idx = col_idx
        self.logic_data = logic_data
        self.claim_idx = claim_idx


def _try_find_match(context, reversal_idx):
    """Attempt to find a matching claim for a reversal. Returns True if match found."""
    # Guard clause: no claims to match against
    if context.claim_idx.size == 0:
        return False
    
    # Find potential matches
    matches = _find_matching_claims(context.logic_data, context.claim_idx, reversal_idx)
    
    # Guard clause: no matches found
    if not np.any(matches):
        return False
    
    # Mark both reversal and matching claim as 'OR'
    context.arr[reversal_idx, context.col_idx["Logic"]] = "OR"
    context.arr[context.claim_idx[matches][0], context.col_idx["Logic"]] = "OR"
    return True


def _find_matching_claims(logic_data, claim_idx, reversal_idx):
    """Find claims that match the reversal based on NDC, member, quantity, and date."""
    matches = (
        (logic_data["ndc"][claim_idx] == logic_data["ndc"][reversal_idx])
        & (logic_data["member"][claim_idx] == logic_data["member"][reversal_idx])
        & (logic_data["abs_qty"][claim_idx] == logic_data["abs_qty"][reversal_idx])
    )
    
    # Add date constraint (within 30 days)
    date_diffs = np.abs(
        (logic_data["datefilled"][claim_idx] - logic_data["datefilled"][reversal_idx]).days
    )
    matches &= date_diffs <= 30
    
    return matches


if __name__ == "__main__":
    multiprocessing.freeze_support()
    ctk.set_appearance_mode("light")  # Start in light mode
    root = ctk.CTk()  # or tk.Tk() if not using customtkinter
    app = App(root)
    root.mainloop()
import os

import pytest
from openpyxl import Workbook


@pytest.fixture(scope="session", autouse=True)
def create_dummy_excel():
    filename = "./_Rx Repricing_wf.xlsx"
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Claims Table"
        ws.append(
            ["pharmacy_npi", "pharmacy_nabp", "pharmacy_id"]
        )  # add columns as needed
        ws.append([1234567890, "NABP123", "123"])  # add dummy data as needed
        wb.save(filename)
    yield
    # Optionally, remove the file after tests
    # os.remove(filename)
import json
import os
import tkinter as tk

import pandas as pd
import pytest

from app import App, ConfigManager


@pytest.fixture
def tmp_work_dir(tmp_path, monkeypatch):
    """
    Create a temporary working directory and chdir into it, so that ConfigManager
    will create its config.json there.
    """
    monkeypatch.chdir(tmp_path)
    return tmp_path


def test_save_default_creates_config(tmp_work_dir):
    # When no config.json exists, ConfigManager.save_default() should create one
    # under the current directory (tmp_work_dir).
    cm = ConfigManager()
    config_path = tmp_work_dir / "config.json"

    assert config_path.exists(), "Config file was not created."
    with open(config_path, "r") as f:
        data = json.load(f)
    # By default, ConfigManager sets 'last_folder' to the cwd
    assert "last_folder" in data
    assert data["last_folder"] == str(tmp_work_dir)


def test_load_existing_config(tmp_work_dir):
    # Write a custom config.json, then ensure ConfigManager.load() honors it.
    config_path = tmp_work_dir / "config.json"
    custom = {"last_folder": "C:/some/path/xyz"}
    with open(config_path, "w") as f:
        json.dump(custom, f)

    cm = ConfigManager()
    assert (
        cm.config == custom
    ), "ConfigManager did not load the existing config.json correctly."


def test_filter_template_columns_extracts_correct_range():
    # Build a sample DataFrame where columns go: ['A','B','Client Name','X','Y','Logic','Z','W']
    df = pd.DataFrame(
        {
            "A": [1],
            "B": [2],
            "Client Name": ["foo"],
            "X": [3],
            "Y": [4],
            "Logic": [5],
            "Z": [6],
            "W": [7],
        }
    )

    # We only expect columns from 'Client Name' up through 'Logic' (inclusive).
    root = tk.Tk()
    root.withdraw()
    app = App(root)
    filtered = app.filter_template_columns(df)
    root.destroy()

    assert list(filtered.columns) == [
        "Client Name",
        "X",
        "Y",
        "Logic",
    ], f"Expected columns from 'Client Name' to 'Logic', got {list(filtered.columns)}"


def test_filter_template_columns_fallback_to_full_df_if_missing_logic():
    # If 'Client Name' or 'Logic' aren't found, it should return the full DataFrame unmodified
    df = pd.DataFrame({"Foo": [1, 2], "Bar": [3, 4]})

    root = tk.Tk()
    root.withdraw()
    app = App(root)
    result = app.filter_template_columns(df)
    root.destroy()

    # Since 'Client Name' or 'Logic' are not present, filter_template_columns should catch ValueError
    # and return the original DataFrame
    pd.testing.assert_frame_equal(result, df)


def test_format_dataframe_converts_datetimes_and_handles_na():
    # Build a DataFrame with one datetime column and one column containing a None
    orig = pd.DataFrame(
        {
            "dt1": [
                pd.to_datetime("2020-12-31 13:45:00"),
                pd.to_datetime("2021-01-01 00:00:00"),
            ],
            "value": [10, None],
        }
    )

    root = tk.Tk()
    root.withdraw()
    app = App(root)
    formatted = app.format_dataframe(orig)
    root.destroy()

    # 'dt1' should now be strings in format '%Y-%m-%d %H:%M:%S'
    assert formatted["dt1"].dtype == object
    assert formatted["dt1"].iloc[0] == "2020-12-31 13:45:00"
    assert formatted["dt1"].iloc[1] == "2021-01-01 00:00:00"

    # The None in 'value' should become an empty string
    assert formatted["value"].iloc[1] == ""


# (Optional) If you want to guard GUI‐dependent tests to skip automatically when no display is available:
def is_display_available():
    try:
        root = tk.Tk()
        root.destroy()
        return True
    except tk.TclError:
        return False


@pytest.mark.skipif(not is_display_available(), reason="Tkinter display not available")
def test_app_instantiation_and_basic_attributes():
    # A minimal smoke–test to ensure that App(root) does not crash immediately,
    # and that certain attributes exist.
    root = tk.Tk()
    root.withdraw()
    app = App(root)

    # Basic sanity checks:
    assert hasattr(app, "file1_path")
    assert hasattr(app, "file2_path")
    assert hasattr(app, "template_file_path")
    assert isinstance(app.progress_bar, type(app.progress_bar))

    root.destroy()
from bg_disruption import process_data


def test_process_bg_runs():
    process_data()
from epls_lbl import main


def test_epls_main_runs():
    main()
from pathlib import Path

import pytest

from merge import merge_files


def test_merge_files(tmp_path):
    file1 = tmp_path / "f1.csv"
    file2 = tmp_path / "f2.csv"
    file1.write_text("DATEFILLED,SOURCERECORDID\n2020-01-01,1")
    file2.write_text("SOURCERECORDID,Total AWP (Historical)\n1,50.0")
    merge_files(str(file1), str(file2))
    assert Path("merged_file.xlsx").exists()
from openmdf_bg import process_data


def test_process_openmdf_bg_runs():
    process_data()
"""
UI Builder module for constructing the main application interface.
Extracted from app.py to reduce file size and improve organization.
"""

import customtkinter as ctk
from tkinter import messagebox
import random
import pyjokes
import emoji
import getpass
from pathlib import Path

from ui.ui_components import UIFactory, LIGHT_COLORS, DARK_COLORS
from config.app_config import DisruptionConfig, AppConstants


class UIBuilder:
    """Handles the construction of the main application UI."""
    
    def __init__(self, app_instance):
        self.app = app_instance
        
    def build_complete_ui(self):
        """Build the complete user interface."""
        self._setup_window()
        self._build_ui_components()
        self._setup_default_template()
        self._show_welcome_message()
        
    def _setup_window(self):
        """Configure the main window properties."""
        self.app.root.title("Reprice Automation")
        self.app.root.configure(fg_color=LIGHT_COLORS["dark_blue"])
        self.app.root.resizable(True, True)
        self.app.root.geometry("900x900")
        
    def _build_ui_components(self):
        """Build all UI components."""
        self._create_title()
        self._create_button_frame()
        self._create_notes_frame()
        self._create_disruption_frame()
        self._create_progress_frame()
        
    def _create_title(self):
        """Create the title label."""
        self.app.title_label = ctk.CTkLabel(
            self.app.root, text="Repricing Automation", font=("Cambria", 26, "bold")
        )
        self.app.title_label.grid(row=0, column=0, sticky="w", pady=20, padx=20)

    def _create_button_frame(self):
        """Create the main button frame with all action buttons."""
        self.app.button_frame = UIFactory.create_standard_frame(self.app.root)
        self.app.button_frame.grid(
            row=2, column=0, columnspan=3, sticky="ew", pady=10, padx=10
        )

        # Headers
        file_name_title = UIFactory.create_standard_label(
            self.app.button_frame, "File Name"
        )
        file_name_title.grid(row=0, column=2, pady=10, padx=10)

        # Create sub-components
        self._create_file_import_buttons()
        self._create_action_buttons()
        self._create_process_buttons()

    def _create_file_import_buttons(self):
        """Create file import buttons and labels."""
        # Import File 1
        self.app.file1_button = UIFactory.create_standard_button(
            self.app.button_frame, "Import File Uploaded to Tool", self.app.import_file1
        )
        self.app.file1_button.grid(row=1, column=0, pady=10, padx=10, sticky="ew")
        self.app.file1_label = UIFactory.create_standard_label(
            self.app.button_frame, "", width=350
        )
        self.app.file1_label.grid(row=1, column=2, pady=20, padx=10)

        # Import File 2
        self.app.file2_button = UIFactory.create_standard_button(
            self.app.button_frame, "Import File From Tool", self.app.import_file2
        )
        self.app.file2_button.grid(row=2, column=0, pady=10, padx=10, sticky="ew")
        self.app.file2_label = UIFactory.create_standard_label(self.app.button_frame, "")
        self.app.file2_label.grid(row=2, column=2, pady=20, padx=10)

        # Select Template
        self.app.template_button = UIFactory.create_standard_button(
            self.app.button_frame, "Select Template File", self.app.import_template_file
        )
        self.app.template_button.grid(row=3, column=0, pady=10, padx=10, sticky="ew")
        self.app.template_label = UIFactory.create_standard_label(self.app.button_frame, "")
        self.app.template_label.grid(row=3, column=2, pady=20, padx=10)

    def _create_action_buttons(self):
        """Create action buttons (cancel, logs, theme)."""
        # Cancel button
        self.app.cancel_button = UIFactory.create_red_button(
            self.app.button_frame, "Cancel", self.app.cancel_process
        )
        self.app.cancel_button.grid(row=4, column=0, pady=10, padx=10, sticky="ew")

        # View Logs button
        self.app.logs_button = UIFactory.create_standard_button(
            self.app.button_frame, "View Logs", self.app.show_log_viewer
        )
        self.app.logs_button.grid(row=4, column=1, pady=10, padx=10, sticky="ew")

        # Toggle Dark Mode button
        self.app.toggle_theme_button = UIFactory.create_standard_button(
            self.app.button_frame, "Switch to Dark Mode", self.toggle_dark_mode
        )
        self.app.toggle_theme_button.grid(row=4, column=2, pady=10, padx=10, sticky="ew")

        # Shared Log Button
        self.app.shared_log_button = UIFactory.create_standard_button(
            self.app.button_frame, "Shared Audit Log", self.app.show_shared_log_viewer
        )
        self.app.shared_log_button.grid(row=6, column=1, pady=10, padx=10, sticky="ew")

        # Exit button
        self.app.exit_button = UIFactory.create_red_button(
            self.app.button_frame, "Exit", self.app.root.quit
        )
        self.app.exit_button.grid(row=6, column=2, pady=10, padx=10, sticky="ew")

    def _create_process_buttons(self):
        """Create processing and LBL generation buttons."""
        # SHARx LBL button
        self.app.sharx_lbl_button = UIFactory.create_standard_button(
            self.app.button_frame, "Generate SHARx LBL", self.app.sharx_lbl
        )
        self.app.sharx_lbl_button.grid(row=5, column=0, pady=10, padx=10, sticky="ew")
        
        # EPLS LBL button
        self.app.epls_lbl_button = UIFactory.create_standard_button(
            self.app.button_frame, "Generate EPLS LBL", self.app.epls_lbl
        )
        self.app.epls_lbl_button.grid(row=5, column=1, pady=10, padx=10, sticky="ew")
        
        # Start Process button
        self.app.start_process_button = ctk.CTkButton(
            self.app.button_frame,
            text="Start Repricing",
            command=self.app.start_process_threaded,
            font=("Cambria", 20, "bold"),
            height=40,
            width=200,
            fg_color=LIGHT_COLORS["mint"],
            text_color="#000000",
        )
        self.app.start_process_button.grid(row=5, column=2, pady=10, padx=10, sticky="ew")

    def _create_notes_frame(self):
        """Create the notes frame with important information."""
        self.app.notes_frame = UIFactory.create_standard_frame(self.app.root)
        self.app.notes_frame.grid(
            row=3, column=0, columnspan=4, sticky="ew", pady=10, padx=10
        )
        notes = UIFactory.create_standard_label(
            self.app.notes_frame, AppConstants.NOTES_TEXT
        )
        notes.configure(justify="left")
        notes.pack(padx=20, pady=10)

    def _create_disruption_frame(self):
        """Create the disruption type selector frame."""
        self.app.dis_frame = UIFactory.create_standard_frame(self.app.root)
        self.app.dis_frame.grid(
            row=4, column=0, columnspan=4, sticky="ew", pady=10, padx=10
        )

        # Create disruption buttons using configuration
        disruption_labels = DisruptionConfig.get_disruption_labels()
        for idx, label in enumerate(disruption_labels):
            btn = UIFactory.create_standard_button(
                self.app.dis_frame, label, 
                lambda label_text=label: self.app.start_disruption(label_text)
            )
            btn.grid(row=0, column=idx, padx=10, pady=10, sticky="ew")

    def _create_progress_frame(self):
        """Create the progress bar frame."""
        self.app.prog_frame = UIFactory.create_standard_frame(self.app.root)
        self.app.prog_frame.grid(
            row=5, column=0, columnspan=4, sticky="ew", pady=10, padx=10
        )
        self.app.progress_bar = ctk.CTkProgressBar(
            self.app.prog_frame, orientation="horizontal", mode="determinate"
        )
        self.app.progress_bar.set(self.app.progress_var.get())
        self.app.progress_bar.pack(padx=10, pady=(10, 5), fill="x")
        self.app.progress_label = ctk.CTkLabel(
            self.app.prog_frame, textvariable=self.app.progress_label_var
        )
        self.app.progress_label.pack(padx=10, pady=(0, 10), anchor="w")
        
    def _setup_default_template(self):
        """Set up default template if it exists."""
        default_template = Path("_Rx Repricing_wf.xlsx")
        if default_template.exists():
            self.app.template_file_path = str(default_template)
            if hasattr(self.app, 'template_label'):
                self.app.template_label.configure(
                    text=default_template.name
                )
        else:
            self.app.template_file_path = None
            
    def _show_welcome_message(self):
        """Show a personalized welcome message with a random joke and emoji."""
        user = getpass.getuser()
        welcome_messages = AppConstants.WELCOME_MESSAGES
        
        msg = welcome_messages.get(
            user, f"Welcome, {user}! Ready to use the Repricing Automation Toolkit?"
        )

        # Add a random joke and emoji
        try:
            joke = pyjokes.get_joke()
        except Exception:
            joke = "Have a great day!"
            
        chosen_emoji = emoji.emojize(random.choice(AppConstants.EMOJIS), language="alias")
        full_msg = f"{msg}\n\n{joke} {chosen_emoji}"
        
        # Show after UI is built
        self.app.root.after(500, lambda: messagebox.showinfo("Welcome", full_msg))
        
    def toggle_dark_mode(self):
        """Toggle between light and dark modes."""
        current = ctk.get_appearance_mode().lower()

        if current == "light":
            self._switch_to_dark_mode()
        else:
            self._switch_to_light_mode()
            
    def _switch_to_dark_mode(self):
        """Switch to dark mode."""
        ctk.set_appearance_mode("dark")
        self.app.apply_theme_colors(DARK_COLORS)
        if self.app.toggle_theme_button:
            self.app.toggle_theme_button.configure(text="Switch to Light Mode")
            
    def _switch_to_light_mode(self):
        """Switch to light mode."""
        ctk.set_appearance_mode("light")
        self.app.apply_theme_colors(LIGHT_COLORS)
        if self.app.toggle_theme_button:
            self.app.toggle_theme_button.configure(text="Switch to Dark Mode")
from openmdf_tier import process_data


def test_process_openmdf_tier_runs():
    process_data()
from tier_disruption import process_data


def test_process_tier_runs():
    process_data()
import json
import os

import pandas as pd

# Standardize IDs


def standardize_pharmacy_ids(df):
    if "PHARMACYNPI" in df.columns:
        df["PHARMACYNPI"] = df["PHARMACYNPI"].astype(str).str.zfill(10)
    if "NABP" in df.columns:
        df["NABP"] = df["NABP"].astype(str).str.zfill(7)
    return df


def standardize_network_ids(network):
    if "pharmacy_npi" in network.columns:
        network["pharmacy_npi"] = network["pharmacy_npi"].astype(str).str.zfill(10)
    if "pharmacy_nabp" in network.columns:
        network["pharmacy_nabp"] = network["pharmacy_nabp"].astype(str).str.zfill(7)
    return network


def merge_with_network(df, network):
    return df.merge(
        network,
        left_on=["PHARMACYNPI", "NABP"],
        right_on=["pharmacy_npi", "pharmacy_nabp"],
        how="left",
    )


def load_file_paths(json_file):
    with open(json_file, "r") as f:
        file_paths = json.load(f)
    return file_paths


def summarize_by_tier(df, col, from_val, to_val):
    filtered = df[(df[col] == from_val) & (df["FormularyTier"] == to_val)]
    pt = pd.pivot_table(
        filtered,
        values=["Rxs", "MemberID"],
        index=["Product Name"],
        aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
    )
    rxs = filtered["Rxs"].sum()
    members = filtered["MemberID"].nunique()
    return pt, rxs, members


def process_data():
    file_paths = load_file_paths("file_paths.json")

    # Only load claims if the path is present and not empty
    claims = None
    if "reprice" in file_paths and file_paths["reprice"]:
        claims = pd.read_excel(file_paths["reprice"], sheet_name="Claims Table")
    else:
        print("No reprice/template file provided. Skipping claims loading.")
        # You can decide to return, raise, or continue with alternate logic here
        return

    medi = pd.read_excel(file_paths["medi_span"])[
        ["NDC", "Maint Drug?", "Product Name"]
    ]
    u = pd.read_excel(file_paths["u_disrupt"], sheet_name="Universal NDC")[
        ["NDC", "Tier"]
    ]
    e = pd.read_excel(file_paths["e_disrupt"], sheet_name="Alternatives NDC")[
        ["NDC", "Tier", "Alternative"]
    ]
    network = pd.read_excel(file_paths["n_disrupt"])[
        ["pharmacy_npi", "pharmacy_nabp", "pharmacy_is_excluded"]
    ]

    df = claims.merge(medi, on="NDC", how="left")
    df = df.merge(u.rename(columns={"Tier": "Universal Tier"}), on="NDC", how="left")
    df = df.merge(e.rename(columns={"Tier": "Exclusive Tier"}), on="NDC", how="left")

    df = standardize_pharmacy_ids(df)
    network = standardize_network_ids(network)
    df = merge_with_network(df, network)

    df["DATEFILLED"] = pd.to_datetime(df["DATEFILLED"], errors="coerce")
    df = df.drop_duplicates()
    df["Logic"] = pd.to_numeric(df["Logic"], errors="coerce")
    df["FormularyTier"] = pd.to_numeric(df["FormularyTier"], errors="coerce")

    latest_date = df["DATEFILLED"].max()
    starting_point = latest_date - pd.DateOffset(months=6) + pd.DateOffset(days=1)
    df = df[(df["DATEFILLED"] >= starting_point) & (df["DATEFILLED"] <= latest_date)]
    df = df[(df["Logic"] >= 5) & (df["Logic"] <= 10) & (df["Maint Drug?"] == "Y")]
    df = df[~df["Product Name"].str.contains(r"\balbuterol\b", case=False)]
    df = df[~df["Product Name"].str.contains(r"\bventolin\b", case=False)]
    df = df[~df["Product Name"].str.contains(r"\bepinephrine\b", case=False)]
    df = df[
        ~df["Alternative"]
        .astype(str)
        .str.contains("Covered|Use different NDC", case=False, regex=True)
    ]

    total_claims = df["Rxs"].sum()
    total_members = df["MemberID"].nunique()

    writer = pd.ExcelWriter("LBL for Disruption.xlsx", engine="xlsxwriter")
    df.to_excel(writer, sheet_name="Data", index=False)

    tiers = [
        ("Universal_Positive 2-1", "Universal Tier", 2, 1),
        ("Universal_Positive 3-1", "Universal Tier", 3, 1),
        ("Universal_Positive 3-2", "Universal Tier", 3, 2),
        ("Universal_Negative 1-2", "Universal Tier", 1, 2),
        ("Universal_Negative 1-3", "Universal Tier", 1, 3),
        ("Universal_Negative 2-3", "Universal Tier", 2, 3),
        ("Exclusive_Positive 2-1", "Exclusive Tier", 2, 1),
        ("Exclusive_Positive 3-1", "Exclusive Tier", 3, 1),
        ("Exclusive_Positive 3-2", "Exclusive Tier", 3, 2),
        ("Exclusive_Negative 1-2", "Exclusive Tier", 1, 2),
        ("Exclusive_Negative 1-3", "Exclusive Tier", 1, 3),
        ("Exclusive_Negative 2-3", "Exclusive Tier", 2, 3),
    ]

    summary_rows = []
    tab_members = {}

    for name, col, from_val, to_val in tiers:
        pt, rxs, members = summarize_by_tier(df, col, from_val, to_val)
        pt.to_excel(writer, sheet_name=name)
        summary_rows.append((name.replace("_", " "), members, rxs))
        tab_members[name] = members

    exclusions = df[df["pharmacy_is_excluded"] == True]
    ex_pt = exclusions.pivot_table(
        values=["Rxs", "MemberID"],
        index=["Product Name"],
        aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
    )
    exc_rxs = exclusions["Rxs"].sum()
    exc_members = exclusions["MemberID"].nunique()
    ex_pt.to_excel(writer, sheet_name="Exclusions")

    summary_rows.append(("Exclusions", exc_members, exc_rxs))
    tab_members["Exclusions"] = exc_members

    # Aggregate utilizer and claims counts for summary
    uni_pos_utilizers = (
        tab_members.get("Universal_Positive 2-1", 0)
        + tab_members.get("Universal_Positive 3-1", 0)
        + tab_members.get("Universal_Positive 3-2", 0)
    )
    uni_neg_utilizers = (
        tab_members.get("Universal_Negative 1-2", 0)
        + tab_members.get("Universal_Negative 1-3", 0)
        + tab_members.get("Universal_Negative 2-3", 0)
    )
    ex_pos_utilizers = (
        tab_members.get("Exclusive_Positive 2-1", 0)
        + tab_members.get("Exclusive_Positive 3-1", 0)
        + tab_members.get("Exclusive_Positive 3-2", 0)
    )
    ex_neg_utilizers = (
        tab_members.get("Exclusive_Negative 1-2", 0)
        + tab_members.get("Exclusive_Negative 1-3", 0)
        + tab_members.get("Exclusive_Negative 2-3", 0)
    )
    exc_utilizers = tab_members.get("Exclusions", 0)

    # Similarly, aggregate claims for each group
    uni_pos_claims = 0
    uni_neg_claims = 0
    ex_pos_claims = 0
    ex_neg_claims = 0
    exc_claims = exc_rxs

    for name, _, _, _ in tiers:
        if name.startswith("Universal_Positive"):
            uni_pos_claims += df[
                (df["Universal Tier"] == int(name.split()[1][0]))
                & (df["FormularyTier"] == int(name.split()[1][2]))
            ]["Rxs"].sum()
        elif name.startswith("Universal_Negative"):
            uni_neg_claims += df[
                (df["Universal Tier"] == int(name.split()[1][0]))
                & (df["FormularyTier"] == int(name.split()[1][2]))
            ]["Rxs"].sum()
        elif name.startswith("Exclusive_Positive"):
            ex_pos_claims += df[
                (df["Exclusive Tier"] == int(name.split()[1][0]))
                & (df["FormularyTier"] == int(name.split()[1][2]))
            ]["Rxs"].sum()
        elif name.startswith("Exclusive_Negative"):
            ex_neg_claims += df[
                (df["Exclusive Tier"] == int(name.split()[1][0]))
                & (df["FormularyTier"] == int(name.split()[1][2]))
            ]["Rxs"].sum()

    # Calculate percentages
    uni_pos_pct = uni_pos_claims / total_claims if total_claims else 0
    uni_neg_pct = uni_neg_claims / total_claims if total_claims else 0
    ex_pos_pct = ex_pos_claims / total_claims if total_claims else 0
    ex_neg_pct = ex_neg_claims / total_claims if total_claims else 0
    exc_pct = exc_claims / total_claims if total_claims else 0

    summary_df = pd.DataFrame(
        {
            "Formulary": [
                "Universal Positive",
                "Universal Negative",
                "Exclusive Positive",
                "Exclusive Negative",
                "Exclusions",
            ],
            "Utilizers": [
                uni_pos_utilizers,
                uni_neg_utilizers,
                ex_pos_utilizers,
                ex_neg_utilizers,
                exc_utilizers,
            ],
            "Rxs": [
                uni_pos_claims,
                uni_neg_claims,
                ex_pos_claims,
                ex_neg_claims,
                exc_claims,
            ],
            "% of claims": [uni_pos_pct, uni_neg_pct, ex_pos_pct, ex_neg_pct, exc_pct],
            "": ["", "", "", "", ""],
            "Totals": [
                f"Members: {total_members}",
                f"Claims: {total_claims}",
                "",
                "",
                "",
            ],
        }
    )

    summary_df.to_excel(writer, sheet_name="Summary", index=False)

    for sheet_name, value in tab_members.items():
        worksheet = writer.sheets[sheet_name]
        worksheet.write("F1", f"Total Members: {value}")

    # Network summary for non-excluded pharmacies
    network_df = df[df["pharmacy_is_excluded"].isna()]
    filter_phrases = [
        "CVS",
        "Walgreens",
        "Kroger",
        "Walmart",
        "Rite Aid",
        "Optum",
        "Express Scripts",
        "DMR",
        "Williams Bro",
        "Publix",
    ]
    regex_pattern = "|".join([f"\b{phrase}\b" for phrase in filter_phrases])
    network_df = network_df[
        ~network_df["Pharmacy Name"].str.contains(regex_pattern, case=False, regex=True)
    ]

    if (
        "PHARMACYNPI" in network_df.columns
        and "NABP" in network_df.columns
        and "Pharmacy Name" in network_df.columns
    ):
        network_pivot = pd.pivot_table(
            network_df,
            values=["Rxs", "MemberID"],
            index=["PHARMACYNPI", "NABP", "Pharmacy Name"],
            aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
        )
        network_pivot.to_excel(writer, sheet_name="Network")
    else:
        print(
            "PHARMACYNPI, NABP, or Pharmacy Name column missing in the data dataframe."
        )

    # Reorder sheets so 'Summary' is right after 'Data'
    workbook = writer.book
    sheets = workbook.worksheets()
    sheet_names = [ws.get_name() for ws in sheets]

    # Move 'Summary' after 'Data'
    if "Data" in sheet_names and "Summary" in sheet_names:
        data_idx = sheet_names.index("Data")
        summary_idx = sheet_names.index("Summary")
        if summary_idx != data_idx + 1:
            summary_ws = sheets[summary_idx]
            sheets.pop(summary_idx)
            sheets.insert(data_idx + 1, summary_ws)

    # Save once at the end
    writer._save()


if __name__ == "__main__":
    process_data()
    print("Data processing complete")
"""
UI Components module for the Repricing Automation application.
This module contains UI-related classes and utilities to improve code organization.
"""

import customtkinter as ctk


# UI styling variables
FONT_SELECT = ("Cambria", 20, "bold")

# Color palettes
LIGHT_COLORS = {
    "dark_blue": "#D9EAF7",
    "grey_blue": "#A3B9CC",
    "mint": "#8FD9A8",
    "button_red": "#D52B2B",
}

DARK_COLORS = {
    "dark_blue": "#223354",
    "grey_blue": "#31476A",
    "mint": "#26A69A",
    "button_red": "#931D1D",
}


class UIFactory:
    """Factory class to create UI components and reduce code duplication."""
    
    @staticmethod
    def _create_button_base(parent, text, command, fg_color):
        """Base method for creating buttons with common styling."""
        return ctk.CTkButton(
            parent,
            text=text,
            command=command,
            font=FONT_SELECT,
            height=40,
            fg_color=fg_color,
            text_color="#000000"
        )
    
    @staticmethod
    def create_standard_button(parent, text, command):
        """Create a standardized button with common styling."""
        return UIFactory._create_button_base(parent, text, command, LIGHT_COLORS["mint"])
    
    @staticmethod
    def create_red_button(parent, text, command):
        """Create a red button (for cancel/exit actions)."""
        return UIFactory._create_button_base(parent, text, command, LIGHT_COLORS["button_red"])
    
    @staticmethod
    def create_standard_frame(parent):
        """Create a standardized frame with common styling."""
        return ctk.CTkFrame(parent, fg_color=LIGHT_COLORS["grey_blue"])
    
    @staticmethod
    def create_standard_label(parent, text, width=None):
        """Create a standardized label."""
        if width:
            return ctk.CTkLabel(parent, text=text, font=FONT_SELECT, width=width)
        return ctk.CTkLabel(parent, text=text, font=FONT_SELECT)


class ThemeManager:
    """Manages theme colors and application of themes to UI components."""
    
    @staticmethod
    def apply_theme_colors(app_instance, colors):
        """Apply theme colors to all UI components."""
        ThemeManager._apply_root_colors(app_instance, colors)
        ThemeManager._apply_frame_colors(app_instance, colors)
        ThemeManager._apply_button_colors(app_instance, colors)
        ThemeManager._apply_special_component_colors(app_instance, colors)
    
    @staticmethod
    def _apply_root_colors(app_instance, colors):
        """Apply colors to the root window."""
        app_instance.root.configure(fg_color=colors["dark_blue"])
    
    @staticmethod
    def _apply_frame_colors(app_instance, colors):
        """Apply colors to frames."""
        frames = ["button_frame", "notes_frame", "dis_frame", "prog_frame"]
        for frame_name in frames:
            frame = getattr(app_instance, frame_name, None)
            if frame:
                frame.configure(fg_color=colors["grey_blue"])
    
    @staticmethod
    def _apply_button_colors(app_instance, colors):
        """Apply colors to standard buttons."""
        button_widgets = [
            "file1_button", "file2_button", "template_button", "cancel_button",
            "logs_button", "toggle_theme_button", "sharx_lbl_button", 
            "epls_lbl_button", "start_process_button"
        ]
        
        for btn_name in button_widgets:
            btn = getattr(app_instance, btn_name, None)
            if btn:
                btn.configure(fg_color=colors["mint"], text_color="#000000")
    
    @staticmethod
    def _apply_special_component_colors(app_instance, colors):
        """Apply colors to special components."""
        # Apply colors to special buttons
        if hasattr(app_instance, "exit_button"):
            app_instance.exit_button.configure(
                fg_color=colors["button_red"], text_color="#000000"
            )
        
        # Apply colors to progress components
        if hasattr(app_instance, "progress_label"):
            app_instance.progress_label.configure(
                bg_color=colors["grey_blue"], text_color="#000000"
            )


class ProgressManager:
    """Manages progress bar updates and calculations."""
    
    @staticmethod
    def calculate_time_estimates(value, start_time):
        """Calculate progress percentage and time estimates."""
        import time
        
        percent = int(value * 100)
        elapsed = time.time() - start_time if start_time else 0
        est = int((elapsed / value) * (1 - value)) if value > 0 else 0
        return percent, est
    
    @staticmethod
    def format_progress_message(percent, estimated_seconds):
        """Format progress message with percentage and time estimate."""
        return f"Progress: {percent}% | Est. {estimated_seconds}s left"
import os
import shutil
import logging
import xlwings as xw
import importlib.util
from typing import Any, Tuple
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed

logger = logging.getLogger(__name__)

# COM fallback via pywin32
EXCEL_COM_AVAILABLE = importlib.util.find_spec("win32com.client") is not None


def open_workbook(path: str, visible: bool = False) -> Tuple[Any, Any, bool]:
    """
    Open workbook via xlwings or COM fallback.
    Returns (wb, app_obj, use_com).
    """
    import time

    max_retries = 3
    delay = 2
    last_exc = None
    for attempt in range(max_retries):
        try:
            app = xw.App(visible=visible, add_book=False)  # Ensure no new book is added
            try:
                wb = app.books.open(path)
            except TypeError:
                # Try with password if provided in path (e.g., path='file.xlsx::password')
                if "::" in path:
                    file_path, password = path.split("::", 1)
                    wb = app.books.open(file_path, password=password)
                else:
                    raise
            return wb, app, False
        except Exception as e:
            last_exc = e
            logger.warning(
                f"Failed to open workbook (attempt {attempt + 1}/{max_retries}): {e}"
            )
            time.sleep(delay)
    if EXCEL_COM_AVAILABLE:
        import win32com.client as win32

        excel: Any = win32.Dispatch("Excel.Application")
        excel.Visible = visible  # Ensure Excel remains hidden
        excel.DisplayAlerts = False  # Suppress alerts
        try:
            if "::" in path:
                file_path, password = path.split("::", 1)
                wb: Any = excel.Workbooks.Open(
                    os.path.abspath(file_path), False, False, None, password
                )
            else:
                wb: Any = excel.Workbooks.Open(os.path.abspath(path))
        except Exception as e:
            logger.error(f"COM fallback failed to open workbook: {e}")
            raise
        return wb, excel, True
    logger.error(f"Failed to open workbook after {max_retries} attempts: {last_exc}")
    if last_exc is not None:
        raise last_exc
    # Should never reach here, but raise as a safeguard
    raise RuntimeError("Failed to open workbook and no exception was captured.")


def write_df_to_sheet_async(
    path: str,
    sheet_name: str,
    df: pd.DataFrame,
    start_cell: str = "A2",
    header: bool = False,
    index: bool = False,
    clear: bool = True,
    visible: bool = False,
    clear_by_label: bool = False,
    max_workers: int = 4,
) -> None:
    """
    Async version of write_df_to_sheet for large DataFrames (xlwings only).
    Splits DataFrame into row blocks and writes in parallel threads.
    """
    logger.info(
        f"[ASYNC] Writing to {path} in sheet '{sheet_name}' from cell {start_cell} with {max_workers} workers"
    )
    wb, app, use_com = open_workbook(path, visible)
    if use_com:
        # COM automation is not thread-safe; fallback to sync
        logger.warning("COM fallback does not support async writes. Using sync write.")
        return write_df_to_sheet(
            path,
            sheet_name,
            df,
            start_cell,
            header,
            index,
            clear,
            visible,
            clear_by_label,
        )
    try:
        ws = wb.sheets[sheet_name]
        cell = ws.range(start_cell)
    except Exception as e:
        close_workbook(wb, app, save=False, use_com=use_com)
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook.") from e

    start_row = cell.row
    start_col = cell.column
    n_rows, n_cols = df.shape
    total_rows = n_rows + (1 if header else 0)
    end_row = start_row + total_rows - 1
    end_col = start_col + n_cols - 1

    # Optionally clear before writing
    target = ws.range((start_row, start_col), (end_row, end_col))
    if clear:
        if clear_by_label:
            for idx, col in enumerate(df.columns, start_col):
                col_range = ws.range((start_row, idx), (end_row, idx))
                col_range.clear_contents()
        else:
            target.clear_contents()

    # Write header if needed
    if header:
        for j, h in enumerate(df.columns, start_col):
            ws.Cells(start_row, j).Value = h
        data_start = start_row + 1
    else:
        data_start = start_row

    # Split DataFrame into blocks for parallel writing
    block_size = max(100, n_rows // max_workers)
    blocks = [(i, min(i + block_size, n_rows)) for i in range(0, n_rows, block_size)]

    def write_block(start, stop):
        for i, row in enumerate(df.values[start:stop], data_start + start):
            for j, val in enumerate(row, start_col):
                ws.Cells(i, j).Value = val

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(write_block, start, stop) for start, stop in blocks]
        for f in as_completed(futures):
            f.result()

    close_workbook(wb, app, save=True, use_com=use_com)


def close_workbook(
    wb: Any, app_obj: Any, save: bool = True, use_com: bool = False
) -> None:
    """
    Close the workbook and quit the application.
    """
    if not use_com:
        if save:
            wb.save()
        wb.close()
        app_obj.quit()
    else:
        if save:
            wb.Save()
        wb.Close(SaveChanges=save)
        app_obj.Quit()


def write_df_to_sheet(
    path: str,
    sheet_name: str,
    df: pd.DataFrame,
    start_cell: str = "A2",
    header: bool = False,
    index: bool = False,
    clear: bool = True,
    visible: bool = False,
    clear_by_label: bool = False,
) -> None:
    """
    Write DataFrame to an Excel sheet without removing any formatting.
    Only clears the cells where values will be written.
    """
    logger.info(f"Writing to {path} in sheet '{sheet_name}' from cell {start_cell}")

    wb, app, use_com = open_workbook(path, visible)

    try:
        if not use_com:
            ws = wb.sheets[sheet_name]
            cell = ws.range(start_cell)

            def clear_func(rng):
                rng.clear_contents()
        else:
            ws: Any = wb.Worksheets(sheet_name)
            cell: Any = ws.Range(start_cell)

            def clear_func(rng):
                rng.ClearContents()
    except Exception as e:
        close_workbook(wb, app, save=False, use_com=use_com)
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook.") from e

    # Determine start row/col and target range
    start_row = cell.row if not use_com else cell.Row
    start_col = cell.column if not use_com else cell.Column
    n_rows, n_cols = df.shape
    total_rows = n_rows + (1 if header else 0)
    end_row = start_row + total_rows - 1
    end_col = start_col + n_cols - 1

    if not use_com:
        target = ws.range((start_row, start_col), (end_row, end_col))
        if clear:
            if clear_by_label:
                # Clear by column label (header row)
                for idx, col in enumerate(df.columns, start_col):
                    col_range = ws.range((start_row, idx), (end_row, idx))
                    col_range.clear_contents()
            else:
                clear_func(target)
        target.options(index=index, header=header).value = df
    else:
        target = ws.Range(ws.Cells(start_row, start_col), ws.Cells(end_row, end_col))
        if clear:
            if clear_by_label:
                # Clear by column label (header row)
                for idx, col in enumerate(df.columns, start_col):
                    col_rng = ws.Range(ws.Cells(start_row, idx), ws.Cells(end_row, idx))
                    col_rng.ClearContents()
            else:
                clear_func(target)
        data_start = start_row
        if header:
            for j, h in enumerate(df.columns, start_col):
                ws.Cells(start_row, j).Value = h
            data_start += 1
        for i, row in enumerate(df.values.tolist(), data_start):
            for j, val in enumerate(row, start_col):
                ws.Cells(i, j).Value = val

    close_workbook(wb, app, save=True, use_com=use_com)


def write_df_to_template(
    template_path: str,
    output_path: str,
    sheet_name: str,
    df: pd.DataFrame,
    start_cell: str = "A2",
    header: bool = False,
    index: bool = False,
    visible: bool = False,
    open_file: bool = False,
) -> None:
    """
    Copy an Excel template and write a DataFrame into it without altering
    any existing formatting, charts, tables, or objects.

    If open_file is True, launch the filled workbook in Excel after writing.
    """
    shutil.copy(template_path, output_path)
    write_df_to_sheet(
        path=output_path,
        sheet_name=sheet_name,
        df=df,
        start_cell=start_cell,
        header=header,
        index=index,
        clear=True,
        visible=visible,
    )
    if open_file:
        os.startfile(output_path)
"""
Logic processing utilities extracted from app.py
Following CodeScene ACE principles for better code organization
"""

import logging
import warnings
from dataclasses import dataclass
from typing import Dict

import numpy as np
import pandas as pd

logger = logging.getLogger(__name__)

# Filter out specific warnings
warnings.filterwarnings("ignore", category=FutureWarning, message=".*swapaxes.*")


@dataclass
class LogicData:
    """Data class to encapsulate logic processing data."""
    qty: np.ndarray
    is_reversal: np.ndarray
    is_claim: np.ndarray
    ndc: np.ndarray
    member: np.ndarray
    datefilled: pd.DatetimeIndex
    abs_qty: np.ndarray


@dataclass
class MatchContext:
    """Context object to encapsulate matching parameters."""
    arr: np.ndarray
    col_idx: Dict[str, int]
    logic_data: LogicData
    claim_idx: np.ndarray


class LogicProcessor:
    """Handles logic processing for reversal matching."""
    
    @staticmethod
    def process_logic_block(df_block: pd.DataFrame) -> pd.DataFrame:
        """
        Vectorized numpy logic to mark 'OR' in 'Logic' for reversals with matching claims.
        Refactored to reduce nesting complexity and improve readability.
        """
        arr = df_block.to_numpy()
        col_idx = {col: i for i, col in enumerate(df_block.columns)}
        
        # Extract and prepare data
        logic_data = LogicProcessor._extract_logic_data(arr, col_idx)
        
        # Early return if no reversals to process
        if not np.any(logic_data.is_reversal):
            return pd.DataFrame(arr, columns=df_block.columns)
        
        # Process reversals with reduced nesting
        LogicProcessor._process_reversals(arr, col_idx, logic_data)
        
        return pd.DataFrame(arr, columns=df_block.columns)

    @staticmethod
    def _extract_logic_data(arr: np.ndarray, col_idx: Dict[str, int]) -> LogicData:
        """Extract and prepare data for logic processing."""
        qty = arr[:, col_idx["QUANTITY"]].astype(float)
        
        return LogicData(
            qty=qty,
            is_reversal=qty < 0,
            is_claim=qty > 0,
            ndc=arr[:, col_idx["NDC"]].astype(str),
            member=arr[:, col_idx["MemberID"]].astype(str),
            datefilled=pd.to_datetime(arr[:, col_idx["DATEFILLED"]], errors="coerce"),
            abs_qty=np.abs(qty)
        )

    @staticmethod
    def _process_reversals(arr: np.ndarray, col_idx: Dict[str, int], logic_data: LogicData):
        """Process reversals with matching logic, using guard clauses to reduce nesting."""
        rev_idx = np.where(logic_data.is_reversal)[0]
        claim_idx = (
            np.where(logic_data.is_claim)[0] 
            if np.any(logic_data.is_claim) 
            else np.array([], dtype=int)
        )
        
        # Create context object to reduce function argument count
        match_context = MatchContext(arr, col_idx, logic_data, claim_idx)
        
        for i in rev_idx:
            found_match = LogicProcessor._try_find_match(match_context, i)
            
            # Mark unmatched reversals as 'OR'
            if not found_match:
                arr[i, col_idx["Logic"]] = "OR"

    @staticmethod
    def _try_find_match(context: MatchContext, reversal_idx: int) -> bool:
        """Attempt to find a matching claim for a reversal. Returns True if match found."""
        # Guard clause: no claims to match against
        if context.claim_idx.size == 0:
            return False
        
        # Find potential matches
        matches = LogicProcessor._find_matching_claims(
            context.logic_data, context.claim_idx, reversal_idx
        )
        
        # Guard clause: no matches found
        if not np.any(matches):
            return False
        
        # Mark both reversal and matching claim as 'OR'
        context.arr[reversal_idx, context.col_idx["Logic"]] = "OR"
        context.arr[context.claim_idx[matches][0], context.col_idx["Logic"]] = "OR"
        return True

    @staticmethod
    def _find_matching_claims(logic_data: LogicData, claim_idx: np.ndarray, 
                            reversal_idx: int) -> np.ndarray:
        """Find claims that match the reversal based on NDC, member, quantity, and date."""
        matches = (
            (logic_data.ndc[claim_idx] == logic_data.ndc[reversal_idx])
            & (logic_data.member[claim_idx] == logic_data.member[reversal_idx])
            & (logic_data.abs_qty[claim_idx] == logic_data.abs_qty[reversal_idx])
        )
        
        # Add date constraint (within 30 days)
        try:
            date_diffs = np.abs(
                (logic_data.datefilled[claim_idx] - logic_data.datefilled[reversal_idx]).days
            )
            matches &= date_diffs <= 30
        except Exception as e:
            logger.warning(f"Date filtering failed: {e}")
            # Continue without date constraint
        
        return matches


# Backwards compatibility functions
def process_logic_block(df_block: pd.DataFrame) -> pd.DataFrame:
    """Backwards compatibility wrapper."""
    return LogicProcessor.process_logic_block(df_block)
from sharx_lbl import main


def test_sharx_main_runs():
    main()
import pandas as pd
import json
import logging
import os
import csv
from pathlib import Path
import getpass
from datetime import datetime
from dataclasses import dataclass

shared_log_path = os.path.expandvars(
    r"%OneDrive%/True Community - Data Analyst/Python Repricing Automation Program/Logs/audit_log.csv"
)


@dataclass
class LogicMaintenanceConfig:
    """Configuration for logic and maintenance filtering."""
    logic_col: str = "Logic"
    min_logic: int = 5
    max_logic: int = 10
    maint_col: str = "Maint Drug?"


def ensure_directory_exists(path):
    """
    Ensures the directory for the given path exists.
    """
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
    except Exception as e:
        print(f"[ensure_directory_exists] Error: {e}")


def write_shared_log(script_name, message, status="INFO"):
    """
    Appends a log entry to the shared audit log in OneDrive. Rotates log if too large.
    """
    try:
        username = getpass.getuser()
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry = [timestamp, username, script_name, message, status]

        write_header = not os.path.exists(shared_log_path)
        ensure_directory_exists(shared_log_path)

        # Log rotation: if file > 5MB, rotate (keep 3 backups)
        max_size = 5 * 1024 * 1024
        if (
            os.path.exists(shared_log_path)
            and os.path.getsize(shared_log_path) > max_size
        ):
            for i in range(2, 0, -1):
                prev = f"{shared_log_path}.{i}"
                prev2 = f"{shared_log_path}.{i + 1}"
                if os.path.exists(prev):
                    os.replace(prev, prev2)
            os.replace(shared_log_path, f"{shared_log_path}.1")

        with open(shared_log_path, mode="a", newline="", encoding="utf-8") as file:
            writer = csv.writer(file)
            if write_header:
                writer.writerow(["Timestamp", "User", "Script", "Message", "Status"])
            writer.writerow(log_entry)
    except Exception as e:
        print(f"[Shared Log] Error: {e}")


def log_exception(script_name, exc, status="ERROR"):
    """
    Standardized exception logging to shared log and console.
    """
    import traceback

    tb = traceback.format_exc()
    msg = f"{exc}: {tb}"
    print(f"[Exception] {msg}")
    write_shared_log(script_name, msg, status)


def load_file_paths(json_file="file_paths.json"):
    """
    Loads a JSON config file, replacing %OneDrive% with the user's OneDrive path.
    Returns a dictionary mapping keys to resolved absolute file paths.
    """
    try:
        with open(json_file, "r") as f:
            paths = json.load(f)

        # Resolve the user's OneDrive path
        onedrive_path = os.environ.get("OneDrive")
        if not onedrive_path:
            raise EnvironmentError(
                "OneDrive environment variable not found. Please ensure OneDrive is set up."
            )

        resolved_paths = {}
        for key, path in paths.items():
            if path.startswith("%OneDrive%"):
                path = path.replace("%OneDrive%", onedrive_path)
            resolved_paths[key] = str(Path(path).resolve())

        return resolved_paths

    except Exception:
        logging.exception(f"Failed to load or resolve file paths from {json_file}")
        raise


def standardize_pharmacy_ids(df):
    """
    Pads 'PHARMACYNPI' to 10 digits and 'NABP' to 7 digits in the DataFrame.

    Args:
        df (pd.DataFrame): Claims DataFrame.

    Returns:
        pd.DataFrame: Updated DataFrame with padded ID columns.
    """
    if "PHARMACYNPI" in df.columns:
        df["PHARMACYNPI"] = df["PHARMACYNPI"].astype(str).str.zfill(10)
    if "NABP" in df.columns:
        df["NABP"] = df["NABP"].astype(str).str.zfill(7)
    return df


def standardize_network_ids(network):
    """
    Pads 'pharmacy_npi' to 10 digits and 'pharmacy_nabp' to 7 digits in the network DataFrame.

    Args:
        network (pd.DataFrame): Network DataFrame.

    Returns:
        pd.DataFrame: Updated network DataFrame with padded ID columns.
    """
    if "pharmacy_npi" in network.columns:
        network["pharmacy_npi"] = network["pharmacy_npi"].astype(str).str.zfill(10)
    if "pharmacy_nabp" in network.columns:
        network["pharmacy_nabp"] = network["pharmacy_nabp"].astype(str).str.zfill(7)
    return network


def merge_with_network(df, network):
    """
    Performs a left join of df with network on ['PHARMACYNPI','NABP'] ⟷ ['pharmacy_npi','pharmacy_nabp'].

    Args:
        df (pd.DataFrame): Claims DataFrame.
        network (pd.DataFrame): Network DataFrame.

    Returns:
        pd.DataFrame: Merged DataFrame.
    """
    return df.merge(
        network,
        left_on=["PHARMACYNPI", "NABP"],
        right_on=["pharmacy_npi", "pharmacy_nabp"],
        how="left",
    )


def drop_duplicates_df(df):
    """
    Drops duplicate rows from the DataFrame.

    Args:
        df (pd.DataFrame): DataFrame to deduplicate.

    Returns:
        pd.DataFrame: Deduplicated DataFrame.
    """
    df = df.drop_duplicates()
    return df.drop_duplicates()


def clean_logic_and_tier(df, logic_col="Logic", tier_col="FormularyTier"):
    """
    Cleans 'Logic' as numeric.
    Cleans 'FormularyTier':
        - If all entries are numeric-like, coerces to numeric
        - Otherwise, strips and uppercases text for brand/generic disruptions
    """
    df[logic_col] = pd.to_numeric(df[logic_col], errors="coerce")

    # Inspect tier values
    sample = df[tier_col].dropna().astype(str).head(10)
    numeric_like = sample.str.replace(".", "", regex=False).str.isnumeric().all()

    if numeric_like:
        df[tier_col] = pd.to_numeric(df[tier_col], errors="coerce")
    else:
        df[tier_col] = df[tier_col].astype(str).str.strip().str.upper()

    return df


def filter_recent_date(df, date_col="DATEFILLED"):
    """
    Keeps only rows where date_col falls in the last 6 months (inclusive).

    Args:
        df (pd.DataFrame): DataFrame with date column.
        date_col (str): Name of the date column.

    Returns:
        pd.DataFrame: Filtered DataFrame.
    """
    latest = df[date_col].max()
    start = latest - pd.DateOffset(months=6) + pd.DateOffset(days=1)
    return df[(df[date_col] >= start) & (df[date_col] <= latest)]


def filter_logic_and_maintenance(df, config=None):
    """
    Filters rows where min_logic ≤ Logic ≤ max_logic and 'Maint Drug?' == 'Y'.

    Args:
        df (pd.DataFrame): DataFrame with logic and maintenance columns.
        config (LogicMaintenanceConfig, optional): Configuration object with filtering parameters.
                                                 If None, uses default configuration.

    Returns:
        pd.DataFrame: Filtered DataFrame.
    """
    if config is None:
        config = LogicMaintenanceConfig()
    
    return df[
        (df[config.logic_col] >= config.min_logic)
        & (df[config.logic_col] <= config.max_logic)
        & (df[config.maint_col] == "Y")
    ]


def filter_products_and_alternative(
    df, product_col="Product Name", alternative_col="Alternative"
):
    """
    Excludes rows where 'Product Name' contains albuterol, ventolin, epinephrine,
    or where 'Alternative' contains 'Covered' or 'Use different NDC'.

    Args:
        df (pd.DataFrame): DataFrame with product/alternative columns.
        product_col (str): Name of the product column.
        alternative_col (str): Name of the alternative column.

    Returns:
        pd.DataFrame: Filtered DataFrame.
    """
    exclude_pats = [r"\balbuterol\b", r"\bventolin\b", r"\bepinephrine\b"]
    for pat in exclude_pats:
        df = df[~df[product_col].str.contains(pat, case=False, na=False)]
    df = df[
        ~df[alternative_col]
        .astype(str)
        .str.contains(r"Covered|Use different NDC", case=False, regex=True, na=False)
    ]
    return df
import pandas as pd
import sys
import re
import logging
from pathlib import Path
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.utils import (
    load_file_paths,
    standardize_pharmacy_ids,
    standardize_network_ids,
    merge_with_network,
    drop_duplicates_df,
    clean_logic_and_tier,
    filter_recent_date,
    filter_logic_and_maintenance,
    filter_products_and_alternative,
    write_shared_log,
)

# Set up logger
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

try:
    import importlib.util

    if importlib.util.find_spec("xlsxwriter") is None:
        print(
            "The 'xlsxwriter' module is not installed. Please install it using 'pip install xlsxwriter'."
        )
        sys.exit(1)
except Exception:
    print("Error checking for 'xlsxwriter' module.")
    sys.exit(1)

# Load NABP/NPI list
included_nabp_npi = {
    "4528874": "1477571404",
    "2365422": "1659313435",
    "3974157": "1972560688",
    "320793": "1164437406",
    "4591055": "1851463087",
    "2348046": "1942303110",
    "4023610": "1407879588",
    "4025385": "1588706212",
    "4025311": "1588705446",
    "4026806": "1285860312",
    "4931350": "1750330775",
    "4024585": "1396768461",
    "4028026": "1497022438",
    "2643749": "1326490376",
}


# ---------------------------------------------------------------------------
# Tier summarization helper
# ---------------------------------------------------------------------------
def summarize_by_tier(df, col, from_val, to_val):
    filtered = df[(df[col] == from_val) & (df["FormularyTier"] == to_val)]
    pt = pd.pivot_table(
        filtered,
        values=["Rxs", "MemberID"],
        index=["Product Name"],
        aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
    )
    pt = pt.rename(columns={"Rxs": "Total Rxs", "MemberID": "Unique Members"})
    rxs = filtered["Rxs"].sum()
    members = filtered["MemberID"].nunique()
    return pt, rxs, members


def load_tier_disruption_data(file_paths):
    """Load all required data files for tier disruption processing."""
    # Load claims with fallback
    if not file_paths.get("reprice"):
        write_shared_log(
            "tier_disruption.py",
            "No reprice/template file provided.",
            status="ERROR",
        )
        print("No reprice/template file provided. Skipping claims loading.")
        return None

    try:
        claims = pd.read_excel(file_paths["reprice"], sheet_name="Claims Table")
    except Exception as e:
        logger.error(f"Error loading claims: {e}")
        claims = pd.read_excel(file_paths["reprice"], sheet_name=0)

    print(f"claims shape: {claims.shape}")
    claims.info()

    # Load reference tables
    medi = pd.read_excel(
        file_paths["medi_span"], usecols=["NDC", "Maint Drug?", "Product Name"]
    )
    print(f"medi shape: {medi.shape}")

    u = pd.read_excel(
        file_paths["u_disrupt"], sheet_name="Universal NDC", usecols=["NDC", "Tier"]
    )
    print(f"u shape: {u.shape}")

    e = pd.read_excel(
        file_paths["e_disrupt"],
        sheet_name="Alternatives NDC",
        usecols=["NDC", "Tier", "Alternative"],
    )
    print(f"e shape: {e.shape}")

    network = pd.read_excel(
        file_paths["n_disrupt"],
        usecols=["pharmacy_npi", "pharmacy_nabp", "pharmacy_is_excluded"],
    )
    print(f"network shape: {network.shape}")

    return claims, medi, u, e, network


def process_tier_data_pipeline(claims, reference_data, network):
    """Process the data pipeline for tier disruption."""
    medi, u, e = reference_data
    
    # Merge reference data
    df = claims.merge(medi, on="NDC", how="left")
    print(f"After merge with medi: {df.shape}")

    df = df.merge(
        u.rename(columns={"Tier": "Universal Tier"}), on="NDC", how="left"
    )
    print(f"After merge with u: {df.shape}")

    df = df.merge(
        e.rename(columns={"Tier": "Exclusive Tier"}), on="NDC", how="left"
    )
    print(f"After merge with e: {df.shape}")

    # Standardize IDs and perform network merge
    df = standardize_pharmacy_ids(df)
    print(f"After standardize_pharmacy_ids: {df.shape}")
    network = standardize_network_ids(network)
    print(f"After standardize_network_ids: {network.shape}")
    df = merge_with_network(df, network)
    print(f"After merge_with_network: {df.shape}")

    # Date parsing, deduplication, type cleaning, and filters
    df["DATEFILLED"] = pd.to_datetime(df["DATEFILLED"], errors="coerce")
    print(f"After DATEFILLED to_datetime: {df.shape}")

    df = drop_duplicates_df(df)
    print(f"After drop_duplicates_df: {df.shape}")

    df = clean_logic_and_tier(df)
    print(f"After clean_logic_and_tier: {df.shape}")

    df = filter_recent_date(df)
    print(f"After filter_recent_date: {df.shape}")

    df = filter_logic_and_maintenance(df)
    print(f"After filter_logic_and_maintenance: {df.shape}")

    df = filter_products_and_alternative(df)
    print(f"After filter_products_and_alternative: {df.shape}")

    return df


def handle_tier_pharmacy_exclusions(df, file_paths):
    """Handle pharmacy exclusions for tier disruption."""
    # Ensure 'pharmacy_is_excluded' column contains actual boolean values with type inference
    if "pharmacy_is_excluded" in df.columns:
        df["pharmacy_is_excluded"] = (
            df["pharmacy_is_excluded"]
            .astype(str)
            .str.lower()
            .map({"true": True, "false": False})
            .fillna(False)
            .infer_objects(copy=False)
        )
        logger.info(
            f"pharmacy_is_excluded value counts: {df['pharmacy_is_excluded'].value_counts().to_dict()}"
        )

        # Identify rows where pharmacy_is_excluded is "NA"
        na_pharmacies = df[df["pharmacy_is_excluded"].isna()]
        logger.info(f"NA pharmacies count: {na_pharmacies.shape[0]}")

        if not na_pharmacies.empty:
            output_file_path = file_paths["pharmacy_validation"]
            na_pharmacies_output = na_pharmacies[["PHARMACYNPI", "NABP"]].fillna("N/A")
            
            # Update existing template or create new file
            try:
                from openpyxl import load_workbook
                import os
                
                # Check if template exists
                if os.path.exists(output_file_path):
                    # Load existing workbook
                    wb = load_workbook(output_file_path)
                    logger.info(f"Loading existing template: {output_file_path}")
                else:
                    # Create new workbook
                    from openpyxl import Workbook
                    wb = Workbook()
                    logger.info(f"Creating new validation file: {output_file_path}")
                
                # Check if Validations sheet exists, if not create it
                if "Validations" in wb.sheetnames:
                    ws = wb["Validations"]
                    # Clear existing data
                    ws.delete_rows(1, ws.max_row)
                else:
                    ws = wb.create_sheet("Validations")
                    # Remove default sheet if it exists and is empty
                    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                        wb.remove(wb["Sheet"])
                
                # Write headers
                ws.append(["PHARMACYNPI", "NABP"])
                
                # Write data
                for _, row in na_pharmacies_output.iterrows():
                    ws.append([row["PHARMACYNPI"], row["NABP"]])
                
                # Save the workbook
                wb.save(output_file_path)
                logger.info(f"NA pharmacies written to '{output_file_path}' Validations sheet.")
                
            except Exception as e:
                logger.error(f"Error updating pharmacy validation file: {e}")
                # Fallback to original method
                writer = pd.ExcelWriter(output_file_path, engine="openpyxl")
                na_pharmacies_output.to_excel(
                    writer, sheet_name="Validations", index=False, engine="openpyxl"
                )
                writer.close()
                logger.info(f"NA pharmacies written to '{output_file_path}' sheet (fallback).")

    return df


def create_tier_definitions():
    """Create the tier definitions for analysis."""
    return [
        ("Universal_Positive 2-1", "Universal Tier", 1, 2),
        ("Universal_Positive 3-1", "Universal Tier", 1, 3),
        ("Universal_Positive 3-2", "Universal Tier", 2, 3),
        ("Universal_Negative 1-2", "Universal Tier", 2, 1),
        ("Universal_Negative 1-3", "Universal Tier", 3, 1),
        ("Universal_Negative 2-3", "Universal Tier", 3, 2),
        ("Exclusive_Positive 2-1", "Exclusive Tier", 1, 2),
        ("Exclusive_Positive 3-1", "Exclusive Tier", 1, 3),
        ("Exclusive_Positive 3-2", "Exclusive Tier", 2, 3),
        ("Exclusive_Negative 1-2", "Exclusive Tier", 2, 1),
        ("Exclusive_Negative 1-3", "Exclusive Tier", 3, 1),
        ("Exclusive_Negative 2-3", "Exclusive Tier", 3, 2),
    ]


def process_tier_pivots(df, tiers):
    """Process tier pivot tables and collect statistics."""
    tab_members = {}
    tab_rxs = {}
    tier_pivots = []

    for name, col, from_val, to_val in tiers:
        pt, rxs, members = summarize_by_tier(df, col, from_val, to_val)
        tier_pivots.append((name, pt, members, rxs))
        tab_members[name] = members
        tab_rxs[name] = rxs

    return tier_pivots, tab_members, tab_rxs


def process_exclusions(df):
    """Process exclusions data and create pivot table."""
    exclusions = df[df["Exclusive Tier"] == "Nonformulary"]
    ex_pt = exclusions.pivot_table(
        values=["Rxs", "MemberID"],
        index=["Product Name", "Alternative"],
        aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
    )
    ex_pt = ex_pt.rename(columns={"Rxs": "Total Rxs", "MemberID": "Unique Members"})
    exc_rxs = exclusions["Rxs"].sum()
    exc_members = exclusions["MemberID"].nunique()
    
    return ex_pt, exc_rxs, exc_members


def create_summary_dataframe(tab_members, tab_rxs, total_claims, total_members):
    """Create the summary DataFrame with calculated statistics."""
    uni_pos_keys = [
        "Universal_Positive 2-1",
        "Universal_Positive 3-1",
        "Universal_Positive 3-2",
    ]
    uni_neg_keys = [
        "Universal_Negative 1-2",
        "Universal_Negative 1-3",
        "Universal_Negative 2-3",
    ]
    ex_pos_keys = [
        "Exclusive_Positive 2-1",
        "Exclusive_Positive 3-1",
        "Exclusive_Positive 3-2",
    ]
    ex_neg_keys = [
        "Exclusive_Negative 1-2",
        "Exclusive_Negative 1-3",
        "Exclusive_Negative 2-3",
    ]

    uni_pos_utilizers = sum(tab_members[k] for k in uni_pos_keys)
    uni_pos_claims = sum(tab_rxs[k] for k in uni_pos_keys)
    uni_pos_pct = uni_pos_claims / total_claims if total_claims else 0

    uni_neg_utilizers = sum(tab_members[k] for k in uni_neg_keys)
    uni_neg_claims = sum(tab_rxs[k] for k in uni_neg_keys)
    uni_neg_pct = uni_neg_claims / total_claims if total_claims else 0

    ex_pos_utilizers = sum(tab_members[k] for k in ex_pos_keys)
    ex_pos_claims = sum(tab_rxs[k] for k in ex_pos_keys)
    ex_pos_pct = ex_pos_claims / total_claims if total_claims else 0

    ex_neg_utilizers = sum(tab_members[k] for k in ex_neg_keys)
    ex_neg_claims = sum(tab_rxs[k] for k in ex_neg_keys)
    ex_neg_pct = ex_neg_claims / total_claims if total_claims else 0

    exc_utilizers = tab_members["Exclusions"]
    exc_claims = tab_rxs["Exclusions"]
    exc_pct = exc_claims / total_claims if total_claims else 0

    return pd.DataFrame(
        {
            "Formulary": [
                "Universal Positive",
                "Universal Negative",
                "Exclusive Positive",
                "Exclusive Negative",
                "Exclusions",
            ],
            "Utilizers": [
                uni_pos_utilizers,
                uni_neg_utilizers,
                ex_pos_utilizers,
                ex_neg_utilizers,
                exc_utilizers,
            ],
            "Rxs": [
                uni_pos_claims,
                uni_neg_claims,
                ex_pos_claims,
                ex_neg_claims,
                exc_claims,
            ],
            "% of claims": [
                uni_pos_pct,
                uni_neg_pct,
                ex_pos_pct,
                ex_neg_pct,
                exc_pct,
            ],
            "": ["", "", "", "", ""],
            "Totals": [
                f"Members: {total_members}",
                f"Claims: {total_claims}",
                "",
                "",
                "",
            ],
        }
    )


def create_network_analysis(df):
    """Create network analysis for excluded pharmacies."""
    network_df = df[df["pharmacy_is_excluded"]]
    logger.debug(f"Initial network_df shape: {network_df.shape}")
    logger.debug(f"Initial network_df contents: {network_df.head(10).to_dict()}")

    filter_phrases = [
        "CVS",
        "Walgreens",
        "Kroger",
        "Walmart",
        "Rite Aid",
        "Optum",
        "Express Scripts",
        "DMR",
        "Williams Bro",
        "Publix",
    ]
    filter_phrases_escaped = [
        re.escape(phrase.lower()) for phrase in filter_phrases
    ]
    regex_pattern = "|".join([f"\\b{p}\\b" for p in filter_phrases_escaped])
    network_df = network_df[
        ~network_df["Pharmacy Name"]
        .str.lower()
        .str.contains(regex_pattern, case=False, regex=True, na=False)
    ]
    logger.debug(f"Filtered network_df shape: {network_df.shape}")
    logger.debug(f"Filtered network_df contents: {network_df.head(10).to_dict()}")

    if {"PHARMACYNPI", "NABP", "Pharmacy Name"}.issubset(network_df.columns):
        network_df["PHARMACYNPI"] = network_df["PHARMACYNPI"].fillna("N/A")
        network_df["NABP"] = network_df["NABP"].fillna("NABP")
        # Create Unique Identifier based on whichever identifier is available (PHARMACYNPI or NABP)
        network_df["Unique Identifier"] = network_df.apply(
            lambda row: row["PHARMACYNPI"]
            if row["PHARMACYNPI"] != "N/A"
            else row["NABP"],
            axis=1,
        )
        network_pivot = pd.pivot_table(
            network_df,
            values=["Rxs", "MemberID"],
            index=["Unique Identifier"],
            aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
        )
        network_pivot = network_pivot.rename(
            columns={"Rxs": "Total Rxs", "MemberID": "Unique Members"}
        )
        network_pivot.reset_index(
            inplace=True
        )  # Ensure index columns are included in the output
        logger.debug(f"Network pivot shape: {network_pivot.shape}")
        logger.debug(f"Network pivot contents: {network_pivot.head(10).to_dict()}")
        
        return network_df, network_pivot
    
    return network_df, None


def write_excel_sheets(writer, df, summary_df, tier_pivots, ex_pt, exc_members, network_df, network_pivot):
    """Write all sheets to the Excel file."""
    # Write Summary sheet
    summary_df.to_excel(writer, sheet_name="Summary", index=False)

    # Write tier pivots
    for name, pt, members, _ in tier_pivots:
        pt.to_excel(writer, sheet_name=name)
        writer.sheets[name].write("F1", f"Total Members: {members}")

    # Write Exclusions sheet
    ex_pt.to_excel(writer, sheet_name="Exclusions")
    writer.sheets["Exclusions"].write("F1", f"Total Members: {exc_members}")

    # Write Data sheet
    data_sheet_df = df.copy()
    data_sheet_df.to_excel(writer, sheet_name="Data", index=False)

    # Write Network sheet
    if network_pivot is not None:
        network_pivot.to_excel(writer, sheet_name="Network", index=False)

    # Write filtered network data
    selected_columns = [
        "PHARMACYNPI",
        "NABP",
        "MemberID",
        "Pharmacy Name",
        "pharmacy_is_excluded",
        "Unique Identifier",
    ]
    network_df[selected_columns].to_excel(writer, sheet_name="Network", index=False)
    logger.info(
        f"Network sheet updated with {network_df.shape[0]} excluded pharmacy records (minus major chains) and selected columns"
    )


def reorder_excel_sheets(writer):
    """Reorder sheets so Summary follows Data."""
    sheets = writer.sheets  # This is a dict: {sheet_name: worksheet_object}
    names = list(sheets.keys())
    if "Data" in names and "Summary" in names:
        data_idx = names.index("Data")
        summary_idx = names.index("Summary")
        if summary_idx != data_idx + 1:
            # Move "Summary" sheet immediately after "Data"
            items = list(sheets.items())
            summary_item = items.pop(summary_idx)
            items.insert(data_idx + 1, summary_item)
            writer.sheets.clear()
            writer.sheets.update(dict(items))


def show_completion_message(output_path):
    """Show completion message and popup."""
    write_shared_log("tier_disruption.py", "Processing complete.")
    print(f"Processing complete. Output file: {output_path}")
    try:
        import tkinter as tk
        from tkinter import messagebox

        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo("Processing Complete", "Processing complete")
        root.destroy()
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Main processing pipeline
# ---------------------------------------------------------------------------
def process_data():
    write_shared_log("tier_disruption.py", "Processing started.")
    # Output filename from CLI arg or default
    output_filename = "LBL for Disruption.xlsx"
    if len(sys.argv) > 1:
        output_filename = sys.argv[1]
    output_path = Path(output_filename).resolve()
    try:
        # Get the config file path relative to the project root
        config_path = Path(__file__).parent.parent / "config" / "file_paths.json"
        file_paths = load_file_paths(str(config_path))

        result = load_tier_disruption_data(file_paths)
        if result is None:
            return  # Early exit if claims loading failed
        claims, medi, u, e, network = result

        reference_data = (medi, u, e)
        df = process_tier_data_pipeline(claims, reference_data, network)

        df = handle_tier_pharmacy_exclusions(df, file_paths)

        # Totals for summary
        total_claims = df["Rxs"].sum()
        total_members = df["MemberID"].nunique()

        # Excel writer setup
        writer = pd.ExcelWriter(output_path, engine="xlsxwriter")

        # Summary calculations (must be written immediately after Data)
        tiers = create_tier_definitions()

        tier_pivots, tab_members, tab_rxs = process_tier_pivots(df, tiers)

        # Exclusions sheet (Nonformulary)
        ex_pt, exc_rxs, exc_members = process_exclusions(df)
        tab_members["Exclusions"] = exc_members
        tab_rxs["Exclusions"] = exc_rxs

        # Summary calculations
        summary_df = create_summary_dataframe(tab_members, tab_rxs, total_claims, total_members)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # Write tier pivots and Exclusions after Summary
        for name, pt, members, _ in tier_pivots:
            pt.to_excel(writer, sheet_name=name)
            writer.sheets[name].write("F1", f"Total Members: {members}")

        ex_pt.to_excel(writer, sheet_name="Exclusions")
        writer.sheets["Exclusions"].write("F1", f"Total Members: {exc_members}")

        # Write the 'Data Sheet' with excluded and non-excluded pharmacies
        data_sheet_df = df.copy()
        data_sheet_df.to_excel(writer, sheet_name="Data", index=False)

        # Network summary for excluded pharmacies (pharmacy_is_excluded=True)
        network_df, network_pivot = create_network_analysis(df)
        logger.info(f"Total pharmacies in dataset: {df.shape[0]}")
        logger.info(
            f"Excluded pharmacies (pharmacy_is_excluded=True): {df['pharmacy_is_excluded'].sum()}"
        )
        logger.info(
            f"Non-excluded pharmacies (pharmacy_is_excluded=False): {(~df['pharmacy_is_excluded']).sum()}"
        )
        logger.info(
            f"Network sheet will show {network_df.shape[0]} excluded pharmacy records (minus major chains)"
        )

        # Write Network sheet
        if network_pivot is not None:
            network_pivot.to_excel(writer, sheet_name="Network", index=False)

        # Write filtered network data
        selected_columns = [
            "PHARMACYNPI",
            "NABP",
            "MemberID",
            "Pharmacy Name",
            "pharmacy_is_excluded",
            "Unique Identifier",
        ]
        network_df[selected_columns].to_excel(writer, sheet_name="Network", index=False)
        logger.info(
            f"Network sheet updated with {network_df.shape[0]} excluded pharmacy records (minus major chains) and selected columns"
        )

        # Reorder sheets so Summary follows Data
        reorder_excel_sheets(writer)

        writer.close()
        show_completion_message(output_path)
    except Exception as e:
        write_shared_log(
            "tier_disruption.py", f"Processing failed: {e}", status="ERROR"
        )
        raise


if __name__ == "__main__":
    process_data()
    print("Data processing complete")
import logging
import tkinter as tk
from pathlib import Path
from tkinter import messagebox
import os
import sys

import pandas as pd
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.excel_utils import write_df_to_template
from utils.utils import load_file_paths, write_shared_log

CLAIMS_SHEET = "Claims Table"
OUTPUT_SHEET = "Line By Line"

# Setup logging
logging.basicConfig(
    filename="sharx_lbl.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


def show_message(awp: float, ing: float, total: float, rxs: int) -> None:
    messagebox.showinfo(
        "Process Complete",
        f"SHARx LBL has been created\n\n"
        f"Total AWP: ${awp:.2f}\n\n"
        f"Total Ing Cost: ${ing:.2f}\n\n"
        f"Total Gross Cost: ${total:.2f}\n\n"
        f"Total Claim Count: {rxs}",
    )


def main():
    root = tk.Tk()
    root.withdraw()
    write_shared_log("sharx_lbl.py", "Processing started.")

    try:
        # Get the config file path relative to the project root
        config_path = Path(__file__).parent.parent / "config" / "file_paths.json"
        paths = load_file_paths(str(config_path))
        # Fallback to file dialogs if required keys are missing
        if "reprice" not in paths:
            from tkinter import filedialog

            paths["reprice"] = filedialog.askopenfilename(title="Select Claims File")
        if "sharx" not in paths:
            from tkinter import filedialog

            paths["sharx"] = filedialog.askopenfilename(
                title="Select SHARx Template File"
            )

        template_path = Path(paths["sharx"])
        try:
            df = pd.read_excel(paths["reprice"], sheet_name=CLAIMS_SHEET)
        except FileNotFoundError:
            logger.error(f"Claims file not found: {paths['reprice']}")
            raise FileNotFoundError(f"Claims file not found: {paths['reprice']}")
        except ValueError as e:
            logger.error(f"Sheet loading failed: {e}")
            raise ValueError(f"Sheet loading failed: {e}")

        # Debug print and log
        print("Columns in claims DataFrame:")
        print(df.columns)
        logger.info(f"Columns in claims DataFrame: {df.columns.tolist()}")

        df["Logic"] = pd.to_numeric(df["Logic"], errors="coerce")
        df = df[df["Logic"].between(1, 10)]

        awp = df["Total AWP (Historical)"].sum()
        ing = df["Rx Sense Ing Cost"].sum()
        total = df["RxSense Total Cost"].sum()
        rxs = df["Rxs"].sum()

        columns_to_keep = [
            "MONY",
            "Rxs",
            "Rx Sense Ing Cost",
            "RxSense Dispense Fee",
            "RxSense Total Cost",
            "Total AWP (Historical)",
            "GrossCost",
            "Pharmacy Name",
            "INPUTFILECHANNEL",
            "DATEFILLED",
            "MemberID",
            "DAYSUPPLY",
            "QUANTITY",
            "NDC",
            "DST Drug Name",
        ]
        missing_cols = [col for col in columns_to_keep if col not in df.columns]
        if missing_cols:
            raise ValueError(f"Missing columns in input data: {missing_cols}")
        df = df[columns_to_keep]

        output_path = Path("_Rx Claims for SHARx.xlsx")

        write_df_to_template(
            str(template_path),
            str(output_path),
            sheet_name=OUTPUT_SHEET,
            df=df,
            start_cell="A2",
            header=False,
            index=False,
            open_file=False,
            visible=False,
        )

        logger.info(f"SHARx output saved to: {output_path}")
        logger.info("SHARx LBL file created successfully.")
        write_shared_log("sharx_lbl.py", "SHARx LBL file created successfully.")
        show_message(awp, ing, total, rxs)
        messagebox.showinfo("Processing Complete", "Processing complete")

    except Exception as e:
        logger.exception("An error occurred during SHARx LBL processing")
        write_shared_log("sharx_lbl.py", f"An error occurred: {e}", status="ERROR")
        messagebox.showerror("Error", f"An error occurred: {e}")
    finally:
        root.quit()


if __name__ == "__main__":
    main()
"""
Configuration module for the Repricing Automation application.
Contains configuration classes and constants.
"""

import multiprocessing
from pathlib import Path


class ProcessingConfig:
    """Configuration class for processing settings and validation."""
    
    REQUIRED_COLUMNS = [
        "DATEFILLED", "SOURCERECORDID", "QUANTITY", "DAYSUPPLY", "NDC",
        "MemberID", "Drug Name", "Pharmacy Name", "Total AWP (Historical)"
    ]
    
    FILE_TYPES = [
        ("All files", "*.*"),
        ("CSV files", "*.csv"),
        ("Excel files", "*.xlsx"),
    ]
    
    TEMPLATE_FILE_TYPES = [("Excel files", "*.xlsx")]
    
    DEFAULT_OPPORTUNITY_NAME = "claims detail PCU"
    
    @classmethod
    def get_multiprocessing_workers(cls):
        """Get the optimal number of workers for multiprocessing."""
        return min(4, max(1, multiprocessing.cpu_count() // 2))
    
    @classmethod
    def validate_required_columns(cls, df):
        """Validate that all required columns are present in the DataFrame."""
        missing_cols = [col for col in cls.REQUIRED_COLUMNS if col not in df.columns]
        if missing_cols:
            raise ValueError(f"Missing required columns: {missing_cols}")
        return True


class DisruptionConfig:
    """Configuration for disruption types to reduce conditional complexity."""
    
    DISRUPTION_TYPES = {
        "Tier Disruption": {
            "module": "modules.tier_disruption",
            "file": "tier_disruption.py"
        },
        "B/G Disruption": {
            "module": "modules.bg_disruption", 
            "file": "bg_disruption.py"
        },
        "OpenMDF (Tier)": {
            "module": "modules.openmdf_tier",
            "file": "openmdf_tier.py"
        },
        "OpenMDF (B/G)": {
            "module": "modules.openmdf_bg",
            "file": "openmdf_bg.py"
        },
        "Full Claims File": {
            "module": "modules.full_claims",
            "file": "full_claims.py"
        }
    }
    
    @classmethod
    def get_program_file(cls, disruption_type):
        """Get the program file for a disruption type."""
        config = cls.DISRUPTION_TYPES.get(disruption_type)
        return config["file"] if config else None
    
    @classmethod
    def get_disruption_labels(cls):
        """Get list of available disruption types (excluding Full Claims File)."""
        return [
            label for label in cls.DISRUPTION_TYPES.keys() 
            if label != "Full Claims File"
        ]


class AppConstants:
    """Application constants and configuration values."""
    
    # Configuration and audit log files
    CONFIG_FILE = Path("config.json")
    AUDIT_LOG = Path("audit_log.csv")
    
    # Template handling constants
    BACKUP_SUFFIX = "_backup.xlsx"
    UPDATED_TEMPLATE_NAME = "_Rx Repricing_wf.xlsx"
    
    # Welcome messages for different users
    WELCOME_MESSAGES = {
        "DamionMorrison": "Welcome back, Damion! Ready to reprice?",
        "DannyBushnell": "Hello Danny! Let's get started.",
        "BrettBauer": "Hi Brett, your automation awaits!",
        "BrendanReamer": "Welcome Brendan! Ready to optimize?",
    }
    
    # Emoji options for welcome message
    EMOJIS = [
        ":rocket:", ":sunglasses:", ":star:", 
        ":tada:", ":computer:", ":chart_with_upwards_trend:"
    ]
    
    # Notes text for UI
    NOTES_TEXT = (
        "Note:\n\n"
        "Ensure FormularyTier is set before running disruption.\n"
        "Close any open Excel instances.\n"
        "Keep template name as _Rx Repricing_wf until done."
    )
"""
Configuration Management - CodeScene ACE Improvement
Centralized configuration handling with better error management
"""

import json
import logging
from pathlib import Path
from typing import Dict, Any, Optional
from dataclasses import dataclass, asdict

logger = logging.getLogger(__name__)


@dataclass
class AppSettings:
    """Application settings data class."""
    last_folder: str
    theme: str = "light"
    auto_save: bool = True
    log_level: str = "INFO"
    max_workers: int = 4
    backup_enabled: bool = True
    
    @classmethod
    def default(cls) -> 'AppSettings':
        """Create default settings."""
        return cls(
            last_folder=str(Path.cwd()),
            theme="light",
            auto_save=True,
            log_level="INFO",
            max_workers=4,
            backup_enabled=True
        )


class ConfigurationManager:
    """
    Improved configuration management following CodeScene ACE principles.
    - Single responsibility: Only handles configuration
    - Better error handling
    - Type safety with dataclasses
    - Clear separation of concerns
    """
    
    def __init__(self, config_file: Path = None):
        self.config_file = config_file or Path("config.json")
        self._settings: Optional[AppSettings] = None
        self._load_configuration()
    
    def _load_configuration(self) -> None:
        """Load configuration from file or create default."""
        try:
            if self.config_file.exists():
                self._settings = self._load_from_file()
                logger.info(f"Configuration loaded from {self.config_file}")
            else:
                self._settings = AppSettings.default()
                self._save_configuration()
                logger.info("Default configuration created")
        except Exception as e:
            logger.error(f"Failed to load configuration: {e}")
            self._settings = AppSettings.default()
    
    def _load_from_file(self) -> AppSettings:
        """Load settings from JSON file."""
        try:
            with open(self.config_file, 'r') as f:
                data = json.load(f)
                return AppSettings(**data)
        except (json.JSONDecodeError, TypeError) as e:
            logger.warning(f"Invalid configuration file: {e}")
            return AppSettings.default()
    
    def _save_configuration(self) -> None:
        """Save current settings to file."""
        try:
            if self._settings is not None:
                with open(self.config_file, 'w') as f:
                    json.dump(asdict(self._settings), f, indent=4)
                logger.debug(f"Configuration saved to {self.config_file}")
            else:
                logger.error("No settings to save: self._settings is None")
                return
        except Exception as e:
            logger.error(f"Failed to save configuration: {e}")
    
    @property
    def settings(self) -> AppSettings:
        """Get current settings."""
        return self._settings
    
    def update_setting(self, key: str, value: Any) -> bool:
        """Update a single setting."""
        try:
            if hasattr(self._settings, key):
                setattr(self._settings, key, value)
                self._save_configuration()
                logger.info(f"Setting updated: {key} = {value}")
                return True
            else:
                logger.warning(f"Unknown setting: {key}")
                return False
        except Exception as e:
            logger.error(f"Failed to update setting {key}: {e}")
            return False
    
    def update_settings(self, **kwargs) -> bool:
        """Update multiple settings."""
        try:
            for key, value in kwargs.items():
                if not hasattr(self._settings, key):
                    logger.warning(f"Unknown setting: {key}")
                    continue
                setattr(self._settings, key, value)
            
            self._save_configuration()
            logger.info(f"Settings updated: {list(kwargs.keys())}")
            return True
        except Exception as e:
            logger.error(f"Failed to update settings: {e}")
            return False
    
    def reset_to_defaults(self) -> None:
        """Reset settings to defaults."""
        self._settings = AppSettings.default()
        self._save_configuration()
        logger.info("Settings reset to defaults")
    
    def get_setting(self, key: str, default: Any = None) -> Any:
        """Get a specific setting value."""
        return getattr(self._settings, key, default)


# Usage example for improved app.py
class ImprovedConfigManager(ConfigurationManager):
    """
    Backwards compatible configuration manager.
    Maintains the same interface as the original while providing improvements.
    """
    
    def __init__(self):
        super().__init__(Path("config.json"))
        # Maintain backwards compatibility
        self.config = self._get_legacy_config()
    
    def _get_legacy_config(self) -> Dict[str, Any]:
        """Get configuration in legacy format."""
        return {
            "last_folder": self.settings.last_folder,
            "theme": self.settings.theme,
            "auto_save": self.settings.auto_save,
            "log_level": self.settings.log_level,
            "max_workers": self.settings.max_workers,
            "backup_enabled": self.settings.backup_enabled
        }
    
    def save_default(self) -> None:
        """Legacy method for saving defaults."""
        self.reset_to_defaults()
        self.config = self._get_legacy_config()
    
    def load(self) -> None:
        """Legacy method for loading configuration."""
        self._load_configuration()
        self.config = self._get_legacy_config()
    
    def save(self) -> None:
        """Legacy method for saving configuration."""
        self._save_configuration()


# Example usage in app_improved.py
def example_usage():
    """Example of how to use the improved configuration manager."""
    
    # Create configuration manager
    config_manager = ImprovedConfigManager()
    
    # Access settings
    print(f"Last folder: {config_manager.settings.last_folder}")
    print(f"Theme: {config_manager.settings.theme}")
    
    # Update settings
    config_manager.update_setting("last_folder", "/new/path")
    config_manager.update_settings(theme="dark", auto_save=False)
    
    # Use in app initialization
    app_settings = config_manager.settings
    print(f"Starting app with theme: {app_settings.theme}")
    print(f"Auto-save enabled: {app_settings.auto_save}")


if __name__ == "__main__":
    example_usage()
import logging
import os
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.utils import write_shared_log


def make_audit_entry(script_name, message, status="INFO"):
    try:
        write_shared_log(script_name, message, status)
    except Exception as e:
        logging.error(f"[AUDIT FAIL] {script_name} audit failed: {e}")
        try:
            with open("local_fallback_log.txt", "a") as f:
                f.write(f"{script_name}: {message} [{status}]\n")
        except Exception as inner:
            logging.error(f"[FALLBACK FAIL] Could not write fallback log: {inner}")
import pandas as pd
import logging
import os
import sys
from pathlib import Path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.utils import (
    load_file_paths,
    filter_logic_and_maintenance,
    filter_products_and_alternative,
    standardize_pharmacy_ids,
    standardize_network_ids,
    merge_with_network,
    drop_duplicates_df,
    clean_logic_and_tier,
    filter_recent_date,
    write_shared_log,
)

# Logging setup
logging.basicConfig(
    filename="bg_disruption.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


def load_data_files(file_paths):
    """Load and return all required data files."""
    logger.info("Loading data files...")
    
    # Load claims data
    try:
        claims = pd.read_excel(
            file_paths["reprice"],
            sheet_name="Claims Table",
            usecols=[
                "SOURCERECORDID",
                "NDC",
                "MemberID",
                "DATEFILLED",
                "FormularyTier",
                "Rxs",
                "Logic",
                "PHARMACYNPI",
                "NABP",
                "Pharmacy Name",
                "Universal Rebates",
                "Exclusive Rebates",
            ],
        )
    except Exception as e:
        logger.warning(f"Claims Table fallback: {e}")
        write_shared_log(
            "bg_disruption.py", f"Claims Table fallback: {e}", status="WARNING"
        )
        claims = pd.read_excel(file_paths["reprice"], sheet_name=0)
    
    logger.info(f"claims shape: {claims.shape}")
    claims.info()

    # Load other data files
    medi = pd.read_excel(file_paths["medi_span"])[
        ["NDC", "Maint Drug?", "Product Name"]
    ]
    logger.info(f"medi shape: {medi.shape}")
    
    uni = pd.read_excel(file_paths["u_disrupt"], sheet_name="Universal NDC")[
        ["NDC", "Tier"]
    ]
    logger.info(f"uni shape: {uni.shape}")
    
    exl = pd.read_excel(file_paths["e_disrupt"], sheet_name="Alternatives NDC")[
        ["NDC", "Tier", "Alternative"]
    ]
    logger.info(f"exl shape: {exl.shape}")
    
    network = pd.read_excel(file_paths["n_disrupt"])[
        ["pharmacy_npi", "pharmacy_nabp", "pharmacy_is_excluded"]
    ]
    logger.info(f"network shape: {network.shape}")
    
    return claims, medi, uni, exl, network


def merge_data_files(claims, reference_data, network):
    """Merge all data files into a single DataFrame."""
    logger.info("Merging data files...")
    
    medi, uni, exl = reference_data
    
    df = claims.merge(medi, on="NDC", how="left")
    logger.info(f"After merge with medi: {df.shape}")
    
    df = df.merge(uni.rename(columns={"Tier": "Universal Tier"}), on="NDC", how="left")
    logger.info(f"After merge with uni: {df.shape}")
    
    df = df.merge(exl.rename(columns={"Tier": "Exclusive Tier"}), on="NDC", how="left")
    logger.info(f"After merge with exl: {df.shape}")

    # Standardize IDs and perform network merge
    df = standardize_pharmacy_ids(df)
    logger.info(f"After standardize_pharmacy_ids: {df.shape}")
    network = standardize_network_ids(network)
    logger.info(f"After standardize_network_ids: {network.shape}")
    df = merge_with_network(df, network)
    logger.info(f"After merge_with_network: {df.shape}")
    
    return df


def process_and_filter_data(df):
    """Process and filter the merged data."""
    logger.info("Processing and filtering data...")
    
    # Date parsing, deduplication, type cleaning, and filters
    df["DATEFILLED"] = pd.to_datetime(df["DATEFILLED"], errors="coerce")
    logger.info(f"After DATEFILLED to_datetime: {df.shape}")
    
    df = drop_duplicates_df(df)
    logger.info(f"After drop_duplicates_df: {df.shape}")
    
    df = clean_logic_and_tier(df)
    logger.info(f"After clean_logic_and_tier: {df.shape}")
    
    df = filter_recent_date(df)
    logger.info(f"After filter_recent_date: {df.shape}")
    
    df = filter_logic_and_maintenance(df)
    logger.info(f"After filter_logic_and_maintenance: {df.shape}")
    
    df = filter_products_and_alternative(df)
    logger.info(f"After filter_products_and_alternative: {df.shape}")

    df["FormularyTier"] = df["FormularyTier"].astype(str).str.strip().str.upper()
    df["Alternative"] = df["Alternative"].astype(str)
    
    return df


def handle_pharmacy_exclusions(df, file_paths):
    """Handle pharmacy exclusions and validation."""
    logger.info("Handling pharmacy exclusions...")
    
    # Ensure 'pharmacy_is_excluded' column contains actual boolean values with type inference
    if "pharmacy_is_excluded" in df.columns:
        df["pharmacy_is_excluded"] = (
            df["pharmacy_is_excluded"]
            .astype(str)
            .str.lower()
            .map({"true": True, "false": False})
            .fillna(False)
            .infer_objects(copy=False)
        )
        logger.info(
            f"pharmacy_is_excluded value counts: {df['pharmacy_is_excluded'].value_counts().to_dict()}"
        )
        
        # Identify rows where pharmacy_is_excluded is "NA"
        na_pharmacies = df[df["pharmacy_is_excluded"].isna()]
        logger.info(f"NA pharmacies count: {na_pharmacies.shape[0]}")

        if not na_pharmacies.empty:
            # Define the writer before using it
            output_file_path = file_paths["pharmacy_validation"]
            na_pharmacies_output = na_pharmacies[["PHARMACYNPI", "NABP"]].fillna("N/A")
            
            # Update existing template or create new file
            try:
                from openpyxl import load_workbook
                import os
                
                # Check if template exists
                if os.path.exists(output_file_path):
                    # Load existing workbook
                    wb = load_workbook(output_file_path)
                    logger.info(f"Loading existing template: {output_file_path}")
                else:
                    # Create new workbook
                    from openpyxl import Workbook
                    wb = Workbook()
                    logger.info(f"Creating new validation file: {output_file_path}")
                
                # Check if Validations sheet exists, if not create it
                if "Validations" in wb.sheetnames:
                    ws = wb["Validations"]
                    # Clear existing data
                    ws.delete_rows(1, ws.max_row)
                else:
                    ws = wb.create_sheet("Validations")
                    # Remove default sheet if it exists and is empty
                    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                        wb.remove(wb["Sheet"])
                
                # Write headers
                ws.append(["PHARMACYNPI", "NABP"])
                
                # Write data
                for _, row in na_pharmacies_output.iterrows():
                    ws.append([row["PHARMACYNPI"], row["NABP"]])
                
                # Save the workbook
                wb.save(output_file_path)
                logger.info(f"NA pharmacies written to '{output_file_path}' Validations sheet.")
                
            except Exception as e:
                logger.error(f"Error updating pharmacy validation file: {e}")
                # Fallback to original method
                writer = pd.ExcelWriter(output_file_path, engine="openpyxl")
                na_pharmacies_output.to_excel(
                    writer, sheet_name="Validations", index=False, engine="openpyxl"
                )
                writer.close()
                logger.info(f"NA pharmacies written to '{output_file_path}' sheet (fallback).")
    
    return df


def create_data_filters(df):
    """Create filtered datasets for different scenarios."""
    logger.info("Creating data filters...")
    
    uni_pos = df[(df["Universal Tier"] == 1) & df["FormularyTier"].isin(["B", "BRAND"])]
    logger.info(f"uni_pos shape: {uni_pos.shape}")
    
    uni_neg = df[
        df["Universal Tier"].isin([2, 3]) & df["FormularyTier"].isin(["G", "GENERIC"])
    ]
    logger.info(f"uni_neg shape: {uni_neg.shape}")
    
    ex_pos = df[(df["Exclusive Tier"] == 1) & df["FormularyTier"].isin(["B", "BRAND"])]
    logger.info(f"ex_pos shape: {ex_pos.shape}")
    
    ex_neg = df[
        df["Exclusive Tier"].isin([2, 3]) & df["FormularyTier"].isin(["G", "GENERIC"])
    ]
    logger.info(f"ex_neg shape: {ex_neg.shape}")
    
    ex_ex = df[df["Exclusive Tier"] == "Nonformulary"]
    logger.info(f"ex_ex shape: {ex_ex.shape}")
    
    return uni_pos, uni_neg, ex_pos, ex_neg, ex_ex


def create_pivot_tables(filtered_data):
    """Create pivot tables and calculate member counts."""
    logger.info("Creating pivot tables...")
    
    uni_pos, uni_neg, ex_pos, ex_neg, ex_ex = filtered_data
    
    def pivot(d, include_alternative=False):
        index_cols = ["Product Name"]
        if include_alternative and "Alternative" in d.columns:
            index_cols.append("Alternative")
        pt = pd.pivot_table(
            d,
            values=["Rxs", "MemberID"],
            index=index_cols,
            aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
        )
        pt = pt.rename(columns={"Rxs": "Total Rxs", "MemberID": "Unique Members"})
        return pt

    def count(d):
        return 0 if d.empty or d["Rxs"].sum() == 0 else d["MemberID"].nunique()

    tabs = {
        "Universal_Positive": (uni_pos, pivot(uni_pos), count(uni_pos)),
        "Universal_Negative": (uni_neg, pivot(uni_neg), count(uni_neg)),
        "Exclusive_Positive": (ex_pos, pivot(ex_pos), count(ex_pos)),
        "Exclusive_Negative": (ex_neg, pivot(ex_neg), count(ex_neg)),
        "Exclusions": (ex_ex, pivot(ex_ex, include_alternative=True), count(ex_ex)),
    }
    
    return tabs


def create_summary_data(df, tabs):
    """Create summary data for the report."""
    logger.info("Creating summary data...")
    
    total_members = df["MemberID"].nunique()
    total_claims = df["Rxs"].sum()

    summary = pd.DataFrame(
        {
            "Formulary": [
                "Universal Positive",
                "Universal Negative",
                "Exclusive Positive",
                "Exclusive Negative",
                "Exclusions",
            ],
            "Utilizers": [v[2] for v in tabs.values()],
            "Rxs": [v[0]["Rxs"].sum() for v in tabs.values()],
            "% of claims": [
                v[0]["Rxs"].sum() / total_claims if total_claims else 0
                for v in tabs.values()
            ],
            "": ["" for _ in tabs],
            "Totals": [
                f"Members: {total_members}",
                f"Claims: {total_claims}",
                "",
                "",
                "",
            ],
        }
    )
    
    return summary


def create_network_data(df):
    """Create network data for excluded pharmacies."""
    logger.info("Creating network data...")
    
    import re
    
    # Network summary for excluded pharmacies (pharmacy_is_excluded=True)
    network_df = df[df["pharmacy_is_excluded"]]
    filter_phrases = [
        "CVS",
        "Walgreens",
        "Kroger",
        "Walmart",
        "Rite Aid",
        "Optum",
        "Express Scripts",
        "DMR",
        "Williams Bro",
        "Publix",
    ]
    
    # Regex safety: escape and lower-case all phrases and names
    filter_phrases_escaped = [re.escape(phrase.lower()) for phrase in filter_phrases]
    regex_pattern = "|".join([f"\\b{p}\\b" for p in filter_phrases_escaped])
    network_df = network_df[
        ~network_df["Pharmacy Name"]
        .str.lower()
        .str.contains(regex_pattern, case=False, regex=True, na=False)
    ]
    logger.info(f"network_df shape after exclusion: {network_df.shape}")
    
    network_pivot = None
    if {"PHARMACYNPI", "NABP", "Pharmacy Name"}.issubset(network_df.columns):
        network_df["PHARMACYNPI"] = network_df["PHARMACYNPI"].fillna("N/A")
        network_df["NABP"] = network_df["NABP"].fillna("N/A")
        network_pivot = pd.pivot_table(
            network_df,
            values=["Rxs", "MemberID"],
            index=["PHARMACYNPI", "NABP", "Pharmacy Name"],
            aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
        )
        network_pivot = network_pivot.rename(
            columns={"Rxs": "Total Rxs", "MemberID": "Unique Members"}
        )
    
    # Log debug info to verify the filtering
    logger.info(f"Total pharmacies in dataset: {df.shape[0]}")
    logger.info(
        f"Excluded pharmacies (pharmacy_is_excluded=True): {df['pharmacy_is_excluded'].sum()}"
    )
    logger.info(
        f"Non-excluded pharmacies (pharmacy_is_excluded=False): {(~df['pharmacy_is_excluded']).sum()}"
    )
    logger.info(
        f"Network sheet will show {network_df.shape[0]} excluded pharmacy records (minus major chains)"
    )
    
    return network_pivot


def write_excel_report(report_data, output_filename):
    """Write the final Excel report."""
    logger.info("Writing Excel report...")
    
    df, summary, tabs, network_pivot = report_data
    
    writer = pd.ExcelWriter(output_filename, engine="xlsxwriter")
    df.to_excel(writer, sheet_name="Data", index=False)
    summary.to_excel(writer, sheet_name="Summary", index=False)

    for sheet, (_, pt, mems) in tabs.items():
        pt.to_excel(writer, sheet_name=sheet)
        writer.sheets[sheet].write("F1", f"Total Members: {mems}")

    if network_pivot is not None:
        network_pivot.to_excel(writer, sheet_name="Network")

    # Reorder sheets so Summary follows Data
    sheets = writer.sheets  # This is a dict: {sheet_name: worksheet_object}
    names = list(sheets.keys())
    if "Data" in names and "Summary" in names:
        data_idx = names.index("Data")
        summary_idx = names.index("Summary")
        if summary_idx != data_idx + 1:
            # Move "Summary" sheet immediately after "Data"
            items = list(sheets.items())
            summary_item = items.pop(summary_idx)
            items.insert(data_idx + 1, summary_item)
            writer.sheets.clear()
            writer.sheets.update(dict(items))

    writer.close()


def show_completion_notification():
    """Show completion notification popup."""
    try:
        import tkinter as tk
        from tkinter import messagebox

        root = tk.Tk()
        root.withdraw()  # Hide the main window
        messagebox.showinfo("Notification", "Processing complete")
        root.destroy()
    except Exception as e:
        logger.warning(f"Popup notification failed: {e}")


def process_data():
    """Main processing function - coordinates all data processing steps."""
    write_shared_log("bg_disruption.py", "Processing started.")
    import sys

    output_filename = "LBL for Disruption.xlsx"
    if len(sys.argv) > 1:
        output_filename = sys.argv[1]
    
    # Get the config file path relative to the project root
    config_path = Path(__file__).parent.parent / "config" / "file_paths.json"
    file_paths = load_file_paths(str(config_path))
    if "reprice" not in file_paths or not file_paths["reprice"]:
        write_shared_log(
            "bg_disruption.py", "No reprice/template file provided.", status="ERROR"
        )
        print("Error: No reprice/template file provided.")
        return

    # Load all data files
    claims, medi, uni, exl, network = load_data_files(file_paths)
    
    # Merge all data files
    reference_data = (medi, uni, exl)
    df = merge_data_files(claims, reference_data, network)
    
    # Process and filter data
    df = process_and_filter_data(df)
    
    # Handle pharmacy exclusions
    df = handle_pharmacy_exclusions(df, file_paths)
    
    # Create filtered datasets
    uni_pos, uni_neg, ex_pos, ex_neg, ex_ex = create_data_filters(df)
    
    # Create pivot tables
    filtered_data = (uni_pos, uni_neg, ex_pos, ex_neg, ex_ex)
    tabs = create_pivot_tables(filtered_data)
    
    # Create summary data
    summary = create_summary_data(df, tabs)
    
    # Create network data
    network_pivot = create_network_data(df)
    
    # Write Excel report
    report_data = (df, summary, tabs, network_pivot)
    write_excel_report(report_data, output_filename)
    
    write_shared_log("bg_disruption.py", "Processing complete.")
    print("Processing complete")
    
    # Show completion notification
    show_completion_notification()


if __name__ == "__main__":
    process_data()
"""
Data processing module for handling CSV/Excel data operations.
Extracted from app.py to improve cohesion and reduce file size.
"""

import pandas as pd
import numpy as np
import logging
import multiprocessing
from pathlib import Path
import re
import os
import sys

from config.app_config import ProcessingConfig
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.utils import write_shared_log


class DataProcessor:
    """Handles data processing and validation operations."""
    
    def __init__(self, app_instance):
        self.app = app_instance
        
    def load_and_validate_data(self, file_path):
        """Load and validate the merged file data."""
        try:
            df = pd.read_excel(file_path)
            logging.info(f"Loaded {len(df)} records from {file_path}")
            
            # Validate required columns using configuration
            ProcessingConfig.validate_required_columns(df)
            
            # Prepare data for processing
            df = df.sort_values(by=["DATEFILLED", "SOURCERECORDID"], ascending=True)
            df["Logic"] = ""
            df["RowID"] = np.arange(len(df))
            
            return df
            
        except Exception as e:
            error_msg = f"Error loading data from {file_path}: {str(e)}"
            logging.error(error_msg)
            write_shared_log("DataProcessor", error_msg, "ERROR")
            raise
    
    def process_data_multiprocessing(self, df):
        """Process data using multiprocessing for improved performance."""
        try:
            # Import and use multiprocessing helpers
            from modules import mp_helpers
            
            num_workers = ProcessingConfig.get_multiprocessing_workers()
            df_blocks = np.array_split(df, num_workers)
            out_queue = multiprocessing.Queue()
            processes = []
            
            # Start worker processes
            for block in df_blocks:
                p = multiprocessing.Process(
                    target=mp_helpers.worker, args=(block, out_queue)
                )
                p.start()
                processes.append(p)
            
            # Collect results
            results = [out_queue.get() for _ in processes]
            for p in processes:
                p.join()
            
            processed_df = pd.concat(results)
            logging.info(f"Processed {len(processed_df)} records using {num_workers} workers")
            return processed_df
            
        except Exception as e:
            error_msg = f"Error processing data with multiprocessing: {str(e)}"
            logging.error(error_msg)
            write_shared_log("DataProcessor", error_msg, "ERROR")
            raise
    
    def save_processed_outputs(self, df, output_dir=None):
        """Save processed data to various output formats."""
        try:
            if output_dir is None:
                output_dir = Path.cwd()
            else:
                output_dir = Path(output_dir)
            
            # Sort and filter data
            df_sorted = pd.concat([df[df["Logic"] == ""], df[df["Logic"] == "OR"]])
            
            # Create row mapping for highlighting
            row_mapping = {
                row["RowID"]: i + 2 for i, (_, row) in enumerate(df_sorted.iterrows())
            }
            excel_rows_to_highlight = [
                row_mapping[rid] for rid in [] if rid in row_mapping
            ]  # Placeholder for unmatched reversals
            
            # Clean up data
            df_sorted.drop(columns=["RowID"], inplace=True, errors="ignore")
            output_file = output_dir / "merged_file_with_OR.xlsx"
            
            # Save to multiple formats
            self._save_to_parquet(df_sorted, output_dir)
            self._save_to_excel(df_sorted, output_file)
            self._save_to_csv(df_sorted, output_dir)
            self._save_unmatched_reversals(excel_rows_to_highlight, output_dir)
            
            return output_file
            
        except Exception as e:
            error_msg = f"Error saving processed outputs: {str(e)}"
            logging.error(error_msg)
            write_shared_log("DataProcessor", error_msg, "ERROR")
            raise
    
    def _save_to_parquet(self, df, output_dir):
        """Save data to Parquet format for large DataFrames."""
        try:
            parquet_path = output_dir / "merged_file_with_OR.parquet"
            df.drop_duplicates().to_parquet(parquet_path, index=False)
            logging.info(f"Saved intermediate Parquet file: {parquet_path}")
        except Exception as e:
            logging.warning(f"Could not save Parquet: {e}")
    
    def _save_to_excel(self, df, output_file):
        """Save data to Excel format."""
        df.drop_duplicates().to_excel(output_file, index=False)
        logging.info(f"Saved Excel file: {output_file}")
    
    def _save_to_csv(self, df, output_dir):
        """Save data to CSV format with opportunity name."""
        try:
            opportunity_name = self._extract_opportunity_name()
            csv_path = output_dir / f"{opportunity_name} Claim Detail.csv"
            df.drop_duplicates().to_csv(csv_path, index=False)
            logging.info(f"Saved CSV file: {csv_path}")
        except Exception as e:
            logging.warning(f"Could not save CSV: {e}")
    
    def _save_unmatched_reversals(self, excel_rows_to_highlight, output_dir):
        """Save unmatched reversals information."""
        try:
            unmatched_path = output_dir / "unmatched_reversals.txt"
            with open(unmatched_path, "w") as f:
                f.write(",".join(map(str, excel_rows_to_highlight)))
            logging.info(f"Saved unmatched reversals info: {unmatched_path}")
        except Exception as e:
            logging.warning(f"Could not save unmatched reversals: {e}")
    
    def _extract_opportunity_name(self):
        """Extract opportunity name from file1_path."""
        opportunity_name = ProcessingConfig.DEFAULT_OPPORTUNITY_NAME
        try:
            if self.app.file1_path:
                if self.app.file1_path.lower().endswith(".xlsx"):
                    df_file1 = pd.read_excel(self.app.file1_path)
                else:
                    df_file1 = pd.read_csv(self.app.file1_path)
                
                if df_file1.shape[1] >= 2:
                    # Get the value from the first row, second column
                    raw_name = str(df_file1.iloc[0, 1])
                    # Clean for filename
                    opportunity_name = re.sub(r'[\\/*?:"<>|]', "_", raw_name)
        except Exception as e:
            logging.warning(f"Could not extract opportunity name from file1: {e}")
        
        return opportunity_name
    
    def validate_merge_inputs(self, file1_path, file2_path):
        """Validate that merge inputs are valid."""
        # Check if both file paths are set and files exist
        if not file1_path or not file2_path:
            return False, "Both input files must be selected."
        
        if not os.path.isfile(file1_path):
            return False, f"File not found: {file1_path}"
        
        if not os.path.isfile(file2_path):
            return False, f"File not found: {file2_path}"
        
        return True, "Inputs are valid"
    
    def validate_gross_cost_template(self, file_path):
        """Validate GrossCost column and suggest template type."""
        try:
            df = pd.read_csv(file_path)
            
            # Guard clause: no GrossCost column
            if "GrossCost" not in df.columns:
                return None
            
            return self._determine_template_type(df["GrossCost"])
            
        except Exception as e:
            logging.warning(f"Could not validate GrossCost column: {e}")
            return None
    
    def _determine_template_type(self, gross_cost_series):
        """Determine which template type to recommend based on GrossCost data."""
        if gross_cost_series.isna().all() or (gross_cost_series == 0).all():
            return "The GrossCost column is blank or all zero. Please use the Blind template."
        else:
            return "The GrossCost column contains data. Please use the Standard template."
"""
Template processing module for handling Excel template operations.
Extracted from app.py to improve cohesion and reduce file size.

This module provides:
- Template backup creation
- Excel data formatting
- Column filtering for templates
- Data preparation for Excel export
"""

import pandas as pd
import shutil
import os
import sys
from pathlib import Path
import logging

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.utils import write_shared_log


class TemplateProcessor:
    """
    Handles Excel template operations with a focus on simplicity and reliability.
    
    This class manages template backup, data formatting, and Excel export operations
    while maintaining separation of concerns from the main application logic.
    """
    
    def __init__(self, app_instance):
        """Initialize with reference to the main application instance."""
        self.app = app_instance
        
    def create_template_backup(self, paths):
        """Create backup of template and prepare output file."""
        try:
            # Backup original template
            shutil.copy(paths["template"], paths["backup"])
            logging.info(f"Template backed up to {paths['backup']}")
            
            # Remove old output if it exists
            if paths["output"].exists():
                try:
                    os.remove(paths["output"])
                except PermissionError:
                    raise RuntimeError(
                        f"Cannot overwrite {paths['output']} — please close it in Excel."
                    )
            
            # Copy template to output location
            shutil.copy(paths["template"], paths["output"])
            write_shared_log("TemplateProcessor", f"Template backup created: {paths['backup']}")
            
        except Exception as e:
            error_msg = f"Failed to create template backup: {str(e)}"
            logging.error(error_msg)
            write_shared_log("TemplateProcessor", error_msg, "ERROR")
            raise
    
    def format_dataframe(self, df):
        """Format DataFrame for Excel export."""
        # Format datetime columns
        datetime_columns = df.select_dtypes(include=["datetime64"]).columns
        for col in datetime_columns:
            df[col] = df[col].dt.strftime("%Y-%m-%d %H:%M:%S")
        
        # Fill NaN values
        return df.fillna("")
    
    def filter_template_columns(self, df):
        """Filter columns for template pasting."""
        try:
            # Ensure 'ClientName' and 'Logic' columns exist and are in the correct order
            if "ClientName" in df.columns and "Logic" in df.columns:
                client_name_idx = df.columns.get_loc("ClientName")
                logic_idx = df.columns.get_loc("Logic")
                
                if client_name_idx <= logic_idx:
                    # Select columns from 'ClientName' to 'Logic' (inclusive)
                    selected_columns = df.columns[client_name_idx : logic_idx + 1]
                    logging.info(f"Pasting only these columns: {selected_columns.tolist()}")
                    return df[selected_columns]
                else:
                    logging.warning(
                        "'Logic' column appears before 'ClientName'; returning full DataFrame."
                    )
                    return df
            else:
                raise ValueError("Required columns 'ClientName' or 'Logic' are missing.")
                
        except Exception as e:
            logging.warning(f"Error filtering columns: {e}. Using full DataFrame.")
            return df
    
    def prepare_template_data(self, processed_file):
        """Prepare data for template pasting."""
        try:
            df = pd.read_excel(processed_file)
            df = self.format_dataframe(df)
            
            return {
                "data": df.values,
                "nrows": df.shape[0],
                "ncols": df.shape[1]
            }
        except Exception as e:
            error_msg = f"Failed to prepare template data: {str(e)}"
            logging.error(error_msg)
            write_shared_log("TemplateProcessor", error_msg, "ERROR")
            raise
    
    def prepare_excel_data(self, paste_data, formulas):
        """Prepare data for Excel, preserving formulas."""
        data_to_write = []
        
        for i in range(paste_data["nrows"]):
            row = []
            for j in range(paste_data["ncols"]):
                if formulas[i][j] == "":
                    row.append(paste_data["data"][i][j])
                else:
                    row.append(None)
            data_to_write.append(row)
        
        return data_to_write
    
    def validate_template_requirements(self, template_path):
        """Validate that template meets requirements."""
        if not template_path:
            raise ValueError("Template file path is not set.")
            
        template = Path(template_path)
        if not template.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")
            
        if template.suffix != '.xlsx':
            raise ValueError("Template must be an Excel file (.xlsx)")
            
        return True
    
    def show_toast(self, message, duration=3000):
        """Show a toast notification."""
        try:
            import tkinter as tk
            from tkinter import messagebox
            
            toast = tk.Toplevel(self.app.root)
            toast.overrideredirect(True)
            toast.configure(bg="black")

            # Position bottom-right
            self.app.root.update_idletasks()
            screen_width = toast.winfo_screenwidth()
            screen_height = toast.winfo_screenheight()
            x = screen_width - 320
            y = screen_height - 100
            toast.geometry(f"300x50+{x}+{y}")

            label = tk.Label(
                toast, text=message, bg="black", fg="white", font=("Arial", 11)
            )
            label.pack(fill="both", expand=True)

            toast.after(duration, toast.destroy)
            
        except Exception as e:
            logging.warning(f"Toast notification failed: {e}")
            # Fallback to messagebox
            from tkinter import messagebox
            messagebox.showinfo("Notification", message)
import logging
import tkinter as tk
from pathlib import Path
from tkinter import messagebox
import os
import sys

import pandas as pd
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.excel_utils import write_df_to_template

from utils.utils import load_file_paths, write_shared_log

# Setup logging
logging.basicConfig(
    filename="epls_lbl.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


def show_message(awp, ing, total, rxs):
    messagebox.showinfo(
        "Process Complete",
        f"EPLS LBL has been created\n\n"
        f"Total AWP: ${awp:.2f}\n\n"
        f"Total Ing Cost: ${ing:.2f}\n\n"
        f"Total Gross Cost: ${total:.2f}\n\n"
        f"Total Claim Count: {rxs}",
    )


def main() -> None:
    tk.Tk().withdraw()
    write_shared_log("epls_lbl.py", "Processing started.")
    try:
        # Get the config file path relative to the project root
        config_path = Path(__file__).parent.parent / "config" / "file_paths.json"
        paths = load_file_paths(str(config_path))

        # Failsafe: check that both input and template files exist
        for key in ["reprice", "epls"]:
            if not Path(paths[key]).exists():
                raise FileNotFoundError(f"{key} path not found: {paths[key]}")

        template_path = Path(paths["epls"])
        df = pd.read_excel(paths["reprice"], sheet_name="Claims Table")

        # Debug print and log
        print("Columns in claims DataFrame:")
        print(df.columns)
        logger.info(f"Columns in claims DataFrame: {df.columns.tolist()}")

        if "Logic" not in df.columns:
            raise KeyError("Missing 'Logic' column in Claims Table.")

        df["Logic"] = pd.to_numeric(df["Logic"], errors="coerce")
        df = df[df["Logic"].between(1, 10)]

        awp = df["Total AWP (Historical)"].sum()
        ing = df["Rx Sense Ing Cost"].sum()
        total = df["RxSense Total Cost"].sum()
        rxs = df["Rxs"].sum()

        columns_to_keep = [
            "MONY",
            "Rxs",
            "Rx Sense Ing Cost",
            "RxSense Dispense Fee",
            "RxSense Total Cost",
            "Total AWP (Historical)",
            "Pharmacy Name",
            "INPUTFILECHANNEL",
            "DATEFILLED",
            "MemberID",
            "DAYSUPPLY",
            "QUANTITY",
            "NDC",
            "DST Drug Name",
            "GrossCost",
            "Universal Rebates",
            "Exclusive Rebates",
            "Specialty Vlookup",
        ]

        missing_cols = [col for col in columns_to_keep if col not in df.columns]
        if missing_cols:
            raise KeyError(f"Missing expected columns: {missing_cols}")

        df = df[columns_to_keep]
        logger.info(f"Filtered DataFrame shape: {df.shape}")

        df["Specialty Vlookup"] = df["Specialty Vlookup"].map({"No": "N"}).fillna("Y")

        logger.info(f"AWP: {awp:.2f}, Ing: {ing:.2f}, Total: {total:.2f}, Rxs: {rxs}")

        output_path = Path("_Rx Claims for EPLS.xlsx")
        write_df_to_template(
            str(template_path),
            str(output_path),
            sheet_name="Line By Line",
            df=df,
            start_cell="A2",
            header=False,
            index=False,
            open_file=False,
            visible=False,
        )

        logger.info("EPLS LBL file created successfully.")
        write_shared_log("epls_lbl.py", "EPLS LBL file created successfully.")
        show_message(awp, ing, total, rxs)
        messagebox.showinfo("Processing Complete", "Processing complete")
    except Exception as e:
        logger.exception("An error occurred during EPLS LBL processing")
        write_shared_log("epls_lbl.py", f"An error occurred: {e}", status="ERROR")
        messagebox.showerror("Error", f"An error occurred: {e}")


if __name__ == "__main__":
    main()
"""
Log management module for handling various logging and viewer operations.
Extracted from app.py to reduce file size and improve organization.
"""

import tkinter as tk
from tkinter import scrolledtext
import csv
import os
import sys
import logging

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.utils import write_shared_log


class LogManager:
    """Handles log viewing and management operations."""
    
    def __init__(self, app_instance):
        self.app = app_instance
        self.shared_log_path = os.path.expandvars(
            r"%OneDrive%/True Community - Data Analyst/Python Repricing Automation Program/Logs/audit_log.csv"
        )
        
    def show_log_viewer(self):
        """Show the live log viewer window."""
        log_win = tk.Toplevel(self.app.root)
        log_win.title("Live Log Viewer")
        text_area = scrolledtext.ScrolledText(log_win, width=100, height=30)
        text_area.pack(fill="both", expand=True)

        def update_logs():
            try:
                with open("repricing_log.log", "r") as f:
                    text_area.delete(1.0, tk.END)
                    text_area.insert(tk.END, f.read())
            except FileNotFoundError:
                text_area.delete(1.0, tk.END)
                text_area.insert(tk.END, "No log file found.")
            except Exception as e:
                text_area.delete(1.0, tk.END)
                text_area.insert(tk.END, f"Error reading log file: {e}")
            log_win.after(3000, update_logs)

        update_logs()
        
    def show_shared_log_viewer(self):
        """Show the shared audit log viewer with search functionality."""
        log_win = tk.Toplevel(self.app.root)
        log_win.title("Shared Audit Log Viewer")
        log_win.geometry("1000x600")

        # Create filter frame
        filter_frame = tk.Frame(log_win)
        filter_frame.pack(fill="x")
        tk.Label(filter_frame, text="Search:").pack(side="left", padx=5)
        filter_entry = tk.Entry(filter_frame)
        filter_entry.pack(side="left", fill="x", expand=True, padx=5)

        # Create text area
        text_area = scrolledtext.ScrolledText(log_win, width=150, height=30)
        text_area.pack(fill="both", expand=True)

        def refresh():
            """Refresh the log display with optional filtering."""
            try:
                if not os.path.exists(self.shared_log_path):
                    text_area.delete(1.0, tk.END)
                    text_area.insert(tk.END, f"Shared log file not found at: {self.shared_log_path}")
                    return
                    
                with open(self.shared_log_path, "r", newline="", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    rows = list(reader)

                search_term = filter_entry.get().lower()
                if search_term:
                    filtered = [
                        row for row in rows
                        if any(search_term in str(cell).lower() for cell in row)
                    ]
                else:
                    filtered = rows

                text_area.delete(1.0, tk.END)
                for row in filtered:
                    text_area.insert(tk.END, " | ".join(row) + "\n")
                    
            except Exception as e:
                text_area.delete(1.0, tk.END)
                text_area.insert(tk.END, f"[ERROR] Could not read shared log:\n{e}")
                logging.error(f"Error reading shared log: {e}")

            # Auto-refresh every 5 seconds
            log_win.after(5000, refresh)

        # Bind search on Enter key
        filter_entry.bind('<Return>', lambda event: refresh())
        
        # Initial load
        refresh()
        
    def initialize_logging(self):
        """Initialize logging configuration."""
        # Clear existing log
        log_file = "repricing_log.log"
        try:
            open(log_file, "w").close()  # Clear the file
        except Exception as e:
            logging.warning(f"Could not clear log file: {e}")
            
        # Configure logging
        logging.basicConfig(
            filename=log_file,
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
            filemode='w'  # Overwrite mode
        )
        logging.info("Logging initialized")
        
    def log_application_start(self):
        """Log application startup."""
        logging.info("Repricing Automation application started")
        write_shared_log("LogManager", "Application started")
        
    def log_application_shutdown(self):
        """Log application shutdown."""
        logging.info("Repricing Automation application shutting down")
        write_shared_log("LogManager", "Application shutdown")


class ThemeController:
    """Controls theme switching functionality."""
    
    def __init__(self, app_instance):
        self.app = app_instance
        self.current_theme = "light"
        
    def toggle_dark_mode(self):
        """Toggle between light and dark themes."""
        import customtkinter as ctk
        from ui.ui_components import ThemeManager, LIGHT_COLORS, DARK_COLORS
        
        if self.current_theme == "light":
            # Switch to Dark mode
            ctk.set_appearance_mode("dark")
            ThemeManager.apply_theme_colors(self.app, DARK_COLORS)
            self.app.toggle_theme_button.configure(text="Switch to Light Mode")
            self.current_theme = "dark"
        else:
            # Switch to Light mode
            ctk.set_appearance_mode("light")
            ThemeManager.apply_theme_colors(self.app, LIGHT_COLORS)
            self.app.toggle_theme_button.configure(text="Switch to Dark Mode")
            self.current_theme = "light"
            
        logging.info(f"Theme switched to {self.current_theme} mode")
        write_shared_log("ThemeController", f"Theme changed to {self.current_theme}")
        
    def apply_initial_theme(self):
        """Apply the initial light theme."""
        from ui.ui_components import ThemeManager, LIGHT_COLORS
        ThemeManager.apply_theme_colors(self.app, LIGHT_COLORS)
import sys
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
import logging
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.utils import write_shared_log

# Configure logging
logging.basicConfig(
    filename="merge_log.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

MERGED_FILENAME = "merged_file.xlsx"
REQUIRED_COLUMNS = [
    "DATEFILLED",
    "SOURCERECORDID",
    "QUANTITY",
    "DAYSUPPLY",
    "NDC",
    "MemberID",
    "Drug Name",
    "Pharmacy Name",
    "Total AWP (Historical)",
]


def merge_files(file1_path, file2_path):
    file1 = Path(file1_path)
    file2 = Path(file2_path)
    try:
        logger.info(f"Starting merge: {file1} + {file2}")
        write_shared_log("merge.py", f"Starting merge: {file1} + {file2}")

        if not file1.exists():
            logger.error(f"File not found: {file1}")
            write_shared_log("merge.py", f"File not found: {file1}", status="ERROR")
            return False
        if not file2.exists():
            logger.error(f"File not found: {file2}")
            write_shared_log("merge.py", f"File not found: {file2}", status="ERROR")
            return False

        # Load data (support Excel or CSV for both files)
        try:
            if file1.suffix == ".csv":
                df1 = pd.read_csv(file1, parse_dates=["DATEFILLED"], dayfirst=False)
            else:
                df1 = pd.read_excel(file1, parse_dates=["DATEFILLED"])
        except Exception as e:
            logger.error(f"Failed to load file1: {e}")
            write_shared_log("merge.py", f"Failed to load file1: {e}", status="ERROR")
            return False
        try:
            if file2.suffix == ".csv":
                df2 = pd.read_csv(file2)
            else:
                df2 = pd.read_excel(file2)
        except Exception as e:
            logger.error(f"Failed to load file2: {e}")
            write_shared_log("merge.py", f"Failed to load file2: {e}", status="ERROR")
            return False

        # Log data source details
        logger.info(f"df1 shape: {df1.shape}, df2 shape: {df2.shape}")
        write_shared_log("merge.py", f"df1 shape: {df1.shape}, df2 shape: {df2.shape}")
        logger.info(f"df1 columns: {list(df1.columns)}")
        logger.info(f"df2 columns: {list(df2.columns)}")

        # Clean up and standardize column names
        df2.columns = [col.strip() for col in df2.columns]
        if "Source Record ID" in df2.columns:
            df2.rename(columns={"Source Record ID": "SOURCERECORDID"}, inplace=True)

        # Merge
        try:
            df_merged = pd.merge(df1, df2, on="SOURCERECORDID", how="outer")
        except Exception as e:
            logger.error(f"Failed to merge: {e}")
            write_shared_log("merge.py", f"Failed to merge: {e}", status="ERROR")
            return False
        if "Total AWP (Historical)" in df_merged.columns:
            df_merged["Total AWP (Historical)"] = pd.to_numeric(
                df_merged["Total AWP (Historical)"], errors="coerce"
            ).round(2)
        else:
            df_merged["Total AWP (Historical)"] = 0.00
        df_merged["MemberID"] = df_merged["MemberID"].fillna(0)

        # Log merged file row count and checksum
        row_count = len(df_merged)
        try:
            import hashlib

            sample = df_merged.head(1000).to_csv(index=False).encode("utf-8")
            checksum = hashlib.md5(sample).hexdigest()
        except Exception as e:
            checksum = f"ERROR: {e}"
        logger.info(f"Merged row count: {row_count}, sample checksum: {checksum}")
        write_shared_log(
            "merge.py", f"Merged row count: {row_count}, sample checksum: {checksum}"
        )

        # Log missing required columns
        for col in REQUIRED_COLUMNS:
            if col not in df_merged.columns:
                logger.warning(f"Missing expected column: {col}")
                write_shared_log(
                    "merge.py", f"Missing expected column: {col}", status="WARNING"
                )

        # Drop the DATEFILLED_DIFF column if it exists before saving the merged file
        if "DATEFILLED_DIFF" in df_merged.columns:
            df_merged.drop(columns=["DATEFILLED_DIFF"], inplace=True)

        merged_path = Path.cwd() / MERGED_FILENAME
        try:
            df_merged.to_excel(merged_path, index=False)
        except Exception as e:
            logger.error(f"Failed to write merged Excel: {e}")
            write_shared_log(
                "merge.py", f"Failed to write merged Excel: {e}", status="ERROR"
            )
            return False
        logger.info(f"Merged file saved to: {merged_path}")
        write_shared_log("merge.py", f"Merged file saved to: {merged_path}")

        # Apply Excel formatting
        try:
            wb = load_workbook(merged_path)
            ws = wb.active
            date_style = NamedStyle(name="date_style", number_format="MM/DD/YYYY")

            if ws is not None and ws.max_row >= 1:
                header = [
                    cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))
                ]
                if "DATEFILLED" in header:
                    date_col_index = header.index("DATEFILLED") + 1
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=date_col_index).style = date_style
                    wb.save(merged_path)
                    logger.info("Applied date formatting successfully.")
                    write_shared_log(
                        "merge.py", "Applied date formatting successfully."
                    )
                else:
                    logger.warning("DATEFILLED column not found for formatting.")
                    write_shared_log(
                        "merge.py",
                        "DATEFILLED column not found for formatting.",
                        status="WARNING",
                    )
            else:
                logger.warning(
                    "Worksheet is empty or not loaded, cannot apply formatting."
                )
                write_shared_log(
                    "merge.py",
                    "Worksheet is empty or not loaded, cannot apply formatting.",
                    status="WARNING",
                )

        except Exception as ex:
            logger.warning(f"Failed to apply formatting: {ex}")
            write_shared_log(
                "merge.py", f"Failed to apply formatting: {ex}", status="WARNING"
            )

        return True
    except Exception as e:
        logger.exception(f"Merge failed: {e}")
        write_shared_log("merge.py", f"Merge failed: {e}", status="ERROR")
        return False


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python merge.py <file1_path> <file2_path>")
        sys.exit(1)

    merge_files(sys.argv[1], sys.argv[2])
import pandas as pd
import numpy as np


def process_logic_block(df_block):
    """
    Vectorized numpy logic to mark 'OR' in 'Logic' for reversals with matching claims.
    Refactored to reduce nesting complexity and improve readability.
    """
    arr = df_block.to_numpy()
    col_idx = {col: i for i, col in enumerate(df_block.columns)}
    
    # Extract and prepare data
    logic_data = _extract_logic_data(arr, col_idx)
    
    # Early return if no reversals to process
    if not np.any(logic_data["is_reversal"]):
        return pd.DataFrame(arr, columns=df_block.columns)
    
    # Process reversals with reduced nesting
    _process_reversals(arr, col_idx, logic_data)
    
    return pd.DataFrame(arr, columns=df_block.columns)


def _extract_logic_data(arr, col_idx):
    """Extract and prepare data for logic processing."""
    qty = arr[:, col_idx["QUANTITY"]].astype(float)
    return {
        "qty": qty,
        "is_reversal": qty < 0,
        "is_claim": qty > 0,
        "ndc": arr[:, col_idx["NDC"]].astype(str),
        "member": arr[:, col_idx["MemberID"]].astype(str),
        "datefilled": pd.to_datetime(arr[:, col_idx["DATEFILLED"]], errors="coerce"),
        "abs_qty": np.abs(qty)
    }


def _process_reversals(arr, col_idx, logic_data):
    """Process reversals with matching logic, using guard clauses to reduce nesting."""
    rev_idx = np.where(logic_data["is_reversal"])[0]
    claim_idx = (
        np.where(logic_data["is_claim"])[0] 
        if np.any(logic_data["is_claim"]) 
        else np.array([], dtype=int)
    )
    
    match_context = {
        "arr": arr,
        "col_idx": col_idx,
        "logic_data": logic_data,
        "claim_idx": claim_idx
    }
    
    for i in rev_idx:
        found_match = _try_find_match(match_context, i)
        
        # Mark unmatched reversals as 'OR'
        if not found_match:
            arr[i, col_idx["Logic"]] = "OR"


def _try_find_match(match_context, reversal_idx):
    """Attempt to find a matching claim for a reversal. Returns True if match found."""
    arr = match_context["arr"]
    col_idx = match_context["col_idx"]
    logic_data = match_context["logic_data"]
    claim_idx = match_context["claim_idx"]
    
    # Guard clause: no claims to match against
    if claim_idx.size == 0:
        return False
    
    # Find potential matches
    matches = _find_matching_claims(logic_data, claim_idx, reversal_idx)
    
    # Guard clause: no matches found
    if not np.any(matches):
        return False
    
    # Mark both reversal and matching claim as 'OR'
    arr[reversal_idx, col_idx["Logic"]] = "OR"
    arr[claim_idx[matches][0], col_idx["Logic"]] = "OR"
    return True


def _find_matching_claims(logic_data, claim_idx, reversal_idx):
    """Find claims that match the reversal based on NDC, member, quantity, and date."""
    matches = (
        (logic_data["ndc"][claim_idx] == logic_data["ndc"][reversal_idx])
        & (logic_data["member"][claim_idx] == logic_data["member"][reversal_idx])
        & (logic_data["abs_qty"][claim_idx] == logic_data["abs_qty"][reversal_idx])
    )
    
    # Add date constraint (within 30 days)
    date_diffs = np.abs(
        (logic_data["datefilled"][claim_idx] - logic_data["datefilled"][reversal_idx]).days
    )
    matches &= date_diffs <= 30
    
    return matches


def worker(df_block, out_queue):
    """Worker function for multiprocessing."""
    result = process_logic_block(df_block)
    out_queue.put(result)
import pandas as pd
import logging
import os
import sys
from pathlib import Path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.utils import (
    load_file_paths,
    standardize_pharmacy_ids,
    standardize_network_ids,
    merge_with_network,
    drop_duplicates_df,
    clean_logic_and_tier,
    filter_recent_date,
    filter_logic_and_maintenance,
    filter_products_and_alternative,
    write_shared_log,
)

# Setup logging
logging.basicConfig(
    filename="openmdf_bg.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# Load NABP/NPI list
included_nabp_npi = {
    "4528874": "1477571404",
    "2365422": "1659313435",
    "3974157": "1972560688",
    "320793": "1164437406",
    "4591055": "1851463087",
    "2348046": "1942303110",
    "4023610": "1407879588",
    "4025385": "1588706212",
    "4025311": "1588705446",
    "4026806": "1285860312",
    "4931350": "1750330775",
    "4024585": "1396768461",
    "4028026": "1497022438",
    "2643749": "1326490376",
}


def process_data():
    write_shared_log("openmdf_bg.py", "Processing started.")

    import sys

    # Get the config file path relative to the project root
    config_path = Path(__file__).parent.parent / "config" / "file_paths.json"
    paths = load_file_paths(str(config_path))

    if "reprice" not in paths or not paths["reprice"]:
        logger.warning("No reprice/template file provided.")
        write_shared_log(
            "openmdf_bg.py", "No reprice/template file provided.", status="ERROR"
        )
        print("No reprice/template file provided.")
        return False

    # Check for required sheet name in reprice file
    try:
        xl = pd.ExcelFile(paths["reprice"])
        if "Claims Table" not in xl.sheet_names:
            logger.error(
                f"Sheet 'Claims Table' not found in {paths['reprice']}. Sheets: {xl.sheet_names}"
            )
            write_shared_log(
                "openmdf_bg.py",
                f"Sheet 'Claims Table' not found in {paths['reprice']}. Sheets: {xl.sheet_names}",
                status="ERROR",
            )
            return False
        claims = xl.parse(
            "Claims Table",
            usecols=[
                "SOURCERECORDID",
                "NDC",
                "MemberID",
                "DATEFILLED",
                "FormularyTier",
                "Rxs",
                "Logic",
                "PHARMACYNPI",
                "NABP",
                "Pharmacy Name",
                "Universal Rebates",
                "Exclusive Rebates",
            ],
        )
    except Exception as e:
        logger.error(f"Failed to read Claims Table: {e}")
        write_shared_log(
            "openmdf_bg.py", f"Failed to read Claims Table: {e}", status="ERROR"
        )
        return False

    # Log claim count before any filtering
    logger.info(f"Initial claims count: {claims.shape[0]}")
    write_shared_log("openmdf_bg.py", f"Initial claims count: {claims.shape[0]}")

    try:
        medi = pd.read_excel(paths["medi_span"])[["NDC", "Maint Drug?", "Product Name"]]
    except Exception as e:
        logger.error(f"Failed to read medi_span file: {paths['medi_span']} | {e}")
        write_shared_log(
            "openmdf_bg.py",
            f"Failed to read medi_span file: {paths['medi_span']} | {e}",
            status="ERROR",
        )
        return False
    try:
        mdf = pd.read_excel(paths["mdf_disrupt"], sheet_name="Open MDF NDC")[
            ["NDC", "Tier"]
        ]
    except Exception as e:
        logger.error(f"Failed to read mdf_disrupt file: {paths['mdf_disrupt']} | {e}")
        write_shared_log(
            "openmdf_bg.py",
            f"Failed to read mdf_disrupt file: {paths['mdf_disrupt']} | {e}",
            status="ERROR",
        )
        return False
    try:
        network = pd.read_excel(paths["n_disrupt"])[
            ["pharmacy_npi", "pharmacy_nabp", "pharmacy_is_excluded"]
        ]
    except Exception as e:
        logger.error(f"Failed to read n_disrupt file: {paths['n_disrupt']} | {e}")
        write_shared_log(
            "openmdf_bg.py",
            f"Failed to read n_disrupt file: {paths['n_disrupt']} | {e}",
            status="ERROR",
        )
        return False
        # Read Alternatives NDC for 'Alternative' column
    try:
        exclusive = pd.read_excel(paths["e_disrupt"], sheet_name="Alternatives NDC")[
            ["NDC", "Tier", "Alternative"]
        ]
    except Exception as e:
        logger.error(f"Failed to read e_disrupt file: {paths['e_disrupt']} | {e}")
        write_shared_log(
            "openmdf_bg.py",
            f"Failed to read e_disrupt file: {paths['e_disrupt']} | {e}",
            status="ERROR",
        )
        return False

    df = claims.merge(medi, on="NDC", how="left")
    print(f"After merge with medi: {df.shape}")
    logger.info(f"After merge with medi: {df.shape}")
    df = df.merge(mdf, on="NDC", how="left")
    print(f"After merge with mdf: {df.shape}")
    logger.info(f"After merge with mdf: {df.shape}")
    # Merge in Alternatives for 'Alternative' column
    df = df.merge(
        exclusive.rename(columns={"Tier": "Exclusive Tier"}), on="NDC", how="left"
    )
    print(f"After merge with exclusive: {df.shape}")
    logger.info(f"After merge with exclusive: {df.shape}")
    df = standardize_pharmacy_ids(df)
    print(f"After standardize_pharmacy_ids: {df.shape}")
    logger.info(f"After standardize_pharmacy_ids: {df.shape}")
    network = standardize_network_ids(network)
    print(f"After standardize_network_ids: {network.shape}")
    logger.info(f"After standardize_network_ids: {network.shape}")

    # Ensure pharmacy_id exists
    if "pharmacy_id" not in df.columns:
        df["pharmacy_id"] = df.apply(
            lambda row: row["PHARMACYNPI"]
            if pd.notna(row["PHARMACYNPI"])
            else row["NABP"],
            axis=1,
        )

    logger.info(f"Columns in df before merging: {df.columns.tolist()}")
    print(f"Columns in df before merging: {df.columns.tolist()}")

    # Log claim count after merge
    print(f"Claims after merge: {df.shape}")
    logger.info(f"Claims after merge: {df.shape}")
    write_shared_log("openmdf_bg.py", f"Claims after merge: {df.shape[0]}")

    df = merge_with_network(df, network)
    print(f"After merge_with_network: {df.shape}")
    logger.info(f"After merge_with_network: {df.shape}")
    write_shared_log("openmdf_bg.py", f"Claims after merge_with_network: {df.shape[0]}")

    df = drop_duplicates_df(df)
    print(f"After drop_duplicates_df: {df.shape}")
    logger.info(f"After drop_duplicates_df: {df.shape}")
    write_shared_log("openmdf_bg.py", f"Claims after drop_duplicates_df: {df.shape[0]}")

    df = clean_logic_and_tier(df)
    print(f"After clean_logic_and_tier: {df.shape}")
    logger.info(f"After clean_logic_and_tier: {df.shape}")
    write_shared_log(
        "openmdf_bg.py", f"Claims after clean_logic_and_tier: {df.shape[0]}"
    )

    df = filter_products_and_alternative(df)
    print(f"After filter_products_and_alternative: {df.shape}")
    logger.info(f"After filter_products_and_alternative: {df.shape}")
    write_shared_log(
        "openmdf_bg.py", f"Claims after filter_products_and_alternative: {df.shape[0]}"
    )

    df["DATEFILLED"] = pd.to_datetime(df["DATEFILLED"], errors="coerce")
    print(f"After DATEFILLED to_datetime: {df.shape}")
    df = filter_recent_date(df)
    print(f"After filter_recent_date: {df.shape}")
    logger.info(f"After filter_recent_date: {df.shape}")
    write_shared_log("openmdf_bg.py", f"Claims after filter_recent_date: {df.shape[0]}")

    df["Logic"] = pd.to_numeric(df["Logic"], errors="coerce")
    df = filter_logic_and_maintenance(df)
    print(f"After filter_logic_and_maintenance: {df.shape}")
    logger.info(f"After filter_logic_and_maintenance: {df.shape}")
    write_shared_log(
        "openmdf_bg.py", f"Claims after filter_logic_and_maintenance: {df.shape[0]}"
    )

    df = df[
        ~df["Product Name"].str.contains(
            r"albuterol|ventolin|epinephrine", case=False, regex=True
        )
    ]
    print(f"After final product exclusion: {df.shape}")
    logger.info(f"After final product exclusion: {df.shape}")
    write_shared_log(
        "openmdf_bg.py", f"Claims after final product exclusion: {df.shape[0]}"
    )

    df["FormularyTier"] = df["FormularyTier"].astype(str).str.strip().str.upper()
    total_members = df["MemberID"].nunique()
    total_claims = df["Rxs"].sum()

    uni_pos = df[(df["Tier"] == 1) & (df["FormularyTier"].isin(["B", "BRAND"]))]
    uni_neg = df[
        (df["Tier"].isin([2, 3])) & (df["FormularyTier"].isin(["G", "GENERIC"]))
    ]

    def pivot_and_count(data):
        if data.empty:
            return pd.DataFrame([[0] * len(df.columns)], columns=df.columns), 0
        return data, data["MemberID"].nunique()

    uni_pos, uni_pos_members = pivot_and_count(uni_pos)
    uni_neg, uni_neg_members = pivot_and_count(uni_neg)

    # Output filename from CLI arg or default
    import re

    output_filename = "LBL for Disruption.xlsx"
    output_path = output_filename  # Default assignment
    for i, arg in enumerate(sys.argv):
        if arg in ("--output", "-o") and i + 1 < len(sys.argv):
            output_filename = sys.argv[i + 1]
            output_path = output_filename

    # Write LBL output unconditionally (no --output-lbl flag required)
    writer = pd.ExcelWriter(output_path, engine="xlsxwriter")

    # Ensure 'pharmacy_is_excluded' column contains actual boolean values with type inference
    if "pharmacy_is_excluded" in df.columns:
        df["pharmacy_is_excluded"] = (
            df["pharmacy_is_excluded"]
            .astype(str)
            .str.lower()
            .map({"true": True, "false": False})
            .fillna(False)
            .infer_objects(copy=False)
        )
        logger.info(
            f"pharmacy_is_excluded value counts: {df['pharmacy_is_excluded'].value_counts().to_dict()}"
        )
        # Identify rows where pharmacy_is_excluded is "NA"
        na_pharmacies = df[df["pharmacy_is_excluded"].isna()]
        logger.info(f"NA pharmacies count: {na_pharmacies.shape[0]}")

        if not na_pharmacies.empty:
            output_file_path = paths["pharmacy_validation"]
            na_pharmacies_output = na_pharmacies[["PHARMACYNPI", "NABP"]].fillna("N/A")
            
            # Update existing template or create new file
            try:
                from openpyxl import load_workbook
                import os
                
                # Check if template exists
                if os.path.exists(output_file_path):
                    # Load existing workbook
                    wb = load_workbook(output_file_path)
                    logger.info(f"Loading existing template: {output_file_path}")
                else:
                    # Create new workbook
                    from openpyxl import Workbook
                    wb = Workbook()
                    logger.info(f"Creating new validation file: {output_file_path}")
                
                # Check if Validations sheet exists, if not create it
                if "Validations" in wb.sheetnames:
                    ws = wb["Validations"]
                    # Clear existing data
                    ws.delete_rows(1, ws.max_row)
                else:
                    ws = wb.create_sheet("Validations")
                    # Remove default sheet if it exists and is empty
                    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                        wb.remove(wb["Sheet"])
                
                # Write headers
                ws.append(["PHARMACYNPI", "NABP"])
                
                # Write data
                for _, row in na_pharmacies_output.iterrows():
                    ws.append([row["PHARMACYNPI"], row["NABP"]])
                
                # Save the workbook
                wb.save(output_file_path)
                logger.info(f"NA pharmacies written to '{output_file_path}' Validations sheet.")
                
            except Exception as e:
                logger.error(f"Error updating pharmacy validation file: {e}")
                # Fallback to original method
                writer = pd.ExcelWriter(output_file_path, engine="openpyxl")
                na_pharmacies_output.to_excel(
                    writer, sheet_name="Validations", index=False, engine="openpyxl"
                )
                writer.close()
                logger.info(f"NA pharmacies written to '{output_file_path}' sheet (fallback).")

    summary = pd.DataFrame(
        {
            "Formulary": ["Open MDF Positive", "Open MDF Negative"],
            "Utilizers": [uni_pos_members, uni_neg_members],
            "Rxs": [uni_pos["Rxs"].sum(), uni_neg["Rxs"].sum()],
            "% of claims": [
                uni_pos["Rxs"].sum() / total_claims,
                uni_neg["Rxs"].sum() / total_claims,
            ],
            "": ["", ""],
            "Totals": [f"Members: {total_members}", f"Claims: {total_claims}"],
        }
    )
    summary.to_excel(writer, sheet_name="Summary", index=False)

    pt_pos = pd.pivot_table(
        uni_pos,
        values=["Rxs", "MemberID"],
        index="Product Name",
        aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
    )
    pt_pos = pt_pos.rename(columns={"Rxs": "Total Rxs", "MemberID": "Unique Members"})
    pt_pos.to_excel(writer, sheet_name="OpenMDF_Positive")

    pt_neg = pd.pivot_table(
        uni_neg,
        values=["Rxs", "MemberID"],
        index="Product Name",
        aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
    )
    pt_neg = pt_neg.rename(columns={"Rxs": "Total Rxs", "MemberID": "Unique Members"})
    pt_neg.to_excel(writer, sheet_name="OpenMDF_Negative")

    writer.sheets["OpenMDF_Positive"].write("F1", f"Total Members: {uni_pos_members}")
    writer.sheets["OpenMDF_Negative"].write("F1", f"Total Members: {uni_neg_members}")

    # Network summary for excluded pharmacies (pharmacy_is_excluded=True)
    network_df = df[df["pharmacy_is_excluded"]]
    logger.debug(f"Initial network_df shape: {network_df.shape}")
    logger.debug(f"Initial network_df contents: {network_df.head(10).to_dict()}")

    filter_phrases = [
        "CVS",
        "Walgreens",
        "Kroger",
        "Walmart",
        "Rite Aid",
        "Optum",
        "Express Scripts",
        "DMR",
        "Williams Bro",
        "Publix",
    ]
    filter_phrases_escaped = [re.escape(phrase.lower()) for phrase in filter_phrases]
    regex_pattern = "|".join([f"\\b{p}\\b" for p in filter_phrases_escaped])
    network_df = network_df[
        ~network_df["Pharmacy Name"]
        .str.lower()
        .str.contains(regex_pattern, case=False, regex=True, na=False)
    ]
    logger.debug(f"Filtered network_df shape: {network_df.shape}")
    logger.debug(f"Filtered network_df contents: {network_df.head(10).to_dict()}")

    if {"PHARMACYNPI", "NABP", "Pharmacy Name"}.issubset(network_df.columns):
        network_df["PHARMACYNPI"] = network_df["PHARMACYNPI"].fillna("N/A")
        network_df["NABP"] = network_df["NABP"].fillna("NABP")
        # Create Unique Identifier based on whichever identifier is available (PHARMACYNPI or NABP)
        network_df["Unique Identifier"] = network_df.apply(
            lambda row: row["PHARMACYNPI"]
            if row["PHARMACYNPI"] != "N/A"
            else row["NABP"],
            axis=1,
        )
        network_pivot = pd.pivot_table(
            network_df,
            values=["Rxs", "MemberID"],
            index=["Unique Identifier"],
            aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
        )
        network_pivot = network_pivot.rename(
            columns={"Rxs": "Total Rxs", "MemberID": "Unique Members"}
        )
        network_pivot.reset_index(
            inplace=True
        )  # Ensure index columns are included in the output
        logger.debug(f"Network pivot shape: {network_pivot.shape}")
        logger.debug(f"Network pivot contents: {network_pivot.head(10).to_dict()}")
        network_pivot.to_excel(writer, sheet_name="Network", index=False)

    # Write the filtered network_df directly to the 'Network Sheet' with selected columns
    selected_columns = [
        "PHARMACYNPI",
        "NABP",
        "MemberID",
        "Pharmacy Name",
        "pharmacy_is_excluded",
        "Unique Identifier",
    ]
    network_df[selected_columns].to_excel(writer, sheet_name="Network", index=False)
    logger.info(
        f"Network sheet updated with {network_df.shape[0]} excluded pharmacy records (minus major chains) and selected columns"
    )

    logger.info(f"Total pharmacies in dataset: {df.shape[0]}")
    logger.info(
        f"Excluded pharmacies (pharmacy_is_excluded=True): {df['pharmacy_is_excluded'].sum()}"
    )
    logger.info(
        f"Non-excluded pharmacies (pharmacy_is_excluded=False): {(~df['pharmacy_is_excluded']).sum()}"
    )
    logger.info(
        f"Network sheet will show {network_df.shape[0]} excluded pharmacy records (minus major chains)"
    )

    # Reorder sheets so Summary follows Data
    sheets = writer.sheets  # This is a dict: {sheet_name: worksheet_object}
    names = list(sheets.keys())
    if "Data" in names and "Summary" in names:
        data_idx = names.index("Data")
        summary_idx = names.index("Summary")
        if summary_idx != data_idx + 1:
            # Move "Summary" sheet immediately after "Data"
            items = list(sheets.items())
            summary_item = items.pop(summary_idx)
            items.insert(data_idx + 1, summary_item)
            writer.sheets.clear()
            writer.sheets.update(dict(items))

    writer.close()
    logger.info(f"Open MDF BG processing completed. Output file: {output_path}")
    write_shared_log("openmdf_bg.py", "Processing complete.")
    try:
        import tkinter as tk
        from tkinter import messagebox

        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo("Processing Complete", "Processing complete")
        root.destroy()
    except Exception:
        pass
    return True


if __name__ == "__main__":
    process_data()
import pandas as pd
import logging
import os
import sys
from pathlib import Path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.utils import (
    load_file_paths,
    standardize_pharmacy_ids,
    standardize_network_ids,
    merge_with_network,
    drop_duplicates_df,
    clean_logic_and_tier,
    filter_recent_date,
    filter_logic_and_maintenance,
    filter_products_and_alternative,
    write_shared_log,
)

# Logging setup
logging.basicConfig(
    filename="openmdf_tier.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# Load NABP/NPI list
included_nabp_npi = {
    "4528874": "1477571404",
    "2365422": "1659313435",
    "3974157": "1972560688",
    "320793": "1164437406",
    "4591055": "1851463087",
    "2348046": "1942303110",
    "4023610": "1407879588",
    "4025385": "1588706212",
    "4025311": "1588705446",
    "4026806": "1285860312",
    "4931350": "1750330775",
    "4024585": "1396768461",
    "4028026": "1497022438",
    "2643749": "1326490376",
}


# ---------------------------------------------------------------------------
# Open MDF Tier processing functions
# ---------------------------------------------------------------------------
def load_openmdf_tier_data(file_paths):
    """Load all required data files for Open MDF tier disruption processing."""
    # Load claims with fallback
    if not file_paths.get("reprice"):
        write_shared_log(
            "openmdf_tier.py",
            "No reprice/template file provided.",
            status="ERROR",
        )
        print("No reprice/template file provided. Skipping claims loading.")
        return None

    try:
        claims = pd.read_excel(file_paths["reprice"], sheet_name="Claims Table")
    except Exception as e:
        logger.error(f"Error loading claims: {e}")
        claims = pd.read_excel(file_paths["reprice"], sheet_name=0)

    print(f"claims shape: {claims.shape}")
    claims.info()

    # Load reference tables
    try:
        medi = pd.read_excel(file_paths["medi_span"])[["NDC", "Maint Drug?", "Product Name"]]
        print(f"medi shape: {medi.shape}")
    except Exception as e:
        logger.error(f"Failed to read medi_span file: {file_paths['medi_span']} | {e}")
        write_shared_log(
            "openmdf_tier.py",
            f"Failed to read medi_span file: {file_paths['medi_span']} | {e}",
            status="ERROR",
        )
        return None

    try:
        mdf = pd.read_excel(file_paths["mdf_disrupt"], sheet_name="Open MDF NDC")[
            ["NDC", "Tier"]
        ]
        print(f"mdf shape: {mdf.shape}")
    except Exception as e:
        logger.error(f"Failed to read mdf_disrupt file: {file_paths['mdf_disrupt']} | {e}")
        write_shared_log(
            "openmdf_tier.py",
            f"Failed to read mdf_disrupt file: {file_paths['mdf_disrupt']} | {e}",
            status="ERROR",
        )
        return None

    try:
        exclusive = pd.read_excel(file_paths["e_disrupt"], sheet_name="Alternatives NDC")[
            ["NDC", "Tier", "Alternative"]
        ]
        print(f"exclusive shape: {exclusive.shape}")
    except Exception as e:
        logger.error(f"Failed to read e_disrupt file: {file_paths['e_disrupt']} | {e}")
        write_shared_log(
            "openmdf_tier.py",
            f"Failed to read e_disrupt file: {file_paths['e_disrupt']} | {e}",
            status="ERROR",
        )
        return None

    try:
        network = pd.read_excel(file_paths["n_disrupt"])[
            ["pharmacy_npi", "pharmacy_nabp", "pharmacy_is_excluded"]
        ]
        print(f"network shape: {network.shape}")
    except Exception as e:
        logger.error(f"Failed to read n_disrupt file: {file_paths['n_disrupt']} | {e}")
        write_shared_log(
            "openmdf_tier.py",
            f"Failed to read n_disrupt file: {file_paths['n_disrupt']} | {e}",
            status="ERROR",
        )
        return None

    return claims, medi, mdf, exclusive, network


def process_openmdf_data_pipeline(claims, reference_data, network):
    """Process the data pipeline for Open MDF tier disruption."""
    medi, mdf, exclusive = reference_data
    
    # Merge reference data
    df = claims.merge(medi, on="NDC", how="left")
    print(f"After merge with medi: {df.shape}")

    df = df.merge(mdf.rename(columns={"Tier": "Open MDF Tier"}), on="NDC", how="left")
    print(f"After merge with mdf: {df.shape}")

    df = df.merge(
        exclusive.rename(columns={"Tier": "Exclusive Tier"}), on="NDC", how="left"
    )
    print(f"After merge with exclusive: {df.shape}")

    # Standardize IDs and perform network merge
    df = standardize_pharmacy_ids(df)
    print(f"After standardize_pharmacy_ids: {df.shape}")
    network = standardize_network_ids(network)
    print(f"After standardize_network_ids: {network.shape}")
    df = merge_with_network(df, network)
    print(f"After merge_with_network: {df.shape}")

    if "pharmacy_id" not in df.columns:
        df["pharmacy_id"] = df.apply(
            lambda row: row["PHARMACYNPI"]
            if pd.notna(row["PHARMACYNPI"])
            else row["NABP"],
            axis=1,
        )

    print("Columns in df before further processing:")
    print(df.columns)

    # Date parsing, deduplication, type cleaning, and filters
    df["DATEFILLED"] = pd.to_datetime(df["DATEFILLED"], errors="coerce")
    print(f"After DATEFILLED to_datetime: {df.shape}")

    df = drop_duplicates_df(df)
    print(f"After drop_duplicates_df: {df.shape}")

    df = clean_logic_and_tier(df)
    print(f"After clean_logic_and_tier: {df.shape}")

    df = filter_recent_date(df)
    print(f"After filter_recent_date: {df.shape}")

    df = filter_logic_and_maintenance(df)
    print(f"After filter_logic_and_maintenance: {df.shape}")

    df = filter_products_and_alternative(df)
    print(f"After filter_products_and_alternative: {df.shape}")

    return df


def handle_openmdf_pharmacy_exclusions(df, file_paths):
    """Handle pharmacy exclusions for Open MDF tier disruption."""
    # Ensure 'pharmacy_is_excluded' column contains actual boolean values with type inference
    if "pharmacy_is_excluded" in df.columns:
        df["pharmacy_is_excluded"] = (
            df["pharmacy_is_excluded"]
            .astype(str)
            .str.lower()
            .map({"true": True, "false": False})
            .fillna(False)
            .infer_objects(copy=False)
        )
        logger.info(
            f"pharmacy_is_excluded value counts: {df['pharmacy_is_excluded'].value_counts().to_dict()}"
        )

        # Identify rows where pharmacy_is_excluded is "NA"
        na_pharmacies = df[df["pharmacy_is_excluded"].isna()]
        logger.info(f"NA pharmacies count: {na_pharmacies.shape[0]}")

        if not na_pharmacies.empty:
            output_file_path = file_paths["pharmacy_validation"]
            na_pharmacies_output = na_pharmacies[["PHARMACYNPI", "NABP"]].fillna("N/A")
            
            # Update existing template or create new file
            try:
                from openpyxl import load_workbook
                import os
                
                # Check if template exists
                if os.path.exists(output_file_path):
                    # Load existing workbook
                    wb = load_workbook(output_file_path)
                    logger.info(f"Loading existing template: {output_file_path}")
                else:
                    # Create new workbook
                    from openpyxl import Workbook
                    wb = Workbook()
                    logger.info(f"Creating new validation file: {output_file_path}")
                
                # Check if Validations sheet exists, if not create it
                if "Validations" in wb.sheetnames:
                    ws = wb["Validations"]
                    # Clear existing data
                    ws.delete_rows(1, ws.max_row)
                else:
                    ws = wb.create_sheet("Validations")
                    # Remove default sheet if it exists and is empty
                    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                        wb.remove(wb["Sheet"])
                
                # Write headers
                ws.append(["PHARMACYNPI", "NABP"])
                
                # Write data
                for _, row in na_pharmacies_output.iterrows():
                    ws.append([row["PHARMACYNPI"], row["NABP"]])
                
                # Save the workbook
                wb.save(output_file_path)
                logger.info(f"NA pharmacies written to '{output_file_path}' Validations sheet.")
                
            except Exception as e:
                logger.error(f"Error updating pharmacy validation file: {e}")
                # Fallback to original method
                writer = pd.ExcelWriter(output_file_path, engine="openpyxl")
                na_pharmacies_output.to_excel(
                    writer, sheet_name="Validations", index=False, engine="openpyxl"
                )
                writer.close()
                logger.info(f"NA pharmacies written to '{output_file_path}' sheet (fallback).")

    return df


def create_openmdf_tier_definitions():
    """Create the Open MDF tier definitions for analysis."""
    return [
        ("OpenMDF_Positive 2-1", "Open MDF Tier", 1, 2),
        ("OpenMDF_Positive 3-1", "Open MDF Tier", 1, 3),
        ("OpenMDF_Positive 3-2", "Open MDF Tier", 2, 3),
        ("OpenMDF_Negative 1-2", "Open MDF Tier", 2, 1),
        ("OpenMDF_Negative 1-3", "Open MDF Tier", 3, 1),
        ("OpenMDF_Negative 2-3", "Open MDF Tier", 3, 2),
    ]


def summarize_by_openmdf_tier(df, col, from_val, to_val):
    """Summarize Open MDF tier data."""
    filtered = df[(df[col] == from_val) & (df["FormularyTier"] == to_val)]
    pt = pd.pivot_table(
        filtered,
        values=["Rxs", "MemberID"],
        index=["Product Name"],
        aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
    )
    pt = pt.rename(columns={"Rxs": "Total Rxs", "MemberID": "Unique Members"})
    rxs = filtered["Rxs"].sum()
    members = filtered["MemberID"].nunique()
    return pt, rxs, members


def process_openmdf_tier_pivots(df, tiers):
    """Process Open MDF tier pivot tables and collect statistics."""
    tab_members = {}
    tab_rxs = {}
    tier_pivots = []

    for name, col, from_val, to_val in tiers:
        pt, rxs, members = summarize_by_openmdf_tier(df, col, from_val, to_val)
        tier_pivots.append((name, pt, members, rxs))
        tab_members[name] = members
        tab_rxs[name] = rxs

    return tier_pivots, tab_members, tab_rxs


def create_openmdf_summary_dataframe(tab_members, tab_rxs, total_claims, total_members):
    """Create the Open MDF summary DataFrame with calculated statistics."""
    pos_keys = [
        "OpenMDF_Positive 2-1",
        "OpenMDF_Positive 3-1",
        "OpenMDF_Positive 3-2",
    ]
    neg_keys = [
        "OpenMDF_Negative 1-2",
        "OpenMDF_Negative 1-3",
        "OpenMDF_Negative 2-3",
    ]

    pos_utilizers = sum(tab_members[k] for k in pos_keys)
    pos_claims = sum(tab_rxs[k] for k in pos_keys)
    pos_pct = pos_claims / total_claims if total_claims else 0

    neg_utilizers = sum(tab_members[k] for k in neg_keys)
    neg_claims = sum(tab_rxs[k] for k in neg_keys)
    neg_pct = neg_claims / total_claims if total_claims else 0

    return pd.DataFrame(
        {
            "Formulary": [
                "Open MDF Positive",
                "Open MDF Negative",
            ],
            "Utilizers": [
                pos_utilizers,
                neg_utilizers,
            ],
            "Rxs": [
                pos_claims,
                neg_claims,
            ],
            "% of claims": [
                pos_pct,
                neg_pct,
            ],
            "": ["", ""],
            "Totals": [
                f"Members: {total_members}",
                f"Claims: {total_claims}",
            ],
        }
    )


def create_openmdf_network_analysis(df):
    """Create network analysis for excluded pharmacies."""
    network_df = df[df["pharmacy_is_excluded"]]
    logger.debug(f"Initial network_df shape: {network_df.shape}")
    logger.debug(f"Initial network_df contents: {network_df.head(10).to_dict()}")

    import re
    filter_phrases = [
        "CVS",
        "Walgreens",
        "Kroger",
        "Walmart",
        "Rite Aid",
        "Optum",
        "Express Scripts",
        "DMR",
        "Williams Bro",
        "Publix",
    ]
    filter_phrases_escaped = [
        re.escape(phrase.lower()) for phrase in filter_phrases
    ]
    regex_pattern = "|".join([f"\\b{p}\\b" for p in filter_phrases_escaped])
    network_df = network_df[
        ~network_df["Pharmacy Name"]
        .str.lower()
        .str.contains(regex_pattern, case=False, regex=True, na=False)
    ]
    logger.debug(f"Filtered network_df shape: {network_df.shape}")
    logger.debug(f"Filtered network_df contents: {network_df.head(10).to_dict()}")

    if {"PHARMACYNPI", "NABP", "Pharmacy Name"}.issubset(network_df.columns):
        network_df["PHARMACYNPI"] = network_df["PHARMACYNPI"].fillna("N/A")
        network_df["NABP"] = network_df["NABP"].fillna("NABP")
        # Create Unique Identifier based on whichever identifier is available (PHARMACYNPI or NABP)
        network_df["Unique Identifier"] = network_df.apply(
            lambda row: row["PHARMACYNPI"]
            if row["PHARMACYNPI"] != "N/A"
            else row["NABP"],
            axis=1,
        )
        network_pivot = pd.pivot_table(
            network_df,
            values=["Rxs", "MemberID"],
            index=["Unique Identifier"],
            aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
        )
        network_pivot = network_pivot.rename(
            columns={"Rxs": "Total Rxs", "MemberID": "Unique Members"}
        )
        network_pivot.reset_index(
            inplace=True
        )  # Ensure index columns are included in the output
        logger.debug(f"Network pivot shape: {network_pivot.shape}")
        logger.debug(f"Network pivot contents: {network_pivot.head(10).to_dict()}")
        
        return network_df, network_pivot
    
    return network_df, None


def write_openmdf_excel_sheets(writer, df, summary_df, tier_pivots, network_df, network_pivot):
    """Write all sheets to the Excel file."""
    # Write Summary sheet
    summary_df.to_excel(writer, sheet_name="Summary", index=False)

    # Write tier pivots
    for name, pt, members, _ in tier_pivots:
        pt.to_excel(writer, sheet_name=name)
        writer.sheets[name].write("F1", f"Total Members: {members}")

    # Write Data sheet
    data_sheet_df = df.copy()
    data_sheet_df.to_excel(writer, sheet_name="Data", index=False)

    # Write Network sheet
    if network_pivot is not None:
        network_pivot.to_excel(writer, sheet_name="Network", index=False)

    # Write filtered network data
    selected_columns = [
        "PHARMACYNPI",
        "NABP",
        "MemberID",
        "Pharmacy Name",
        "pharmacy_is_excluded",
        "Unique Identifier",
    ]
    network_df[selected_columns].to_excel(writer, sheet_name="Network", index=False)
    logger.info(
        f"Network sheet updated with {network_df.shape[0]} excluded pharmacy records (minus major chains) and selected columns"
    )


def reorder_openmdf_excel_sheets(writer):
    """Reorder sheets so Summary follows Data."""
    sheets = writer.sheets  # This is a dict: {sheet_name: worksheet_object}
    names = list(sheets.keys())
    if "Data" in names and "Summary" in names:
        data_idx = names.index("Data")
        summary_idx = names.index("Summary")
        if summary_idx != data_idx + 1:
            # Move "Summary" sheet immediately after "Data"
            items = list(sheets.items())
            summary_item = items.pop(summary_idx)
            items.insert(data_idx + 1, summary_item)
            writer.sheets.clear()
            writer.sheets.update(dict(items))


def show_openmdf_completion_message(output_path):
    """Show completion message and popup."""
    write_shared_log("openmdf_tier.py", "Processing complete.")
    print(f"Processing complete. Output file: {output_path}")
    try:
        import tkinter as tk
        from tkinter import messagebox

        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo("Processing Complete", "Processing complete")
        root.destroy()
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Main processing pipeline
# ---------------------------------------------------------------------------


def process_data():
    write_shared_log("openmdf_tier.py", "Processing started.")
    # Output filename from CLI arg or default
    output_filename = "LBL for Disruption.xlsx"
    if len(sys.argv) > 1:
        output_filename = sys.argv[1]
    output_path = Path(output_filename).resolve()
    try:
        # Get the config file path relative to the project root
        config_path = Path(__file__).parent.parent / "config" / "file_paths.json"
        file_paths = load_file_paths(str(config_path))

        result = load_openmdf_tier_data(file_paths)
        if result is None:
            return  # Early exit if claims loading failed
        claims, medi, mdf, exclusive, network = result

        reference_data = (medi, mdf, exclusive)
        df = process_openmdf_data_pipeline(claims, reference_data, network)

        df = handle_openmdf_pharmacy_exclusions(df, file_paths)

        # Convert FormularyTier to numeric for proper filtering
        df["FormularyTier"] = pd.to_numeric(df["FormularyTier"], errors="coerce")

        # Totals for summary
        total_claims = df["Rxs"].sum()
        total_members = df["MemberID"].nunique()

        # Excel writer setup
        writer = pd.ExcelWriter(output_path, engine="xlsxwriter")

        # Summary calculations (must be written immediately after Data)
        tiers = create_openmdf_tier_definitions()

        tier_pivots, tab_members, tab_rxs = process_openmdf_tier_pivots(df, tiers)

        # Summary calculations
        summary_df = create_openmdf_summary_dataframe(tab_members, tab_rxs, total_claims, total_members)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # Write tier pivots after Summary
        for name, pt, members, _ in tier_pivots:
            pt.to_excel(writer, sheet_name=name)
            writer.sheets[name].write("F1", f"Total Members: {members}")

        # Write the 'Data Sheet' with excluded and non-excluded pharmacies
        data_sheet_df = df.copy()
        data_sheet_df.to_excel(writer, sheet_name="Data", index=False)

        # Network summary for excluded pharmacies (pharmacy_is_excluded=True)
        network_df, network_pivot = create_openmdf_network_analysis(df)
        logger.info(f"Total pharmacies in dataset: {df.shape[0]}")
        logger.info(
            f"Excluded pharmacies (pharmacy_is_excluded=True): {df['pharmacy_is_excluded'].sum()}"
        )
        logger.info(
            f"Non-excluded pharmacies (pharmacy_is_excluded=False): {(~df['pharmacy_is_excluded']).sum()}"
        )
        logger.info(
            f"Network sheet will show {network_df.shape[0]} excluded pharmacy records (minus major chains)"
        )

        # Write Network sheet
        if network_pivot is not None:
            network_pivot.to_excel(writer, sheet_name="Network", index=False)

        # Write filtered network data
        selected_columns = [
            "PHARMACYNPI",
            "NABP",
            "MemberID",
            "Pharmacy Name",
            "pharmacy_is_excluded",
            "Unique Identifier",
        ]
        network_df[selected_columns].to_excel(writer, sheet_name="Network", index=False)
        logger.info(
            f"Network sheet updated with {network_df.shape[0]} excluded pharmacy records (minus major chains) and selected columns"
        )

        # Reorder sheets so Summary follows Data
        reorder_openmdf_excel_sheets(writer)

        writer.close()
        show_openmdf_completion_message(output_path)
    except Exception as e:
        write_shared_log(
            "openmdf_tier.py", f"Processing failed: {e}", status="ERROR"
        )
        raise


if __name__ == "__main__":
    process_data()
    print("Data processing complete")
"""
Process management module for handling automation workflows.
Extracted from app.py to improve cohesion and reduce file size.
"""

import subprocess
import threading
import time
import logging
import os
import sys
from tkinter import messagebox

from config.app_config import DisruptionConfig
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.utils import write_shared_log


class ProcessManager:
    """Handles process management and workflow orchestration."""
    
    def __init__(self, app_instance):
        self.app = app_instance
        
    def start_process_threaded(self):
        """Start the main repricing process in a separate thread."""
        threading.Thread(target=self._start_process_internal).start()
        write_shared_log("ProcessManager", "Repricing process started")
        
    def _start_process_internal(self):
        """Internal method to handle the repricing process."""
        try:
            self.app.start_time = time.time()
            self.app.update_progress(0.05)
            
            # Extra safeguard: Remove any accidental LBL/disruption output during repricing
            os.environ["NO_LBL_OUTPUT"] = "1"
            
            # Validate inputs
            if not self.app.validate_merge_inputs():
                self.app.update_progress(0)
                return
                
            # Kill Excel processes
            self.app.update_progress(0.10)
            self._kill_excel_processes()
            
            # Run merge operation
            self.app.update_progress(0.20)
            self._run_merge_process()
            
            # Process merged file
            self.app.update_progress(0.50)
            merged_file = "merged_file.xlsx"
            self.app.process_merged_file(merged_file)
            
            # Complete
            self.app.update_progress(1.0)
            
        except Exception as e:
            self.app.update_progress(0)
            logging.error(f"Process failed: {e}")
            messagebox.showerror("Error", f"Process failed: {e}")
            
    def _kill_excel_processes(self):
        """Kill any running Excel processes."""
        try:
            subprocess.run(
                ["taskkill", "/F", "/IM", "excel.exe"],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )
        except Exception as e:
            logging.warning(f"Could not kill Excel processes: {e}")
            
    def _run_merge_process(self):
        """Run the merge.py script with file inputs."""
        try:
            subprocess.run(
                ["python", "merge.py", self.app.file1_path, self.app.file2_path], 
                check=True
            )
        except subprocess.CalledProcessError as e:
            logging.error(f"Merge process failed: {e}")
            raise
            
    def start_disruption(self, disruption_type=None):
        """Start disruption processing using configuration-driven approach."""
        if disruption_type is None:
            disruption_type = self.app.selected_disruption_type.get().strip()
        
        program_file = DisruptionConfig.get_program_file(disruption_type)
        if not program_file:
            messagebox.showerror("Error", f"Unknown disruption type: {disruption_type}")
            return
        
        self._execute_disruption_process(disruption_type, program_file)
        
    def _execute_disruption_process(self, disruption_type, program_file):
        """Execute the disruption process with error handling."""
        try:
            args = ["python", program_file]
            if self.app.template_file_path:
                args.append(str(self.app.template_file_path))
            
            # Use subprocess to run the disruption script
            subprocess.Popen(args)
            messagebox.showinfo(
                "Success",
                f"{disruption_type} disruption started in a separate process.",
            )
            
        except Exception as e:
            logging.error(f"Failed to start {program_file}: {e}")
            messagebox.showerror("Error", f"{disruption_type} disruption failed: {e}")
            
    def run_label_generation(self, label_type):
        """Run label generation scripts (SHARx or EPLS)."""
        try:
            script_name = f"{label_type.lower()}_lbl.py"
            subprocess.run(["python", script_name], check=True)
            write_shared_log("ProcessManager", f"{label_type} LBL generation completed")
            
        except subprocess.CalledProcessError as e:
            logging.error(f"{label_type} LBL generation failed: {e}")
            messagebox.showerror("Error", f"{label_type} LBL generation failed: {e}")
            
    def cancel_process(self):
        """Cancel the current process."""
        logging.info("Process cancellation requested")
        write_shared_log("ProcessManager", "Process cancelled")
        messagebox.showinfo("Cancelled", "Process cancellation requested.")
        
    def finish_notification(self):
        """Show completion notification."""
        try:
            from plyer import notification
            if hasattr(notification, "notify") and callable(notification.notify):
                notification.notify(
                    title="Repricing Automation",
                    message="Batch processing completed.",
                    timeout=5,
                )
        except ImportError:
            pass  # Notification not available
            
        write_shared_log("ProcessManager", "Batch processing completed")
        messagebox.showinfo("Completed", "Batch processing finished!")
import cProfile
import pstats
import sys

if __name__ == "__main__":
    script_path = (
        sys.argv[1] if len(sys.argv) > 1 else "app.py"
    )  # Allow dynamic script path input
    profile_output = "profile_stats.prof"

    print(f"Profiling {script_path}...\n")
    cProfile.runctx(
        "exec(compile(open(script_path).read(), script_path, 'exec'))",
        globals(),
        locals(),
        profile_output,
    )

    # Optional: print top 20 cumulative time functions
    stats = pstats.Stats(profile_output)
    stats.sort_stats(pstats.SortKey.CUMULATIVE).print_stats(20)
"""
File processing module for handling file operations and validation.
Extracted from app.py to improve cohesion and reduce file size.
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
import csv
import logging
import os
import sys
from tkinter import messagebox

from config.app_config import ProcessingConfig, AppConstants
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.utils import write_shared_log


class FileProcessor:
    """Handles file operations and validation."""
    
    def __init__(self, app_instance):
        self.app = app_instance
        
    def check_template(self, file_path):
        """Check if template file exists and is valid."""
        return Path(file_path).exists() and Path(file_path).suffix == '.xlsx'
    
    def import_file(self, file_type="File"):
        """Import and validate a file."""
        file_path = self.app.ui_factory.create_file_dialog(
            title=f"Select {file_type}",
            filetypes=ProcessingConfig.FILE_TYPES
        )
        
        if not file_path:
            return None
            
        try:
            # Validate file exists
            if not Path(file_path).exists():
                messagebox.showerror("Error", f"{file_type} not found.")
                return None
                
            # Load and validate the file
            df = pd.read_csv(file_path) if file_path.endswith('.csv') else pd.read_excel(file_path)
            
            # Log the import
            write_shared_log("FileProcessor", f"{file_type} imported successfully: {file_path}")
            
            return file_path, df
            
        except Exception as e:
            error_msg = f"Error importing {file_type}: {str(e)}"
            messagebox.showerror("Error", error_msg)
            write_shared_log("FileProcessor", error_msg, "ERROR")
            return None
    
    def validate_file_structure(self, df, required_columns=None):
        """Validate that the file has the required structure."""
        if required_columns is None:
            required_columns = ProcessingConfig.REQUIRED_COLUMNS
            
        try:
            ProcessingConfig.validate_required_columns(df)
            return True
        except ValueError as e:
            messagebox.showerror("Validation Error", str(e))
            return False
    
    def write_audit_log(self, file1, file2, status):
        """Write an entry to the audit log."""
        entry = [datetime.now().isoformat(), str(file1), str(file2), status]
        write_header = not AppConstants.AUDIT_LOG.exists()
        
        try:
            with open(AppConstants.AUDIT_LOG, "a", newline="") as csvfile:
                writer = csv.writer(csvfile)
                if write_header:
                    writer.writerow(["Timestamp", "File1", "File2", "Status"])
                writer.writerow(entry)
        except Exception as e:
            logging.error(f"Failed to write audit log: {e}")
    
    def prepare_file_paths(self, template_path):
        """Prepare file paths for template operations."""
        if not template_path:
            raise ValueError("Template file path is not set.")
            
        template = Path(template_path)
        backup_name = template.stem + AppConstants.BACKUP_SUFFIX
        
        return {
            "template": template,
            "backup": Path.cwd() / backup_name,
            "output": Path.cwd() / AppConstants.UPDATED_TEMPLATE_NAME
        }
    
    def safe_file_operation(self, operation, *args, **kwargs):
        """Safely perform file operations with error handling."""
        try:
            return operation(*args, **kwargs)
        except Exception as e:
            error_msg = f"File operation failed: {str(e)}"
            messagebox.showerror("File Error", error_msg)
            write_shared_log("FileProcessor", error_msg, "ERROR")
            return None
"""
Configuration module for the Repricing Automation application.
Contains configuration classes and constants.
"""

import multiprocessing
from pathlib import Path


class ProcessingConfig:
    """Configuration class for processing settings and validation."""
    
    REQUIRED_COLUMNS = [
        "DATEFILLED", "SOURCERECORDID", "QUANTITY", "DAYSUPPLY", "NDC",
        "MemberID", "Drug Name", "Pharmacy Name", "Total AWP (Historical)"
    ]
    
    FILE_TYPES = [
        ("All files", "*.*"),
        ("CSV files", "*.csv"),
        ("Excel files", "*.xlsx"),
    ]
    
    TEMPLATE_FILE_TYPES = [("Excel files", "*.xlsx")]
    
    DEFAULT_OPPORTUNITY_NAME = "claims detail PCU"
    
    @classmethod
    def get_multiprocessing_workers(cls):
        """Get the optimal number of workers for multiprocessing."""
        return min(4, max(1, multiprocessing.cpu_count() // 2))
    
    @classmethod
    def validate_required_columns(cls, df):
        """Validate that all required columns are present in the DataFrame."""
        missing_cols = [col for col in cls.REQUIRED_COLUMNS if col not in df.columns]
        if missing_cols:
            raise ValueError(f"Missing required columns: {missing_cols}")
        return True


class DisruptionConfig:
    """Configuration for disruption types to reduce conditional complexity."""
    
    DISRUPTION_TYPES = {
        "Tier Disruption": {
            "module": "modules.tier_disruption",
            "file": "tier_disruption.py"
        },
        "B/G Disruption": {
            "module": "modules.bg_disruption", 
            "file": "bg_disruption.py"
        },
        "OpenMDF (Tier)": {
            "module": "modules.openmdf_tier",
            "file": "openmdf_tier.py"
        },
        "OpenMDF (B/G)": {
            "module": "modules.openmdf_bg",
            "file": "openmdf_bg.py"
        },
        "Full Claims File": {
            "module": "modules.full_claims",
            "file": "full_claims.py"
        }
    }
    
    @classmethod
    def get_program_file(cls, disruption_type):
        """Get the program file for a disruption type."""
        config = cls.DISRUPTION_TYPES.get(disruption_type)
        return config["file"] if config else None
    
    @classmethod
    def get_disruption_labels(cls):
        """Get list of available disruption types (excluding Full Claims File)."""
        return [
            label for label in cls.DISRUPTION_TYPES.keys() 
            if label != "Full Claims File"
        ]


class AppConstants:
    """Application constants and configuration values."""
    
    # Configuration and audit log files
    CONFIG_FILE = Path("config.json")
    AUDIT_LOG = Path("audit_log.csv")
    
    # Template handling constants
    BACKUP_SUFFIX = "_backup.xlsx"
    UPDATED_TEMPLATE_NAME = "_Rx Repricing_wf.xlsx"
    
    # Welcome messages for different users
    WELCOME_MESSAGES = {
        "DamionMorrison": "Welcome back, Damion! Ready to reprice?",
        "DannyBushnell": "Hello Danny! Let's get started.",
        "BrettBauer": "Hi Brett, your automation awaits!",
        "BrendanReamer": "Welcome Brendan! Ready to optimize?",
    }
    
    # Emoji options for welcome message
    EMOJIS = [
        ":rocket:", ":sunglasses:", ":star:", 
        ":tada:", ":computer:", ":chart_with_upwards_trend:"
    ]
    
    # Notes text for UI
    NOTES_TEXT = (
        "Note:\n\n"
        "Ensure FormularyTier is set before running disruption.\n"
        "Close any open Excel instances.\n"
        "Keep template name as _Rx Repricing_wf until done."
    )
"""
Configuration Management - CodeScene ACE Improvement
Centralized configuration handling with better error management
"""

import json
import logging
from pathlib import Path
from typing import Dict, Any, Optional
from dataclasses import dataclass, asdict

logger = logging.getLogger(__name__)


@dataclass
class AppSettings:
    """Application settings data class."""
    last_folder: str
    theme: str = "light"
    auto_save: bool = True
    log_level: str = "INFO"
    max_workers: int = 4
    backup_enabled: bool = True
    
    @classmethod
    def default(cls) -> 'AppSettings':
        """Create default settings."""
        return cls(
            last_folder=str(Path.cwd()),
            theme="light",
            auto_save=True,
            log_level="INFO",
            max_workers=4,
            backup_enabled=True
        )


class ConfigurationManager:
    """
    Improved configuration management following CodeScene ACE principles.
    - Single responsibility: Only handles configuration
    - Better error handling
    - Type safety with dataclasses
    - Clear separation of concerns
    """
    
    def __init__(self, config_file: Path = None):
        self.config_file = config_file or Path("config.json")
        self._settings: Optional[AppSettings] = None
        self._load_configuration()
    
    def _load_configuration(self) -> None:
        """Load configuration from file or create default."""
        try:
            if self.config_file.exists():
                self._settings = self._load_from_file()
                logger.info(f"Configuration loaded from {self.config_file}")
            else:
                self._settings = AppSettings.default()
                self._save_configuration()
                logger.info("Default configuration created")
        except Exception as e:
            logger.error(f"Failed to load configuration: {e}")
            self._settings = AppSettings.default()
    
    def _load_from_file(self) -> AppSettings:
        """Load settings from JSON file."""
        try:
            with open(self.config_file, 'r') as f:
                data = json.load(f)
                return AppSettings(**data)
        except (json.JSONDecodeError, TypeError) as e:
            logger.warning(f"Invalid configuration file: {e}")
            return AppSettings.default()
    
    def _save_configuration(self) -> None:
        """Save current settings to file."""
        try:
            if self._settings is not None:
                with open(self.config_file, 'w') as f:
                    json.dump(asdict(self._settings), f, indent=4)
                logger.debug(f"Configuration saved to {self.config_file}")
            else:
                logger.error("No settings to save: self._settings is None")
                return
        except Exception as e:
            logger.error(f"Failed to save configuration: {e}")
    
    @property
    def settings(self) -> AppSettings:
        """Get current settings."""
        return self._settings
    
    def update_setting(self, key: str, value: Any) -> bool:
        """Update a single setting."""
        try:
            if hasattr(self._settings, key):
                setattr(self._settings, key, value)
                self._save_configuration()
                logger.info(f"Setting updated: {key} = {value}")
                return True
            else:
                logger.warning(f"Unknown setting: {key}")
                return False
        except Exception as e:
            logger.error(f"Failed to update setting {key}: {e}")
            return False
    
    def update_settings(self, **kwargs) -> bool:
        """Update multiple settings."""
        try:
            for key, value in kwargs.items():
                if not hasattr(self._settings, key):
                    logger.warning(f"Unknown setting: {key}")
                    continue
                setattr(self._settings, key, value)
            
            self._save_configuration()
            logger.info(f"Settings updated: {list(kwargs.keys())}")
            return True
        except Exception as e:
            logger.error(f"Failed to update settings: {e}")
            return False
    
    def reset_to_defaults(self) -> None:
        """Reset settings to defaults."""
        self._settings = AppSettings.default()
        self._save_configuration()
        logger.info("Settings reset to defaults")
    
    def get_setting(self, key: str, default: Any = None) -> Any:
        """Get a specific setting value."""
        return getattr(self._settings, key, default)


# Usage example for improved app.py
class ImprovedConfigManager(ConfigurationManager):
    """
    Backwards compatible configuration manager.
    Maintains the same interface as the original while providing improvements.
    """
    
    def __init__(self):
        super().__init__(Path("config.json"))
        # Maintain backwards compatibility
        self.config = self._get_legacy_config()
    
    def _get_legacy_config(self) -> Dict[str, Any]:
        """Get configuration in legacy format."""
        return {
            "last_folder": self.settings.last_folder,
            "theme": self.settings.theme,
            "auto_save": self.settings.auto_save,
            "log_level": self.settings.log_level,
            "max_workers": self.settings.max_workers,
            "backup_enabled": self.settings.backup_enabled
        }
    
    def save_default(self) -> None:
        """Legacy method for saving defaults."""
        self.reset_to_defaults()
        self.config = self._get_legacy_config()
    
    def load(self) -> None:
        """Legacy method for loading configuration."""
        self._load_configuration()
        self.config = self._get_legacy_config()
    
    def save(self) -> None:
        """Legacy method for saving configuration."""
        self._save_configuration()


# Example usage in app_improved.py
def example_usage():
    """Example of how to use the improved configuration manager."""
    
    # Create configuration manager
    config_manager = ImprovedConfigManager()
    
    # Access settings
    print(f"Last folder: {config_manager.settings.last_folder}")
    print(f"Theme: {config_manager.settings.theme}")
    
    # Update settings
    config_manager.update_setting("last_folder", "/new/path")
    config_manager.update_settings(theme="dark", auto_save=False)
    
    # Use in app initialization
    app_settings = config_manager.settings
    print(f"Starting app with theme: {app_settings.theme}")
    print(f"Auto-save enabled: {app_settings.auto_save}")


if __name__ == "__main__":
    example_usage()
import logging
import os
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.utils import write_shared_log


def make_audit_entry(script_name, message, status="INFO"):
    try:
        write_shared_log(script_name, message, status)
    except Exception as e:
        logging.error(f"[AUDIT FAIL] {script_name} audit failed: {e}")
        try:
            with open("local_fallback_log.txt", "a") as f:
                f.write(f"{script_name}: {message} [{status}]\n")
        except Exception as inner:
            logging.error(f"[FALLBACK FAIL] Could not write fallback log: {inner}")
import pandas as pd
import logging
import os
import sys
from pathlib import Path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.utils import (
    load_file_paths,
    filter_logic_and_maintenance,
    filter_products_and_alternative,
    standardize_pharmacy_ids,
    standardize_network_ids,
    merge_with_network,
    drop_duplicates_df,
    clean_logic_and_tier,
    filter_recent_date,
    write_shared_log,
)

# Logging setup
logging.basicConfig(
    filename="bg_disruption.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


def load_data_files(file_paths):
    """Load and return all required data files."""
    logger.info("Loading data files...")
    
    # Load claims data
    try:
        claims = pd.read_excel(
            file_paths["reprice"],
            sheet_name="Claims Table",
            usecols=[
                "SOURCERECORDID",
                "NDC",
                "MemberID",
                "DATEFILLED",
                "FormularyTier",
                "Rxs",
                "Logic",
                "PHARMACYNPI",
                "NABP",
                "Pharmacy Name",
                "Universal Rebates",
                "Exclusive Rebates",
            ],
        )
    except Exception as e:
        logger.warning(f"Claims Table fallback: {e}")
        write_shared_log(
            "bg_disruption.py", f"Claims Table fallback: {e}", status="WARNING"
        )
        claims = pd.read_excel(file_paths["reprice"], sheet_name=0)
    
    logger.info(f"claims shape: {claims.shape}")
    claims.info()

    # Load other data files
    medi = pd.read_excel(file_paths["medi_span"])[
        ["NDC", "Maint Drug?", "Product Name"]
    ]
    logger.info(f"medi shape: {medi.shape}")
    
    uni = pd.read_excel(file_paths["u_disrupt"], sheet_name="Universal NDC")[
        ["NDC", "Tier"]
    ]
    logger.info(f"uni shape: {uni.shape}")
    
    exl = pd.read_excel(file_paths["e_disrupt"], sheet_name="Alternatives NDC")[
        ["NDC", "Tier", "Alternative"]
    ]
    logger.info(f"exl shape: {exl.shape}")
    
    network = pd.read_excel(file_paths["n_disrupt"])[
        ["pharmacy_npi", "pharmacy_nabp", "pharmacy_is_excluded"]
    ]
    logger.info(f"network shape: {network.shape}")
    
    return claims, medi, uni, exl, network


def merge_data_files(claims, reference_data, network):
    """Merge all data files into a single DataFrame."""
    logger.info("Merging data files...")
    
    medi, uni, exl = reference_data
    
    df = claims.merge(medi, on="NDC", how="left")
    logger.info(f"After merge with medi: {df.shape}")
    
    df = df.merge(uni.rename(columns={"Tier": "Universal Tier"}), on="NDC", how="left")
    logger.info(f"After merge with uni: {df.shape}")
    
    df = df.merge(exl.rename(columns={"Tier": "Exclusive Tier"}), on="NDC", how="left")
    logger.info(f"After merge with exl: {df.shape}")

    # Standardize IDs and perform network merge
    df = standardize_pharmacy_ids(df)
    logger.info(f"After standardize_pharmacy_ids: {df.shape}")
    network = standardize_network_ids(network)
    logger.info(f"After standardize_network_ids: {network.shape}")
    df = merge_with_network(df, network)
    logger.info(f"After merge_with_network: {df.shape}")
    
    return df


def process_and_filter_data(df):
    """Process and filter the merged data."""
    logger.info("Processing and filtering data...")
    
    # Date parsing, deduplication, type cleaning, and filters
    df["DATEFILLED"] = pd.to_datetime(df["DATEFILLED"], errors="coerce")
    logger.info(f"After DATEFILLED to_datetime: {df.shape}")
    
    df = drop_duplicates_df(df)
    logger.info(f"After drop_duplicates_df: {df.shape}")
    
    df = clean_logic_and_tier(df)
    logger.info(f"After clean_logic_and_tier: {df.shape}")
    
    df = filter_recent_date(df)
    logger.info(f"After filter_recent_date: {df.shape}")
    
    df = filter_logic_and_maintenance(df)
    logger.info(f"After filter_logic_and_maintenance: {df.shape}")
    
    df = filter_products_and_alternative(df)
    logger.info(f"After filter_products_and_alternative: {df.shape}")

    df["FormularyTier"] = df["FormularyTier"].astype(str).str.strip().str.upper()
    df["Alternative"] = df["Alternative"].astype(str)
    
    return df


def handle_pharmacy_exclusions(df, file_paths):
    """Handle pharmacy exclusions and validation."""
    logger.info("Handling pharmacy exclusions...")
    
    # Ensure 'pharmacy_is_excluded' column contains actual boolean values with type inference
    if "pharmacy_is_excluded" in df.columns:
        df["pharmacy_is_excluded"] = (
            df["pharmacy_is_excluded"]
            .astype(str)
            .str.lower()
            .map({"true": True, "false": False})
            .fillna(False)
            .infer_objects(copy=False)
        )
        logger.info(
            f"pharmacy_is_excluded value counts: {df['pharmacy_is_excluded'].value_counts().to_dict()}"
        )
        
        # Identify rows where pharmacy_is_excluded is "NA"
        na_pharmacies = df[df["pharmacy_is_excluded"].isna()]
        logger.info(f"NA pharmacies count: {na_pharmacies.shape[0]}")

        if not na_pharmacies.empty:
            # Define the writer before using it
            output_file_path = file_paths["pharmacy_validation"]
            na_pharmacies_output = na_pharmacies[["PHARMACYNPI", "NABP"]].fillna("N/A")
            
            # Update existing template or create new file
            try:
                from openpyxl import load_workbook
                import os
                
                # Check if template exists
                if os.path.exists(output_file_path):
                    # Load existing workbook
                    wb = load_workbook(output_file_path)
                    logger.info(f"Loading existing template: {output_file_path}")
                else:
                    # Create new workbook
                    from openpyxl import Workbook
                    wb = Workbook()
                    logger.info(f"Creating new validation file: {output_file_path}")
                
                # Check if Validations sheet exists, if not create it
                if "Validations" in wb.sheetnames:
                    ws = wb["Validations"]
                    # Clear existing data
                    ws.delete_rows(1, ws.max_row)
                else:
                    ws = wb.create_sheet("Validations")
                    # Remove default sheet if it exists and is empty
                    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                        wb.remove(wb["Sheet"])
                
                # Write headers
                ws.append(["PHARMACYNPI", "NABP"])
                
                # Write data
                for _, row in na_pharmacies_output.iterrows():
                    ws.append([row["PHARMACYNPI"], row["NABP"]])
                
                # Save the workbook
                wb.save(output_file_path)
                logger.info(f"NA pharmacies written to '{output_file_path}' Validations sheet.")
                
            except Exception as e:
                logger.error(f"Error updating pharmacy validation file: {e}")
                # Fallback to original method
                writer = pd.ExcelWriter(output_file_path, engine="openpyxl")
                na_pharmacies_output.to_excel(
                    writer, sheet_name="Validations", index=False, engine="openpyxl"
                )
                writer.close()
                logger.info(f"NA pharmacies written to '{output_file_path}' sheet (fallback).")
    
    return df


def create_data_filters(df):
    """Create filtered datasets for different scenarios."""
    logger.info("Creating data filters...")
    
    uni_pos = df[(df["Universal Tier"] == 1) & df["FormularyTier"].isin(["B", "BRAND"])]
    logger.info(f"uni_pos shape: {uni_pos.shape}")
    
    uni_neg = df[
        df["Universal Tier"].isin([2, 3]) & df["FormularyTier"].isin(["G", "GENERIC"])
    ]
    logger.info(f"uni_neg shape: {uni_neg.shape}")
    
    ex_pos = df[(df["Exclusive Tier"] == 1) & df["FormularyTier"].isin(["B", "BRAND"])]
    logger.info(f"ex_pos shape: {ex_pos.shape}")
    
    ex_neg = df[
        df["Exclusive Tier"].isin([2, 3]) & df["FormularyTier"].isin(["G", "GENERIC"])
    ]
    logger.info(f"ex_neg shape: {ex_neg.shape}")
    
    ex_ex = df[df["Exclusive Tier"] == "Nonformulary"]
    logger.info(f"ex_ex shape: {ex_ex.shape}")
    
    return uni_pos, uni_neg, ex_pos, ex_neg, ex_ex


def create_pivot_tables(filtered_data):
    """Create pivot tables and calculate member counts."""
    logger.info("Creating pivot tables...")
    
    uni_pos, uni_neg, ex_pos, ex_neg, ex_ex = filtered_data
    
    def pivot(d, include_alternative=False):
        index_cols = ["Product Name"]
        if include_alternative and "Alternative" in d.columns:
            index_cols.append("Alternative")
        pt = pd.pivot_table(
            d,
            values=["Rxs", "MemberID"],
            index=index_cols,
            aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
        )
        pt = pt.rename(columns={"Rxs": "Total Rxs", "MemberID": "Unique Members"})
        return pt

    def count(d):
        return 0 if d.empty or d["Rxs"].sum() == 0 else d["MemberID"].nunique()

    tabs = {
        "Universal_Positive": (uni_pos, pivot(uni_pos), count(uni_pos)),
        "Universal_Negative": (uni_neg, pivot(uni_neg), count(uni_neg)),
        "Exclusive_Positive": (ex_pos, pivot(ex_pos), count(ex_pos)),
        "Exclusive_Negative": (ex_neg, pivot(ex_neg), count(ex_neg)),
        "Exclusions": (ex_ex, pivot(ex_ex, include_alternative=True), count(ex_ex)),
    }
    
    return tabs


def create_summary_data(df, tabs):
    """Create summary data for the report."""
    logger.info("Creating summary data...")
    
    total_members = df["MemberID"].nunique()
    total_claims = df["Rxs"].sum()

    summary = pd.DataFrame(
        {
            "Formulary": [
                "Universal Positive",
                "Universal Negative",
                "Exclusive Positive",
                "Exclusive Negative",
                "Exclusions",
            ],
            "Utilizers": [v[2] for v in tabs.values()],
            "Rxs": [v[0]["Rxs"].sum() for v in tabs.values()],
            "% of claims": [
                v[0]["Rxs"].sum() / total_claims if total_claims else 0
                for v in tabs.values()
            ],
            "": ["" for _ in tabs],
            "Totals": [
                f"Members: {total_members}",
                f"Claims: {total_claims}",
                "",
                "",
                "",
            ],
        }
    )
    
    return summary


def create_network_data(df):
    """Create network data for excluded pharmacies."""
    logger.info("Creating network data...")
    
    import re
    
    # Network summary for excluded pharmacies (pharmacy_is_excluded=True)
    network_df = df[df["pharmacy_is_excluded"]]
    filter_phrases = [
        "CVS",
        "Walgreens",
        "Kroger",
        "Walmart",
        "Rite Aid",
        "Optum",
        "Express Scripts",
        "DMR",
        "Williams Bro",
        "Publix",
    ]
    
    # Regex safety: escape and lower-case all phrases and names
    filter_phrases_escaped = [re.escape(phrase.lower()) for phrase in filter_phrases]
    regex_pattern = "|".join([f"\\b{p}\\b" for p in filter_phrases_escaped])
    network_df = network_df[
        ~network_df["Pharmacy Name"]
        .str.lower()
        .str.contains(regex_pattern, case=False, regex=True, na=False)
    ]
    logger.info(f"network_df shape after exclusion: {network_df.shape}")
    
    network_pivot = None
    if {"PHARMACYNPI", "NABP", "Pharmacy Name"}.issubset(network_df.columns):
        network_df["PHARMACYNPI"] = network_df["PHARMACYNPI"].fillna("N/A")
        network_df["NABP"] = network_df["NABP"].fillna("N/A")
        network_pivot = pd.pivot_table(
            network_df,
            values=["Rxs", "MemberID"],
            index=["PHARMACYNPI", "NABP", "Pharmacy Name"],
            aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
        )
        network_pivot = network_pivot.rename(
            columns={"Rxs": "Total Rxs", "MemberID": "Unique Members"}
        )
    
    # Log debug info to verify the filtering
    logger.info(f"Total pharmacies in dataset: {df.shape[0]}")
    logger.info(
        f"Excluded pharmacies (pharmacy_is_excluded=True): {df['pharmacy_is_excluded'].sum()}"
    )
    logger.info(
        f"Non-excluded pharmacies (pharmacy_is_excluded=False): {(~df['pharmacy_is_excluded']).sum()}"
    )
    logger.info(
        f"Network sheet will show {network_df.shape[0]} excluded pharmacy records (minus major chains)"
    )
    
    return network_pivot


def write_excel_report(report_data, output_filename):
    """Write the final Excel report."""
    logger.info("Writing Excel report...")
    
    df, summary, tabs, network_pivot = report_data
    
    writer = pd.ExcelWriter(output_filename, engine="xlsxwriter")
    df.to_excel(writer, sheet_name="Data", index=False)
    summary.to_excel(writer, sheet_name="Summary", index=False)

    for sheet, (_, pt, mems) in tabs.items():
        pt.to_excel(writer, sheet_name=sheet)
        writer.sheets[sheet].write("F1", f"Total Members: {mems}")

    if network_pivot is not None:
        network_pivot.to_excel(writer, sheet_name="Network")

    # Reorder sheets so Summary follows Data
    sheets = writer.sheets  # This is a dict: {sheet_name: worksheet_object}
    names = list(sheets.keys())
    if "Data" in names and "Summary" in names:
        data_idx = names.index("Data")
        summary_idx = names.index("Summary")
        if summary_idx != data_idx + 1:
            # Move "Summary" sheet immediately after "Data"
            items = list(sheets.items())
            summary_item = items.pop(summary_idx)
            items.insert(data_idx + 1, summary_item)
            writer.sheets.clear()
            writer.sheets.update(dict(items))

    writer.close()


def show_completion_notification():
    """Show completion notification popup."""
    try:
        import tkinter as tk
        from tkinter import messagebox

        root = tk.Tk()
        root.withdraw()  # Hide the main window
        messagebox.showinfo("Notification", "Processing complete")
        root.destroy()
    except Exception as e:
        logger.warning(f"Popup notification failed: {e}")


def process_data():
    """Main processing function - coordinates all data processing steps."""
    write_shared_log("bg_disruption.py", "Processing started.")
    import sys

    output_filename = "LBL for Disruption.xlsx"
    if len(sys.argv) > 1:
        output_filename = sys.argv[1]
    
    # Get the config file path relative to the project root
    config_path = Path(__file__).parent.parent / "config" / "file_paths.json"
    file_paths = load_file_paths(str(config_path))
    if "reprice" not in file_paths or not file_paths["reprice"]:
        write_shared_log(
            "bg_disruption.py", "No reprice/template file provided.", status="ERROR"
        )
        print("Error: No reprice/template file provided.")
        return

    # Load all data files
    claims, medi, uni, exl, network = load_data_files(file_paths)
    
    # Merge all data files
    reference_data = (medi, uni, exl)
    df = merge_data_files(claims, reference_data, network)
    
    # Process and filter data
    df = process_and_filter_data(df)
    
    # Handle pharmacy exclusions
    df = handle_pharmacy_exclusions(df, file_paths)
    
    # Create filtered datasets
    uni_pos, uni_neg, ex_pos, ex_neg, ex_ex = create_data_filters(df)
    
    # Create pivot tables
    filtered_data = (uni_pos, uni_neg, ex_pos, ex_neg, ex_ex)
    tabs = create_pivot_tables(filtered_data)
    
    # Create summary data
    summary = create_summary_data(df, tabs)
    
    # Create network data
    network_pivot = create_network_data(df)
    
    # Write Excel report
    report_data = (df, summary, tabs, network_pivot)
    write_excel_report(report_data, output_filename)
    
    write_shared_log("bg_disruption.py", "Processing complete.")
    print("Processing complete")
    
    # Show completion notification
    show_completion_notification()


if __name__ == "__main__":
    process_data()
"""
Data processing module for handling CSV/Excel data operations.
Extracted from app.py to improve cohesion and reduce file size.
"""

import pandas as pd
import numpy as np
import logging
import multiprocessing
from pathlib import Path
import re
import os
import sys

from config.app_config import ProcessingConfig
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.utils import write_shared_log


class DataProcessor:
    """Handles data processing and validation operations."""
    
    def __init__(self, app_instance):
        self.app = app_instance
        
    def load_and_validate_data(self, file_path):
        """Load and validate the merged file data."""
        try:
            df = pd.read_excel(file_path)
            logging.info(f"Loaded {len(df)} records from {file_path}")
            
            # Validate required columns using configuration
            ProcessingConfig.validate_required_columns(df)
            
            # Prepare data for processing
            df = df.sort_values(by=["DATEFILLED", "SOURCERECORDID"], ascending=True)
            df["Logic"] = ""
            df["RowID"] = np.arange(len(df))
            
            return df
            
        except Exception as e:
            error_msg = f"Error loading data from {file_path}: {str(e)}"
            logging.error(error_msg)
            write_shared_log("DataProcessor", error_msg, "ERROR")
            raise
    
    def process_data_multiprocessing(self, df):
        """Process data using multiprocessing for improved performance."""
        try:
            # Import and use multiprocessing helpers
            from modules import mp_helpers
            
            num_workers = ProcessingConfig.get_multiprocessing_workers()
            df_blocks = np.array_split(df, num_workers)
            out_queue = multiprocessing.Queue()
            processes = []
            
            # Start worker processes
            for block in df_blocks:
                p = multiprocessing.Process(
                    target=mp_helpers.worker, args=(block, out_queue)
                )
                p.start()
                processes.append(p)
            
            # Collect results
            results = [out_queue.get() for _ in processes]
            for p in processes:
                p.join()
            
            processed_df = pd.concat(results)
            logging.info(f"Processed {len(processed_df)} records using {num_workers} workers")
            return processed_df
            
        except Exception as e:
            error_msg = f"Error processing data with multiprocessing: {str(e)}"
            logging.error(error_msg)
            write_shared_log("DataProcessor", error_msg, "ERROR")
            raise
    
    def save_processed_outputs(self, df, output_dir=None):
        """Save processed data to various output formats."""
        try:
            if output_dir is None:
                output_dir = Path.cwd()
            else:
                output_dir = Path(output_dir)
            
            # Sort and filter data
            df_sorted = pd.concat([df[df["Logic"] == ""], df[df["Logic"] == "OR"]])
            
            # Create row mapping for highlighting
            row_mapping = {
                row["RowID"]: i + 2 for i, (_, row) in enumerate(df_sorted.iterrows())
            }
            excel_rows_to_highlight = [
                row_mapping[rid] for rid in [] if rid in row_mapping
            ]  # Placeholder for unmatched reversals
            
            # Clean up data
            df_sorted.drop(columns=["RowID"], inplace=True, errors="ignore")
            output_file = output_dir / "merged_file_with_OR.xlsx"
            
            # Save to multiple formats
            self._save_to_parquet(df_sorted, output_dir)
            self._save_to_excel(df_sorted, output_file)
            self._save_to_csv(df_sorted, output_dir)
            self._save_unmatched_reversals(excel_rows_to_highlight, output_dir)
            
            return output_file
            
        except Exception as e:
            error_msg = f"Error saving processed outputs: {str(e)}"
            logging.error(error_msg)
            write_shared_log("DataProcessor", error_msg, "ERROR")
            raise
    
    def _save_to_parquet(self, df, output_dir):
        """Save data to Parquet format for large DataFrames."""
        try:
            parquet_path = output_dir / "merged_file_with_OR.parquet"
            df.drop_duplicates().to_parquet(parquet_path, index=False)
            logging.info(f"Saved intermediate Parquet file: {parquet_path}")
        except Exception as e:
            logging.warning(f"Could not save Parquet: {e}")
    
    def _save_to_excel(self, df, output_file):
        """Save data to Excel format."""
        df.drop_duplicates().to_excel(output_file, index=False)
        logging.info(f"Saved Excel file: {output_file}")
    
    def _save_to_csv(self, df, output_dir):
        """Save data to CSV format with opportunity name."""
        try:
            opportunity_name = self._extract_opportunity_name()
            csv_path = output_dir / f"{opportunity_name} Claim Detail.csv"
            df.drop_duplicates().to_csv(csv_path, index=False)
            logging.info(f"Saved CSV file: {csv_path}")
        except Exception as e:
            logging.warning(f"Could not save CSV: {e}")
    
    def _save_unmatched_reversals(self, excel_rows_to_highlight, output_dir):
        """Save unmatched reversals information."""
        try:
            unmatched_path = output_dir / "unmatched_reversals.txt"
            with open(unmatched_path, "w") as f:
                f.write(",".join(map(str, excel_rows_to_highlight)))
            logging.info(f"Saved unmatched reversals info: {unmatched_path}")
        except Exception as e:
            logging.warning(f"Could not save unmatched reversals: {e}")
    
    def _extract_opportunity_name(self):
        """Extract opportunity name from file1_path."""
        opportunity_name = ProcessingConfig.DEFAULT_OPPORTUNITY_NAME
        try:
            if self.app.file1_path:
                if self.app.file1_path.lower().endswith(".xlsx"):
                    df_file1 = pd.read_excel(self.app.file1_path)
                else:
                    df_file1 = pd.read_csv(self.app.file1_path)
                
                if df_file1.shape[1] >= 2:
                    # Get the value from the first row, second column
                    raw_name = str(df_file1.iloc[0, 1])
                    # Clean for filename
                    opportunity_name = re.sub(r'[\\/*?:"<>|]', "_", raw_name)
        except Exception as e:
            logging.warning(f"Could not extract opportunity name from file1: {e}")
        
        return opportunity_name
    
    def validate_merge_inputs(self, file1_path, file2_path):
        """Validate that merge inputs are valid."""
        # Check if both file paths are set and files exist
        if not file1_path or not file2_path:
            return False, "Both input files must be selected."
        
        if not os.path.isfile(file1_path):
            return False, f"File not found: {file1_path}"
        
        if not os.path.isfile(file2_path):
            return False, f"File not found: {file2_path}"
        
        return True, "Inputs are valid"
    
    def validate_gross_cost_template(self, file_path):
        """Validate GrossCost column and suggest template type."""
        try:
            df = pd.read_csv(file_path)
            
            # Guard clause: no GrossCost column
            if "GrossCost" not in df.columns:
                return None
            
            return self._determine_template_type(df["GrossCost"])
            
        except Exception as e:
            logging.warning(f"Could not validate GrossCost column: {e}")
            return None
    
    def _determine_template_type(self, gross_cost_series):
        """Determine which template type to recommend based on GrossCost data."""
        if gross_cost_series.isna().all() or (gross_cost_series == 0).all():
            return "The GrossCost column is blank or all zero. Please use the Blind template."
        else:
            return "The GrossCost column contains data. Please use the Standard template."
import logging
import tkinter as tk
from pathlib import Path
from tkinter import messagebox
import os
import sys

import pandas as pd
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.excel_utils import write_df_to_template

from utils.utils import load_file_paths, write_shared_log

# Setup logging
logging.basicConfig(
    filename="epls_lbl.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


def show_message(awp, ing, total, rxs):
    messagebox.showinfo(
        "Process Complete",
        f"EPLS LBL has been created\n\n"
        f"Total AWP: ${awp:.2f}\n\n"
        f"Total Ing Cost: ${ing:.2f}\n\n"
        f"Total Gross Cost: ${total:.2f}\n\n"
        f"Total Claim Count: {rxs}",
    )


def main() -> None:
    tk.Tk().withdraw()
    write_shared_log("epls_lbl.py", "Processing started.")
    try:
        # Get the config file path relative to the project root
        config_path = Path(__file__).parent.parent / "config" / "file_paths.json"
        paths = load_file_paths(str(config_path))

        # Failsafe: check that both input and template files exist
        for key in ["reprice", "epls"]:
            if not Path(paths[key]).exists():
                raise FileNotFoundError(f"{key} path not found: {paths[key]}")

        template_path = Path(paths["epls"])
        df = pd.read_excel(paths["reprice"], sheet_name="Claims Table")

        # Debug print and log
        print("Columns in claims DataFrame:")
        print(df.columns)
        logger.info(f"Columns in claims DataFrame: {df.columns.tolist()}")

        if "Logic" not in df.columns:
            raise KeyError("Missing 'Logic' column in Claims Table.")

        df["Logic"] = pd.to_numeric(df["Logic"], errors="coerce")
        df = df[df["Logic"].between(1, 10)]

        awp = df["Total AWP (Historical)"].sum()
        ing = df["Rx Sense Ing Cost"].sum()
        total = df["RxSense Total Cost"].sum()
        rxs = df["Rxs"].sum()

        columns_to_keep = [
            "MONY",
            "Rxs",
            "Rx Sense Ing Cost",
            "RxSense Dispense Fee",
            "RxSense Total Cost",
            "Total AWP (Historical)",
            "Pharmacy Name",
            "INPUTFILECHANNEL",
            "DATEFILLED",
            "MemberID",
            "DAYSUPPLY",
            "QUANTITY",
            "NDC",
            "DST Drug Name",
            "GrossCost",
            "Universal Rebates",
            "Exclusive Rebates",
            "Specialty Vlookup",
        ]

        missing_cols = [col for col in columns_to_keep if col not in df.columns]
        if missing_cols:
            raise KeyError(f"Missing expected columns: {missing_cols}")

        df = df[columns_to_keep]
        logger.info(f"Filtered DataFrame shape: {df.shape}")

        df["Specialty Vlookup"] = df["Specialty Vlookup"].map({"No": "N"}).fillna("Y")

        logger.info(f"AWP: {awp:.2f}, Ing: {ing:.2f}, Total: {total:.2f}, Rxs: {rxs}")

        output_path = Path("_Rx Claims for EPLS.xlsx")
        write_df_to_template(
            str(template_path),
            str(output_path),
            sheet_name="Line By Line",
            df=df,
            start_cell="A2",
            header=False,
            index=False,
            open_file=False,
            visible=False,
        )

        logger.info("EPLS LBL file created successfully.")
        write_shared_log("epls_lbl.py", "EPLS LBL file created successfully.")
        show_message(awp, ing, total, rxs)
        messagebox.showinfo("Processing Complete", "Processing complete")
    except Exception as e:
        logger.exception("An error occurred during EPLS LBL processing")
        write_shared_log("epls_lbl.py", f"An error occurred: {e}", status="ERROR")
        messagebox.showerror("Error", f"An error occurred: {e}")


if __name__ == "__main__":
    main()
"""
File processing module for handling file operations and validation.
Extracted from app.py to improve cohesion and reduce file size.
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
import csv
import logging
import os
import sys
from tkinter import messagebox

from config.app_config import ProcessingConfig, AppConstants
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.utils import write_shared_log


class FileProcessor:
    """Handles file operations and validation."""
    
    def __init__(self, app_instance):
        self.app = app_instance
        
    def check_template(self, file_path):
        """Check if template file exists and is valid."""
        return Path(file_path).exists() and Path(file_path).suffix == '.xlsx'
    
    def import_file(self, file_type="File"):
        """Import and validate a file."""
        file_path = self.app.ui_factory.create_file_dialog(
            title=f"Select {file_type}",
            filetypes=ProcessingConfig.FILE_TYPES
        )
        
        if not file_path:
            return None
            
        try:
            # Validate file exists
            if not Path(file_path).exists():
                messagebox.showerror("Error", f"{file_type} not found.")
                return None
                
            # Load and validate the file
            df = pd.read_csv(file_path) if file_path.endswith('.csv') else pd.read_excel(file_path)
            
            # Log the import
            write_shared_log("FileProcessor", f"{file_type} imported successfully: {file_path}")
            
            return file_path, df
            
        except Exception as e:
            error_msg = f"Error importing {file_type}: {str(e)}"
            messagebox.showerror("Error", error_msg)
            write_shared_log("FileProcessor", error_msg, "ERROR")
            return None
    
    def validate_file_structure(self, df, required_columns=None):
        """Validate that the file has the required structure."""
        if required_columns is None:
            required_columns = ProcessingConfig.REQUIRED_COLUMNS
            
        try:
            ProcessingConfig.validate_required_columns(df)
            return True
        except ValueError as e:
            messagebox.showerror("Validation Error", str(e))
            return False
    
    def write_audit_log(self, file1, file2, status):
        """Write an entry to the audit log."""
        entry = [datetime.now().isoformat(), str(file1), str(file2), status]
        write_header = not AppConstants.AUDIT_LOG.exists()
        
        try:
            with open(AppConstants.AUDIT_LOG, "a", newline="") as csvfile:
                writer = csv.writer(csvfile)
                if write_header:
                    writer.writerow(["Timestamp", "File1", "File2", "Status"])
                writer.writerow(entry)
        except Exception as e:
            logging.error(f"Failed to write audit log: {e}")
    
    def prepare_file_paths(self, template_path):
        """Prepare file paths for template operations."""
        if not template_path:
            raise ValueError("Template file path is not set.")
            
        template = Path(template_path)
        backup_name = template.stem + AppConstants.BACKUP_SUFFIX
        
        return {
            "template": template,
            "backup": Path.cwd() / backup_name,
            "output": Path.cwd() / AppConstants.UPDATED_TEMPLATE_NAME
        }
    
    def safe_file_operation(self, operation, *args, **kwargs):
        """Safely perform file operations with error handling."""
        try:
            return operation(*args, **kwargs)
        except Exception as e:
            error_msg = f"File operation failed: {str(e)}"
            messagebox.showerror("File Error", error_msg)
            write_shared_log("FileProcessor", error_msg, "ERROR")
            return None
"""
Log management module for handling various logging and viewer operations.
Extracted from app.py to reduce file size and improve organization.
"""

import tkinter as tk
from tkinter import scrolledtext
import csv
import os
import sys
import logging

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.utils import write_shared_log


class LogManager:
    """Handles log viewing and management operations."""
    
    def __init__(self, app_instance):
        self.app = app_instance
        self.shared_log_path = os.path.expandvars(
            r"%OneDrive%/True Community - Data Analyst/Python Repricing Automation Program/Logs/audit_log.csv"
        )
        
    def show_log_viewer(self):
        """Show the live log viewer window."""
        log_win = tk.Toplevel(self.app.root)
        log_win.title("Live Log Viewer")
        text_area = scrolledtext.ScrolledText(log_win, width=100, height=30)
        text_area.pack(fill="both", expand=True)

        def update_logs():
            try:
                with open("repricing_log.log", "r") as f:
                    text_area.delete(1.0, tk.END)
                    text_area.insert(tk.END, f.read())
            except FileNotFoundError:
                text_area.delete(1.0, tk.END)
                text_area.insert(tk.END, "No log file found.")
            except Exception as e:
                text_area.delete(1.0, tk.END)
                text_area.insert(tk.END, f"Error reading log file: {e}")
            log_win.after(3000, update_logs)

        update_logs()
        
    def show_shared_log_viewer(self):
        """Show the shared audit log viewer with search functionality."""
        log_win = tk.Toplevel(self.app.root)
        log_win.title("Shared Audit Log Viewer")
        log_win.geometry("1000x600")

        # Create filter frame
        filter_frame = tk.Frame(log_win)
        filter_frame.pack(fill="x")
        tk.Label(filter_frame, text="Search:").pack(side="left", padx=5)
        filter_entry = tk.Entry(filter_frame)
        filter_entry.pack(side="left", fill="x", expand=True, padx=5)

        # Create text area
        text_area = scrolledtext.ScrolledText(log_win, width=150, height=30)
        text_area.pack(fill="both", expand=True)

        def refresh():
            """Refresh the log display with optional filtering."""
            try:
                if not os.path.exists(self.shared_log_path):
                    text_area.delete(1.0, tk.END)
                    text_area.insert(tk.END, f"Shared log file not found at: {self.shared_log_path}")
                    return
                    
                with open(self.shared_log_path, "r", newline="", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    rows = list(reader)

                search_term = filter_entry.get().lower()
                if search_term:
                    filtered = [
                        row for row in rows
                        if any(search_term in str(cell).lower() for cell in row)
                    ]
                else:
                    filtered = rows

                text_area.delete(1.0, tk.END)
                for row in filtered:
                    text_area.insert(tk.END, " | ".join(row) + "\n")
                    
            except Exception as e:
                text_area.delete(1.0, tk.END)
                text_area.insert(tk.END, f"[ERROR] Could not read shared log:\n{e}")
                logging.error(f"Error reading shared log: {e}")

            # Auto-refresh every 5 seconds
            log_win.after(5000, refresh)

        # Bind search on Enter key
        filter_entry.bind('<Return>', lambda event: refresh())
        
        # Initial load
        refresh()
        
    def initialize_logging(self):
        """Initialize logging configuration."""
        # Clear existing log
        log_file = "repricing_log.log"
        try:
            open(log_file, "w").close()  # Clear the file
        except Exception as e:
            logging.warning(f"Could not clear log file: {e}")
            
        # Configure logging
        logging.basicConfig(
            filename=log_file,
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
            filemode='w'  # Overwrite mode
        )
        logging.info("Logging initialized")
        
    def log_application_start(self):
        """Log application startup."""
        logging.info("Repricing Automation application started")
        write_shared_log("LogManager", "Application started")
        
    def log_application_shutdown(self):
        """Log application shutdown."""
        logging.info("Repricing Automation application shutting down")
        write_shared_log("LogManager", "Application shutdown")


class ThemeController:
    """Controls theme switching functionality."""
    
    def __init__(self, app_instance):
        self.app = app_instance
        self.current_theme = "light"
        
    def toggle_dark_mode(self):
        """Toggle between light and dark themes."""
        import customtkinter as ctk
        from ui.ui_components import ThemeManager, LIGHT_COLORS, DARK_COLORS
        
        if self.current_theme == "light":
            # Switch to Dark mode
            ctk.set_appearance_mode("dark")
            ThemeManager.apply_theme_colors(self.app, DARK_COLORS)
            self.app.toggle_theme_button.configure(text="Switch to Light Mode")
            self.current_theme = "dark"
        else:
            # Switch to Light mode
            ctk.set_appearance_mode("light")
            ThemeManager.apply_theme_colors(self.app, LIGHT_COLORS)
            self.app.toggle_theme_button.configure(text="Switch to Dark Mode")
            self.current_theme = "light"
            
        logging.info(f"Theme switched to {self.current_theme} mode")
        write_shared_log("ThemeController", f"Theme changed to {self.current_theme}")
        
    def apply_initial_theme(self):
        """Apply the initial light theme."""
        from ui.ui_components import ThemeManager, LIGHT_COLORS
        ThemeManager.apply_theme_colors(self.app, LIGHT_COLORS)
import sys
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
import logging
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.utils import write_shared_log

# Configure logging
logging.basicConfig(
    filename="merge_log.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

MERGED_FILENAME = "merged_file.xlsx"
REQUIRED_COLUMNS = [
    "DATEFILLED",
    "SOURCERECORDID",
    "QUANTITY",
    "DAYSUPPLY",
    "NDC",
    "MemberID",
    "Drug Name",
    "Pharmacy Name",
    "Total AWP (Historical)",
]


def merge_files(file1_path, file2_path):
    file1 = Path(file1_path)
    file2 = Path(file2_path)
    try:
        logger.info(f"Starting merge: {file1} + {file2}")
        write_shared_log("merge.py", f"Starting merge: {file1} + {file2}")

        if not file1.exists():
            logger.error(f"File not found: {file1}")
            write_shared_log("merge.py", f"File not found: {file1}", status="ERROR")
            return False
        if not file2.exists():
            logger.error(f"File not found: {file2}")
            write_shared_log("merge.py", f"File not found: {file2}", status="ERROR")
            return False

        # Load data (support Excel or CSV for both files)
        try:
            if file1.suffix == ".csv":
                df1 = pd.read_csv(file1, parse_dates=["DATEFILLED"], dayfirst=False)
            else:
                df1 = pd.read_excel(file1, parse_dates=["DATEFILLED"])
        except Exception as e:
            logger.error(f"Failed to load file1: {e}")
            write_shared_log("merge.py", f"Failed to load file1: {e}", status="ERROR")
            return False
        try:
            if file2.suffix == ".csv":
                df2 = pd.read_csv(file2)
            else:
                df2 = pd.read_excel(file2)
        except Exception as e:
            logger.error(f"Failed to load file2: {e}")
            write_shared_log("merge.py", f"Failed to load file2: {e}", status="ERROR")
            return False

        # Log data source details
        logger.info(f"df1 shape: {df1.shape}, df2 shape: {df2.shape}")
        write_shared_log("merge.py", f"df1 shape: {df1.shape}, df2 shape: {df2.shape}")
        logger.info(f"df1 columns: {list(df1.columns)}")
        logger.info(f"df2 columns: {list(df2.columns)}")

        # Clean up and standardize column names
        df2.columns = [col.strip() for col in df2.columns]
        if "Source Record ID" in df2.columns:
            df2.rename(columns={"Source Record ID": "SOURCERECORDID"}, inplace=True)

        # Merge
        try:
            df_merged = pd.merge(df1, df2, on="SOURCERECORDID", how="outer")
        except Exception as e:
            logger.error(f"Failed to merge: {e}")
            write_shared_log("merge.py", f"Failed to merge: {e}", status="ERROR")
            return False
        if "Total AWP (Historical)" in df_merged.columns:
            df_merged["Total AWP (Historical)"] = pd.to_numeric(
                df_merged["Total AWP (Historical)"], errors="coerce"
            ).round(2)
        else:
            df_merged["Total AWP (Historical)"] = 0.00
        df_merged["MemberID"] = df_merged["MemberID"].fillna(0)

        # Log merged file row count and checksum
        row_count = len(df_merged)
        try:
            import hashlib

            sample = df_merged.head(1000).to_csv(index=False).encode("utf-8")
            checksum = hashlib.md5(sample).hexdigest()
        except Exception as e:
            checksum = f"ERROR: {e}"
        logger.info(f"Merged row count: {row_count}, sample checksum: {checksum}")
        write_shared_log(
            "merge.py", f"Merged row count: {row_count}, sample checksum: {checksum}"
        )

        # Log missing required columns
        for col in REQUIRED_COLUMNS:
            if col not in df_merged.columns:
                logger.warning(f"Missing expected column: {col}")
                write_shared_log(
                    "merge.py", f"Missing expected column: {col}", status="WARNING"
                )

        # Drop the DATEFILLED_DIFF column if it exists before saving the merged file
        if "DATEFILLED_DIFF" in df_merged.columns:
            df_merged.drop(columns=["DATEFILLED_DIFF"], inplace=True)

        merged_path = Path.cwd() / MERGED_FILENAME
        try:
            df_merged.to_excel(merged_path, index=False)
        except Exception as e:
            logger.error(f"Failed to write merged Excel: {e}")
            write_shared_log(
                "merge.py", f"Failed to write merged Excel: {e}", status="ERROR"
            )
            return False
        logger.info(f"Merged file saved to: {merged_path}")
        write_shared_log("merge.py", f"Merged file saved to: {merged_path}")

        # Apply Excel formatting
        try:
            wb = load_workbook(merged_path)
            ws = wb.active
            date_style = NamedStyle(name="date_style", number_format="MM/DD/YYYY")

            if ws is not None and ws.max_row >= 1:
                header = [
                    cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))
                ]
                if "DATEFILLED" in header:
                    date_col_index = header.index("DATEFILLED") + 1
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=date_col_index).style = date_style
                    wb.save(merged_path)
                    logger.info("Applied date formatting successfully.")
                    write_shared_log(
                        "merge.py", "Applied date formatting successfully."
                    )
                else:
                    logger.warning("DATEFILLED column not found for formatting.")
                    write_shared_log(
                        "merge.py",
                        "DATEFILLED column not found for formatting.",
                        status="WARNING",
                    )
            else:
                logger.warning(
                    "Worksheet is empty or not loaded, cannot apply formatting."
                )
                write_shared_log(
                    "merge.py",
                    "Worksheet is empty or not loaded, cannot apply formatting.",
                    status="WARNING",
                )

        except Exception as ex:
            logger.warning(f"Failed to apply formatting: {ex}")
            write_shared_log(
                "merge.py", f"Failed to apply formatting: {ex}", status="WARNING"
            )

        return True
    except Exception as e:
        logger.exception(f"Merge failed: {e}")
        write_shared_log("merge.py", f"Merge failed: {e}", status="ERROR")
        return False


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python merge.py <file1_path> <file2_path>")
        sys.exit(1)

    merge_files(sys.argv[1], sys.argv[2])
import pandas as pd
import numpy as np


def process_logic_block(df_block):
    """
    Vectorized numpy logic to mark 'OR' in 'Logic' for reversals with matching claims.
    Refactored to reduce nesting complexity and improve readability.
    """
    arr = df_block.to_numpy()
    col_idx = {col: i for i, col in enumerate(df_block.columns)}
    
    # Extract and prepare data
    logic_data = _extract_logic_data(arr, col_idx)
    
    # Early return if no reversals to process
    if not np.any(logic_data["is_reversal"]):
        return pd.DataFrame(arr, columns=df_block.columns)
    
    # Process reversals with reduced nesting
    _process_reversals(arr, col_idx, logic_data)
    
    return pd.DataFrame(arr, columns=df_block.columns)


def _extract_logic_data(arr, col_idx):
    """Extract and prepare data for logic processing."""
    qty = arr[:, col_idx["QUANTITY"]].astype(float)
    return {
        "qty": qty,
        "is_reversal": qty < 0,
        "is_claim": qty > 0,
        "ndc": arr[:, col_idx["NDC"]].astype(str),
        "member": arr[:, col_idx["MemberID"]].astype(str),
        "datefilled": pd.to_datetime(arr[:, col_idx["DATEFILLED"]], errors="coerce"),
        "abs_qty": np.abs(qty)
    }


def _process_reversals(arr, col_idx, logic_data):
    """Process reversals with matching logic, using guard clauses to reduce nesting."""
    rev_idx = np.where(logic_data["is_reversal"])[0]
    claim_idx = (
        np.where(logic_data["is_claim"])[0] 
        if np.any(logic_data["is_claim"]) 
        else np.array([], dtype=int)
    )
    
    match_context = {
        "arr": arr,
        "col_idx": col_idx,
        "logic_data": logic_data,
        "claim_idx": claim_idx
    }
    
    for i in rev_idx:
        found_match = _try_find_match(match_context, i)
        
        # Mark unmatched reversals as 'OR'
        if not found_match:
            arr[i, col_idx["Logic"]] = "OR"


def _try_find_match(match_context, reversal_idx):
    """Attempt to find a matching claim for a reversal. Returns True if match found."""
    arr = match_context["arr"]
    col_idx = match_context["col_idx"]
    logic_data = match_context["logic_data"]
    claim_idx = match_context["claim_idx"]
    
    # Guard clause: no claims to match against
    if claim_idx.size == 0:
        return False
    
    # Find potential matches
    matches = _find_matching_claims(logic_data, claim_idx, reversal_idx)
    
    # Guard clause: no matches found
    if not np.any(matches):
        return False
    
    # Mark both reversal and matching claim as 'OR'
    arr[reversal_idx, col_idx["Logic"]] = "OR"
    arr[claim_idx[matches][0], col_idx["Logic"]] = "OR"
    return True


def _find_matching_claims(logic_data, claim_idx, reversal_idx):
    """Find claims that match the reversal based on NDC, member, quantity, and date."""
    matches = (
        (logic_data["ndc"][claim_idx] == logic_data["ndc"][reversal_idx])
        & (logic_data["member"][claim_idx] == logic_data["member"][reversal_idx])
        & (logic_data["abs_qty"][claim_idx] == logic_data["abs_qty"][reversal_idx])
    )
    
    # Add date constraint (within 30 days)
    date_diffs = np.abs(
        (logic_data["datefilled"][claim_idx] - logic_data["datefilled"][reversal_idx]).days
    )
    matches &= date_diffs <= 30
    
    return matches


def worker(df_block, out_queue):
    """Worker function for multiprocessing."""
    result = process_logic_block(df_block)
    out_queue.put(result)
import pandas as pd
import logging
import os
import sys
from pathlib import Path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.utils import (
    load_file_paths,
    standardize_pharmacy_ids,
    standardize_network_ids,
    merge_with_network,
    drop_duplicates_df,
    clean_logic_and_tier,
    filter_recent_date,
    filter_logic_and_maintenance,
    filter_products_and_alternative,
    write_shared_log,
)

# Setup logging
logging.basicConfig(
    filename="openmdf_bg.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# Load NABP/NPI list
included_nabp_npi = {
    "4528874": "1477571404",
    "2365422": "1659313435",
    "3974157": "1972560688",
    "320793": "1164437406",
    "4591055": "1851463087",
    "2348046": "1942303110",
    "4023610": "1407879588",
    "4025385": "1588706212",
    "4025311": "1588705446",
    "4026806": "1285860312",
    "4931350": "1750330775",
    "4024585": "1396768461",
    "4028026": "1497022438",
    "2643749": "1326490376",
}


def process_data():
    write_shared_log("openmdf_bg.py", "Processing started.")

    import sys

    # Get the config file path relative to the project root
    config_path = Path(__file__).parent.parent / "config" / "file_paths.json"
    paths = load_file_paths(str(config_path))

    if "reprice" not in paths or not paths["reprice"]:
        logger.warning("No reprice/template file provided.")
        write_shared_log(
            "openmdf_bg.py", "No reprice/template file provided.", status="ERROR"
        )
        print("No reprice/template file provided.")
        return False

    # Check for required sheet name in reprice file
    try:
        xl = pd.ExcelFile(paths["reprice"])
        if "Claims Table" not in xl.sheet_names:
            logger.error(
                f"Sheet 'Claims Table' not found in {paths['reprice']}. Sheets: {xl.sheet_names}"
            )
            write_shared_log(
                "openmdf_bg.py",
                f"Sheet 'Claims Table' not found in {paths['reprice']}. Sheets: {xl.sheet_names}",
                status="ERROR",
            )
            return False
        claims = xl.parse(
            "Claims Table",
            usecols=[
                "SOURCERECORDID",
                "NDC",
                "MemberID",
                "DATEFILLED",
                "FormularyTier",
                "Rxs",
                "Logic",
                "PHARMACYNPI",
                "NABP",
                "Pharmacy Name",
                "Universal Rebates",
                "Exclusive Rebates",
            ],
        )
    except Exception as e:
        logger.error(f"Failed to read Claims Table: {e}")
        write_shared_log(
            "openmdf_bg.py", f"Failed to read Claims Table: {e}", status="ERROR"
        )
        return False

    # Log claim count before any filtering
    logger.info(f"Initial claims count: {claims.shape[0]}")
    write_shared_log("openmdf_bg.py", f"Initial claims count: {claims.shape[0]}")

    try:
        medi = pd.read_excel(paths["medi_span"])[["NDC", "Maint Drug?", "Product Name"]]
    except Exception as e:
        logger.error(f"Failed to read medi_span file: {paths['medi_span']} | {e}")
        write_shared_log(
            "openmdf_bg.py",
            f"Failed to read medi_span file: {paths['medi_span']} | {e}",
            status="ERROR",
        )
        return False
    try:
        mdf = pd.read_excel(paths["mdf_disrupt"], sheet_name="Open MDF NDC")[
            ["NDC", "Tier"]
        ]
    except Exception as e:
        logger.error(f"Failed to read mdf_disrupt file: {paths['mdf_disrupt']} | {e}")
        write_shared_log(
            "openmdf_bg.py",
            f"Failed to read mdf_disrupt file: {paths['mdf_disrupt']} | {e}",
            status="ERROR",
        )
        return False
    try:
        network = pd.read_excel(paths["n_disrupt"])[
            ["pharmacy_npi", "pharmacy_nabp", "pharmacy_is_excluded"]
        ]
    except Exception as e:
        logger.error(f"Failed to read n_disrupt file: {paths['n_disrupt']} | {e}")
        write_shared_log(
            "openmdf_bg.py",
            f"Failed to read n_disrupt file: {paths['n_disrupt']} | {e}",
            status="ERROR",
        )
        return False
        # Read Alternatives NDC for 'Alternative' column
    try:
        exclusive = pd.read_excel(paths["e_disrupt"], sheet_name="Alternatives NDC")[
            ["NDC", "Tier", "Alternative"]
        ]
    except Exception as e:
        logger.error(f"Failed to read e_disrupt file: {paths['e_disrupt']} | {e}")
        write_shared_log(
            "openmdf_bg.py",
            f"Failed to read e_disrupt file: {paths['e_disrupt']} | {e}",
            status="ERROR",
        )
        return False

    df = claims.merge(medi, on="NDC", how="left")
    print(f"After merge with medi: {df.shape}")
    logger.info(f"After merge with medi: {df.shape}")
    df = df.merge(mdf, on="NDC", how="left")
    print(f"After merge with mdf: {df.shape}")
    logger.info(f"After merge with mdf: {df.shape}")
    # Merge in Alternatives for 'Alternative' column
    df = df.merge(
        exclusive.rename(columns={"Tier": "Exclusive Tier"}), on="NDC", how="left"
    )
    print(f"After merge with exclusive: {df.shape}")
    logger.info(f"After merge with exclusive: {df.shape}")
    df = standardize_pharmacy_ids(df)
    print(f"After standardize_pharmacy_ids: {df.shape}")
    logger.info(f"After standardize_pharmacy_ids: {df.shape}")
    network = standardize_network_ids(network)
    print(f"After standardize_network_ids: {network.shape}")
    logger.info(f"After standardize_network_ids: {network.shape}")

    # Ensure pharmacy_id exists
    if "pharmacy_id" not in df.columns:
        df["pharmacy_id"] = df.apply(
            lambda row: row["PHARMACYNPI"]
            if pd.notna(row["PHARMACYNPI"])
            else row["NABP"],
            axis=1,
        )

    logger.info(f"Columns in df before merging: {df.columns.tolist()}")
    print(f"Columns in df before merging: {df.columns.tolist()}")

    # Log claim count after merge
    print(f"Claims after merge: {df.shape}")
    logger.info(f"Claims after merge: {df.shape}")
    write_shared_log("openmdf_bg.py", f"Claims after merge: {df.shape[0]}")

    df = merge_with_network(df, network)
    print(f"After merge_with_network: {df.shape}")
    logger.info(f"After merge_with_network: {df.shape}")
    write_shared_log("openmdf_bg.py", f"Claims after merge_with_network: {df.shape[0]}")

    df = drop_duplicates_df(df)
    print(f"After drop_duplicates_df: {df.shape}")
    logger.info(f"After drop_duplicates_df: {df.shape}")
    write_shared_log("openmdf_bg.py", f"Claims after drop_duplicates_df: {df.shape[0]}")

    df = clean_logic_and_tier(df)
    print(f"After clean_logic_and_tier: {df.shape}")
    logger.info(f"After clean_logic_and_tier: {df.shape}")
    write_shared_log(
        "openmdf_bg.py", f"Claims after clean_logic_and_tier: {df.shape[0]}"
    )

    df = filter_products_and_alternative(df)
    print(f"After filter_products_and_alternative: {df.shape}")
    logger.info(f"After filter_products_and_alternative: {df.shape}")
    write_shared_log(
        "openmdf_bg.py", f"Claims after filter_products_and_alternative: {df.shape[0]}"
    )

    df["DATEFILLED"] = pd.to_datetime(df["DATEFILLED"], errors="coerce")
    print(f"After DATEFILLED to_datetime: {df.shape}")
    df = filter_recent_date(df)
    print(f"After filter_recent_date: {df.shape}")
    logger.info(f"After filter_recent_date: {df.shape}")
    write_shared_log("openmdf_bg.py", f"Claims after filter_recent_date: {df.shape[0]}")

    df["Logic"] = pd.to_numeric(df["Logic"], errors="coerce")
    df = filter_logic_and_maintenance(df)
    print(f"After filter_logic_and_maintenance: {df.shape}")
    logger.info(f"After filter_logic_and_maintenance: {df.shape}")
    write_shared_log(
        "openmdf_bg.py", f"Claims after filter_logic_and_maintenance: {df.shape[0]}"
    )

    df = df[
        ~df["Product Name"].str.contains(
            r"albuterol|ventolin|epinephrine", case=False, regex=True
        )
    ]
    print(f"After final product exclusion: {df.shape}")
    logger.info(f"After final product exclusion: {df.shape}")
    write_shared_log(
        "openmdf_bg.py", f"Claims after final product exclusion: {df.shape[0]}"
    )

    df["FormularyTier"] = df["FormularyTier"].astype(str).str.strip().str.upper()
    total_members = df["MemberID"].nunique()
    total_claims = df["Rxs"].sum()

    uni_pos = df[(df["Tier"] == 1) & (df["FormularyTier"].isin(["B", "BRAND"]))]
    uni_neg = df[
        (df["Tier"].isin([2, 3])) & (df["FormularyTier"].isin(["G", "GENERIC"]))
    ]

    def pivot_and_count(data):
        if data.empty:
            return pd.DataFrame([[0] * len(df.columns)], columns=df.columns), 0
        return data, data["MemberID"].nunique()

    uni_pos, uni_pos_members = pivot_and_count(uni_pos)
    uni_neg, uni_neg_members = pivot_and_count(uni_neg)

    # Output filename from CLI arg or default
    import re

    output_filename = "LBL for Disruption.xlsx"
    output_path = output_filename  # Default assignment
    for i, arg in enumerate(sys.argv):
        if arg in ("--output", "-o") and i + 1 < len(sys.argv):
            output_filename = sys.argv[i + 1]
            output_path = output_filename

    # Write LBL output unconditionally (no --output-lbl flag required)
    writer = pd.ExcelWriter(output_path, engine="xlsxwriter")

    # Ensure 'pharmacy_is_excluded' column contains actual boolean values with type inference
    if "pharmacy_is_excluded" in df.columns:
        df["pharmacy_is_excluded"] = (
            df["pharmacy_is_excluded"]
            .astype(str)
            .str.lower()
            .map({"true": True, "false": False})
            .fillna(False)
            .infer_objects(copy=False)
        )
        logger.info(
            f"pharmacy_is_excluded value counts: {df['pharmacy_is_excluded'].value_counts().to_dict()}"
        )
        # Identify rows where pharmacy_is_excluded is "NA"
        na_pharmacies = df[df["pharmacy_is_excluded"].isna()]
        logger.info(f"NA pharmacies count: {na_pharmacies.shape[0]}")

        if not na_pharmacies.empty:
            output_file_path = paths["pharmacy_validation"]
            na_pharmacies_output = na_pharmacies[["PHARMACYNPI", "NABP"]].fillna("N/A")
            
            # Update existing template or create new file
            try:
                from openpyxl import load_workbook
                import os
                
                # Check if template exists
                if os.path.exists(output_file_path):
                    # Load existing workbook
                    wb = load_workbook(output_file_path)
                    logger.info(f"Loading existing template: {output_file_path}")
                else:
                    # Create new workbook
                    from openpyxl import Workbook
                    wb = Workbook()
                    logger.info(f"Creating new validation file: {output_file_path}")
                
                # Check if Validations sheet exists, if not create it
                if "Validations" in wb.sheetnames:
                    ws = wb["Validations"]
                    # Clear existing data
                    ws.delete_rows(1, ws.max_row)
                else:
                    ws = wb.create_sheet("Validations")
                    # Remove default sheet if it exists and is empty
                    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                        wb.remove(wb["Sheet"])
                
                # Write headers
                ws.append(["PHARMACYNPI", "NABP"])
                
                # Write data
                for _, row in na_pharmacies_output.iterrows():
                    ws.append([row["PHARMACYNPI"], row["NABP"]])
                
                # Save the workbook
                wb.save(output_file_path)
                logger.info(f"NA pharmacies written to '{output_file_path}' Validations sheet.")
                
            except Exception as e:
                logger.error(f"Error updating pharmacy validation file: {e}")
                # Fallback to original method
                writer = pd.ExcelWriter(output_file_path, engine="openpyxl")
                na_pharmacies_output.to_excel(
                    writer, sheet_name="Validations", index=False, engine="openpyxl"
                )
                writer.close()
                logger.info(f"NA pharmacies written to '{output_file_path}' sheet (fallback).")

    summary = pd.DataFrame(
        {
            "Formulary": ["Open MDF Positive", "Open MDF Negative"],
            "Utilizers": [uni_pos_members, uni_neg_members],
            "Rxs": [uni_pos["Rxs"].sum(), uni_neg["Rxs"].sum()],
            "% of claims": [
                uni_pos["Rxs"].sum() / total_claims,
                uni_neg["Rxs"].sum() / total_claims,
            ],
            "": ["", ""],
            "Totals": [f"Members: {total_members}", f"Claims: {total_claims}"],
        }
    )
    summary.to_excel(writer, sheet_name="Summary", index=False)

    pt_pos = pd.pivot_table(
        uni_pos,
        values=["Rxs", "MemberID"],
        index="Product Name",
        aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
    )
    pt_pos = pt_pos.rename(columns={"Rxs": "Total Rxs", "MemberID": "Unique Members"})
    pt_pos.to_excel(writer, sheet_name="OpenMDF_Positive")

    pt_neg = pd.pivot_table(
        uni_neg,
        values=["Rxs", "MemberID"],
        index="Product Name",
        aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
    )
    pt_neg = pt_neg.rename(columns={"Rxs": "Total Rxs", "MemberID": "Unique Members"})
    pt_neg.to_excel(writer, sheet_name="OpenMDF_Negative")

    writer.sheets["OpenMDF_Positive"].write("F1", f"Total Members: {uni_pos_members}")
    writer.sheets["OpenMDF_Negative"].write("F1", f"Total Members: {uni_neg_members}")

    # Network summary for excluded pharmacies (pharmacy_is_excluded=True)
    network_df = df[df["pharmacy_is_excluded"]]
    logger.debug(f"Initial network_df shape: {network_df.shape}")
    logger.debug(f"Initial network_df contents: {network_df.head(10).to_dict()}")

    filter_phrases = [
        "CVS",
        "Walgreens",
        "Kroger",
        "Walmart",
        "Rite Aid",
        "Optum",
        "Express Scripts",
        "DMR",
        "Williams Bro",
        "Publix",
    ]
    filter_phrases_escaped = [re.escape(phrase.lower()) for phrase in filter_phrases]
    regex_pattern = "|".join([f"\\b{p}\\b" for p in filter_phrases_escaped])
    network_df = network_df[
        ~network_df["Pharmacy Name"]
        .str.lower()
        .str.contains(regex_pattern, case=False, regex=True, na=False)
    ]
    logger.debug(f"Filtered network_df shape: {network_df.shape}")
    logger.debug(f"Filtered network_df contents: {network_df.head(10).to_dict()}")

    if {"PHARMACYNPI", "NABP", "Pharmacy Name"}.issubset(network_df.columns):
        network_df["PHARMACYNPI"] = network_df["PHARMACYNPI"].fillna("N/A")
        network_df["NABP"] = network_df["NABP"].fillna("NABP")
        # Create Unique Identifier based on whichever identifier is available (PHARMACYNPI or NABP)
        network_df["Unique Identifier"] = network_df.apply(
            lambda row: row["PHARMACYNPI"]
            if row["PHARMACYNPI"] != "N/A"
            else row["NABP"],
            axis=1,
        )
        network_pivot = pd.pivot_table(
            network_df,
            values=["Rxs", "MemberID"],
            index=["Unique Identifier"],
            aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
        )
        network_pivot = network_pivot.rename(
            columns={"Rxs": "Total Rxs", "MemberID": "Unique Members"}
        )
        network_pivot.reset_index(
            inplace=True
        )  # Ensure index columns are included in the output
        logger.debug(f"Network pivot shape: {network_pivot.shape}")
        logger.debug(f"Network pivot contents: {network_pivot.head(10).to_dict()}")
        network_pivot.to_excel(writer, sheet_name="Network", index=False)

    # Write the filtered network_df directly to the 'Network Sheet' with selected columns
    selected_columns = [
        "PHARMACYNPI",
        "NABP",
        "MemberID",
        "Pharmacy Name",
        "pharmacy_is_excluded",
        "Unique Identifier",
    ]
    network_df[selected_columns].to_excel(writer, sheet_name="Network", index=False)
    logger.info(
        f"Network sheet updated with {network_df.shape[0]} excluded pharmacy records (minus major chains) and selected columns"
    )

    logger.info(f"Total pharmacies in dataset: {df.shape[0]}")
    logger.info(
        f"Excluded pharmacies (pharmacy_is_excluded=True): {df['pharmacy_is_excluded'].sum()}"
    )
    logger.info(
        f"Non-excluded pharmacies (pharmacy_is_excluded=False): {(~df['pharmacy_is_excluded']).sum()}"
    )
    logger.info(
        f"Network sheet will show {network_df.shape[0]} excluded pharmacy records (minus major chains)"
    )

    # Reorder sheets so Summary follows Data
    sheets = writer.sheets  # This is a dict: {sheet_name: worksheet_object}
    names = list(sheets.keys())
    if "Data" in names and "Summary" in names:
        data_idx = names.index("Data")
        summary_idx = names.index("Summary")
        if summary_idx != data_idx + 1:
            # Move "Summary" sheet immediately after "Data"
            items = list(sheets.items())
            summary_item = items.pop(summary_idx)
            items.insert(data_idx + 1, summary_item)
            writer.sheets.clear()
            writer.sheets.update(dict(items))

    writer.close()
    logger.info(f"Open MDF BG processing completed. Output file: {output_path}")
    write_shared_log("openmdf_bg.py", "Processing complete.")
    try:
        import tkinter as tk
        from tkinter import messagebox

        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo("Processing Complete", "Processing complete")
        root.destroy()
    except Exception:
        pass
    return True


if __name__ == "__main__":
    process_data()
import pandas as pd
import logging
import os
import sys
from pathlib import Path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.utils import (
    load_file_paths,
    standardize_pharmacy_ids,
    standardize_network_ids,
    merge_with_network,
    drop_duplicates_df,
    clean_logic_and_tier,
    filter_recent_date,
    filter_logic_and_maintenance,
    filter_products_and_alternative,
    write_shared_log,
)

# Logging setup
logging.basicConfig(
    filename="openmdf_tier.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# Load NABP/NPI list
included_nabp_npi = {
    "4528874": "1477571404",
    "2365422": "1659313435",
    "3974157": "1972560688",
    "320793": "1164437406",
    "4591055": "1851463087",
    "2348046": "1942303110",
    "4023610": "1407879588",
    "4025385": "1588706212",
    "4025311": "1588705446",
    "4026806": "1285860312",
    "4931350": "1750330775",
    "4024585": "1396768461",
    "4028026": "1497022438",
    "2643749": "1326490376",
}


# ---------------------------------------------------------------------------
# Open MDF Tier processing functions
# ---------------------------------------------------------------------------
def load_openmdf_tier_data(file_paths):
    """Load all required data files for Open MDF tier disruption processing."""
    # Load claims with fallback
    if not file_paths.get("reprice"):
        write_shared_log(
            "openmdf_tier.py",
            "No reprice/template file provided.",
            status="ERROR",
        )
        print("No reprice/template file provided. Skipping claims loading.")
        return None

    try:
        claims = pd.read_excel(file_paths["reprice"], sheet_name="Claims Table")
    except Exception as e:
        logger.error(f"Error loading claims: {e}")
        claims = pd.read_excel(file_paths["reprice"], sheet_name=0)

    print(f"claims shape: {claims.shape}")
    claims.info()

    # Load reference tables
    try:
        medi = pd.read_excel(file_paths["medi_span"])[["NDC", "Maint Drug?", "Product Name"]]
        print(f"medi shape: {medi.shape}")
    except Exception as e:
        logger.error(f"Failed to read medi_span file: {file_paths['medi_span']} | {e}")
        write_shared_log(
            "openmdf_tier.py",
            f"Failed to read medi_span file: {file_paths['medi_span']} | {e}",
            status="ERROR",
        )
        return None

    try:
        mdf = pd.read_excel(file_paths["mdf_disrupt"], sheet_name="Open MDF NDC")[
            ["NDC", "Tier"]
        ]
        print(f"mdf shape: {mdf.shape}")
    except Exception as e:
        logger.error(f"Failed to read mdf_disrupt file: {file_paths['mdf_disrupt']} | {e}")
        write_shared_log(
            "openmdf_tier.py",
            f"Failed to read mdf_disrupt file: {file_paths['mdf_disrupt']} | {e}",
            status="ERROR",
        )
        return None

    try:
        exclusive = pd.read_excel(file_paths["e_disrupt"], sheet_name="Alternatives NDC")[
            ["NDC", "Tier", "Alternative"]
        ]
        print(f"exclusive shape: {exclusive.shape}")
    except Exception as e:
        logger.error(f"Failed to read e_disrupt file: {file_paths['e_disrupt']} | {e}")
        write_shared_log(
            "openmdf_tier.py",
            f"Failed to read e_disrupt file: {file_paths['e_disrupt']} | {e}",
            status="ERROR",
        )
        return None

    try:
        network = pd.read_excel(file_paths["n_disrupt"])[
            ["pharmacy_npi", "pharmacy_nabp", "pharmacy_is_excluded"]
        ]
        print(f"network shape: {network.shape}")
    except Exception as e:
        logger.error(f"Failed to read n_disrupt file: {file_paths['n_disrupt']} | {e}")
        write_shared_log(
            "openmdf_tier.py",
            f"Failed to read n_disrupt file: {file_paths['n_disrupt']} | {e}",
            status="ERROR",
        )
        return None

    return claims, medi, mdf, exclusive, network


def process_openmdf_data_pipeline(claims, reference_data, network):
    """Process the data pipeline for Open MDF tier disruption."""
    medi, mdf, exclusive = reference_data
    
    # Merge reference data
    df = claims.merge(medi, on="NDC", how="left")
    print(f"After merge with medi: {df.shape}")

    df = df.merge(mdf.rename(columns={"Tier": "Open MDF Tier"}), on="NDC", how="left")
    print(f"After merge with mdf: {df.shape}")

    df = df.merge(
        exclusive.rename(columns={"Tier": "Exclusive Tier"}), on="NDC", how="left"
    )
    print(f"After merge with exclusive: {df.shape}")

    # Standardize IDs and perform network merge
    df = standardize_pharmacy_ids(df)
    print(f"After standardize_pharmacy_ids: {df.shape}")
    network = standardize_network_ids(network)
    print(f"After standardize_network_ids: {network.shape}")
    df = merge_with_network(df, network)
    print(f"After merge_with_network: {df.shape}")

    if "pharmacy_id" not in df.columns:
        df["pharmacy_id"] = df.apply(
            lambda row: row["PHARMACYNPI"]
            if pd.notna(row["PHARMACYNPI"])
            else row["NABP"],
            axis=1,
        )

    print("Columns in df before further processing:")
    print(df.columns)

    # Date parsing, deduplication, type cleaning, and filters
    df["DATEFILLED"] = pd.to_datetime(df["DATEFILLED"], errors="coerce")
    print(f"After DATEFILLED to_datetime: {df.shape}")

    df = drop_duplicates_df(df)
    print(f"After drop_duplicates_df: {df.shape}")

    df = clean_logic_and_tier(df)
    print(f"After clean_logic_and_tier: {df.shape}")

    df = filter_recent_date(df)
    print(f"After filter_recent_date: {df.shape}")

    df = filter_logic_and_maintenance(df)
    print(f"After filter_logic_and_maintenance: {df.shape}")

    df = filter_products_and_alternative(df)
    print(f"After filter_products_and_alternative: {df.shape}")

    return df


def handle_openmdf_pharmacy_exclusions(df, file_paths):
    """Handle pharmacy exclusions for Open MDF tier disruption."""
    # Ensure 'pharmacy_is_excluded' column contains actual boolean values with type inference
    if "pharmacy_is_excluded" in df.columns:
        df["pharmacy_is_excluded"] = (
            df["pharmacy_is_excluded"]
            .astype(str)
            .str.lower()
            .map({"true": True, "false": False})
            .fillna(False)
            .infer_objects(copy=False)
        )
        logger.info(
            f"pharmacy_is_excluded value counts: {df['pharmacy_is_excluded'].value_counts().to_dict()}"
        )

        # Identify rows where pharmacy_is_excluded is "NA"
        na_pharmacies = df[df["pharmacy_is_excluded"].isna()]
        logger.info(f"NA pharmacies count: {na_pharmacies.shape[0]}")

        if not na_pharmacies.empty:
            output_file_path = file_paths["pharmacy_validation"]
            na_pharmacies_output = na_pharmacies[["PHARMACYNPI", "NABP"]].fillna("N/A")
            
            # Update existing template or create new file
            try:
                from openpyxl import load_workbook
                import os
                
                # Check if template exists
                if os.path.exists(output_file_path):
                    # Load existing workbook
                    wb = load_workbook(output_file_path)
                    logger.info(f"Loading existing template: {output_file_path}")
                else:
                    # Create new workbook
                    from openpyxl import Workbook
                    wb = Workbook()
                    logger.info(f"Creating new validation file: {output_file_path}")
                
                # Check if Validations sheet exists, if not create it
                if "Validations" in wb.sheetnames:
                    ws = wb["Validations"]
                    # Clear existing data
                    ws.delete_rows(1, ws.max_row)
                else:
                    ws = wb.create_sheet("Validations")
                    # Remove default sheet if it exists and is empty
                    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                        wb.remove(wb["Sheet"])
                
                # Write headers
                ws.append(["PHARMACYNPI", "NABP"])
                
                # Write data
                for _, row in na_pharmacies_output.iterrows():
                    ws.append([row["PHARMACYNPI"], row["NABP"]])
                
                # Save the workbook
                wb.save(output_file_path)
                logger.info(f"NA pharmacies written to '{output_file_path}' Validations sheet.")
                
            except Exception as e:
                logger.error(f"Error updating pharmacy validation file: {e}")
                # Fallback to original method
                writer = pd.ExcelWriter(output_file_path, engine="openpyxl")
                na_pharmacies_output.to_excel(
                    writer, sheet_name="Validations", index=False, engine="openpyxl"
                )
                writer.close()
                logger.info(f"NA pharmacies written to '{output_file_path}' sheet (fallback).")

    return df


def create_openmdf_tier_definitions():
    """Create the Open MDF tier definitions for analysis."""
    return [
        ("OpenMDF_Positive 2-1", "Open MDF Tier", 1, 2),
        ("OpenMDF_Positive 3-1", "Open MDF Tier", 1, 3),
        ("OpenMDF_Positive 3-2", "Open MDF Tier", 2, 3),
        ("OpenMDF_Negative 1-2", "Open MDF Tier", 2, 1),
        ("OpenMDF_Negative 1-3", "Open MDF Tier", 3, 1),
        ("OpenMDF_Negative 2-3", "Open MDF Tier", 3, 2),
    ]


def summarize_by_openmdf_tier(df, col, from_val, to_val):
    """Summarize Open MDF tier data."""
    filtered = df[(df[col] == from_val) & (df["FormularyTier"] == to_val)]
    pt = pd.pivot_table(
        filtered,
        values=["Rxs", "MemberID"],
        index=["Product Name"],
        aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
    )
    pt = pt.rename(columns={"Rxs": "Total Rxs", "MemberID": "Unique Members"})
    rxs = filtered["Rxs"].sum()
    members = filtered["MemberID"].nunique()
    return pt, rxs, members


def process_openmdf_tier_pivots(df, tiers):
    """Process Open MDF tier pivot tables and collect statistics."""
    tab_members = {}
    tab_rxs = {}
    tier_pivots = []

    for name, col, from_val, to_val in tiers:
        pt, rxs, members = summarize_by_openmdf_tier(df, col, from_val, to_val)
        tier_pivots.append((name, pt, members, rxs))
        tab_members[name] = members
        tab_rxs[name] = rxs

    return tier_pivots, tab_members, tab_rxs


def create_openmdf_summary_dataframe(tab_members, tab_rxs, total_claims, total_members):
    """Create the Open MDF summary DataFrame with calculated statistics."""
    pos_keys = [
        "OpenMDF_Positive 2-1",
        "OpenMDF_Positive 3-1",
        "OpenMDF_Positive 3-2",
    ]
    neg_keys = [
        "OpenMDF_Negative 1-2",
        "OpenMDF_Negative 1-3",
        "OpenMDF_Negative 2-3",
    ]

    pos_utilizers = sum(tab_members[k] for k in pos_keys)
    pos_claims = sum(tab_rxs[k] for k in pos_keys)
    pos_pct = pos_claims / total_claims if total_claims else 0

    neg_utilizers = sum(tab_members[k] for k in neg_keys)
    neg_claims = sum(tab_rxs[k] for k in neg_keys)
    neg_pct = neg_claims / total_claims if total_claims else 0

    return pd.DataFrame(
        {
            "Formulary": [
                "Open MDF Positive",
                "Open MDF Negative",
            ],
            "Utilizers": [
                pos_utilizers,
                neg_utilizers,
            ],
            "Rxs": [
                pos_claims,
                neg_claims,
            ],
            "% of claims": [
                pos_pct,
                neg_pct,
            ],
            "": ["", ""],
            "Totals": [
                f"Members: {total_members}",
                f"Claims: {total_claims}",
            ],
        }
    )


def create_openmdf_network_analysis(df):
    """Create network analysis for excluded pharmacies."""
    network_df = df[df["pharmacy_is_excluded"]]
    logger.debug(f"Initial network_df shape: {network_df.shape}")
    logger.debug(f"Initial network_df contents: {network_df.head(10).to_dict()}")

    import re
    filter_phrases = [
        "CVS",
        "Walgreens",
        "Kroger",
        "Walmart",
        "Rite Aid",
        "Optum",
        "Express Scripts",
        "DMR",
        "Williams Bro",
        "Publix",
    ]
    filter_phrases_escaped = [
        re.escape(phrase.lower()) for phrase in filter_phrases
    ]
    regex_pattern = "|".join([f"\\b{p}\\b" for p in filter_phrases_escaped])
    network_df = network_df[
        ~network_df["Pharmacy Name"]
        .str.lower()
        .str.contains(regex_pattern, case=False, regex=True, na=False)
    ]
    logger.debug(f"Filtered network_df shape: {network_df.shape}")
    logger.debug(f"Filtered network_df contents: {network_df.head(10).to_dict()}")

    if {"PHARMACYNPI", "NABP", "Pharmacy Name"}.issubset(network_df.columns):
        network_df["PHARMACYNPI"] = network_df["PHARMACYNPI"].fillna("N/A")
        network_df["NABP"] = network_df["NABP"].fillna("NABP")
        # Create Unique Identifier based on whichever identifier is available (PHARMACYNPI or NABP)
        network_df["Unique Identifier"] = network_df.apply(
            lambda row: row["PHARMACYNPI"]
            if row["PHARMACYNPI"] != "N/A"
            else row["NABP"],
            axis=1,
        )
        network_pivot = pd.pivot_table(
            network_df,
            values=["Rxs", "MemberID"],
            index=["Unique Identifier"],
            aggfunc={"Rxs": "sum", "MemberID": pd.Series.nunique},
        )
        network_pivot = network_pivot.rename(
            columns={"Rxs": "Total Rxs", "MemberID": "Unique Members"}
        )
        network_pivot.reset_index(
            inplace=True
        )  # Ensure index columns are included in the output
        logger.debug(f"Network pivot shape: {network_pivot.shape}")
        logger.debug(f"Network pivot contents: {network_pivot.head(10).to_dict()}")
        
        return network_df, network_pivot
    
    return network_df, None


def write_openmdf_excel_sheets(writer, df, summary_df, tier_pivots, network_df, network_pivot):
    """Write all sheets to the Excel file."""
    # Write Summary sheet
    summary_df.to_excel(writer, sheet_name="Summary", index=False)

    # Write tier pivots
    for name, pt, members, _ in tier_pivots:
        pt.to_excel(writer, sheet_name=name)
        writer.sheets[name].write("F1", f"Total Members: {members}")

    # Write Data sheet
    data_sheet_df = df.copy()
    data_sheet_df.to_excel(writer, sheet_name="Data", index=False)

    # Write Network sheet
    if network_pivot is not None:
        network_pivot.to_excel(writer, sheet_name="Network", index=False)

    # Write filtered network data
    selected_columns = [
        "PHARMACYNPI",
        "NABP",
        "MemberID",
        "Pharmacy Name",
        "pharmacy_is_excluded",
        "Unique Identifier",
    ]
    network_df[selected_columns].to_excel(writer, sheet_name="Network", index=False)
    logger.info(
        f"Network sheet updated with {network_df.shape[0]} excluded pharmacy records (minus major chains) and selected columns"
    )


def reorder_openmdf_excel_sheets(writer):
    """Reorder sheets so Summary follows Data."""
    sheets = writer.sheets  # This is a dict: {sheet_name: worksheet_object}
    names = list(sheets.keys())
    if "Data" in names and "Summary" in names:
        data_idx = names.index("Data")
        summary_idx = names.index("Summary")
        if summary_idx != data_idx + 1:
            # Move "Summary" sheet immediately after "Data"
            items = list(sheets.items())
            summary_item = items.pop(summary_idx)
            items.insert(data_idx + 1, summary_item)
            writer.sheets.clear()
            writer.sheets.update(dict(items))


def show_openmdf_completion_message(output_path):
    """Show completion message and popup."""
    write_shared_log("openmdf_tier.py", "Processing complete.")
    print(f"Processing complete. Output file: {output_path}")
    try:
        import tkinter as tk
        from tkinter import messagebox

        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo("Processing Complete", "Processing complete")
        root.destroy()
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Main processing pipeline
# ---------------------------------------------------------------------------


def process_data():
    write_shared_log("openmdf_tier.py", "Processing started.")
    # Output filename from CLI arg or default
    output_filename = "LBL for Disruption.xlsx"
    if len(sys.argv) > 1:
        output_filename = sys.argv[1]
    output_path = Path(output_filename).resolve()
    try:
        # Get the config file path relative to the project root
        config_path = Path(__file__).parent.parent / "config" / "file_paths.json"
        file_paths = load_file_paths(str(config_path))

        result = load_openmdf_tier_data(file_paths)
        if result is None:
            return  # Early exit if claims loading failed
        claims, medi, mdf, exclusive, network = result

        reference_data = (medi, mdf, exclusive)
        df = process_openmdf_data_pipeline(claims, reference_data, network)

        df = handle_openmdf_pharmacy_exclusions(df, file_paths)

        # Convert FormularyTier to numeric for proper filtering
        df["FormularyTier"] = pd.to_numeric(df["FormularyTier"], errors="coerce")

        # Totals for summary
        total_claims = df["Rxs"].sum()
        total_members = df["MemberID"].nunique()

        # Excel writer setup
        writer = pd.ExcelWriter(output_path, engine="xlsxwriter")

        # Summary calculations (must be written immediately after Data)
        tiers = create_openmdf_tier_definitions()

        tier_pivots, tab_members, tab_rxs = process_openmdf_tier_pivots(df, tiers)

        # Summary calculations
        summary_df = create_openmdf_summary_dataframe(tab_members, tab_rxs, total_claims, total_members)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # Write tier pivots after Summary
        for name, pt, members, _ in tier_pivots:
            pt.to_excel(writer, sheet_name=name)
            writer.sheets[name].write("F1", f"Total Members: {members}")

        # Write the 'Data Sheet' with excluded and non-excluded pharmacies
        data_sheet_df = df.copy()
        data_sheet_df.to_excel(writer, sheet_name="Data", index=False)

        # Network summary for excluded pharmacies (pharmacy_is_excluded=True)
        network_df, network_pivot = create_openmdf_network_analysis(df)
        logger.info(f"Total pharmacies in dataset: {df.shape[0]}")
        logger.info(
            f"Excluded pharmacies (pharmacy_is_excluded=True): {df['pharmacy_is_excluded'].sum()}"
        )
        logger.info(
            f"Non-excluded pharmacies (pharmacy_is_excluded=False): {(~df['pharmacy_is_excluded']).sum()}"
        )
        logger.info(
            f"Network sheet will show {network_df.shape[0]} excluded pharmacy records (minus major chains)"
        )

        # Write Network sheet
        if network_pivot is not None:
            network_pivot.to_excel(writer, sheet_name="Network", index=False)

        # Write filtered network data
        selected_columns = [
            "PHARMACYNPI",
            "NABP",
            "MemberID",
            "Pharmacy Name",
            "pharmacy_is_excluded",
            "Unique Identifier",
        ]
        network_df[selected_columns].to_excel(writer, sheet_name="Network", index=False)
        logger.info(
            f"Network sheet updated with {network_df.shape[0]} excluded pharmacy records (minus major chains) and selected columns"
        )

        # Reorder sheets so Summary follows Data
        reorder_openmdf_excel_sheets(writer)

        writer.close()
        show_openmdf_completion_message(output_path)
    except Exception as e:
        write_shared_log(
            "openmdf_tier.py", f"Processing failed: {e}", status="ERROR"
        )
        raise


if __name__ == "__main__":
    process_data()
    print("Data processing complete")
"""
Process management module for handling automation workflows.
Extracted from app.py to improve cohesion and reduce file size.
"""

import subprocess
import threading
import time
import logging
import os
import sys
from tkinter import messagebox

from config.app_config import DisruptionConfig
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.utils import write_shared_log


class ProcessManager:
    """Handles process management and workflow orchestration."""
    
    def __init__(self, app_instance):
        self.app = app_instance
        
    def start_process_threaded(self):
        """Start the main repricing process in a separate thread."""
        threading.Thread(target=self._start_process_internal).start()
        write_shared_log("ProcessManager", "Repricing process started")
        
    def _start_process_internal(self):
        """Internal method to handle the repricing process."""
        try:
            self.app.start_time = time.time()
            self.app.update_progress(0.05)
            
            # Extra safeguard: Remove any accidental LBL/disruption output during repricing
            os.environ["NO_LBL_OUTPUT"] = "1"
            
            # Validate inputs
            if not self.app.validate_merge_inputs():
                self.app.update_progress(0)
                return
                
            # Kill Excel processes
            self.app.update_progress(0.10)
            self._kill_excel_processes()
            
            # Run merge operation
            self.app.update_progress(0.20)
            self._run_merge_process()
            
            # Process merged file
            self.app.update_progress(0.50)
            merged_file = "merged_file.xlsx"
            self.app.process_merged_file(merged_file)
            
            # Complete
            self.app.update_progress(1.0)
            
        except Exception as e:
            self.app.update_progress(0)
            logging.error(f"Process failed: {e}")
            messagebox.showerror("Error", f"Process failed: {e}")
            
    def _kill_excel_processes(self):
        """Kill any running Excel processes."""
        try:
            subprocess.run(
                ["taskkill", "/F", "/IM", "excel.exe"],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )
        except Exception as e:
            logging.warning(f"Could not kill Excel processes: {e}")
            
    def _run_merge_process(self):
        """Run the merge.py script with file inputs."""
        try:
            subprocess.run(
                ["python", "merge.py", self.app.file1_path, self.app.file2_path], 
                check=True
            )
        except subprocess.CalledProcessError as e:
            logging.error(f"Merge process failed: {e}")
            raise
            
    def start_disruption(self, disruption_type=None):
        """Start disruption processing using configuration-driven approach."""
        if disruption_type is None:
            disruption_type = self.app.selected_disruption_type.get().strip()
        
        program_file = DisruptionConfig.get_program_file(disruption_type)
        if not program_file:
            messagebox.showerror("Error", f"Unknown disruption type: {disruption_type}")
            return
        
        self._execute_disruption_process(disruption_type, program_file)
        
    def _execute_disruption_process(self, disruption_type, program_file):
        """Execute the disruption process with error handling."""
        try:
            args = ["python", program_file]
            if self.app.template_file_path:
                args.append(str(self.app.template_file_path))
            
            # Use subprocess to run the disruption script
            subprocess.Popen(args)
            messagebox.showinfo(
                "Success",
                f"{disruption_type} disruption started in a separate process.",
            )
            
        except Exception as e:
            logging.error(f"Failed to start {program_file}: {e}")
            messagebox.showerror("Error", f"{disruption_type} disruption failed: {e}")
            
    def run_label_generation(self, label_type):
        """Run label generation scripts (SHARx or EPLS)."""
        try:
            script_name = f"{label_type.lower()}_lbl.py"
            subprocess.run(["python", script_name], check=True)
            write_shared_log("ProcessManager", f"{label_type} LBL generation completed")
            
        except subprocess.CalledProcessError as e:
            logging.error(f"{label_type} LBL generation failed: {e}")
            messagebox.showerror("Error", f"{label_type} LBL generation failed: {e}")
            
    def cancel_process(self):
        """Cancel the current process."""
        logging.info("Process cancellation requested")
        write_shared_log("ProcessManager", "Process cancelled")
        messagebox.showinfo("Cancelled", "Process cancellation requested.")
        
    def finish_notification(self):
        """Show completion notification."""
        try:
            from plyer import notification
            if hasattr(notification, "notify") and callable(notification.notify):
                notification.notify(
                    title="Repricing Automation",
                    message="Batch processing completed.",
                    timeout=5,
                )
        except ImportError:
            pass  # Notification not available
            
        write_shared_log("ProcessManager", "Batch processing completed")
        messagebox.showinfo("Completed", "Batch processing finished!")
import cProfile
import pstats
import sys

if __name__ == "__main__":
    script_path = (
        sys.argv[1] if len(sys.argv) > 1 else "app.py"
    )  # Allow dynamic script path input
    profile_output = "profile_stats.prof"

    print(f"Profiling {script_path}...\n")
    cProfile.runctx(
        "exec(compile(open(script_path).read(), script_path, 'exec'))",
        globals(),
        locals(),
        profile_output,
    )

    # Optional: print top 20 cumulative time functions
    stats = pstats.Stats(profile_output)
    stats.sort_stats(pstats.SortKey.CUMULATIVE).print_stats(20)
import pandas as pd
import json
import logging
import os
import csv
from pathlib import Path
import getpass
from datetime import datetime
from dataclasses import dataclass

shared_log_path = os.path.expandvars(
    r"%OneDrive%/True Community - Data Analyst/Python Repricing Automation Program/Logs/audit_log.csv"
)


@dataclass
class LogicMaintenanceConfig:
    """Configuration for logic and maintenance filtering."""
    logic_col: str = "Logic"
    min_logic: int = 5
    max_logic: int = 10
    maint_col: str = "Maint Drug?"


def ensure_directory_exists(path):
    """
    Ensures the directory for the given path exists.
    """
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
    except Exception as e:
        print(f"[ensure_directory_exists] Error: {e}")


def write_shared_log(script_name, message, status="INFO"):
    """
    Appends a log entry to the shared audit log in OneDrive. Rotates log if too large.
    """
    try:
        username = getpass.getuser()
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry = [timestamp, username, script_name, message, status]

        write_header = not os.path.exists(shared_log_path)
        ensure_directory_exists(shared_log_path)

        # Log rotation: if file > 5MB, rotate (keep 3 backups)
        max_size = 5 * 1024 * 1024
        if (
            os.path.exists(shared_log_path)
            and os.path.getsize(shared_log_path) > max_size
        ):
            for i in range(2, 0, -1):
                prev = f"{shared_log_path}.{i}"
                prev2 = f"{shared_log_path}.{i + 1}"
                if os.path.exists(prev):
                    os.replace(prev, prev2)
            os.replace(shared_log_path, f"{shared_log_path}.1")

        with open(shared_log_path, mode="a", newline="", encoding="utf-8") as file:
            writer = csv.writer(file)
            if write_header:
                writer.writerow(["Timestamp", "User", "Script", "Message", "Status"])
            writer.writerow(log_entry)
    except Exception as e:
        print(f"[Shared Log] Error: {e}")


def log_exception(script_name, exc, status="ERROR"):
    """
    Standardized exception logging to shared log and console.
    """
    import traceback

    tb = traceback.format_exc()
    msg = f"{exc}: {tb}"
    print(f"[Exception] {msg}")
    write_shared_log(script_name, msg, status)


def load_file_paths(json_file="file_paths.json"):
    """
    Loads a JSON config file, replacing %OneDrive% with the user's OneDrive path.
    Returns a dictionary mapping keys to resolved absolute file paths.
    """
    try:
        with open(json_file, "r") as f:
            paths = json.load(f)

        # Resolve the user's OneDrive path
        onedrive_path = os.environ.get("OneDrive")
        if not onedrive_path:
            raise EnvironmentError(
                "OneDrive environment variable not found. Please ensure OneDrive is set up."
            )

        resolved_paths = {}
        for key, path in paths.items():
            if path.startswith("%OneDrive%"):
                path = path.replace("%OneDrive%", onedrive_path)
            resolved_paths[key] = str(Path(path).resolve())

        return resolved_paths

    except Exception:
        logging.exception(f"Failed to load or resolve file paths from {json_file}")
        raise


def standardize_pharmacy_ids(df):
    """
    Pads 'PHARMACYNPI' to 10 digits and 'NABP' to 7 digits in the DataFrame.

    Args:
        df (pd.DataFrame): Claims DataFrame.

    Returns:
        pd.DataFrame: Updated DataFrame with padded ID columns.
    """
    if "PHARMACYNPI" in df.columns:
        df["PHARMACYNPI"] = df["PHARMACYNPI"].astype(str).str.zfill(10)
    if "NABP" in df.columns:
        df["NABP"] = df["NABP"].astype(str).str.zfill(7)
    return df


def standardize_network_ids(network):
    """
    Pads 'pharmacy_npi' to 10 digits and 'pharmacy_nabp' to 7 digits in the network DataFrame.

    Args:
        network (pd.DataFrame): Network DataFrame.

    Returns:
        pd.DataFrame: Updated network DataFrame with padded ID columns.
    """
    if "pharmacy_npi" in network.columns:
        network["pharmacy_npi"] = network["pharmacy_npi"].astype(str).str.zfill(10)
    if "pharmacy_nabp" in network.columns:
        network["pharmacy_nabp"] = network["pharmacy_nabp"].astype(str).str.zfill(7)
    return network


def merge_with_network(df, network):
    """
    Performs a left join of df with network on ['PHARMACYNPI','NABP'] ⟷ ['pharmacy_npi','pharmacy_nabp'].

    Args:
        df (pd.DataFrame): Claims DataFrame.
        network (pd.DataFrame): Network DataFrame.

    Returns:
        pd.DataFrame: Merged DataFrame.
    """
    return df.merge(
        network,
        left_on=["PHARMACYNPI", "NABP"],
        right_on=["pharmacy_npi", "pharmacy_nabp"],
        how="left",
    )


def drop_duplicates_df(df):
    """
    Drops duplicate rows from the DataFrame.

    Args:
        df (pd.DataFrame): DataFrame to deduplicate.

    Returns:
        pd.DataFrame: Deduplicated DataFrame.
    """
    df = df.drop_duplicates()
    return df.drop_duplicates()


def clean_logic_and_tier(df, logic_col="Logic", tier_col="FormularyTier"):
    """
    Cleans 'Logic' as numeric.
    Cleans 'FormularyTier':
        - If all entries are numeric-like, coerces to numeric
        - Otherwise, strips and uppercases text for brand/generic disruptions
    """
    df[logic_col] = pd.to_numeric(df[logic_col], errors="coerce")

    # Inspect tier values
    sample = df[tier_col].dropna().astype(str).head(10)
    numeric_like = sample.str.replace(".", "", regex=False).str.isnumeric().all()

    if numeric_like:
        df[tier_col] = pd.to_numeric(df[tier_col], errors="coerce")
    else:
        df[tier_col] = df[tier_col].astype(str).str.strip().str.upper()

    return df


def filter_recent_date(df, date_col="DATEFILLED"):
    """
    Keeps only rows where date_col falls in the last 6 months (inclusive).

    Args:
        df (pd.DataFrame): DataFrame with date column.
        date_col (str): Name of the date column.

    Returns:
        pd.DataFrame: Filtered DataFrame.
    """
    latest = df[date_col].max()
    start = latest - pd.DateOffset(months=6) + pd.DateOffset(days=1)
    return df[(df[date_col] >= start) & (df[date_col] <= latest)]


def filter_logic_and_maintenance(df, config=None):
    """
    Filters rows where min_logic ≤ Logic ≤ max_logic and 'Maint Drug?' == 'Y'.

    Args:
        df (pd.DataFrame): DataFrame with logic and maintenance columns.
        config (LogicMaintenanceConfig, optional): Configuration object with filtering parameters.
                                                 If None, uses default configuration.

    Returns:
        pd.DataFrame: Filtered DataFrame.
    """
    if config is None:
        config = LogicMaintenanceConfig()
    
    return df[
        (df[config.logic_col] >= config.min_logic)
        & (df[config.logic_col] <= config.max_logic)
        & (df[config.maint_col] == "Y")
    ]


def filter_products_and_alternative(
    df, product_col="Product Name", alternative_col="Alternative"
):
    """
    Excludes rows where 'Product Name' contains albuterol, ventolin, epinephrine,
    or where 'Alternative' contains 'Covered' or 'Use different NDC'.

    Args:
        df (pd.DataFrame): DataFrame with product/alternative columns.
        product_col (str): Name of the product column.
        alternative_col (str): Name of the alternative column.

    Returns:
        pd.DataFrame: Filtered DataFrame.
    """
    exclude_pats = [r"\balbuterol\b", r"\bventolin\b", r"\bepinephrine\b"]
    for pat in exclude_pats:
        df = df[~df[product_col].str.contains(pat, case=False, na=False)]
    df = df[
        ~df[alternative_col]
        .astype(str)
        .str.contains(r"Covered|Use different NDC", case=False, regex=True, na=False)
    ]
    return df
